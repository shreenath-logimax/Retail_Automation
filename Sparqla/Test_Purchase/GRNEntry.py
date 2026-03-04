from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import UnexpectedAlertPresentException, TimeoutException
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
import os
import re
import unittest

FILE_PATH = ExcelUtils.file_path


class GRNEntry(unittest.TestCase):
    """
    GRN Entry Module Automation - Create, Edit, Cancel GRN Entries
    Follows Sparqla framework rules: Function_Call only, ExcelUtils only, No raw Selenium

    Business Context:
    - GRN Entry records goods received from karigar/supplier
    - GRN Type: 1=Bill, 2=Receipt, 3=Charges
    - Item table only for Bill or Receipt (NOT for Charges type)
    - grn_bill_status=2 on cancellation
    - Stock and Ledger are updated at GRN stage
    """

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)

    def test_grn_entry(self):
        """Main entry point for GRN Entry automation"""
        driver = self.driver
        wait = self.wait

        # Navigate to GRN Entry List
        try:
            wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
            Function_Call.click(self, "//span[contains(text(), 'Purchase Module')]")
            Function_Call.click(self, "//span[contains(text(), 'GRN Entry')]")
            sleep(2)
            print("✅ Navigated to GRN Entry list page")
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            return

        # Read Excel data
        sheet_name = "GRNEntry"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return

        workbook = load_workbook(FILE_PATH)
        sheet = workbook[sheet_name]

        for row_num in range(2, valid_rows):
            # Column mapping
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,
                "GRNType": 4, "Karigar": 5, "RefNo": 6,
                "PurchaseType": 7, "RefDate": 8, "EWayBillNo": 9,
                "IRNNo": 10, "DispatchThrough": 11, "Category": 12,
                "Pcs": 13, "GrossWt": 14, "NetWt": 15,
                "Wastage": 16, "RatePerGram": 17, "RateType": 18,
                "TDSPercent": 19, "TCSPercent": 20, "GRNNo": 21,
                "ExpectedStatus": 22, "CancelReason": 23, "Remark": 24
            }

            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}

            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']}")
            print(f"{'='*80}")

            # Dispatch to correct scenario based on TestCaseId keyword
            tc_id = str(row_data["TestCaseId"]).upper()

            try:
                if "CANCEL" in tc_id:
                    print("🔄 Executing Scenario: Cancel Flow")
                    result = self.test_grn_cancel_flow(row_data, row_num, sheet_name)
                elif "EDIT" in tc_id:
                    print("🔄 Executing Scenario: Edit/Update Flow")
                    result = self.test_grn_edit_update(row_data, row_num, sheet_name)
                elif "MULTIPLE" in tc_id:
                    print("🔄 Executing Scenario: Multiple Items")
                    result = self.test_multiple_item_grn(row_data, row_num, sheet_name)
                else:
                    print("🔄 Executing Scenario: Standard GRN Creation")
                    result = self.test_GRN_creation_save(row_data, row_num, sheet_name)

                print(f"🏁 Test Result: {result[0]} - {result[1]}")
                grn_no = result[2] if (len(result) > 2 and result[2]) else ""
                self._update_excel_status(row_num, result[0], result[1], sheet_name, grn_no)

                # If save successful, verify in list page
                if result[0] == "Pass" and "CANCEL" not in tc_id and "EDIT" not in tc_id:
                    print(f"🔍 Verifying GRN {grn_no} in List Page...")
                    list_result = self.test_GRN_list_verification(grn_no, row_data)
                    print(f"📊 List Page Verification: {list_result[0]} - {list_result[1]}")

            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed with exception: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
                self._take_screenshot(f"Exception_TC{row_data['TestCaseId']}")

        workbook.close()
        print(f"\n{'='*80}")
        print(f"✅ GRN Entry Automation Completed")
        print(f"{'='*80}")

    # ─────────────────────────────────────────────────────────────────
    # STEP 1: STANDARD GRN CREATION (SAVE)
    # ─────────────────────────────────────────────────────────────────
    def test_GRN_creation_save(self, row_data, row_num, sheet_name):
        """
        Test Method: Create GRN Entry
        Validates: GRN creation, grn_ref_no generation
        Business: GRN Type Bill/Receipt requires item table; Charges does not
        """
        driver = self.driver
        wait = self.wait
        current_field = "Initial Setup"

        try:
            # Handle any pending alerts
            Function_Call.alert(self)

            # Click Add button
            current_field = "Add Button"
            Function_Call.click(self, '//a[@id="add_Order"]')
            sleep(3)

            # ── Step 1: Select GRN Type (radio) ──
            current_field = f"GRN Type ({row_data['GRNType']})"
            grn_type_map = {
                "Bill":     '//input[@id="oranment_type"]',
                "Receipt":  '//input[@id="mt_type"]',
                "Charges":  '//input[@id="st_type"]'
            }
            grn_type = str(row_data["GRNType"]).strip() if row_data["GRNType"] else "Bill"
            xpath = grn_type_map.get(grn_type, '//input[@id="oranment_type"]')
            Function_Call.click(self, xpath)
            sleep(1)
            print(f"✅ Selected GRN Type: {grn_type}")

            # ── Step 2: Select Karigar (MANDATORY) ──
            if row_data["Karigar"]:
                current_field = f"Karigar ({row_data['Karigar']})"
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="select_karigar"]/following-sibling::span',
                    row_data["Karigar"],
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)
                print(f"✅ Selected Karigar: {row_data['Karigar']}")
            else:
                msg = "Karigar is mandatory ⚠️"
                self._take_screenshot(f"MissingField_Karigar_TC{row_data['TestCaseId']}")
                Function_Call.Remark(self, row_num, msg, sheet_name)
                self._cancel_form()
                return ("Fail", msg)

            # ── Step 3: Enter Ref No (MANDATORY) ──
            if row_data["RefNo"]:
                current_field = f"Ref No ({row_data['RefNo']})"
                Function_Call.fill_input(
                    self, wait,
                    locator=(By.XPATH, '//input[@name="order[po_supplier_ref_no]"]'),
                    value=str(row_data["RefNo"]),
                    field_name="RefNo",
                    row_num=row_num,
                    pattern=r"^.+$",
                    Sheet_name=sheet_name
                )
                print(f"✅ Entered Ref No: {row_data['RefNo']}")
            else:
                msg = "Ref No is mandatory ⚠️"
                self._take_screenshot(f"MissingField_RefNo_TC{row_data['TestCaseId']}")
                Function_Call.Remark(self, row_num, msg, sheet_name)
                self._cancel_form()
                return ("Fail", msg)

            # ── Step 4: Select Purchase Type (MANDATORY) ──
            if row_data["PurchaseType"]:
                current_field = f"Purchase Type ({row_data['PurchaseType']})"
                try:
                    Select(driver.find_element(By.ID, "purchase_type")).select_by_visible_text(
                        str(row_data["PurchaseType"])
                    )
                    print(f"✅ Selected Purchase Type: {row_data['PurchaseType']}")
                except Exception as e:
                    print(f"⚠️ Purchase Type selection failed: {e}")
            else:
                msg = "Purchase Type is mandatory ⚠️"
                self._take_screenshot(f"MissingField_PurchaseType_TC{row_data['TestCaseId']}")
                Function_Call.Remark(self, row_num, msg, sheet_name)
                self._cancel_form()
                return ("Fail", msg)

            # ── Step 5: Enter Ref Date (MANDATORY, datepicker) ──
            if row_data["RefDate"]:
                current_field = f"Ref Date ({row_data['RefDate']})"
                Data = Function_Call.fill_input(
                    self, wait,
                    locator=(By.XPATH, '//input[@name="order[po_ref_date]"]'),
                    value=str(row_data["RefDate"]),
                    field_name="RefDate",
                    row_num=row_num,
                    pattern=r"^\d{2}[-/]\d{2}[-/]\d{4}$",
                    screenshot_prefix="RefDate",
                    extra_keys=Keys.TAB,
                    Sheet_name=sheet_name,
                    Date_range='past_or_current'
                )
                print(Data)
                if "Fail" in str(Data):
                    self._cancel_form()
                    return Data
                print(f"✅ Entered Ref Date: {row_data['RefDate']}")
            else:
                msg = "Ref Date is mandatory ⚠️"
                self._take_screenshot(f"MissingField_RefDate_TC{row_data['TestCaseId']}")
                Function_Call.Remark(self, row_num, msg, sheet_name)
                self._cancel_form()
                return ("Fail", msg)

            # ── Step 6: E-Way Bill No (OPTIONAL) ──
            if row_data["EWayBillNo"]:
                current_field = f"E-Way Bill No ({row_data['EWayBillNo']})"
                Function_Call.fill_input(
                    self, wait,
                    locator=(By.ID, "ewaybillno"),
                    value=str(row_data["EWayBillNo"]),
                    field_name="EWayBillNo",
                    row_num=row_num,
                    pattern=r"^.+$",
                    Sheet_name=sheet_name
                )
                print(f"✅ Entered E-Way Bill No: {row_data['EWayBillNo']}")

            # ── Step 7: IRN No (OPTIONAL) ──
            if row_data["IRNNo"]:
                current_field = f"IRN No ({row_data['IRNNo']})"
                Function_Call.fill_input(
                    self, wait,
                    locator=(By.ID, "invoice_ref_no"),
                    value=str(row_data["IRNNo"]),
                    field_name="IRNNo",
                    row_num=row_num,
                    pattern=r"^.+$",
                    Sheet_name=sheet_name
                )
                print(f"✅ Entered IRN No: {row_data['IRNNo']}")

            # ── Step 8: Select Dispatch Through (MANDATORY) ──
            if row_data["DispatchThrough"]:
                current_field = f"Dispatch Through ({row_data['DispatchThrough']})"
                try:
                    Select(driver.find_element(By.ID, "despatch_through")).select_by_visible_text(
                        str(row_data["DispatchThrough"])
                    )
                    print(f"✅ Selected Dispatch Through: {row_data['DispatchThrough']}")
                except Exception as e:
                    print(f"⚠️ Dispatch Through selection failed: {e}")
            else:
                msg = "Dispatch Through is mandatory ⚠️"
                self._take_screenshot(f"MissingField_DispatchThrough_TC{row_data['TestCaseId']}")
                Function_Call.Remark(self, row_num, msg, sheet_name)
                self._cancel_form()
                return ("Fail", msg)

            # ── Step 9: Item Table (only for Bill or Receipt, NOT Charges) ──
            if grn_type in ["Bill", "Receipt"]:
                result = self._fill_item_row(row_data, row_num, sheet_name, grn_type)
                if result and result[0] == "Fail":
                    return result

            # ── Step 10: TDS % (OPTIONAL) ──
            if row_data["TDSPercent"]:
                current_field = f"TDS% ({row_data['TDSPercent']})"
                Function_Call.fill_input(
                    self, wait,
                    locator=(By.ID, "tds_percent"),
                    value=str(row_data["TDSPercent"]),
                    field_name="TDSPercent",
                    row_num=row_num,
                    pattern=r"^\d+(\.\d+)?$",
                    Sheet_name=sheet_name
                )
                sleep(1)
                print(f"✅ Entered TDS%: {row_data['TDSPercent']}")

            # ── Step 11: TCS % (OPTIONAL) ──
            if row_data["TCSPercent"]:
                current_field = f"TCS% ({row_data['TCSPercent']})"
                Function_Call.fill_input(
                    self, wait,
                    locator=(By.ID, "tcs_percent"),
                    value=str(row_data["TCSPercent"]),
                    field_name="TCSPercent",
                    row_num=row_num,
                    pattern=r"^\d+(\.\d+)?$",
                    Sheet_name=sheet_name
                )
                sleep(1)
                print(f"✅ Entered TCS%: {row_data['TCSPercent']}")

            # ── Step 12: Take screenshot before save ──
            self._take_screenshot(f"BeforeSave_TC{row_data['TestCaseId']}")

            # ── Step 13: Click Save button ──
            current_field = "Save Button"
            Function_Call.click(self, '//button[@id="submit_grn_entry"]')
            sleep(3)

            # ── Step 14: Capture Success/Failure ──
            return self._capture_save_result(row_data, "GRN Entry Added Successfully")

        except Exception as e:
            msg = f"❌ Test execution error in {current_field}: {e}"
            print(msg)
            self._take_screenshot(f"ExecutionError_TC{row_data['TestCaseId']}")
            self._cancel_form()
            return ("Fail", msg)

    # ─────────────────────────────────────────────────────────────────
    # HELPER: Fill single item row in grn_item_details table
    # ─────────────────────────────────────────────────────────────────
    def _fill_item_row(self, row_data, row_num, sheet_name,GRNType, item_data=None):
        """
        Clicks 'Add Item', then fills one item row in grn_item_details table.
        Use item_data dict for multiple-item scenario; defaults to row_data for standard.
        """
        driver = self.driver
        wait = self.wait
        data = item_data if item_data else row_data
        current_field = "Add Item Button"

        try:
            # Click Add Item button
            Function_Call.click(self, '//button[@id="add_item_details"]')
            sleep(2)
            print("✅ Clicked Add Item — new row added to table")

            # Get current number of rows to determine index
            rows = driver.find_elements(By.XPATH, '//table[@id="grn_item_details"]/tbody/tr')
            if not rows:
                msg = "Item table row did not appear after clicking Add Item ⚠️"
                self._take_screenshot(f"ItemRowMissing_TC{row_data['TestCaseId']}")
                return ("Fail", msg)
            
            row_index = len(rows)
            last_row = rows[-1]

            # ── Category (MANDATORY) ──
            if data["Category"]:
                current_field = f"Category ({data['Category']})"
                try:
                    # Use dropdown_select with dynamic indexing for Select2
                    Function_Call.dropdown_select(
                        self,
                        f'//span[@title="Select Category"]',
                        str(data["Category"]),
                        '//span[@class="select2-search select2-search--dropdown"]/input'
                    )
                    sleep(1)
                    print(f"✅ Selected Category items {row_index}: {data['Category']}")
                except Exception as e:
                    msg = f"Category selection failed at index {row_index}: {e}"
                    self._take_screenshot(f"MissingField_Category_TC{row_data['TestCaseId']}")
                    Function_Call.Remark(self, row_num, msg, sheet_name)
                    self._cancel_form()
                    return ("Fail", msg)
            else:
                msg = "Category is mandatory for item row ⚠️"
                self._take_screenshot(f"MissingField_Category_TC{row_data['TestCaseId']}")
                Function_Call.Remark(self, row_num, msg, sheet_name)
                self._cancel_form()
                return ("Fail", msg)

            # ── Pcs (MANDATORY) ──
            if data["Pcs"]:
                current_field = f"Pcs ({data['Pcs']})"
                Data = Function_Call.fill_input(
                    self, wait,
                    locator=(By.XPATH, f'(//input[@name="item[pcs][]"])[{row_index}]'),
                    value=str(data["Pcs"]),
                    field_name="Pcs",
                    row_num=row_num,
                    pattern=r"^[0-9]+$",
                    screenshot_prefix="Pcs",
                    Sheet_name=sheet_name
                )
                print(Data)
                if "Fail" in str(Data):
                    self._cancel_form()
                    return Data
                print(f"✅ Entered Pcs items {row_index}: {data['Pcs']}")
            else:
                msg = "Pcs is mandatory ⚠️"
                self._take_screenshot(f"MissingField_Pcs_TC{row_data['TestCaseId']}")
                Function_Call.Remark(self, row_num, msg, sheet_name)
                self._cancel_form()
                return ("Fail", msg)

            # ── Gross Wt (MANDATORY) ──
            if data["GrossWt"]:
                current_field = f"Gross Wt ({data['GrossWt']})"
                Data = Function_Call.fill_input(
                    self, wait,
                    locator=(By.XPATH, f'(//input[@name="item[gross_wt][]"])[{row_index}]'),
                    value=str(data["GrossWt"]),
                    field_name="GrossWt",
                    row_num=row_num,
                    pattern=r"^\d+(\.\d{1,3})?$",
                    screenshot_prefix="GrossWt",
                    Sheet_name=sheet_name
                )
                print(Data)
                if "Fail" in str(Data):
                    self._cancel_form()
                    return Data
                print(f"✅ Entered Gross Wt items {row_index}: {data['GrossWt']} (Net Wt auto-calculated)")
            else:
                msg = "Gross Wt is mandatory ⚠️"
                self._take_screenshot(f"MissingField_GrossWt_TC{row_data['TestCaseId']}")
                Function_Call.Remark(self, row_num, msg, sheet_name)
                self._cancel_form()
                return ("Fail", msg)

            # ── Wastage VA (OPTIONAL) ──
            Wastage = False
            try:
               if data['GRNType']=='Receipt':
                   Wastage=True
            except:
               if GRNType =="Receipt":
                   Wastage=True
            if Wastage:    
                if data["Wastage"]:
                    current_field = f"Wastage ({data['Wastage']})"
                    Data = Function_Call.fill_input(
                        self, wait,
                        locator=(By.XPATH, f'(//input[@name="item[wastage][]"])[{row_index}]'),
                        value=str(data["Wastage"]),
                        field_name="Wastage",
                        row_num=row_num,
                        pattern=r"^\d+(\.\d{1,3})?$",
                        screenshot_prefix="Wastage",
                        Sheet_name=sheet_name
                    )
                    if "Fail" in str(Data):
                        self._cancel_form()
                        return Data
                    print(f"✅ Entered Wastage items {row_index}: {data['Wastage']}")

            # ── Rate Per Gram (MANDATORY) ──
            if data["RatePerGram"]:
                current_field = f"Rate Per Gram ({data['RatePerGram']})"
                Data = Function_Call.fill_input(
                    self, wait,
                    locator=(By.XPATH, f'(//input[@name="item[rate_per_gram][]"])[{row_index}]'),
                    value=str(data["RatePerGram"]),
                    field_name="RatePerGram",
                    row_num=row_num,
                    pattern=r"^\d+(\.\d{1,2})?$",
                    screenshot_prefix="RatePerGram",
                    Sheet_name=sheet_name
                )
                print(Data)
                if "Fail" in str(Data):
                    self._cancel_form()
                    return Data
                print(f"✅ Entered Rate Per Gram items {row_index}: {data['RatePerGram']}")
            else:
                msg = "Rate Per Gram is mandatory ⚠️"
                self._take_screenshot(f"MissingField_RatePerGram_TC{row_data['TestCaseId']}")
                Function_Call.Remark(self, row_num, msg, sheet_name)
                self._cancel_form()
                return ("Fail", msg)

            # ── Rate Type (MANDATORY: Grm=1, Pcs=2) ──
            if data["RateType"]:
                current_field = f"Rate Type ({data['RateType']})"
                try:
                    rate_type_map = {"Grm": "1", "Pcs": "2"}
                    rate_val = rate_type_map.get(str(data["RateType"]), "1")
                    rate_xpath = f'(//select[@class="ratecaltype form-control"])[{row_index}]'
                    rate_select = wait.until(EC.presence_of_element_located((By.XPATH, rate_xpath)))
                    Select(rate_select).select_by_value(rate_val)
                    sleep(0.5)
                    print(f"✅ Selected Rate Type items {row_index}: {data['RateType']}")
                except Exception as e:
                    print(f"⚠️ Rate Type selection failed at index {row_index}: {e}")

            sleep(1)
            return None  # None means success — caller continues

        except Exception as e:
            msg = f"❌ Item row fill error in {current_field}: {e}"
            print(msg)
            self._take_screenshot(f"ItemRowError_TC{row_data['TestCaseId']}")
            self._cancel_form()
            return ("Fail", msg)

    # ─────────────────────────────────────────────────────────────────
    # STEP 2: LIST PAGE VERIFICATION
    # ─────────────────────────────────────────────────────────────────
    def test_GRN_list_verification(self, grn_no, row_data):
        """
        Verifies the saved GRN record on the list page.
        Uses: date range filter (today), DataTable search, row check.
        """
        driver = self.driver
        wait = self.wait

        try:
            # Close any success alert banners
            try:
                close_btns = driver.find_elements(By.XPATH, '//button[@class="close" and @data-dismiss="alert"]')
                for btn in close_btns:
                    if btn.is_displayed():
                        btn.click()
                        sleep(0.3)
            except:
                pass

            # ── Apply Date Range: Today ──
            print("📅 Applying date range filter (Today)...")
            try:
                Function_Call.click(self, '//button[@id="grn-dt-btn"]')
                sleep(2)
                for loc in [
                    '//div[contains(@class,"daterangepicker")]//li[text()="Today"]',
                    '//div[@class="ranges"]//li[text()="Today"]',
                    '//li[contains(text(),"Today")]'
                ]:
                    if driver.find_elements(By.XPATH, loc):
                        Function_Call.click(self, loc)
                        sleep(2)
                        break
            except Exception as e:
                print(f"⚠️ Date range filter failed: {e}")

            # ── Search GRN No in DataTable search box ──
            print(f"🔍 Searching for GRN No: {grn_no}")
            try:
                search_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, '//div[@id="grn_list_filter"]//input'))
                )
                search_input.clear()
                search_input.send_keys(grn_no)
                sleep(2)

                # Verify row appears
                grn_row_xpath = f'//table[@id="grn_list"]//tr[contains(., "{grn_no}")]'
                if driver.find_elements(By.XPATH, grn_row_xpath):
                    print(f"✅ GRN {grn_no} found in list table")
                    # Also confirm supplier and ref no are visible in the row
                    row_el = driver.find_element(By.XPATH, grn_row_xpath)
                    row_text = row_el.text
                    print(f"   Row data: {row_text[:100]}")
                    self._take_screenshot(f"ListVerify_GRN_{grn_no}")
                    return ("Pass", f"✅ GRN {grn_no} verified in list", grn_no)
                else:
                    print(f"⚠️ GRN {grn_no} NOT found in list table")
                    self._take_screenshot(f"ListVerifyFail_GRN_{grn_no}")
                    return ("Fail", f"❌ GRN {grn_no} not found in list after search", grn_no)

            except Exception as e:
                print(f"⚠️ DataTable search failed: {e}")
                self._take_screenshot(f"ListSearchError_GRN_{grn_no}")
                return ("Pass", f"⚠️ List search failed but GRN creation passed: {grn_no}", grn_no)

        except Exception as e:
            print(f"❌ List verification error: {e}")
            self._take_screenshot(f"ListVerifError_{grn_no}")
            return ("Fail", f"List verification error: {str(e)}", grn_no)

    # ─────────────────────────────────────────────────────────────────
    # STEP 3: CANCEL FLOW
    # ─────────────────────────────────────────────────────────────────
    def test_grn_cancel_flow(self, row_data, row_num, sheet_name):
        """
        Cancel Flow: First creates a GRN, then cancels it from the list page.
        Cancel modal: id=confirm-delete, remark: id=cancel_remark, btn: id=grn_cancel
        """
        driver = self.driver
        wait = self.wait
        current_field = "GRN Creation for Cancel"

        try:
            # ── Step 1: Create GRN first ──
            print("📝 Creating GRN to test cancellation...")
            create_result = self.test_GRN_creation_save(row_data, row_num, sheet_name)

            if create_result[0] != "Pass":
                return ("Fail", f"❌ GRN creation failed, cannot test cancel: {create_result[1]}")

            grn_no = create_result[2]
            print(f"✅ GRN created: {grn_no}. Proceeding to cancel...")
            sleep(2)

            # ── Step 2: Search GRN in list ──
            current_field = "Search GRN in List"
            try:
                search_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, '//div[@id="grn_list_filter"]//input'))
                )
                search_input.clear()
                search_input.send_keys(grn_no)
                sleep(2)
                print(f"✅ Searched for GRN: {grn_no}")
            except Exception as e:
                print(f"⚠️ Search failed: {e}")

            # ── Step 3: Click Cancel button in Action column ──
            current_field = "Cancel Action Button"
            cancel_action_xpath = '(//button[@class="btn btn-warning"])[1]'
            try:
                Function_Call.click(self, cancel_action_xpath)
                sleep(2)
                print("✅ Clicked Cancel in Action column")
            except Exception as e:
                # Try generic cancel button in grn row
                try:
                    row_el = driver.find_element(
                        By.XPATH, f'//table[@id="grn_list"]//tr[contains(.,"{grn_no}")]'
                    )
                    cancel_btn = row_el.find_element(By.XPATH, './/a[contains(@class,"btn-danger")]')
                    cancel_btn.click()
                    sleep(2)
                except Exception as e2:
                    self._take_screenshot(f"CancelBtnNotFound_TC{row_data['TestCaseId']}")
                    return ("Fail", f"❌ Could not click cancel action button: {e2}")

            # ── Step 4: Handle Cancel Modal ──
            current_field = "Cancel Modal"
            try:
                wait.until(EC.visibility_of_element_located((By.ID, "confirm-delete")))
                print("✅ Cancel modal appeared")

                # Enter cancel remark (MANDATORY before button enables)
                cancel_reason = str(row_data["CancelReason"]) if row_data["CancelReason"] else "Cancelled for testing"
                remark_field = driver.find_element(By.ID, "cancel_remark")
                remark_field.clear()
                remark_field.send_keys(cancel_reason)
                sleep(1)
                print(f"✅ Entered cancel remark: {cancel_reason}")

                # Wait for Cancel button to become enabled
                wait.until(EC.element_to_be_clickable((By.ID, "grn_cancel")))
                self._take_screenshot(f"CancelModal_TC{row_data['TestCaseId']}")
                Function_Call.click(self, '//button[@id="grn_cancel"]')
                sleep(3)
                print("✅ Clicked Confirm Cancel button")

            except Exception as e:
                self._take_screenshot(f"CancelModalError_TC{row_data['TestCaseId']}")
                return ("Fail", f"❌ Cancel modal interaction failed: {e}")


            try:
                Function_Call.alert(self)
                sleep(2)
            except:
                pass
            # ── Step 5: Verify success message ──
            current_field = "Verify Cancel Success"
            try:
                success_el = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//div[contains(@class,'alert-success')]"))
                )
                msg_text = success_el.text.strip()
                if "Cancelled Successfully" in msg_text or "cancelled" in msg_text.lower():
                    print(f"✅ GRN {grn_no} Cancelled Successfully")
                    self._take_screenshot(f"CancelSuccess_TC{row_data['TestCaseId']}")
                    return ("Pass", f"✅ GRN {grn_no} Cancelled Successfully", grn_no)
                else:
                    self._take_screenshot(f"CancelUnexpectedMsg_TC{row_data['TestCaseId']}")
                    return ("Fail", f"❌ Unexpected message after cancel: {msg_text}")

            except TimeoutException:
                # Also check if status column updated in list
                try:
                    row_el = driver.find_element(
                        By.XPATH, f'//table[@id="grn_list"]//tr[contains(.,"{grn_no}")]'
                    )
                    if "cancel" in row_el.text.lower():
                        return ("Pass", f"✅ GRN {grn_no} cancelled (status updated in list)", grn_no)
                except:
                    pass
                self._take_screenshot(f"CancelNoMsg_TC{row_data['TestCaseId']}")
                return ("Fail", "❌ No success message after cancel")

        except Exception as e:
            msg = f"❌ Cancel flow error in {current_field}: {e}"
            print(msg)
            self._take_screenshot(f"CancelError_TC{row_data['TestCaseId']}")
            return ("Fail", msg)

    # ─────────────────────────────────────────────────────────────────
    # STEP 4: EDIT / UPDATE FLOW
    # ─────────────────────────────────────────────────────────────────
    def test_grn_edit_update(self, row_data, row_num, sheet_name):
        """
        Edit/Update Flow: Creates a GRN, then edits it and verifies update.
        Edit URL: admin_ret_purchase/grnentry/edit/{grn_id}
        """
        driver = self.driver
        wait = self.wait
        current_field = "GRN Creation for Edit"

        try:
            # ── Step 1: Create GRN first ──
            print("📝 Creating GRN to test edit/update...")
            create_result = self.test_GRN_creation_save(row_data, row_num, sheet_name)

            if create_result[0] != "Pass":
                return ("Fail", f"❌ GRN creation failed, cannot test edit: {create_result[1]}")

            grn_no = create_result[2]
            print(f"✅ GRN created: {grn_no}. Proceeding to edit...")
            sleep(2)

            # ── Step 2: Search GRN in list ──
            current_field = "Search GRN in List"
            try:
                search_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, '//div[@id="grn_list_filter"]//input'))
                )
                search_input.clear()
                search_input.send_keys(grn_no)
                sleep(2)
            except Exception as e:
                print(f"⚠️ Search failed: {e}")

            # ── Step 3: Click Edit button in Action column ──
            current_field = "Edit Action Button"
            try:
                edit_xpath = f'//table[@id="grn_list"]//tr[contains(.,"{grn_no}")]//a[contains(@href,"edit") or contains(@class,"btn-warning")]'
                Function_Call.click(self, edit_xpath)
                sleep(3)
                print(f"✅ Clicked Edit for GRN: {grn_no}")
            except Exception as e:
                self._take_screenshot(f"EditBtnNotFound_TC{row_data['TestCaseId']}")
                return ("Fail", f"❌ Could not click Edit button: {e}")

            # ── Step 4: Verify edit form loaded ──
            current_field = "Edit Form Load"
            try:
                wait.until(EC.presence_of_element_located((By.ID, "grn_entry_form")))
                print("✅ Edit form loaded successfully")
            except TimeoutException:
                self._take_screenshot(f"EditFormNotLoaded_TC{row_data['TestCaseId']}")
                return ("Fail", "❌ Edit form did not load")

            # ── Step 5: Modify Ref No (to mark it as edited) ──
            current_field = "Modify Ref No"
            try:
                ref_input = driver.find_element(By.XPATH, '//input[@name="order[po_supplier_ref_no]"]')
                existing_ref = ref_input.get_attribute("value")
                ref_input.clear()
                new_ref = f"{existing_ref}_EDITED"
                ref_input.send_keys(new_ref)
                sleep(0.5)
                print(f"✅ Modified Ref No: {existing_ref} → {new_ref}")
            except Exception as e:
                print(f"⚠️ Ref No modification failed: {e}")

            # Take screenshot before update
            self._take_screenshot(f"BeforeUpdate_TC{row_data['TestCaseId']}")

            # ── Step 6: Click Save/Update button ──
            current_field = "Update Save Button"
            Function_Call.click(self, '//button[@id="save_grn"]')
            sleep(3)

            # ── Step 7: Verify update success ──
            current_field = "Verify Update Success"
            return self._capture_save_result(row_data, "GRN Entry Updated Successfully", grn_no)

        except Exception as e:
            msg = f"❌ Edit/Update flow error in {current_field}: {e}"
            print(msg)
            self._take_screenshot(f"EditError_TC{row_data['TestCaseId']}")
            return ("Fail", msg)

    # ─────────────────────────────────────────────────────────────────
    # STEP 5: MULTIPLE ITEMS GRN
    # ─────────────────────────────────────────────────────────────────
    def test_multiple_item_grn(self, row_data, row_num, sheet_name):
        """
        Multiple Items Flow: Reads item rows from GRNEntry_Items sheet.
        Each row in GRNEntry_Items with matching TestCaseId is one item.
        """
        driver = self.driver
        wait = self.wait
        tc_id = str(row_data["TestCaseId"]).strip()
        items_sheet_name = "GRNEntry_Items"
        current_field = "Read Items Sheet"

        try:
            # ── Read items from GRNEntry_Items sheet ──
            try:
                wb = load_workbook(FILE_PATH)
                items_sheet = wb[items_sheet_name]

                item_col_map = {
                    "TestCaseId": 1, "Category": 2, "Pcs": 3,
                    "GrossWt": 4, "Wastage": 5, "RatePerGram": 6, "RateType": 7
                }

                item_rows = []
                for r in range(2, items_sheet.max_row + 1):
                    cell_val = items_sheet.cell(row=r, column=1).value
                    if cell_val is None:
                        break
                    if str(cell_val).strip() == tc_id:
                        item_data = {key: items_sheet.cell(row=r, column=col).value
                                     for key, col in item_col_map.items()}
                        item_rows.append(item_data)

                wb.close()

                if not item_rows:
                    print(f"⚠️ No items found in '{items_sheet_name}' for TestCaseId: {tc_id}")
                    return ("Fail", f"No items in {items_sheet_name} for {tc_id}")

                print(f"📋 Found {len(item_rows)} items for {tc_id}")

            except Exception as e:
                print(f"❌ Failed to read items sheet: {e}")
                return ("Fail", f"Items sheet read error: {str(e)}")

            # ── Navigate and fill header fields ──
            Function_Call.alert(self)
            current_field = "Add Button"
            Function_Call.click(self, '//a[@id="add_Order"]')
            sleep(3)

            # GRN Type (default Bill for multiple items)
            GRNType=row_data["GRNType"]
            grn_type = str(row_data["GRNType"]).strip() if row_data["GRNType"] else "Bill"
            grn_type_map = {
                "Bill": '//input[@id="oranment_type"]',
                "Receipt": '//input[@id="mt_type"]',
                "Charges": '//input[@id="st_type"]'
            }
            Function_Call.click(self, grn_type_map.get(grn_type, '//input[@id="oranment_type"]'))
            sleep(1)

            # Karigar
            if row_data["Karigar"]:
                Function_Call.dropdown_select(
                    self,
                    '//select[@id="select_karigar"]/following-sibling::span',
                    row_data["Karigar"],
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
            else:
                return ("Fail", "Karigar is mandatory ⚠️")

            # Ref No
            if row_data["RefNo"]:
                Function_Call.fill_input(
                    self, wait,
                    locator=(By.XPATH, '//input[@name="order[po_supplier_ref_no]"]'),
                    value=str(row_data["RefNo"]),
                    field_name="RefNo",
                    row_num=row_num,
                    pattern=r"^.+$",
                    Sheet_name=sheet_name
                )

            # Purchase Type
            if row_data["PurchaseType"]:
                Select(driver.find_element(By.ID, "purchase_type")).select_by_visible_text(
                    str(row_data["PurchaseType"])
                )

            # Ref Date
            if row_data["RefDate"]:
                Data = Function_Call.fill_input(
                    self, wait,
                    locator=(By.XPATH, '//input[@name="order[po_ref_date]"]'),
                    value=str(row_data["RefDate"]),
                    field_name="RefDate",
                    row_num=row_num,
                    pattern=r"^\d{2}[-/]\d{2}[-/]\d{4}$",
                    screenshot_prefix="RefDate",
                    extra_keys=Keys.TAB,
                    Sheet_name=sheet_name,
                    Date_range=True
                )
                if "Fail" in str(Data):
                    self._cancel_form()
                    return Data

            # Dispatch Through
            if row_data["DispatchThrough"]:
                Select(driver.find_element(By.ID, "despatch_through")).select_by_visible_text(
                    str(row_data["DispatchThrough"])
                )

            # ── Add each item row ──
            items_added = 0
            for idx, item in enumerate(item_rows, 1):
                print(f"\n📦 Adding Item {idx}/{len(item_rows)}: Category={item['Category']} Pcs={item['Pcs']}")
                result = self._fill_item_row(row_data, row_num, sheet_name,GRNType, item_data=item)
                if result and result[0] == "Fail":
                    return result
                items_added += 1

            # Verify item count in table
            table_rows = driver.find_elements(By.XPATH, '//table[@id="grn_item_details"]/tbody/tr')
            print(f"📊 Items in table: {len(table_rows)} (Expected: {len(item_rows)})")

            self._take_screenshot(f"MultiItem_BeforeSave_TC{tc_id}")

            # Save
            current_field = "Save Button"
            Function_Call.click(self, '//button[@id="submit_grn_entry"]')
            sleep(3)

            result = self._capture_save_result(row_data, "GRN Entry Added Successfully")
            if result[0] == "Pass":
                grn_no = result[2]
                return ("Pass", f"✅ Multi-item GRN ({items_added} items): {grn_no}", grn_no)
            return result

        except Exception as e:
            msg = f"❌ Multi-item GRN error in {current_field}: {e}"
            print(msg)
            self._take_screenshot(f"MultiItemError_TC{row_data['TestCaseId']}")
            self._cancel_form()
            return ("Fail", msg)

    # ─────────────────────────────────────────────────────────────────
    # PRIVATE HELPERS
    # ─────────────────────────────────────────────────────────────────
    def _capture_save_result(self, row_data, expected_msg, grn_no=""):
        """
        Waits for alert-success or alert-danger after form save.
        Extracts GRN No from grn_list table first row column 2.
        """
        wait = self.wait
        driver = self.driver
        try :
            Function_Call.alert(self)
        except :
            pass

        try:
            success_el = wait.until(
                EC.presence_of_element_located((By.XPATH, "//div[contains(@class,'alert-success')]"))
            )
            msg_text = success_el.text.strip()

            if "successfully" in msg_text.lower() or expected_msg.lower() in msg_text.lower():
                # Extract GRN No from list table
                if not grn_no:
                    try:
                        grn_no = wait.until(
                            EC.presence_of_element_located(
                                (By.XPATH, '//table[@id="grn_list"]/tbody/tr[1]/td[2]')
                            )
                        ).text.strip()
                        print(f"🆕 Extracted GRN No: {grn_no}")
                    except Exception as e:
                        print(f"⚠️ Could not extract GRN No: {e}")
                        grn_no = "Unknown"

                self._take_screenshot(f"SaveSuccess_TC{row_data['TestCaseId']}")
                print(f"✅ {expected_msg}: {grn_no}")
                return ("Pass", f"✅ {msg_text}", grn_no)
            else:
                self._take_screenshot(f"UnexpectedMsg_TC{row_data['TestCaseId']}")
                return ("Fail", f"❌ Unexpected message: {msg_text}")

        except UnexpectedAlertPresentException as e:
            alert_text = e.alert_text or "Validation error"
            print(f"⚠️ Alert: {alert_text}")
            try:
                driver.switch_to.alert.accept()
            except:
                pass
            self._take_screenshot(f"Alert_TC{row_data['TestCaseId']}")
            self._cancel_form()
            return ("Fail", f"❌ Alert: {alert_text}")

        except TimeoutException:
            # Check for danger alert
            try:
                danger_el = driver.find_element(
                    By.XPATH, "//div[contains(@class,'alert-danger')]"
                )
                msg_text = danger_el.text.strip()
                self._take_screenshot(f"SaveFail_TC{row_data['TestCaseId']}")
                return ("Fail", f"❌ {msg_text}")
            except:
                pass
            self._take_screenshot(f"NoSuccessMsg_TC{row_data['TestCaseId']}")
            return ("Fail", "❌ Success message not found after save")

        except Exception as e:
            self._take_screenshot(f"SaveError_TC{row_data['TestCaseId']}")
            return ("Fail", f"❌ Save result error: {str(e)}")

    def _cancel_form(self):
        """Cancels the current GRN form by clicking the Cancel/Back button"""
        try :
            Function_Call.alert(self)
        except :
            pass
        try:
            Function_Call.click(self, '(//button[@class="btn btn-default btn-cancel"])[2]')
            sleep(1)
        except:
            try:
                Function_Call.click(self, '//button[contains(@class,"btn-cancel")]')
                sleep(1)
            except:
                pass

    def _take_screenshot(self, filename):
        """Takes a screenshot and saves to SCREENSHOT_PATH folder"""
        try:
            screenshot_path = os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{filename}.png")
            self.driver.save_screenshot(screenshot_path)
            print(f"📸 Screenshot saved: {screenshot_path}")
        except Exception as e:
            print(f"⚠️ Screenshot failed: {e}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name, grn_no=""):
        """Writes test result back to Excel with color coding"""
        try:
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]

            color = "00B050" if test_status == "Pass" else "FF0000"

            # Col 2: TestStatus
            sheet.cell(row=row_num, column=2, value=test_status).font = Font(bold=True, color=color)
            # Col 3: ActualStatus
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            # Col 21: GRNNo
            if grn_no:
                sheet.cell(row=row_num, column=21, value=grn_no)
            # Col 24: Remark
            sheet.cell(row=row_num, column=24,
                       value=f"{test_status} - {actual_status}").font = Font(color=color)

            workbook.save(FILE_PATH)
            workbook.close()
            print(f"✅ Excel updated: Row {row_num} → {test_status}")

        except Exception as e:
            print(f"⚠️ Excel update failed: {e}")
