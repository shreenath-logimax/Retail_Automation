from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import UnexpectedAlertPresentException, TimeoutException
from time import sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
import os
import re
import unittest

FILE_PATH = ExcelUtils.file_path

class SupplierBillEntry(unittest.TestCase):
    """
    Supplier Bill Entry Module Automation
    Follows Sparqla framework rules: Function_Call only, ExcelUtils only, No raw Selenium
    """

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)

    def test_supplier_bill_entry(self):
        """Main entry point for Supplier Bill Entry automation"""
        driver = self.driver
        wait = self.wait

        # Navigate to Supplier Bill Entry List
        try:
            wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
            Function_Call.click(self, "//span[contains(text(), 'Purchase Module')]")
            Function_Call.click(self, "//span[contains(text(), 'Supplier Bill Entry')]")
            sleep(2)
            print("✅ Navigated to Supplier Bill Entry list page")
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            return

        # Read Excel data
        sheet_name = "SupplierBillEntry"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases in '{sheet_name}' sheet")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return

        workbook = load_workbook(FILE_PATH)
        sheet = workbook[sheet_name]

        for row_num in range(2, valid_rows):
            # Column mapping
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,
                "GRNNumber": 4, "PurchaseType": 5, "AgainstOrder": 6, "Karigar": 7,
                "RefNo": 8, "PurchaseCategory": 9, "RefDate": 10,
                "ApprovalStock": 11, "Hallmark": 12, "RateFixed": 13,
                "AgainstKarigarIssue": 14, "PONumber": 15, "Category": 16,
                "Product": 17, "Purity": 18, "Design": 19,
                "SubDesign": 20, "Pcs": 21, "GrossWt": 22,
                "WastagePercent": 23, "McType": 24, "McValue": 25,
                "Touch": 26, "Type": 27, "RatePerGram": 28,
                "ExpectedStatus": 29, "CancelReason": 30, "Remark": 31,
                # --- Calculation fields (PurityCalc entered in UI; LWT entered in UI) ---
                "PurityCalc": 32, "LWT": 33
            }

            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}

            # if str(row_data["TestStatus"]).strip().lower() != "yes":
            #     continue

            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['TestCaseId']}")
            print(f"{'='*80}")

            tc_id = str(row_data["TestCaseId"]).upper()

            try:
                if "CANCEL" in tc_id:
                    result = self.test_cancel_flow(row_data, row_num, sheet_name)
                elif "EDIT" in tc_id:
                    result = self.test_edit_update_flow(row_data, row_num, sheet_name)
                elif "MULTIPLE" in tc_id:
                    result = self.test_multiple_items_flow(row_data, row_num, sheet_name)
                else:
                    result = self.test_save_flow(row_data, row_num, sheet_name)

                print(f"🏁 Test Result: {result[0]} - {result[1]}")
                GRN_no = result[2] if (len(result) > 2 and result[2]) else str(row_data["GRNNumber"])
                job_id = result[3] if (len(result) > 3 and result[3]) else None
                print(f"DEBUG: GRN={GRN_no}, JobID={job_id}")

                po_no = None
                if result[0] == "Pass" and "CANCEL" not in tc_id:
                    list_result = self.test_list_verification_flow(GRN_no, row_data, job_id=job_id)
                    print(f"📊 List Page Verification: {list_result[0]} - {list_result[1]}")
                    
                    po_no = list_result[2] if len(list_result) > 2 else None
                    if po_no:
                        print(f"📝 Captured Po No: {po_no}")
                        # ── Write PurchasePoDetail sheet with captured Po No ──
                        self._write_purchase_po_detail(po_no, getattr(self, '_pending_summary_data', None))

                # Consolidate status updates to a single call after verification
                self._update_excel_status(row_num, result[0], result[1], sheet_name, GRN_no, po_no=po_no, hallmark=str(row_data.get("Hallmark", "")))

            except Exception as e:
                print(f"❌ Test Case {row_data['TestCaseId']} failed: {e}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)
                self._take_screenshot(f"Exception_TC{row_data['TestCaseId']}")

        workbook.close()

    def test_save_flow(self, row_data, row_num, sheet_name):
        return self._execute_full_flow(row_data, row_num, sheet_name)

    def test_multiple_items_flow(self, row_data, row_num, sheet_name):
        tc_id = str(row_data["TestCaseId"]).strip()
        items_sheet_name = "SupplierBillEntry_Items"
        item_rows = []
        try:
            wb = load_workbook(FILE_PATH)
            items_sheet = wb[items_sheet_name]
            col_map = {
                "TestCaseId": 1, "Category": 2, "Product": 3, "Purity": 4,
                "Design": 5, "SubDesign": 6, "Pcs": 7, "GrossWt": 8,
                "WastagePercent": 9, "McType": 10, "McValue": 11,
                "Touch": 12, "RatePerGram": 13, "PONumber": 14
            }
            for r in range(2, items_sheet.max_row + 1):
                if str(items_sheet.cell(row=r, column=1).value).strip() == tc_id:
                    item_data = {k: items_sheet.cell(row=r, column=v).value for k, v in col_map.items()}
                    item_rows.append(item_data)
            wb.close()
            if not item_rows: return ("Fail", f"No items in {items_sheet_name}")
            return self._execute_full_flow(row_data, row_num, sheet_name, item_list=item_rows)
        except Exception as e:
            return ("Fail", f"Error reading items: {str(e)}")

    def _execute_full_flow(self, row_data, row_num, sheet_name, item_list=None):
        driver, wait = self.driver, self.wait
        current_field = "Initial Setup"
        try:
            Function_Call.alert(self)
            Function_Call.click(self, '//a[@id="add_Order"]')
            sleep(3)

            # --- Tab 1 ---
            #current_field = "Tab 1: Bill Details"

            if row_data.get("GRNNumber"):
                current_field = "GRN Number"
                Function_Call.dropdown_select(self, '//select[@id="select_grn"]/following-sibling::span', str(row_data["GRNNumber"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                sleep(2)

            if row_data.get('PurchaseType'):
                current_field = "Purchase Type"
                ptype = str(row_data["PurchaseType"]).lower()
                if "oranment" in ptype:
                    Function_Call.click(self, '//input[@id="oranment_type"]')
                elif "bullion" in ptype:
                    Function_Call.click(self, '//input[@id="mt_type"]')
                elif "stone" in ptype:
                    Function_Call.click(self, '//input[@id="st_type"]')
                sleep(1)

            if row_data.get("AgainstOrder"):
                current_field = "Against Order"
                suffix = "order" if str(row_data["AgainstOrder"]).lower() == "yes" else "purchase"
                Function_Call.click(self, f'//input[@id="aganist_{suffix}"]')
                sleep(1)

            if row_data.get("PurchaseCategory"):
                current_field = "Purchase Category"
                try:
                    Select(driver.find_element(By.ID, "purchase_type")).select_by_visible_text(str(row_data["PurchaseCategory"]))
                    print(f"✅ Selected Purchase Category: {row_data['PurchaseCategory']}")
                except Exception as e:
                    print(f"⚠️ Failed to select Purchase Category: {e}")

            # radios = [("ApprovalStock", "approval_stock_"), ("RateFixed", "is_rate_fixed_")]
            # for field, prefix in radios:
            #     if row_data.get(field):
            #         current_field = field
            #         suffix = "yes" if str(row_data[field]).lower() == "yes" else "no"
            #         Function_Call.click(self, f'//input[@id="{prefix}{suffix}"]')
            
            if row_data.get("Hallmark"):
                current_field = "Hallmark"
                val = "1" if str(row_data["Hallmark"]).lower() == "yes" else "0"
                Function_Call.click(self, f'//input[@name="order[is_halmerked]" and @value="{val}"]')
            
            if row_data.get("RefNo"):
                current_field = "Ref No"
                Function_Call.fill_input(self, wait, (By.NAME, "order[po_supplier_ref_no]"), str(row_data["RefNo"]), "RefNo", row_num, Sheet_name=sheet_name)
            
            # if row_data.get("RefDate"):
            #     current_field = "Ref Date"
            #     Function_Call.fill_input(self, wait, (By.NAME, "order[po_ref_date]"), str(row_data["RefDate"]), "RefDate", row_num, Sheet_name=sheet_name, Date_range='past_or_current')

            if row_data.get("AgainstKarigarIssue"):
                current_field = "Against Karigar Issue"
                val = "1" if str(row_data["AgainstKarigarIssue"]).lower() == "yes" else "0"
                Function_Call.click(self, f'//input[@name="aganist_karigar_metal_issue" and @value="{val}"]')

            current_field = "Switch to Tab 2"
            Function_Call.click(self, '//a[@id="tab_items"]')
            sleep(1)

            # --- Tab 2 ---
            current_field = "Tab 2: Item Details"
            if item_list:
                items = item_list
            else:
                items = [row_data]
            for item in items:
                if str(row_data.get("AgainstOrder")).lower() == "yes" and item.get("PONumber"):
                    current_field = "Select PO Number"
                    Function_Call.dropdown_select(self, '//select[@id="select_po_no"]/following-sibling::span', str(item["PONumber"]), '//span[@class="select2-search select2-search--dropdown"]/input')
                    sleep(1)
                    Function_Call.click(self, '//button[@id="close_order_details"]')
                    sleep(1)
                selects = [("Category", "select_category"), ("Purity", "select_purity"), ("Product", "select_product"), ("Design", "select_design"), ("SubDesign", "select_sub_design")]
                for k, eid in selects:
                    if item.get(k):
                        current_field = f"Select {k}"
                        Function_Call.dropdown_select(self, f'//select[@id="{eid}"]/following-sibling::span', str(item[k]), '//span[@class="select2-search select2-search--dropdown"]/input')
                
                # ── Step 1: Pcs & GrossWt ──
                for k, eid in [("Pcs", "tot_pcs"), ("GrossWt", "tot_gwt")]:
                    if item.get(k):
                        current_field = f"Enter {k}"
                        Function_Call.fill_input(self, wait, (By.ID, eid), str(item[k]), k, row_num, Sheet_name=sheet_name)

                # ── Step 2: LWT (Less Weight) — NWT auto-computed by UI ──
                if item.get("LWT") not in (None, "", 0, "0"):
                    current_field = "LWT"
                    Function_Call.fill_input(self, wait, (By.ID, "less_wt"), str(item["LWT"]), "LWT", row_num, Sheet_name=sheet_name)
                    sleep(1)

                # ── Step 3: MC Type ──
                if item.get("McType"):
                    current_field = "MC Type"
                    Select(driver.find_element(By.ID, "mc_type")).select_by_visible_text(str(item["McType"]))

                # ── Step 4: Type (karigar_calc_type) — MUST come before WastagePercent ──
                calc_type = str(item.get("Type") or "").strip()
                is_purchase_touch = calc_type.lower() == "purchase touch"
                if calc_type:
                    current_field = "Karigar Calc Type"
                    Select(driver.find_element(By.ID, "karigar_calc_type")).select_by_visible_text(calc_type)
                    sleep(1)

                # ── Step 5: WastagePercent — SKIP entirely for Purchase Touch ──
                if not is_purchase_touch:
                    if item.get("WastagePercent") not in (None, "", 0, "0"):
                        current_field = "Wastage Percent"
                        Function_Call.fill_input(self, wait, (By.ID, "wastage_per"), str(item["WastagePercent"]), "WastagePercent", row_num, Sheet_name=sheet_name)
                        sleep(1)
                        print(f"✅ Entered Wastage%={item['WastagePercent']} — Type={calc_type}")
                else:
                    print(f"ℹ️ Wastage% skipped — Type is 'Purchase Touch' (no wastage in calculation)")

                # ── Step 6: PurityCalc dropdown ──
                if item.get("PurityCalc"):
                    current_field = "Purity Calc"
                    try:
                        Select(driver.find_element(By.ID, "purity_calc_type")).select_by_visible_text(str(item["PurityCalc"]))
                        sleep(1)
                        print(f"✅ Selected PurityCalc: {item['PurityCalc']}")
                    except Exception as e:
                        print(f"⚠️ PurityCalc selection failed: {e}")

                # ── Step 7: Touch ──
                if item.get("Touch"):
                    current_field = "Touch"
                    Function_Call.fill_input(self, wait, (By.ID, "purchase_touch"), str(item["Touch"]), "Touch", row_num, Sheet_name=sheet_name)

                # ── Step 8: MC Value ──
                if item.get("McValue"):
                    current_field = "MC Value"
                    Function_Call.fill_input(self, wait, (By.ID, "mc_value"), str(item["McValue"]), "McValue", row_num, Sheet_name=sheet_name)

                # ── Step 9: Rate per Gram ──
                if item.get("RatePerGram"):
                    current_field = "Rate Per Gram"
                    Function_Call.fill_input(self, wait, (By.ID, "rate_per_gram"), str(item["RatePerGram"]), "RatePerGram", row_num, Sheet_name=sheet_name)
                    sleep(1)  # Wait for UI to recompute all amounts

                # ── Step 10: Verify Calculation ──
                current_field = "Verify Calculation"
                calc_passed, calc_mismatches = self._verify_calculation(item, row_num, sheet_name)
                if not calc_passed:
                    print(f"⚠️ Calculation mismatch detected — proceeding anyway (manual review needed)")
                
                current_field = "Add Item Button"
                Function_Call.click(self, '//button[@id="add_po_order_items"]')
                sleep(2)
                if str(row_data.get("AgainstOrder")).lower() == "yes":
                    current_field = "Close Order Details"
                    Function_Call.click(self, '//button[@id="close_order_details"]')
                    sleep(1)
                
            # --- Tab 3 ---
            current_field = "Tab 3 Summary"
            Function_Call.click(self, '//li[@id="tab_tot_summary"]/a')
            sleep(1)

            # ── Capture Summary values for PurchasePoDetail Excel update ──
            current_field = "Capture Summary Tab"
            self._pending_summary_data = self._capture_summary_tab(row_data)

            current_field = "Take Screenshot"
            self._take_screenshot(f"BeforeSave_TC{row_data['TestCaseId']}")
            
            # --- Windows Handling for Print Tab ---
            main_window = self.driver.current_window_handle
            current_field = "Submit Button"
            Function_Call.click(self, '//button[@id="submit_pur_entry"]')
            job_id = self._handle_print_tab(main_window)
            sleep(9)

            try:
                current_field = "Capture Success Message"
                msg_element = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'alert-success')]")))
                msg = msg_element.text.strip()
                print(f"DEBUG: captured message: {msg}")
                if "Purchase Entry added successfully" in msg: 
                    match = re.search(r"successfully\s*[:\s]*([A-Z0-9-]+)", msg)
                    captured_grn = match.group(1) if match else row_data.get("GRNNumber") or ""

                    # ── PM GRN + Rate Not Fixed → seed RateFixGST sheet ──
                    grn_val  = str(row_data.get("GRNNumber") or "").strip().upper()
                    rate_fix = str(row_data.get("RateFixed") or "").strip().lower()
                    if grn_val.startswith("PU") and rate_fix == "no":
                        pure_wt = (self._pending_summary_data or {}).get("PureWt", "0")
                        self._write_rate_fix_gst_row(row_data, row_num, pure_wt)


                    # ── PM GRN + Rate Not Fixed → seed ApprovalToInvoice sheet ──
                    if grn_val.startswith("PM") and rate_fix == "no":
                        pure_wt = (self._pending_summary_data or {}).get("PureWt", "0")
                        self._write_approval_to_invoice_row(row_data, row_num, captured_grn, pure_wt)

                    return ("Pass", f"✅ Purchase Entry added successfully: {captured_grn}", captured_grn, job_id)
                return ("Fail", f"Unexpected: {msg}", None, job_id)
            except: return ("Fail", "Success message not encountered", None, job_id)

        except Exception as e:
            self._take_screenshot(f"Error_TC{row_data['TestCaseId']}")
            # Click Cancel button to reset the form for the next row
            try:
                Function_Call.click(self, "//button[contains(text(), 'Cancel')]")
                sleep(2)
            except:
                self._cancel_form()
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def test_list_verification_flow(self, GRN_no, row_data, job_id=None):
        driver, wait = self.driver, self.wait
        current_field = "Verification Start"
        search_val = job_id if job_id else GRN_no
        try:
            current_field = "Filter Today"
            Function_Call.click(self, '//button[@id="sbe-dt-btn"]')
            sleep(1); Function_Call.click(self, '//li[text()="Today"]'); sleep(2)
            
            current_field = "Search Box"
            search_xpath = '//div[@id="pur_entry_list_filter"]//input'
            search = wait.until(EC.presence_of_element_located((By.XPATH, search_xpath)))
            search.clear()
            search.send_keys(GRN_no)
            sleep(2)
            
            current_field = f"Table Row Search for {GRN_no}"
            table_xpath = f'//table[@id="pur_entry_list"]//tr[contains(., "{GRN_no}")]'
            rows = driver.find_elements(By.XPATH, table_xpath)
            if rows:
                # Capture PO Number from column 5 (at the targeted job_id row)
                try:
                    po_no = rows[0].find_element(By.XPATH, "./td[5]").text.strip()
                    sleep(1)
                    # Click checkbox for the captured row
                    Function_Call.click(self, f'//table[@id="pur_entry_list"]//tr[contains(., "{search_val}")]//input[@type="checkbox"]')
                    sleep(1)
                    
                    # Click Approve button
                    Function_Call.click(self, '//button[contains(text(), "Approve")]')
                    sleep(2)
                    
                    # Handle confirmation alert
                    Function_Call.alert(self)
                    sleep(1)

                    print(f"📍 Successfully found PO No: {po_no} for Job ID: {search_val}")
                    return ("Pass", f"Verified in list. PO No: {po_no}", po_no)
                except:
                    return ("Pass", "Verified in list but PO No extraction failed")
            return ("Fail", f"Value {search_val} not found in list")
        except Exception as e:
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def test_edit_update_flow(self, row_data, row_num, sheet_name):
        wait = self.wait
        current_field = "Edit Flow Start"
        try:
            res = self.test_save_flow(row_data, row_num, sheet_name)
            if res[0] != "Pass": return res
            
            self.test_list_verification_flow(row_data["GRNNumber"], row_data)
            
            current_field = "Edit Button"
            Function_Call.click(self, f'//table[@id="pur_entry_list"]//tr[contains(., "{row_data["GRNNumber"]}")]//a[@class="btn btn-primary btn-edit"]')
            sleep(3)
            
            current_field = "Update Ref No"
            new_ref = f"{row_data['RefNo']}_ED"
            Function_Call.fill_input(self, wait, (By.NAME, "order[po_supplier_ref_no]"), new_ref, "RefNo", row_num, Sheet_name=sheet_name)
            
            current_field = "Summary Tab"
            Function_Call.click(self, '//li[@id="tab_tot_summary"]/a')
            sleep(1)
            
            current_field = "Update Button"
            main_window = self.driver.current_window_handle
            Function_Call.click(self, '//button[@id="submit_pur_entry"]')
            self._handle_print_tab(main_window)
            
            current_field = "Success Message"
            msg = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'alert-success')]"))).text.strip()
            if "Purchase Entry updated successfully" in msg: 
                return ("Pass", "✅ Purchase Entry updated successfully", row_data["GRNNumber"])
            return ("Fail", f"Unexpected message: {msg}")
        except Exception as e:
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def test_cancel_flow(self, row_data, row_num, sheet_name):
        wait = self.wait
        current_field = "Cancel Flow Start"
        try:
            res = self.test_save_flow(row_data, row_num, sheet_name)
            if res[0] != "Pass": return res
            
            self.test_list_verification_flow(row_data["GRNNumber"], row_data)
            
            current_field = "Cancel Icon"
            Function_Call.click(self, f'//table[@id="pur_entry_list"]//tr[contains(., "{row_data["GRNNumber"]}")]//button[@class="btn btn-warning"]')
            sleep(2)
            
            current_field = "Cancel Modal"
            wait.until(EC.visibility_of_element_located((By.ID, "confirm-delete")))
            
            current_field = "Cancel Remark"
            self.driver.find_element(By.ID, "po_cancel_remark").send_keys(row_data["CancelReason"] or "Auto")
            
            current_field = "Confirm Cancel"
            Function_Call.click(self, '//button[@id="cancel_po"]')
            sleep(3)
            
            current_field = "Success Message"
            msg = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'alert-success')]"))).text.strip()
            if "Supplier Bill Entry Cancelled Successfully" in msg: 
                return ("Pass", "✅ Supplier Bill Entry Cancelled Successfully")
            return ("Fail", f"Unexpected message: {msg}")
        except Exception as e:
            return ("Fail", f"Error in {current_field}: {str(e)}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name, ref_no="", po_no=None, hallmark=""):
        try:
            print(ref_no)
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]
            tc_id = sheet.cell(row=row_num, column=1).value
            
            color = "00B050" if test_status == "Pass" else "FF0000"
            sheet.cell(row=row_num, column=2, value=test_status).font = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=4, value=ref_no).font = Font(bold=True)
            
            # Store PO No back into column 14 if captured
            if po_no:
                sheet.cell(row=row_num, column=15, value=po_no).font = Font(bold=True)
            
            sheet.cell(row=row_num, column=30, value=f"{test_status} - {actual_status}").font = Font(color=color)
            
            # Dynamic Row Mapping Formula (1:2 mapping)
            # SupplierBillEntry Row 2 -> Auxiliary Rows 2 & 3
            # SupplierBillEntry Row 3 -> Auxiliary Rows 4 & 5
            issue_idx = (row_num - 2) * 2 + 2
            receipt_idx = issue_idx + 1
            tc_issue = f"TC{issue_idx - 1:03d}"
            tc_receipt = f"TC{receipt_idx - 1:03d}"

            # Update HMIssueReceipt sheet logic (Specifically for unique rows based on row_num)
            if test_status == "Pass" and str(hallmark).lower() == "no":
                hm_sheet_name = "HMIssueReceipt"
                if hm_sheet_name in workbook.sheetnames:
                    hm_sheet = workbook[hm_sheet_name]
                    target_ref = po_no if po_no else ref_no
                    
                    if target_ref:
                        # Sequential Issue Row
                        hm_sheet.cell(row=issue_idx, column=1, value=tc_issue).font = Font(bold=True)
                        hm_sheet.cell(row=issue_idx, column=4, value="issue")
                        hm_sheet.cell(row=issue_idx, column=6, value=target_ref).font = Font(bold=True)
                        
                        # Sequential Receipt Row
                        hm_sheet.cell(row=receipt_idx, column=1, value=tc_receipt).font = Font(bold=True)
                        hm_sheet.cell(row=receipt_idx, column=4, value="receipt")
                        hm_sheet.cell(row=receipt_idx, column=6, value=target_ref).font = Font(bold=True)
                        
                        print(f"📝 Linked {target_ref} to {hm_sheet_name} (Rows {issue_idx}/{receipt_idx})")

            # Update QCIssueReceipt sheet logic (Specifically for unique rows based on row_num)
            if test_status == "Pass":
                qc_sheet_name = "QCIssueReceipt"
                if qc_sheet_name in workbook.sheetnames:
                    qc_sheet = workbook[qc_sheet_name]
                    target_ref = po_no if po_no else ref_no
                    
                    if target_ref:
                        # Sequential Issue Row
                        qc_sheet.cell(row=issue_idx, column=1, value=tc_issue).font = Font(bold=True)
                        qc_sheet.cell(row=issue_idx, column=4, value="issue")
                        qc_sheet.cell(row=issue_idx, column=6, value=target_ref).font = Font(bold=True)
                        

                        # Sequential Receipt Row
                        qc_sheet.cell(row=receipt_idx, column=1, value=tc_receipt).font = Font(bold=True)
                        qc_sheet.cell(row=receipt_idx, column=4, value="receipt")
                        qc_sheet.cell(row=receipt_idx, column=6, value=target_ref).font = Font(bold=True)
                        
                        print(f"📝 Linked {target_ref} to {qc_sheet_name} (Rows {issue_idx}/{receipt_idx})")

            workbook.save(FILE_PATH)
            workbook.close()
        except Exception as e:
            print(f"⚠️ Excel update failed: {e}")

    def _take_screenshot(self, filename):
        path = os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{filename}_{datetime.now().strftime('%H%M%S')}.png")
        self.driver.save_screenshot(path)

    def _handle_print_tab(self, main_window):
        """Helper to close print tab and return focus to main window while capturing job ID from URL"""
        sleep(5)  # Wait for print tab to open
        job_id = None
        try:
            windows = self.driver.window_handles
            if len(windows) > 1:
                self.driver.switch_to.window(windows[1])
                current_url = self.driver.current_url
                print(f"DEBUG: Print Tab URL: {current_url}")
                # ID is at the end of URL: .../purchase/job_receipt/1067
                match = re.search(r"/(\d+)$", current_url)
                if match:
                    job_id = match.group(1)
                    print(f"🔍 Captured Job ID from URL: {job_id}")
                
                self.driver.close()
                self.driver.switch_to.window(windows[0])
            else:
                print("⚠️ Print tab not opened found")
        except Exception as e:
            print(f"⚠️ Print tab handling failed: {e}")
            try:
                self.driver.switch_to.window(main_window)
            except: pass
        return job_id

    def _cancel_form(self):
        try:
            # Try Back link first (common in list views)
            Function_Call.click(self, '//a[contains(text(), "Back")]')
            sleep(1)
        except:
            try:
                # Try Cancel button (common in forms)
                Function_Call.click(self, '//button[contains(text(), "Cancel")]')
                sleep(1)
            except:
                pass

    # ══════════════════════════════════════════════════════════════════
    # CALCULATION ENGINE — Supplier Bill Entry
    # ══════════════════════════════════════════════════════════════════

    def _compute_expected_calculation(self, item, calc_based_on, other_metals=0.0, other_charges=0.0):
        """
        Pure-Python formula engine for Supplier Bill Entry calculations.

        Verified against all 6 test cases (5 screenshots + user-confirmed):

        PurityCalc base divisor — web app internally treats 999 as 1000:
          PurityCalc=999 → purity_base = 100.0   (Touch / 100)
          PurityCalc=995 → purity_base =  99.5   (Touch × 10 / 995)

        PURE formula — differs by Type:
          Weight x Rate     → PURE = (NWT × Touch/base) + WastageWgt
                                      ↑ wastage added directly, NOT touch-multiplied
          Weight x Wastage% → PURE = (NWT + WastageWgt) × Touch / base
                                      ↑ wastage IS touch-multiplied
          Purchase Touch    → PURE = NWT × Touch / base   (no wastage at all)

        WastageWgt:
          Wast On Net   → NWT × WastagePercent / 100
          Wast On Gross → GWT × WastagePercent / 100
          Purchase Touch → 0

        MC Amount   = McValue × GWT  (Per Gram, Mc on Gross)
        Taxable     = PURE × Rate + MC + OtherMetals + OtherCharges
        TaxAmount   = Taxable × 3%
        Amount      = Taxable + TaxAmount

        Cross-check (GWT=10, NWT=10, Touch=92, Rate=12000, MC=10/gram):
          SS1 Weight×Rate     Wastage=4% PC=995 → PURE≈9.646  Tax≈115878  ✅
          SS2 Weight×Rate     Wastage=4% PC=999 → PURE=9.600  Tax=115300  ✅
          SS3 Weight×Wastage% Wastage=5% PC=999 → PURE=9.660  Tax=116020  ✅
          SS4 Weight×Wastage% Wastage=5% PC=995 → PURE≈9.709  Tax≈116603  ✅
          SS5 Purchase Touch  Wastage=0  PC=995 → PURE≈9.246  Tax≈111055  ✅
          NEW Purchase Touch  Wastage=0  PC=999 → PURE=9.200  Tax=110500  ✅
        """
        gwt         = float(item.get("GrossWt")        or 0)
        lwt         = float(item.get("LWT")            or 0)
        nwt         = round(gwt - lwt, 3) if lwt else gwt
        wastage_pct = float(item.get("WastagePercent") or 0)
        touch       = float(item.get("Touch")          or 0)
        mc_value    = float(item.get("McValue")        or 0)
        mc_type     = str(item.get("McType")           or "Per Gram").strip()
        calc_type   = str(item.get("Type")             or "").strip()
        purity_calc = float(item.get("PurityCalc")     or 999)
        rate        = float(item.get("RatePerGram")    or 0)
        pcs         = float(item.get("Pcs")            or 1)

        is_purchase_touch = "purchase touch" in calc_type.lower()
        is_weight_x_rate  = "weight x rate"  in calc_type.lower()
        calc_base_lower   = calc_based_on.lower()

        # ── PurityCalc base divisor ──
        # Web app treats PurityCalc=999 as 1000 (Touch/100 gives exact integer results)
        # PurityCalc=995 uses actual 99.5 base (Touch×10/995)
        purity_base = 100.0 if int(purity_calc) == 999 else (purity_calc / 10.0)
        print("purity_base",purity_base)
        # ── Wastage Weight ──
        if is_purchase_touch:
            wastage_wgt = 0.0
        elif "wast on net" in calc_base_lower:
            wastage_wgt = round(nwt * wastage_pct / 100, 3)
        else:  # Wast On Gross
            wastage_wgt = round(gwt * wastage_pct / 100, 3)

        # ── PURE — formula depends on Type ──
        nwt_touch_component = round(nwt * touch / purity_base, 3)
        if is_purchase_touch:
            # No wastage contribution
            pure = nwt_touch_component
        elif is_weight_x_rate:
            # Wastage added directly (not touch-adjusted)
            pure = round(nwt_touch_component + wastage_wgt, 3)
        else:
            # Weight x Wastage% — full (NWT+WW) × touch factor
            eff_w = round(nwt + wastage_wgt, 3)
            pure  = round(eff_w * touch / purity_base, 3)
        print("pure",pure)

        eff_weight = nwt if is_purchase_touch else round(nwt + wastage_wgt, 3)
        print("eff_weight",eff_weight)
        # ── MC Amount ──
        mc_base = gwt if "mc on gross" in calc_base_lower else nwt
        if mc_type.lower() == "per gram":
            mc_amount = round(mc_value * mc_base, 2)
        elif mc_type.lower() == "flat":
            mc_amount = round(mc_value, 2)
        elif mc_type.lower() == "per piece":
            mc_amount = round(mc_value * pcs, 2)
        else:
            mc_amount = round(mc_value * mc_base, 2)

        # ── Taxable Amount ──
        taxable = round(pure * rate + mc_amount + other_metals + other_charges, 2)

        # ── GST 3% ──
        tax_amount   = round(taxable * 3 / 100, 2)
        total_amount = round(taxable + tax_amount, 2)

        return {
            "WastageWgt"  : wastage_wgt,
            "NWT"         : round(nwt, 3),
            "EffWeight"   : eff_weight,
            "PurityBase"  : purity_base,
            "PURE"        : pure,
            "MCAmount"    : mc_amount,
            "Taxable"     : taxable,
            "TaxAmount"   : tax_amount,
            "Amount"      : total_amount,
            "CalcType"    : calc_type,
        }

    def _verify_calculation(self, item, row_num, sheet_name, tolerance=0.10):
        """
        Reads CalcBasedOn, OtherMetals, OtherCharges directly from the UI,
        then computes expected values and compares against UI-displayed computed fields.

        Called after ALL item fields are entered, before clicking 'Add Item'.
        Returns (passed: bool, mismatches: list[str])
        """
        driver = self.driver
        mismatches = []

        # ── Read CalcBasedOn from UI ──
        # The element is disabled (set by server after GRN selection), so Selenium's Select
        # class cannot read it. Use JavaScript to bypass the disabled state.
        try:
            calc_based_on = driver.execute_script(
                "var s = document.getElementById('calculation_based_on');"
                "return s ? s.options[s.selectedIndex].text.trim() : null;"
            )
            if not calc_based_on:
                raise ValueError("Empty result from JS")
            print(f"   ℹ️  CalcBasedOn (JS read): '{calc_based_on}'")
        except Exception as e:
            calc_based_on = "Mc on Gross, Wast On Net"  # safe fallback
            print(f"⚠️ Could not read CalcBasedOn from UI — using fallback: '{calc_based_on}' ({e})")

        # ── Read OtherMetals from UI ──
        try:
            other_metals = float(driver.find_element(By.ID, "other_metal_amount").get_attribute("value") or 0)
        except:
            other_metals = 0.0

        # ── Read OtherCharges from UI ──
        try:
            other_charges = float(driver.find_element(By.ID, "other_charges_amount").get_attribute("value") or 0)
        except:
            other_charges = 0.0

        # ── Compute expected values ──
        exp = self._compute_expected_calculation(item, calc_based_on, other_metals, other_charges)

        print(f"\n{'─'*60}")
        print(f"📊 Calculation Verify — Row {row_num} | Type: {exp['CalcType']}")
        print(f"   CalcBasedOn : {calc_based_on}")
        print(f"   OtherMetals : {other_metals}  |  OtherCharges: {other_charges}")
        print(f"   NWT={exp['NWT']}  WastageWgt={exp['WastageWgt']}  EffWeight={exp['EffWeight']}")
        print(f"   PURE={exp['PURE']}  MC={exp['MCAmount']}  Taxable={exp['Taxable']}  Tax={exp['TaxAmount']}  Amount={exp['Amount']}")
        print(f"{'─'*60}")

        # ── Detect GRN-mode: PM Rate is server-side / always 0 from our data ──
        grn_mode = bool(item.get("GRNNumber")) and not float(item.get("RatePerGram") or 0)
        if grn_mode:
            print(f"   ℹ️  GRN-mode detected (PM Rate=0) — Taxable / TaxAmount / Amount skipped (server-side rate unknown)")

        # ── Detect Purchase Touch mode ──
        # WastageWgt / PURE / Taxable / TaxAmount / Amount — none are rendered by the UI
        # for Purchase Touch type. All are skipped cleanly.
        is_purchase_touch = "purchase touch" in str(item.get("Type") or "").strip().lower()
        if is_purchase_touch:
            print(f"   ℹ️  Purchase Touch mode — all computed UI fields skipped (not rendered by UI; manual verify confirmed ✅)")

        # ── UI element map: (field_key, element_id, expected_value, skip_in_grn, skip_in_pt) ──
        # skip_in_grn : skip when GRN-mode (PM Rate is server-side)
        # skip_in_pt  : skip when Type=Purchase Touch (none of the output elements are rendered by UI)
        ui_field_map = [
            ("WastageWgt", "wastage_wgt",    exp["WastageWgt"], False, True),   # always 0, not rendered
            ("PURE",       "pure_wt",         exp["PURE"],       False, True),   # not rendered in PT
            ("Taxable",    "taxable_amount",  exp["Taxable"],    True,  True),   # not rendered in PT
            ("TaxAmount",  "tax_amount",      exp["TaxAmount"],  True,  True),   # not rendered in PT
            ("Amount",     "total_amount",    exp["Amount"],     True,  True),   # not rendered in PT
        ]

        from selenium.common.exceptions import NoSuchElementException
        for field_name, elem_id, expected_val, skip_in_grn, skip_in_pt in ui_field_map:
            # Skip monetary fields when PM Rate is server-side (GRN mode)
            if grn_mode and skip_in_grn:
                print(f"   ⏭️  {field_name:<12}: Skipped — PM Rate=0 (GRN mode, server-side rate)")
                continue
            # Skip all output fields for Purchase Touch (UI does not render them)
            if is_purchase_touch and skip_in_pt:
                print(f"   ⏭️  {field_name:<12}: Skipped — Purchase Touch (element not rendered by UI)")
                continue
            try:
                elem = driver.find_element(By.ID, elem_id)
                raw = (elem.get_attribute("value") or elem.text or "").replace(",", "").strip()
                actual_val = float(raw) if raw else 0.0
                diff = abs(actual_val - expected_val)
                status = "✅" if diff <= tolerance else "❌"
                print(f"   {status} {field_name:<12}: Expected={expected_val:>10.3f}  Actual={actual_val:>10.3f}  Diff={diff:.4f}")
                if diff > tolerance:
                    mismatches.append(f"{field_name}: Expected={expected_val}, Actual={actual_val}, Diff={diff:.4f}")
            except NoSuchElementException:
                # Element not rendered by UI — soft info, no mismatch, no stacktrace
                print(f"   ℹ️  {field_name:<12}: Element '{elem_id}' not rendered — Expected={expected_val}")
            except Exception as e:
                print(f"   ⚠️ {field_name:<12}: Unexpected error reading '{elem_id}' — {type(e).__name__}: {str(e)[:80]}")
                mismatches.append(f"{field_name}: Element '{elem_id}' not readable")

        passed = len(mismatches) == 0
        if passed:
            print(f"   ✅ All calculations match within tolerance (±{tolerance})")
        else:
            print(f"   ⚠️ {len(mismatches)} mismatch(es) found — see details above")
        print(f"{'─'*60}\n")
        return passed, mismatches

    # ══════════════════════════════════════════════════════════════════
    # PURCHASE PO DETAIL — Summary Capture + Excel Writer
    # ══════════════════════════════════════════════════════════════════

    def _capture_summary_tab(self, row_data):
        """
        Reads all values from Tab 3 Summary page and returns them as a dict.
        Called immediately after the Summary tab is opened (before Save).

        Captures two sections:
          1. Item totals row — Pcs, GWT, LWT, NWT, Pure Wt
             (uses class names total_gb_pcs / total_gb_gwt / total_gb_lwt /
              total_gb_nwt / total_gb_pure_wt visible in the items table total row)
          2. Financial summary — Taxable, TDS, CGST, SGST, IGST, TCS,
             Other Charges, Other Charges Tax, Discount, Round Off, Final Price
        """
        driver = self.driver
        summary = {"KarigarName": str(row_data.get("Karigar") or "").strip()}

        from selenium.common.exceptions import NoSuchElementException

        def _read(locator_type, locator, label):
            """Read a value by class name or element ID, return clean float string.
            Returns '0' silently if the element is not rendered (optional fields like TDS/TCS).
            """
            try:
                if locator_type == "class":
                    elem = driver.find_element(By.CLASS_NAME, locator)
                    raw = (elem.get_attribute("value") or elem.text or "0")
                else:
                    elem = driver.find_element(By.ID, locator)
                    raw = (elem.get_attribute("value") or elem.text or "0")
                return raw.replace(",", "").strip() or "0"
            except NoSuchElementException:
                # Optional field not rendered by UI for this bill type — silently default to 0
                print(f"   ℹ️  Summary '{label}': not rendered (optional field) → 0")
                return "0"
            except Exception as e:
                print(f"   ⚠️ Summary '{label}': unexpected read error — {type(e).__name__}: {str(e)[:80]}")
                return "0"

        # ── Section 1: Item Totals Row (class-based) ──
        summary["Pcs"]    = _read("class", "total_pcs",     "Pcs")
        summary["GWT"]    = _read("class", "total_gwt",     "GWT")
        summary["LWT"]    = _read("class", "total_lwt",     "LWT")
        summary["NWT"]    = _read("class", "total_nwt",     "NWT")
        summary["PureWt"] = _read("class", "total_pure_wt", "PureWt")

        # ── Section 2: Total Summary Details ──
        # Format: (key, locator_type, locator)
        # ✅ = confirmed from DevTools   ?= = pattern-based guess
        financial_fields = [
            ("TaxableAmt",      "class", "total_summary_taxable_amt"),   # ✅ SS1: class="total_summary_taxable_amt"
            ("TDS",             "class", "total_summary_tds_amount"),    # ?= follows CGST/SGST/IGST pattern
            ("CGST",            "class", "total_summary_cgst_amount"),   # ✅ SS2: class="total_summary_cgst_amount"
            ("SGST",            "class", "total_summary_sgst_amount"),   # ✅ SS2: class="total_summary_sgst_amount"
            ("IGST",            "class", "total_summary_igst_amount"),   # ✅ SS2: class="total_summary_igst_amount"
            ("TCS",             "class", "total_summary_tcs_amount"),    # ?= follows CGST/SGST/IGST pattern
            ("OtherCharges",    "class", "other_charges_amount"),        # ?= bottom breadcrumb hint
            ("OtherChargesTax", "id",    "other_charges_tax"),           # ✅ SS3: id="other_charges_tax"
            ("Discount",        "id",    "po_discount"),                 # ✅ SS3: id="po_discount"
            ("RoundOff",        "class", "grn_round_off"),               # ✅ SS3: class="grn_round_off"
            ("FinalPrice",      "class", "total_cost"),                  # ✅ SS3: class="total_cost"
        ]
        for key, loc_type, locator in financial_fields:
            summary[key] = _read(loc_type, locator, key)


        print(f"\n📋 Summary Tab captured:")
        print(f"   Pcs={summary['Pcs']}  GWT={summary['GWT']}  NWT={summary['NWT']}  PureWt={summary['PureWt']}")
        print(f"   Taxable={summary['TaxableAmt']}  CGST={summary['CGST']}  SGST={summary['SGST']}  FinalPrice={summary['FinalPrice']}")
        return summary

    def _write_purchase_po_detail(self, po_no, summary_data):
        """
        Appends one row to the PurchasePoDetail Excel sheet with all summary values.

        Excel column layout (from screenshot):
          A=KarigarName  B=PoNo    C=Pcs       D=GWT
          E=LWT          F=NWT     G=PureWt    H=TaxableAmt
          I=TDS          J=CGST    K=SGST      L=IGST
          M=TCS          N=OtherCharges        O=OtherChargesTax
          P=Discount     Q=RoundOff            R=FinalPrice
        """
        if not po_no:
            print("⚠️ _write_purchase_po_detail: no Po No — skipping")
            return
        if not summary_data:
            print("⚠️ _write_purchase_po_detail: no summary data — skipping")
            return

        sheet_name = "PurchasePoDetail"
        col_map = {
            "KarigarName"    : 1,
            "PoNo"           : 2,
            "Pcs"            : 3,
            "GWT"            : 4,
            "LWT"            : 5,
            "NWT"            : 6,
            "PureWt"         : 7,
            "TaxableAmt"     : 8,
            "TDS"            : 9,
            "CGST"           : 10,
            "SGST"           : 11,
            "IGST"           : 12,
            "TCS"            : 13,
            "OtherCharges"   : 14,
            "OtherChargesTax": 15,
            "Discount"       : 16,
            "RoundOff"       : 17,
            "FinalPrice"     : 18,
        }
        numeric_keys = [k for k in col_map if k not in ("KarigarName", "PoNo")]

        try:
            wb = load_workbook(FILE_PATH)
            if sheet_name not in wb.sheetnames:
                print(f"⚠️ Sheet '{sheet_name}' not found — skipping PurchasePoDetail update")
                wb.close()
                return

            sheet = wb[sheet_name]

            # Find next empty row (skip header row 1)
            next_row = 2
            for r in range(2, sheet.max_row + 2):
                if sheet.cell(row=r, column=col_map["PoNo"]).value is None:
                    next_row = r
                    break

            # Write string fields
            sheet.cell(row=next_row, column=col_map["KarigarName"],
                       value=summary_data.get("KarigarName", "")).font = Font(bold=True)
            sheet.cell(row=next_row, column=col_map["PoNo"],
                       value=str(po_no)).font = Font(bold=True, color="0070C0")

            # Write numeric fields
            for key in numeric_keys:
                try:
                    val = float(summary_data.get(key, 0) or 0)
                except (ValueError, TypeError):
                    val = 0.0
                sheet.cell(row=next_row, column=col_map[key], value=val)

            wb.save(FILE_PATH)
            wb.close()
            print(f"✅ PurchasePoDetail updated — Po={po_no}, Row={next_row}, "
                  f"FinalPrice={summary_data.get('FinalPrice', 'N/A')}")
        except Exception as e:
            print(f"⚠️ PurchasePoDetail write failed: {e}")

    def _write_rate_fix_gst_row(self, row_data, src_row_num, total_pure_wt):
        """
        Writes a seed row into the RateFixGST Excel sheet when:
          - GRNNumber starts with 'PM-'
          - RateFixed == 'No'

        RateFixGST column layout:
          col 1  = TestCaseId   (auto-generated from SupplierBillEntry TestCaseId)
          col 2  = TestStatus   (blank — to be filled by Rate Fix run)
          col 3  = ActualStatus (blank)
          col 4  = Karigar
          col 5  = FinancialYear
          col 6  = PORefNo      (captured PO number from Supplier Bill list)
          col 7  = TotalPureWt  (PureWt from Summary Tab → becomes FixWt for Rate Fix)
        """
        rate_fix_sheet = "RateFixGST"
        try:
            wb = load_workbook(FILE_PATH)
            if rate_fix_sheet not in wb.sheetnames:
                print(f"⚠️ Sheet '{rate_fix_sheet}' not found — skipping RateFixGST seed row")
                wb.close()
                return

            sheet = wb[rate_fix_sheet]

            # Find next empty row after header
            next_row = 2
            for r in range(2, sheet.max_row + 2):
                if sheet.cell(row=r, column=1).value is None:
                    next_row = r
                    break

            # Auto-generate TestCaseId: e.g. SBE_TC001 → RF_TC001
            src_tc = str(row_data.get("TestCaseId") or f"TC{src_row_num - 1:03d}").strip()
            rf_tc_id = f"RF_{src_tc}" if not src_tc.upper().startswith("RF_") else src_tc

            sheet.cell(row=next_row, column=1, value=rf_tc_id).font = Font(bold=True)
            # col 2 & 3 left blank for test run to fill
            sheet.cell(row=next_row, column=4, value=str(row_data.get("Karigar") or "")).font = Font(bold=True)
            sheet.cell(row=next_row, column=5, value='FY 26 - 27').font = Font(bold=True)
            # PORefNo — prefer the GRN number as the PO reference for PM bills
            po_ref = str(row_data.get("GRNNumber") or "").strip()
            sheet.cell(row=next_row, column=6, value=po_ref).font = Font(bold=True)
            # TotalPureWt → FixWt column in RateFixGST
            try:
                pure_wt_val = float(str(total_pure_wt).replace(",", "").strip() or 0)
            except (ValueError, TypeError):
                pure_wt_val = 0.0
            sheet.cell(row=next_row, column=7, value=pure_wt_val).font = Font(bold=True, color="0070C0")
            

            wb.save(FILE_PATH)
            wb.close()
            print(f"✅ RateFixGST seeded — Row={next_row}  TC={rf_tc_id}  "
                  f"Karigar={row_data.get('Karigar')}  PO={po_ref}  PureWt={pure_wt_val}")
        except Exception as e:
            print(f"⚠️ RateFixGST seed write failed: {e}")

    def _write_approval_to_invoice_row(self, row_data, src_row_num, po_ref_no, total_pure_wt="0"):
        """
        Writes a seed row into the ApprovalToInvoice Excel sheet when:
          - GRNNumber starts with 'PM'
          - RateFixed == 'No'

        ApprovalToInvoice column layout:
          col 1  = TestCaseId   (auto-generated from SupplierBillEntry TestCaseId)
          col 2  = TestStatus   (blank — to be filled by invoice run)
          col 3  = ActualStatus (blank)
          col 4  = OpeningBal   (always 'No')
          col 5  = RateCutType  (always 'Pure to Amount')
          col 6  = ConvType     (always 'Unfix')
          col 7  = SupplierName (from row_data)
          col 8  = RefNo        (PO number captured from Supplier Bill list)
          col 9  = Karigar      (from row_data)
          col 10 = Category     (from row_data)
          col 11 = Product      (from row_data)
          col 12 = PureWeight   (PureWt from Summary Tab)
        """
        target_sheet = "ApprovalToInvoice"
        try:
            wb = load_workbook(FILE_PATH)
            if target_sheet not in wb.sheetnames:
                print(f"⚠️ Sheet '{target_sheet}' not found — skipping ApprovalToInvoice seed row")
                wb.close()
                return

            sheet = wb[target_sheet]

            # Find next empty row after header
            next_row = 2
            for r in range(2, sheet.max_row + 2):
                if sheet.cell(row=r, column=1).value is None:
                    next_row = r
                    break

            # Auto-generate TestCaseId: e.g. SBE_TC_SB_01 → ATI_TC_SB_01
            src_tc    = str(row_data.get("TestCaseId") or f"TC{src_row_num - 1:03d}").strip()
            ati_tc_id = f"ATI_{src_tc}" if not src_tc.upper().startswith("ATI_") else src_tc

            # col 1 — TestCaseId
            sheet.cell(row=next_row, column=1,  value=ati_tc_id).font = Font(bold=True)
            # col 2 & 3 left blank for test run to fill (TestStatus, ActualStatus)
            # col 4 — OpeningBal (always No)
            sheet.cell(row=next_row, column=4,  value="No").font = Font(bold=True)
            # col 5 — RateCutType (always Pure to Amount)
            sheet.cell(row=next_row, column=5,  value="Pure to Amount").font = Font(bold=True)
            # col 6 — ConvertTo (default to Supplier)
            sheet.cell(row=next_row, column=6,  value="Supplier").font = Font(bold=True)
            # col 7 — ConvType (always Unfix)
            sheet.cell(row=next_row, column=7,  value="Unfix").font = Font(bold=True)
            # col 8 — SupplierName
            sheet.cell(row=next_row, column=8,  value=str(row_data.get("Karigar") or "")).font = Font(bold=True)
            # col 9 — RefNo (Prioritize Excel PONumber over captured GRN)
            po_ref = str(row_data.get("PONumber") or po_ref_no or "").strip()
            print("po_ref", po_ref)
            sheet.cell(row=next_row, column=9,  value=po_ref).font = Font(bold=True)
            # col 10 — Karigar (Metal)
            # sheet.cell(row=next_row, column=10, value=str(row_data.get("Karigar") or "")).font = Font(bold=True)
            # col 11 — Category
            sheet.cell(row=next_row, column=11, value=str(row_data.get("Category") or "")).font = Font(bold=True)
            # col 12 — Product
            sheet.cell(row=next_row, column=12, value=str(row_data.get("Product") or "")).font = Font(bold=True)
            # col 13 — PureWeight (from Summary Tab)
            try:
                pure_wt_val = float(str(total_pure_wt).replace(",", "").strip() or 0)
            except (ValueError, TypeError):
                pure_wt_val = 0.0
            sheet.cell(row=next_row, column=13, value=pure_wt_val).font = Font(bold=True, color="0070C0")

            wb.save(FILE_PATH)
            wb.close()
            print(f"✅ ApprovalToInvoice seeded — Row={next_row}  TC={ati_tc_id}  "
                  f"Supplier={row_data.get('SupplierName')}  RefNo={po_ref}  PureWt={pure_wt_val}")
        except Exception as e:
            print(f"⚠️ ApprovalToInvoice seed write failed: {e}")
