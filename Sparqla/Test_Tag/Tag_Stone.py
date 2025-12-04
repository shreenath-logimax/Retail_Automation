from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import sleep
import unittest,math
from Utils.Excel import ExcelUtils
from openpyxl import load_workbook
from time import sleep

FILE_PATH = ExcelUtils.file_path
class Tag_Stone(unittest.TestCase):  
    def __init__(self,driver):
        self.driver =driver 
        self.wait = WebDriverWait(driver, 10) 
         
    def test_tagStone(self,Sheet_name,test_case_id):
        driver = self.driver
        wait = self.wait
        
        function_XPATH = Sheet_name
        test_case_id = test_case_id
        value =ExcelUtils.Test_case_id_count(FILE_PATH, function_XPATH,test_case_id)
        print(value)
        
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, function_XPATH)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[function_XPATH]
        
        row=1
        count = value
        for row_num in range(2, valid_rows):
            current_id = sheet.cell(row=row_num, column=1).value  # Column 1 = Test Case Id
            if current_id == test_case_id:
                data = {
                    "Test Case Id": 1,
                    "Less Weight": 2,  
                    "Type": 3,
                    "XPATH": 4,
                    "Code": 5,
                    "Pcs": 6,
                    "Wt": 7,
                    "Wt Type": 8,
                    "Cal.Type": 9,
                    "Rate": 10,
                    "Amount": 11
                }
                row_Stonedata = {
                    key: sheet.cell(row=row_num, column=col).value
                    for key, col in data.items()
                }
                print(row_Stonedata)
              
                Lwt = "(//select[@name='est_stones_item[show_in_lwt][]'])[{}]".format(row)
                print(Lwt)
                wait.until(EC.element_to_be_clickable((By.XPATH,Lwt))).click()
                print(row_Stonedata["Less Weight"])
                Select(wait.until(EC.element_to_be_clickable(
                    (By.XPATH,Lwt)))).select_by_visible_text(row_Stonedata["Less Weight"])
                
                wait.until(EC.element_to_be_clickable(
                    (By.XPATH,"(//select[@name='est_stones_item[stones_type][]'])[{}]".format(row)))).click()
                Select(wait.until(EC.element_to_be_clickable(
                    (By.XPATH,"(//select[@name='est_stones_item[stones_type][]'])[{}]".format(row))))).select_by_visible_text(row_Stonedata["Type"])
                wait.until(EC.element_to_be_clickable(
                    (By.XPATH,"(//select[@name='est_stones_item[stone_id][]'])[{}]".format(row)))).click()
                Select(wait.until(EC.element_to_be_clickable(
                    (By.XPATH,"(//select[@name='est_stones_item[stone_id][]'])[{}]".format(row))))).select_by_visible_text(row_Stonedata["XPATH"])
   
                wait.until(EC.element_to_be_clickable(
                    (By.XPATH,"(//select[@name='est_stones_item[quality_id][]'])[{}]".format(row)))).click()
                Select(wait.until(EC.element_to_be_clickable(
                    (By.XPATH,"(//select[@name='est_stones_item[quality_id][]'])[{}]".format(row))))).select_by_visible_text(row_Stonedata["Code"])

                pcs="(//input[@name='est_stones_item[stone_pcs][]'])[{}]".format(row)
                print(pcs)
                print(row_Stonedata["Pcs"])
                wait.until(EC.visibility_of_element_located((By.XPATH,pcs))).click()
                wait.until(EC.visibility_of_element_located((By.XPATH,pcs))).clear()
                wait.until(EC.visibility_of_element_located((By.XPATH,pcs))).send_keys(row_Stonedata["Pcs"])
           
                wait.until(EC.visibility_of_element_located(
                    (By.XPATH,"(//input[@name='est_stones_item[stone_wt][]'])[{}]".format(row)))).click()
                wait.until(EC.visibility_of_element_located(
                    (By.XPATH,"(//input[@name='est_stones_item[stone_wt][]'])[{}]".format(row)))).clear()
                wait.until(EC.visibility_of_element_located(
                    (By.XPATH,"(//input[@name='est_stones_item[stone_wt][]'])[{}]".format(row)))).send_keys(row_Stonedata["Wt"])
          
                wait.until(EC.visibility_of_element_located(
                    (By.XPATH,"(//select[@name='est_stones_item[uom_id][]'])[{}]".format(row)))).click()
                Select(wait.until(EC.visibility_of_element_located(
                    (By.XPATH,"(//select[@name='est_stones_item[uom_id][]'])[{}]".format(row))))).select_by_visible_text(row_Stonedata['Wt Type'])
           
                if row_Stonedata["Cal.Type"]== "Wt":      
                    wait.until(EC.element_to_be_clickable(
                        (By.XPATH, "(//input[@name='est_stones_item[cal_type][{}]'and @value='1'])".format(row)))).click()
                else:
                    wait.until(EC.element_to_be_clickable(
                        (By.XPATH, "(//input[@name='est_stones_item[cal_type][{}]'and @value='2'])".format(row)))).click()               
             
                wait.until(EC.visibility_of_element_located(
                    (By.XPATH,"(//input[@name='est_stones_item[stone_rate][]'])[{}]".format(row)))).clear()
                wait.until(EC.visibility_of_element_located(
                    (By.XPATH,"(//input[@name='est_stones_item[stone_rate][]'])[{}]".format(row)))).send_keys(row_Stonedata["Rate"],Keys.TAB)
                
                amt = wait.until(EC.visibility_of_element_located(
                    (By.XPATH,'(//input[@name="est_stones_item[stone_price][]"])[{}]'.format(row))))
                value_lwt = amt.get_attribute("value")
                Table_amt = ('{:.2f}'.format(float(value_lwt)))
                print(Table_amt)
                if row_Stonedata["Cal.Type"]=='Wt':#stone rate Weight calculation
                    Weight = row_Stonedata['Wt']    
                    Rate = row_Stonedata['Rate']
                    Total_amt = float(Weight)*float(Rate)
                    Amount_wt=('{:.2f}'.format(math.ceil(Total_amt)))
                    if Amount_wt == Table_amt:
                        print('Stone Rate Calculation correct')
                    else:
                        print('Stone Rate Calculation not correct')   
                else:
                    Rate = row_Stonedata['Rate'] #stone rate Pcs calculation
                    Stone_Pcs = row_Stonedata['Pcs']
                    Total_value = float(Rate)*float(Stone_Pcs)
                    Amount_pcs=('{:.2f}'.format(math.ceil(Total_value)))
                    if Amount_pcs == Table_amt:
                        print('Stone Rate Calculation correct')
                    else:
                        print('Stone Rate Calculation not correct')                        
                if count != 1:
                    sleep(5)
                    driver.find_element(By.XPATH,'(//button[@class="btn btn-success btn-xs create_stone_item_details"])[{}]'.format(row)).click()
                    row=row+1
                    count=value-1
                else:
                    rows_wt=driver.find_elements(By.XPATH, '//input[@name="est_stones_item[stone_wt][]"]')
                    wt = Tag_Stone.test_tablevalue(self,rows_wt)
                    print(wt)       
                    Wt_gram = wait.until(EC.visibility_of_element_located(
                        (By.XPATH,"//table[@id='estimation_stone_cus_item_details']/tfoot/tr/td[6]"))).text
                    print(Wt_gram)
                    if row_Stonedata["Wt Type"]=="carat":
                        carat = float(wt)
                        value = carat / 5
                        grams = f"{value:.3f}"
                        # Calculate grams
                        #print(f"{carat} carats is {grams} grams")
                        if Wt_gram == grams:
                            print("Carat to gram calculation correct")
                        else:
                            print("Carat to gram calculation not correct")
                    if row_Stonedata["Wt Type"]=="gram":
                        if Wt_gram == wt:
                           print("Weight total is correct")
                        else:
                           print("Weight total is not correct")  
                    else:
                        print(Wt_gram)       
                    TotalAmount = wait.until(EC.visibility_of_element_located(
                        (By.XPATH,"//table[@id='estimation_stone_cus_item_details']/tfoot/tr/td[13]"))).text
                    TotalAmount = float(TotalAmount)
                    print(TotalAmount)    
                    Table_amount = wait.until(EC.visibility_of_all_elements_located(
                        (By.XPATH,'//input[@name="est_stones_item[stone_price][]"]')))
                    Table_rate = Tag_Stone.test_tablevalue(self,Table_amount)
                    print(Table_rate)
                    if   Table_rate == TotalAmount:
                        print("Amount total is correct")  
                    else:
                        print("Amount total is not correct")    
                    break
            else:
                print('No')         
        wait.until(EC.element_to_be_clickable(
            (By.XPATH,'(//button[@id="update_stone_details"])'))).click()
        Lwt_add = "Less weight detail Add successfully"
        data = Lwt_add,Wt_gram,TotalAmount
        print(data)
        return data
        
    def test_tablevalue(self,rows):
        driver = self.driver
        value=[]
        sleep(4)
        for row in rows:
            val = row.get_attribute("value")  
            if val and val.strip():  # Ensure it's not empty
                value.append(float(val.strip()))
            else:
                print("No value found in input.")  
        print("Collected Values:", value)
        if value:
            total_value = round(sum(value), 3)
        else:
            total_value = 0.0
        return total_value
                  
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import Select
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.webdriver.common.keys import Keys
# from time import sleep
# import unittest,math
# from Utils.Excel import ExcelUtils
# from openpyxl import load_workbook
# from time import sleep

# FILE_PATH = ExcelUtils.file_path
# class Tag_Stone(unittest.TestCase):  
#     def __init__(self,driver):
#         self.driver =driver 
#         self.wait = WebDriverWait(driver, 10)  
#     def test_tagStone(self,Sheet_name):
#         driver = self.driver
#         wait = self.wait
#         sleep(3)
#         function_XPATH = Sheet_name
#         test_case_id = "TC001" 
#         value =ExcelUtils.Test_case_id_count(FILE_PATH, function_XPATH,test_case_id)
#         print(value)
#         valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, function_XPATH)
#         workbook = load_workbook(FILE_PATH)
#         sheet = workbook[function_XPATH]
#         row=1
#         count = value
#         for row_num in range(2, valid_rows):
#             current_id = sheet.cell(row=row_num, column=1).value  # Column 1 = Test Case Id
#             if current_id == test_case_id:
#                 data = {
#                     "Test Case Id": 1,
#                     "Less Weight": 2,  
#                     "Type": 3,
#                     "XPATH": 4,
#                     "Code": 5,
#                     "Pcs": 6,
#                     "Wt": 7,
#                     "Wt Type": 8,
#                     "Cal.Type": 9,
#                     "Rate": 10,
#                     "Amount": 11
#                 }
#                 row_Stonedata = {
#                     key: sheet.cell(row=row_num, column=col).value
#                     for key, col in data.items()
#                 }
#                 print(row_Stonedata)
#                 sleep(3)
#                 Lwt = "(//select[@name='est_stones_item[show_in_lwt][]'])[{}]".format(row)
#                 print(Lwt)
#                 driver.find_element(By.XPATH,Lwt).click()
#                 print(row_Stonedata["Less Weight"])
#                 Select(driver.find_element(By.XPATH,Lwt)).select_by_visible_text(row_Stonedata["Less Weight"])
#                 driver.find_element(By.XPATH,"(//select[@name='est_stones_item[stones_type][]'])[{}]".format(row)).click()
#                 Select(driver.find_element(By.XPATH,"(//select[@name='est_stones_item[stones_type][]'])[{}]".format(row))).select_by_visible_text(row_Stonedata["Type"])
#                 driver.find_element(By.XPATH,"(//select[@name='est_stones_item[stone_id][]'])[{}]".format(row)).click()
#                 Select(driver.find_element(By.XPATH,"(//select[@name='est_stones_item[stone_id][]'])[{}]".format(row))).select_by_visible_text(row_Stonedata["XPATH"])
#                 sleep(2)
#                 driver.find_element(By.XPATH,"(//select[@name='est_stones_item[quality_id][]'])[{}]".format(row)).click()
#                 Select(driver.find_element(By.XPATH,"(//select[@name='est_stones_item[quality_id][]'])[{}]".format(row))).select_by_visible_text(row_Stonedata["Code"])
#                 sleep(8)
#                 pcs="(//input[@name='est_stones_item[stone_pcs][]'])[{}]".format(row)
#                 print(pcs)
#                 print(row_Stonedata["Pcs"])
#                 driver.find_element(By.XPATH,pcs).click()
#                 driver.find_element(By.XPATH,pcs).clear()
#                 driver.find_element(By.XPATH,pcs).send_keys(row_Stonedata["Pcs"])
#                 sleep(2)
#                 driver.find_element(By.XPATH,"(//input[@name='est_stones_item[stone_wt][]'])[{}]".format(row)).click()
#                 driver.find_element(By.XPATH,"(//input[@name='est_stones_item[stone_wt][]'])[{}]".format(row)).clear()
#                 driver.find_element(By.XPATH,"(//input[@name='est_stones_item[stone_wt][]'])[{}]".format(row)).send_keys(row_Stonedata["Wt"])
#                 sleep(2)
#                 driver.find_element(By.XPATH,"(//select[@name='est_stones_item[uom_id][]'])[{}]".format(row)).click()
#                 Select(driver.find_element(By.XPATH,"(//select[@name='est_stones_item[uom_id][]'])[{}]".format(row))).select_by_visible_text(row_Stonedata['Wt Type'])
#                 sleep(2)    
#                 if row_Stonedata["Cal.Type"]== "Wt":      
#                     driver.find_element(By.XPATH, "(//input[@name='est_stones_item[cal_type][{}]'and @value='1'])".format(row)).click()
#                 else:
#                     driver.find_element(By.XPATH, "(//input[@name='est_stones_item[cal_type][{}]'and @value='2'])".format(row)).click()               
#                 sleep(2) 
#                 driver.find_element(By.XPATH,"(//input[@name='est_stones_item[stone_rate][]'])[{}]".format(row)).clear()
#                 driver.find_element(By.XPATH,"(//input[@name='est_stones_item[stone_rate][]'])[{}]".format(row)).send_keys(row_Stonedata["Rate"],Keys.TAB)
#                 # driver.find_element(By.XPATH,"(//input[@name='est_stones_item[stone_rate][]'])[{}]".format(row)).send_keys("Some text", Keys.TAB)
#                 sleep(10)
#                 amt = driver.find_element(By.XPATH,'(//input[@name="est_stones_item[stone_price][]"])[{}]'.format(row))
#                 value_lwt = amt.get_attribute("value")
#                 Table_amt = ('{:.2f}'.format(float(value_lwt)))
#                 print(Table_amt)
#                 if row_Stonedata["Cal.Type"]=='Wt':#stone rate Weight calculation
#                     Weight = row_Stonedata['Wt']    
#                     Rate = row_Stonedata['Rate']
#                     Total_amt = float(Weight)*float(Rate)
#                     Amount_wt=('{:.2f}'.format(math.ceil(Total_amt)))
#                     if Amount_wt == Table_amt:
#                         print('Stone Rate Calculation correct')
#                     else:
#                         print('Stone Rate Calculation not correct')   
#                 else:
#                     Rate = row_Stonedata['Rate'] #stone rate Pcs calculation
#                     Stone_Pcs = row_Stonedata['Pcs']
#                     Total_value = float(Rate)*float(Stone_Pcs)
#                     Amount_pcs=('{:.2f}'.format(math.ceil(Total_value)))
#                     if Amount_pcs == Table_amt:
#                         print('Stone Rate Calculation correct')
#                     else:
#                         print('Stone Rate Calculation not correct')                        
#                 if count != 1:
#                     sleep(5)
#                     driver.find_element(By.XPATH,'(//button[@class="btn btn-success btn-xs create_stone_item_details"])[{}]'.format(row)).click()
#                     sleep(5)
#                     row=row+1
#                     count=value-1
#                 else:
#                     sleep(2)
#                     rows_wt=driver.find_elements(By.XPATH, '//input[@name="est_stones_item[stone_wt][]"]')
#                     wt = Tag_Stone.test_tablevalue(self,rows_wt)
#                     print(wt)       
#                     Wt_gram = driver.find_element(By.XPATH,"//table[@id='estimation_stone_cus_item_details']/tfoot/tr/td[6]").text
#                     print(Wt_gram)
#                     if row_Stonedata["Wt Type"]=="carat":
#                         carat = float(wt)
#                         value = carat / 5
#                         grams = f"{value:.3f}"
#                         # Calculate grams
#                         #print(f"{carat} carats is {grams} grams")
#                         if Wt_gram == grams:
#                             print("Carat to gram calculation correct")
#                         else:
#                             print("Carat to gram calculation not correct")
#                     if row_Stonedata["Wt Type"]=="gram":
#                         if Wt_gram == wt:
#                            print("Weight total is correct")
#                         else:
#                            print("Weight total is not correct")  
#                     else:
#                         print(Wt_gram)       
#                     TotalAmount = driver.find_element(By.XPATH,"//table[@id='estimation_stone_cus_item_details']/tfoot/tr/td[13]").text
#                     TotalAmount = float(TotalAmount)
#                     print(TotalAmount)    
#                     Table_amount = driver.find_elements(By.XPATH,'//input[@name="est_stones_item[stone_price][]"]')
#                     Table_rate = Tag_Stone.test_tablevalue(self,Table_amount)
#                     print(Table_rate)
#                     if   Table_rate == TotalAmount:
#                         print("Amount total is correct")  
#                     else:
#                         print("Amount total is not correct")    
#                     break
#             else:
#                 print('No') 
#         sleep(5)        
#         driver.find_element(By.XPATH,'(//button[@id="update_stone_details"])[2]').click()
#         Lwt_add = "Less weight detail Add successfully"
#         data = Lwt_add,Wt_gram,TotalAmount
#         print(data)
#         return data
        
#     def test_tablevalue(self,rows):
#         driver = self.driver
#         value=[]
#         for row in rows:
#             val = row.get_attribute("value")  
#             if val and val.strip():  # Ensure it's not empty
#                 value.append(float(val.strip()))
#             else:
#                 print("No value found in input.")  
#         print("Collected Values:", value)
#         if value:
#             total_value = round(sum(value), 3)
#         else:
#             total_value = 0.0
#         return total_value
              