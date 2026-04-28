from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from Test_lot.Lot import Lot
from Test_Tag.Tag_Stone import Tag_Stone
from Test_Tag.Tag_othermetal import  Tag_othermetal
from Utils.Function import Function_Call
from time import sleep
import unittest,math
from Utils.Excel import ExcelUtils
from openpyxl import load_workbook
from time import sleep
import re


FILE_PATH = ExcelUtils.file_path #Give Excel Filepath

class Tag(unittest.TestCase):  
    
    def __init__(self,driver):
        self.driver =driver 
        self.wait = WebDriverWait(driver, 30) 
        
    def test_tag(self,Sheet_name=None):
        driver = self.driver
        wait = self.wait
        
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Toggle navigation"))).click()
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Inventory"))).click()
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Tagging"))).click()
        wait.until(EC.element_to_be_clickable((By.ID,'add_tagging'))).click()
        Function_Call.click(self,"//span[@class='header_rate']/b[contains(text(),'INR')]")
        rate_text1 = wait.until(EC.presence_of_element_located((By.XPATH, "//li[@class='user-body rate_block_body']//tr[th[contains(text(),'Gold 22KT 1gm')]]/td"))).text
        rate_text2 = wait.until(EC.presence_of_element_located((By.XPATH, "//li[@class='user-body rate_block_body']//tr[th[contains(text(),'Gold 18KT 1gm')]]/td"))).text
        rate_text3 = wait.until(EC.presence_of_element_located((By.XPATH, "//li[@class='user-body rate_block_body']//tr[th[contains(text(),'Silver 1gm')]]/td"))).text
        rate_text4 = wait.until(EC.presence_of_element_located((By.XPATH, "//li[@class='user-body rate_block_body']//tr[th[contains(text(),'Gold 24KT 1gm')]]/td"))).text
        # Example: "INR 9500"
        gold_rate22KT = int(float(rate_text1.replace("INR", "").strip()))
        print(gold_rate22KT)  
        gold_rate18KT = int(float(rate_text2.replace("INR", "").strip()))
        print(gold_rate18KT)  
        Silver_rate = int(float(rate_text3.replace("INR", "").strip()))
        print(Silver_rate) 
        gold_rate24KT = int(float(rate_text4.replace("INR", "").strip()))
        print(gold_rate24KT)  
        function_name = Sheet_name
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, function_name)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[function_name]
        PCS_Count = 0
        previous_branch = None
        Current_Lot = None
        current_branch = None
        for row_num in range(2, valid_rows):
                # Define columns and dynamically fetch their values
            data = {
                    "Test Case Id": 1,
                    "Branch": 4,
                    "Lot No": 5,
                    "Product": 6,
                    "Section": 7,
                    "Design": 8,
                    "Sub Design": 9,
                    "Pieces": 10,
                    "No.of items": 11,
                    "GWT": 12,
                    "Less Weight": 13,
                    "Other Metal": 14,
                    "MC&VA Available": 15,
                    "Wastage%": 16,
                    "Mc Type": 17,
                    "MC": 18,
                    "Size": 19,
                    "Calc Type": 20,
                    "HUID1": 21,
                    "HUID2": 22,
                    "Rate / MRP": 23,
                    "Attribute": 24,
                    "Attribute Name": 25,
                    "Certification":26,
                    "Certification No": 27,
                    "Certification Image": 28,
                    "Button":29
                }
            row_data = {key: sheet.cell(row=row_num, column=col).value 
                        for key, col in data.items()}
            print(row_data)

            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['Test Case Id']}")
            print(f"{'='*80}")
            
            current_branch = row_data["Branch"]
            if row_num!=2:
                row_no=row_num-1
                before_Lot = sheet.cell(row=row_no, column=5).value
                before_Product= sheet.cell(row=row_no, column=6).value
            else:
                before_Lot=None
                before_Product=None
                
            Current_Lot=row_data["Lot No"]
            Current_Product=row_data["Product"]
    
            if  Current_Lot != before_Lot :
                if row_num != 2:
                    # New lot detected, update BranchTransfer for the PREVIOUS lot
                    self.update_branch_transfer(before_Lot, previous_branch)
                    driver.execute_script("window.scrollBy(0, -600);")
                
                if row_data.get("Branch"):
                    wait.until(EC.element_to_be_clickable((By.ID,"select2-branch_select-container"))).click()
                    Branch=wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@type='search']")))
                    Branch.clear()
                    Branch.send_keys(row_data["Branch"],Keys.ENTER)
                    sleep(5)
                else:
                    print(f"ℹ️ Skipping branch selection for row {row_num}")
                
                wait.until(EC.element_to_be_clickable((By.ID,"select2-tag_lot_received_id-container"))).click()
                Lot_No=wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@type='search']")))
                Lot_No.clear()
                Lot_No.send_keys(str(row_data["Lot No"]))
                Lot_NO=row_data["Lot No"]
                LOT = wait.until(EC.element_to_be_clickable((By.XPATH, f"//li[normalize-space()='{Lot_NO}']")))
                LOT.click()
                previous_branch = row_data["Branch"]
                sleep(2)
            if  Current_Product !=  before_Product: 
                if row_num !=2:
                   driver.execute_script("window.scrollBy(0, -600);")
                wait.until(EC.element_to_be_clickable((By.ID, "select2-tag_lt_prod-container"))).click()
                Product=wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@type='search']")))
                Product.send_keys(row_data["Product"])
                Product.send_keys(Keys.ENTER)
                try:
                    Function_Call.alert(self)
                    Product.send_keys(Keys.ENTER)
                except:
                    pass
                if row_data.get("Section"):
                    wait.until(EC.element_to_be_clickable((By.ID,"select2-section_select-container"))).click()
                    Section=wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@type='search']")))
                    Section.clear()
                    Section.send_keys(row_data["Section"],Keys.ENTER)
                else:
                    print(f"ℹ️ Skipping section selection for row {row_num}")
                
                TestCaseId=row_data["Test Case Id"]
                row_Lotdata=Lot.Lotdetails(self,TestCaseId)
                print (row_Lotdata)
                
                if row_Lotdata:
                    Pcs = row_Lotdata.get("Pcs", 0)
                    GWT = row_Lotdata.get("GWT", 0)
                    DWT = row_Lotdata.get("LWt", "NO")
                else:
                    Pcs, GWT, DWT = 0, 0, "NO"
                    print("⚠️ Lot details not found for TestCaseId:", TestCaseId)
                if DWT.upper()=="NO":
                    DWT=0
                    
                Pieces = wait.until(EC.visibility_of_element_located((By.XPATH,"//table[@id='lt-det']/tbody/tr/td[2]"))).text
                Pieces  = int(float(Pieces))
                print(Pieces)
                if Pieces==Pcs :
                    PCS_Count=Pieces
                    print(PCS_Count)
                    print("same pcs")
                else: 
                    print("Pcs not same")
                
                rawValue = wait.until(EC.visibility_of_element_located((By.XPATH,"//table[@id='lt-det']/tbody/tr/td[3]"))).text
                Gross_weigth  = int(float(rawValue))
                print(Gross_weigth)
                if Gross_weigth ==GWT:
                    print("same Gross Weight")
                else: 
                    print("GWT not same")
                
                rawValue2 = wait.until(EC.visibility_of_element_located((By.XPATH,"//table[@id='lt-det']/tbody/tr/td[5]"))).text
                Dia_weight = int(float(rawValue2))
                print(Dia_weight)
                if Dia_weight== DWT:
                    print("same Diamound Weight")
                else: 
                    print("Diamound Wt not same")
                    
                rawValue1 = wait.until(EC.visibility_of_element_located((By.XPATH,"//table[@id='lt-det']/tbody/tr/td[4]"))).text
                Net_weight = int(float(rawValue1))
                print(Net_weight)
                NWT= Gross_weigth-Dia_weight
                if Net_weight== NWT :
                    print("same Net Weight")
                else: 
                    print("Net Wt not same")
                
            else:
                print(f'{row_num} Row Tag runing') 
            try:   
                Function_Call.alert(self)
            except:   
                pass   
            if row_num != 2:
                driver.execute_script("window.scrollBy(0, -600);") 
            
            # Wait for any blocking overlays (loading spinners/banners) to disappear
            try:
                wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "overlay")))
            except:
                pass # Continue if no overlay or timeout occurs

            # Use Javascript click for select2 container as it is more robust against intercepted clicks
            design_dropdown = wait.until(EC.element_to_be_clickable((By.ID, "select2-des_select-container")))
            try:
                design_dropdown.click()
            except:
                driver.execute_script("arguments[0].click();", design_dropdown)
            Design=wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@type='search']")))
            Design.clear()
            Design.send_keys(str(row_data.get("Design") or ""),Keys.ENTER)
        
            wait.until(EC.element_to_be_clickable((By.ID, "select2-sub_des_select-container"))).click()
            Sub_Design=wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@type='search']")))
            Sub_Design.clear() 
            Sub_Design.send_keys(str(row_data.get("Sub Design") or ""),Keys.ENTER)
            
            Pieces=wait.until(EC.visibility_of_element_located((By.ID,"tag_pcs")))
            Pieces.clear()
            Pieces.send_keys(row_data["Pieces"])
            
            No_of_items=wait.until(EC.visibility_of_element_located((By.ID,"bulk_tag")))
            No_of_items.clear()
            No_of_items.send_keys(row_data["No.of items"])
            
            GWT=wait.until(EC.visibility_of_element_located((By.ID,"tag_gwt")))
            GWT.clear()
            GWT.send_keys(row_data["GWT"])
            GWT.send_keys(Keys.TAB)
            wait.until(EC.element_to_be_clickable((By.XPATH,'(//button[@id="close_stone_details"])[1]'))).click()
            sleep(2)
            test_case_id =row_data["Test Case Id"]
            if row_data["Less Weight"]=="Yes":
                sleep(3)
                Function_Call.click(self,'//form[@id="tag_form"]/div/div/div[2]/div/div/div[8]/div[2]/div/div/span')
                Sheet_name='Tag_LWt'
                LessWeight=Tag_Stone.test_tagStone(self,Sheet_name,test_case_id)
                print(LessWeight)
                Lwt,Wt_gram,TotalAmount=LessWeight
                print(LessWeight)
                print(Lwt)
                print(Wt_gram)
                print(TotalAmount)
            else:
                TotalAmount=0
                Wt_gram=0
                print("There is no Less Weight in this product")
            test_case_id =row_data["Test Case Id"]
            if row_data["Other Metal"]=="Yes":
                wait.until(EC.element_to_be_clickable((By.XPATH,"//form[@id='tag_form']/div/div/div[2]/div/div/div[9]/div[2]/div/div/span"))).click()
                Sheet_name='Tag_othermetal'
                Data=Tag_othermetal.test_othermetal(self,Sheet_name,test_case_id)
                OtherMetal,OtherMetalAmount =Data
                print(OtherMetal)
                print(OtherMetalAmount)
            else:
                OtherMetalAmount =0
                print("There is no Other Metal in this product")         
            sleep(3)
            
            if row_data["MC&VA Available"]=="No":
                wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='tag_wast_perc']"))).clear()
                wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='tag_wast_perc']"))).send_keys(row_data["Wastage%"])
            
            else:
                Tab=wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='tag_wast_perc']")))
                Tab.send_keys(Keys.TAB)
                
            if row_data["Mc Type"]:
                McType=wait.until(EC.element_to_be_clickable((By.XPATH,'//select[@id="tag_id_mc_type"]')))
                McType.click()
                Select(McType).select_by_visible_text(row_data["Mc Type"])
                wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='tag_mc_value']"))).clear()    
                wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='tag_mc_value']"))).send_keys(row_data["MC"])
            
            if row_data["Size"]:
                wait.until(EC.element_to_be_clickable((By.ID,"select2-tag_size-container"))).click()
                Size=wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@type='search']")))
                Size.clear()
                Size.send_keys(row_data["Size"],Keys.ENTER)
           
                   
            if row_data["Calc Type"]:
                CalcType=wait.until(EC.element_to_be_clickable((By.ID,"tag_calculation_based_on")))
                CalcType.click()
                Select(CalcType).select_by_visible_text(row_data["Calc Type"])
                CalculationType=row_data["Calc Type"]
            else:
                select_el = wait.until(EC.presence_of_element_located((By.ID, "tag_calculation_based_on")))
                select = Select(select_el)
                selected_text = select.first_selected_option.text
                print("Calc Type selected ->", selected_text)   # e.g. "Mc & Wast On Gross"
                CalculationType=selected_text

            if  CalculationType in 'MRP' or 'Fixed Rate based on Weight' in CalculationType:   
                if row_data["Rate / MRP"] != None:
                    wait.until(EC.element_to_be_clickable((By.ID,"tag_sell_rate"))).click()
                    wait.until(EC.visibility_of_element_located((By.ID,"tag_sell_rate"))).send_keys(row_data["Rate / MRP"])
                    value = "{:.2f}".format(float(row_data["Rate / MRP"]))
                    print(f'rate{value}')
                 
            else:
                metal_text = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//*[@id='lt_metal']"))
                ).text

                Board_rate = 0
                print("Metal:", metal_text)
                if 'GOLD' in metal_text:
                    if '75.0000' in metal_text:
                        Board_rate=gold_rate18KT
                    elif '80.0000' in metal_text or '91.6000' in metal_text:
                        Board_rate=gold_rate22KT
                    elif '100.0000' in metal_text or'999.0000' in metal_text:
                        Board_rate=gold_rate24KT
                elif 'SILVER' in metal_text:
                    if '92.5000' in metal_text or '91.6000' in metal_text or'999.0000' in metal_text or '100.0000' in metal_text or '80.0000' in metal_text:
                        Board_rate=Silver_rate
                else:
                    print(f"⚠️ Unrecognized metal type: {metal_text}. Using default rate.")
                    Board_rate = gold_rate22KT # Default fallback

                        
                    
                value = Tag.calculation(self,row_data,CalculationType,TotalAmount,Wt_gram,OtherMetalAmount,Board_rate)
                print(value)
                value = "{:.2f}".format(float(value))
                print(f'rate{value}')
            HUID1=(row_data["HUID1"])
            print(type(HUID1))
            if HUID1 != None:
                wait.until(EC.element_to_be_clickable((By.ID,"tag_huid"))).click()
                wait.until(EC.visibility_of_element_located((By.ID,"tag_huid"))).send_keys(row_data["HUID1"])
            HUID2=(row_data.get("HUID2"))
            if HUID2 != None:
                wait.until(EC.element_to_be_clickable((By.ID,"tag_huid2"))).click()
                wait.until(EC.visibility_of_element_located((By.ID,"tag_huid2"))).send_keys(row_data["HUID2"])
            else:
                print('done')
            Attribute = row_data["Attribute Name"]   
            if Attribute != None:
                wait.until(EC.element_to_be_clickable((By.ID,"tag_attribute"))).click()    
                
                wait.until(EC.element_to_be_clickable((By.XPATH,"//span[text()='Select Attribute']"))).click()
                Attribute_Name=wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@type='search']")))
                Attribute_Name.clear()
                Attribute_Name.send_keys(row_data["Attribute Name"],Keys.ENTER)
                
                wait.until(EC.element_to_be_clickable((By.XPATH,"//span[text()='Select Attribute Value']"))).click()
                Value=wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@type='search']")))
                Value.clear()
                Value.send_keys(str(row_data.get("Attribute", "")),Keys.ENTER)
                
                wait.until(EC.element_to_be_clickable((By.ID,"update_attribute_details"))).click()
            else:
                print("There is no Attribute")
                
                
            if row_data["Certification"] == "Yes":
                wait.until(EC.element_to_be_clickable((By.ID,"cert_no"))).click()
                wait.until(EC.visibility_of_element_located((By.ID,"cert_no"))).send_keys(row_data["Certification No"])
        
                wait.until(EC.element_to_be_clickable((By.ID,"cert_img"))).click()
                wait.until(EC.visibility_of_element_located((By.ID,"cert_img"))).send_keys(row_data["Certification Image"])
            
            else:
                print('There is no Certification')
            val = wait.until(EC.presence_of_element_located((By.ID,"tag_sale_value")))
            sell_val = val.get_attribute("value")
            print(sell_val)
            if sell_val == value:
                print(f"Tag Calculation Amount Showing correctly ")
            else :
                print(f"Tag Calculation Amount Showing not correctly Actual Amount ")    
            Button=row_data["Button"]  
            print(Button)  
            if Button == "Add":
                Input_Pcs=row_data.get("Pieces", 0)
                if Input_Pcs:
                    PCS_Count=PCS_Count-int(Input_Pcs)
                wait.until(EC.element_to_be_clickable((By.ID,"addTagToPreview"))).click()
                sleep(2)
                # --- Specific Alert Handling for Tagging ---
                try:
                    alert = WebDriverWait(driver, 5).until(EC.alert_is_present())
                    print("🔔 Found Alert:", alert.text)
                    alert.accept()
                    sleep(1)
                except:
                    print("ℹ️ No alert appeared after Add")
            else:
                wait.until(EC.element_to_be_clickable((By.ID,"addTagToPreviewAndCopy"))).click() 
                sleep(2)
                try:
                    alert = WebDriverWait(driver, 5).until(EC.alert_is_present())
                    print("🔔 Found Alert (Copy):", alert.text)
                    alert.accept()
                except:
                    pass
            Test_Status = 'Pass'
            Actual_Status="Tagged successfully"
            # Reload workbook from disk before saving to preserve Tag_Detail changes
            # written by update_Tagdetails in previous iterations (stale object would wipe them)
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[function_name]
            sheet.cell(row=row_num, column=2).value = Test_Status
            sheet.cell(row=row_num, column=3).value = Actual_Status
            workbook.save(FILE_PATH)
            Status = ExcelUtils.get_Status(FILE_PATH,function_name)  
            print(Status) 
            Tag.update_Tagdetails(self,row_num,function_name)            
            Update_master = ExcelUtils.update_master_status(FILE_PATH,Status,function_name)  
        
        # After loop, update the final Lot in BranchTransfer sheet
        if row_num >= 2:
            self.update_branch_transfer(Current_Lot, row_data["Branch"])
            
        #Tag.update_Tagdetails(self,row_num,function_name)
                # driver.find_element(By.XPATH,"(.//*[normalize-space(text()) and normalize-space(.)='Booking Master'])[1]/following::div[3]").click()
    
    
    def update_Tagdetails(self,row_num,function_name):
        wait = self.wait
        if function_name == 'Tag':
           sheet_name = "Tag_Detail"
        else:
            sheet_name ="Purchase_TagDetail"
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[sheet_name]

        # Wait until the table is present
        table = wait.until(EC.presence_of_element_located((By.XPATH, '//table[@id="lt_item_tag_preview"]')))
        table_rows = table.find_elements(By.XPATH, ".//tbody/tr")
        print(f"Total rows found: {len(table_rows)}")

        # Columns you want to copy from the web table
        columns_to_keep = ['Lot', 'Tag No', 'Product', 'Design', 'Sub Design',
                        'Calc Type', 'Pieces', 'Gross Wgt', 'Less Wgt',
                        'Net Wgt', 'Wast %', 'Making Charge']

        # Get all table headers to find indexes
        all_table_headers = [th.text for th in table.find_elements(By.XPATH, ".//thead/tr/th")]
        # Initialize an empty list to store the indexes
        columns_indexes = []
        columns_indexes = [all_table_headers.index(col) for col in columns_to_keep]
        print('oooooo')
        print(columns_indexes)
        # --- Initialize headers ---
        sheet.cell(row=1, column=1, value="Test Case Id")
        for col_offset, header in enumerate(columns_to_keep):
            sheet.cell(row=1, column=2 + col_offset, value=header)
        print('IIIIIIIIIII')
        print(col_offset)
        
        # Find the last used TC number to continue the sequence
        last_tc_num = 0
        for r in range(sheet.max_row, 1, -1):
            val = sheet.cell(row=r, column=1).value
            if val and str(val).startswith("TC"):
                match = re.search(r'\d+', str(val))
                if match:
                    last_tc_num = int(match.group())
                    break
        
        next_tc_num = last_tc_num + 1
        # New rows are prepended to the TOP of the preview table (newest = index 0).
        # Always grab only the first row — the one just added.
        valu = table_rows[:1]
        # Derive row_idx from last_tc_num: header=row1, TC001=row2, TC002=row3...
        # This avoids openpyxl max_row unreliability caused by phantom formatted rows
        row_idx = last_tc_num + 2
        for table_row in valu:
            table_cells = table_row.find_elements(By.TAG_NAME, "td")
            row_values = [table_cells[i].text.strip() for i in columns_indexes]
            print("Row Values:", row_values)
            # Clean product name (3rd element → index 2)
            row_values[2] = re.sub(r'-\d+$', '', row_values[2]).strip()
            print(row_values)

            # Write Test Case Id to column 1
            tc_id = f"TC{str(next_tc_num).zfill(3)}"
            sheet.cell(row=row_idx, column=1, value=tc_id)
            next_tc_num += 1

            # Numeric column indexes in row_values:
            # 0=Lot, 6=Pieces, 7=Gross Wgt, 8=Less Wgt, 9=Net Wgt, 10=Wast%, 11=Making Charge
            numeric_indexes = {0, 6, 7, 8, 9, 10, 11}

            # Write row data to Excel (starting from column 2)
            for col_offset, value in enumerate(row_values):
                if col_offset in numeric_indexes:
                    try:
                        str_val = str(value).strip() if value is not None else ''
                        # Treat empty, None, or web-returned "NaN"/"nan" as 0.0
                        if str_val == '' or str_val.lower() == 'nan':
                            cell_value = 0.0
                        else:
                            cell_value = float(str_val)
                    except (ValueError, TypeError):
                        cell_value = 0.0  # safe fallback for any unexpected value
                else:
                    cell_value = value  # keep as string (Tag No, Product, Design, Sub Design, Calc Type)
                sheet.cell(row=row_idx, column=2 + col_offset, value=cell_value)
            row_idx += 1  # move to next Excel row

        workbook.save(FILE_PATH)
        print("✅ Data merged successfully!")
    
    
    def calculation(self,row_data,CalculationType,TotalAmount,Wt_gram,OtherMetalAmount,Board_rate):
       
        wait = self.wait 
        Nwt_val = wait.until(EC.presence_of_element_located((By.ID,"tag_nwt")))# TAG weight taken form UI
        Nwt = Nwt_val.get_attribute("value")
        Nwt = float(Nwt) if Nwt and Nwt.strip() else 0.0
        print(Nwt)
        
        Wast_val = wait.until(EC.presence_of_element_located((By.ID,"tag_wast_perc")))# Wastage % taken form UI
        Wast = Wast_val.get_attribute("value")
        Wast = float(Wast) if Wast and Wast.strip() else 0.0
        
        Mc_val = wait.until(EC.presence_of_element_located((By.ID,"tag_mc_value")))# Macking Cost value teken form UI
        Mc = Mc_val.get_attribute("value")
        Mc = float(Mc) if Mc and Mc.strip() else 0.0
        
        Mc_type_val = wait.until(EC.presence_of_element_located((By.ID,'tag_id_mc_type')))
        Mc_type = Mc_type_val.get_attribute("value")# Macking Cost value teken form UI
        print(Mc_type) 
        
        gross_weight =row_data["GWT"]# GWT Taken from Excel
        

        gross_weight=float(gross_weight)
        diamond_weight =float(Wt_gram)  
        net_weight = Nwt  
        wastage_percentage = Wast 
        Making_cost_pergram = Mc 
        diamond_cost =float(TotalAmount)
        ceil_value = "0.00"
        
        if CalculationType=="Mc on Gross,Wast On Net":
           # calculation making cost on gross Wastage% on Net  
            if Mc_type == '1':
                mc =Making_cost_pergram
            else:    
                Mc=Making_cost_pergram*gross_weight
                mc=float('{:.2f}'.format(Mc))
            Va = (wastage_percentage/100)*net_weight
            Va = round(Va, 3)
            total = net_weight+Va
            total = round(total, 3)
            Cal = (total*Board_rate)+diamond_cost+mc
            ceil_value=("{:.2f}".format(math.ceil(Cal)))
            
            
        if CalculationType=="Mc & Wast On Net":
            # calculation making cost & Wastage% on Net  
            if Mc_type == '1':
                mc =Making_cost_pergram
            else:    
                Mc=Making_cost_pergram*net_weight
                mc=float("{:.2f}".format(math.ceil(Mc)))
            Va = (wastage_percentage/100)*net_weight
            Va= round(Va, 3)
            total = net_weight+Va
            total = round(total, 3)
            Cal2 = total*Board_rate+mc+diamond_cost
            ceil_value=("{:.2f}".format(math.ceil(Cal2)))

                        
        if CalculationType== "Mc & Wast On Gross":
            # calculation making cost & Wastage% on Gross 81148.00
            if Mc_type == '1':
                mc =Making_cost_pergram
            else:    
                Mc=Making_cost_pergram*gross_weight
                Mc= gross_weight*Making_cost_pergram
                mc=float("{:.2f}".format(math.ceil(Mc)))
           
            Va = (wastage_percentage/100)*gross_weight
            Va = round(Va, 3)
            total= net_weight+Va
            total = round(total, 3)
            cal3 = total*Board_rate+mc+diamond_cost
            ceil_value=("{:.2f}".format(math.ceil(cal3)))
            
        if CalculationType == "Fixed Rate based on Weight":
            if Mc_type=='1':
                mc = Making_cost_pergram
            else:
                Mc=Making_cost_pergram*gross_weight
                mc=float('{:.2f}'.format(math.ceil(Mc)))
            Va = (wastage_percentage/100)*gross_weight
            Va = round(Va, 3)
            total= net_weight+Va
            total = round(total, 3)
            cal3 = total*Board_rate+mc+diamond_cost
            ceil_value=("{:.2f}".format(math.ceil(cal3)))
        
        if row_data["Calc Type"] == "Fixed Rate":
            ceil_value   
          
        if float(OtherMetalAmount or 0) > 0:
           Total =float(ceil_value) + float(OtherMetalAmount)
           return Total
                        
        else:
            print(ceil_value)
            print(type(ceil_value))
            return ceil_value
        
        
    
    
    def is_element_present(self, how, what):
        try: self.driver.find_element(by=how, value=what)
        except NoSuchElementException as e: return False
        return True
    
    def is_alert_present(self):
        try: self.driver.switch_to_alert()
        except NoAlertPresentException as e: return False
        return True
    
    def close_alert_and_get_its_text(self):
        try:
            alert = self.driver.switch_to_alert()
            alert_text = alert.text
            if self.accept_next_alert:
                alert.accept()
            else:
                alert.dismiss()
            return alert_text
        finally: self.accept_next_alert = True
    
    
    def update_branch_transfer(self, lot_no, branch):
        """
        Updates or appends a 'Tagged' row in the BranchTransfer sheet with the completed Lot Number.
        """
        try:
            workbook = load_workbook(FILE_PATH)
            if "BranchTransfer" not in workbook.sheetnames:
                print("⚠️ BranchTransfer sheet not found")
                return
            
            bt_sheet = workbook["BranchTransfer"]
            target_row = None
            last_tc_suffix = 0
            
            # Find last TC suffix and search for an available 'Tagged' row
            for row in range(2, 500):
                tc_id = bt_sheet.cell(row=row, column=1).value
                if not tc_id:
                    if target_row is None: target_row = row
                    break
                
                # Track max TC ID suffix (e.g. TC_BT_01 -> 1)
                if str(tc_id).startswith("TC_BT_"):
                    try:
                        match = re.search(r'\d+', str(tc_id))
                        if match:
                            suffix = int(match.group())
                            if suffix > last_tc_suffix:
                                last_tc_suffix = suffix
                    except:
                        pass
                
                # Use existing row if it is 'Tagged' but has no LotNo
                t_type = str(bt_sheet.cell(row=row, column=4).value or "").strip()
                l_no = bt_sheet.cell(row=row, column=8).value
                if t_type == "Tagged" and (l_no is None or str(l_no).strip() == ""):
                    target_row = row
                    break

            if target_row is None:
                target_row = bt_sheet.max_row + 1

            # Generate TC ID if row is new/empty
            if not bt_sheet.cell(row=target_row, column=1).value:
                next_id = f"TC_BT_{str(last_tc_suffix + 1).zfill(2)}"
                bt_sheet.cell(row=target_row, column=1).value = next_id
            
            # Fill Tagged transfer details
            bt_sheet.cell(row=target_row, column=4).value = "Tagged"
            
            # FromBranch defaults to HEAD OFFICE if not explicitly set in Tag sheet
            if not branch:
                branch = "HEAD OFFICE"
            
            bt_sheet.cell(row=target_row, column=5).value = branch
            
            # Inherit ToBranch from the previous row if current is empty
            if not bt_sheet.cell(row=target_row, column=6).value and target_row > 2:
                bt_sheet.cell(row=target_row, column=6).value = bt_sheet.cell(row=target_row-1, column=6).value
            
            bt_sheet.cell(row=target_row, column=8).value = lot_no
            
            workbook.save(FILE_PATH)
            print(f"✅ BranchTransfer updated row {target_row} with Lot: {lot_no}")
        except Exception as e:
            print(f"❌ BranchTransfer update error: {str(e)}")

    def tearDown(self):
        self.driver.quit()
        self.assertEqual([], self.verificationErrors)

if __name__ == "__main__":
    unittest.main()
