from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
import unittest
from Utils.Excel import ExcelUtils
from openpyxl import load_workbook
from time import sleep

FILE_PATH = ExcelUtils.file_path
class Tag_othermetal(unittest.TestCase):  
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 100)
        
    def test_othermetal(self,Sheet_name,test_case_id):
        driver = self.driver
        wait = self.wait
        
        function_name = Sheet_name
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, function_name)
        test_case_id = test_case_id
        value =ExcelUtils.Test_case_id_count(FILE_PATH, function_name,test_case_id)
        print(value)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[function_name]
        row=1
        count = value
        for row_num in range(2, valid_rows):
            current_id = sheet.cell(row=row_num, column=1).value  # Column 1 = Test Case Id
            if current_id == test_case_id:
                data = {
                        "Test Case Id": 1,
                        "Other Metals": 2,
                        "Metal": 3,
                        "Purity": 4,
                        "Nwt": 5,
                        "V.A(%)": 6,
                        "Mc Type": 7,
                        "Mc": 8,
                        "Rate": 9
                    }
                row_metaldata = {key: sheet.cell(row=row_num, column=col).value 
                            for key, col in data.items()}
                print(row_metaldata)   
                wait.until(EC.element_to_be_clickable((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td/select)".format(row)))).click()
                Select(wait.until(EC.element_to_be_clickable((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td/select)".format(row))))).select_by_visible_text(row_metaldata["Metal"])
               
                wait.until(EC.element_to_be_clickable((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[2]/select)".format(row)))).click()
                Select(wait.until(EC.element_to_be_clickable((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[2]/select)".format(row))))).select_by_visible_text(row_metaldata["Purity"])
              
                Nwt=wait.until(EC.element_to_be_clickable((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[3]/input)".format(row))))
                Nwt.clear()
                Nwt.send_keys(row_metaldata["Nwt"])
                
                V_A=wait.until(EC.element_to_be_clickable((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[4]/input)".format(row))))
                V_A.clear()
                V_A.send_keys(row_metaldata["V.A(%)"])
   
                McType=wait.until(EC.element_to_be_clickable((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[5]/select)".format(row))))
                McType.click()
                Select(McType).select_by_visible_text(row_metaldata["Mc Type"])
      
                Mc=wait.until(EC.element_to_be_clickable((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[6]/input)".format(row))))
                Mc.click()
                Mc.clear()
                Mc.send_keys(row_metaldata["Mc"])

                Rate=wait.until(EC.element_to_be_clickable((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[7]/input)".format(row))))
                Rate.click()
                Rate.clear()
                Rate.send_keys(row_metaldata["Rate"])

                if count != 1:
                    driver.find_element(By.XPATH,'//button[@class="btn btn-success add_other_metal"]').click()
                    row=row+1
                    count=value-1
                    sleep(4)
                else:
                    break      
            else:
                print(f'This row is not {test_case_id}') 
        TotalAmount = wait.until(EC.visibility_of_element_located((By.XPATH,"//table[@id='other_metal_table']/tfoot/tr/td[8]"))).text
        TotalAmount = float(TotalAmount)
        print(TotalAmount)                           
        wait.until(EC.element_to_be_clickable((By.ID,"update_other_metal_details"))).click()
        sleep(4)
        OtherMetal = "Other Metal detail Add successfully"       
        return OtherMetal,TotalAmount
             