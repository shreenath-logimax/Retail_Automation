import os
import sys
import datetime
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException
# Import established framework utilities
from Utils.Function import Function_Call
from Utils.Excel import ExcelUtils
from openpyxl import load_workbook
from openpyxl.styles import Font
import unittest

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class OrderLink(unittest.TestCase):
    """
    Automates the Order Link flow under Tagging > Order Link
    """
    def __init__(self, driver):
        super().__init__('test_order_link')
        self.driver = driver
        self.wait = WebDriverWait(driver, 15)
        self.fc = Function_Call(driver)


    def test_order_link(self):
        driver, wait = self.driver, self.wait

        # Navigate to Order Link
        print("🧭 Navigating to Order Link...")
        try:
            if "admin_ret_tagging/order_link" not in driver.current_url:
                wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
                sleep(1)
                self.fc.click("//span[contains(text(), 'Inventory')]")
                sleep(1)
                self.fc.click("(//span[contains(normalize-space(), 'Order Link')])") 
                sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            pass
        
        # Always force via URL to be perfectly safe as menu structures vary
        driver.get(BASE_URL + "index.php/admin_ret_tagging/tagging/tag_link")
        sleep(2)

        sheet_name = "OrderLink"
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 2} test cases in '{sheet_name}' sheet\n")
            self._resolve_tags(sheet_name, valid_rows)
        except Exception as e:
            print(f"❌ Failed to read valid rows from '{sheet_name}': {e}")
            return []

        for row_num in range(2, valid_rows):
            workbook = load_workbook(FILE_PATH)
            sheet    = workbook[sheet_name]

            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus":3,
                "Branch":     4, "FinYear":    5, "OrderNo":    6, 
                "TagNo":      7, "OldTagNo":   8, "Product":9,
                "Design":     10,"SubDesign":  11,"ExpectedMsg": 12, 
                "Remarks":    13
            }

            row_data = {
                key: sheet.cell(row=row_num, column=col).value
                for key, col in data_map.items()
            }
            workbook.close()

            print(f"\n{'='*80}")
            print(f"🧪 TC: {row_data.get('TestCaseId')}  |  Order: {row_data.get('OrderNo')}")
            print(f"{'='*80}")

            try:
                result = self._run_order_link(row_data)
                print(f"🏁 Result: {result[0]} — {result[1]}")
                self._update_excel_status(row_num, result[0], result[1], sheet_name)
                
                if result[0] == "Pass":
                    tag_no = str(row_data.get("TagNo", "")).strip()
                    order_no = str(row_data.get("OrderNo", "")).strip()
                    if tag_no and tag_no != "None":
                        self._update_detail_sheet_tag_status(tag_no, order_no)
                        
                    customer_mobile = result[2] if len(result) > 2 else ""
                    self._update_billing_sheet(row_data, customer_mobile)
            except Exception as e:
                print(f"❌ TC {row_data.get('TestCaseId')} exception: {e}")
                self._take_screenshot(f"EX_{row_data.get('TestCaseId')}")
                self._update_excel_status(row_num, "Fail", f"Exception: {str(e)}", sheet_name)

    def _run_order_link(self, row_data):
        """Core automation for a single row's Order Link test case."""
        driver, wait = self.driver, self.wait
        current_field = "Initialization"
        
        try:
            # Ensure we start on the correct page for each test case
            driver.get(BASE_URL + "index.php/admin_ret_tagging/tagging/tag_link")
            sleep(2)

            # 1. Select Branch
            branch = str(row_data.get("Branch", "")).strip()
            if branch:
                current_field = "Branch"
                # Using standard select2 locating mechanism
                self.fc.dropdown_select2(
                    '//select[@id="branch_select"]/following-sibling::span',
                    branch,
                    '//span[@class="select2-search select2-search--dropdown"]/input'
                )
                sleep(1)

            # 2. Select Financial Year
            fin_year = str(row_data.get("FinYear", "")).strip()
            if fin_year:
                current_field = "FinYear"
                fin_select_el = wait.until(EC.presence_of_element_located((By.XPATH, '//select[@id="order_fin_year_select"]')))
                Select(fin_select_el).select_by_visible_text(fin_year)
                sleep(1)

            # 3. Search and Select Order
            order_no = str(row_data.get("OrderNo", "")).strip()
            if not order_no:
                raise ValueError("OrderNo cannot be empty.")
            
            current_field = "Search Order"
            self.fc.dropdown_select2(
                '//select[@id="select_order"]/following-sibling::span',
                order_no,
                '//span[@class="select2-search select2-search--dropdown"]/input'
            )
            print(f"  -> Searched and selected Order: {order_no}")
            
            # Wait for table to dynamically load AJAX
            sleep(2)
            wait.until(EC.presence_of_element_located((By.XPATH, '//table[contains(@class,"dataTable") or contains(@class,"table")]/tbody/tr')))

            # 4. Table Interactions
            current_field = "Table Row Processing"
            self._take_screenshot(f"TableLoaded_{row_data.get('TestCaseId')}")

            # Make sure main checkbox is selected (Targeting first visual row)
            checkboxes = driver.find_elements(By.XPATH, '//table/tbody/tr//input[@type="checkbox"]')
            if checkboxes:
                cb = checkboxes[0]
                if not cb.is_selected():
                    driver.execute_script("arguments[0].click();", cb)
            
            tag_val = row_data.get("TagNo")
            tag_no = str(tag_val).strip() if tag_val is not None else ""
            
            old_tag_val = row_data.get("OldTagNo")
            old_tag_no = str(old_tag_val).strip() if old_tag_val is not None else ""

            if tag_no or old_tag_no:
                # Find all valid text inputs in the row
                row_inputs = driver.find_elements(By.XPATH, '//table/tbody/tr[1]//input[@type="text" or not(@type)]')
                
                # We assume Tag No is the first text input and Old Tag No is the second or specific by name
                if tag_no and len(row_inputs) >= 1:
                    xpath_tag_input = '(//table/tbody/tr[1]//input[@type="text" or not(@type)])[1]'
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", row_inputs[0])
                    self.fc.fill_autocomplete_field2(xpath_tag_input, tag_no)
                    print(f"  -> Autocompleted TagNo: {tag_no}")
                    
                    
                if old_tag_no and len(row_inputs) >= 2:
                    xpath_old_tag_input = '(//table/tbody/tr[1]//input[@type="text" or not(@type)])[2]'
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", row_inputs[1])
                    self.fc.fill_autocomplete_field2(xpath_old_tag_input, old_tag_no)
                    print(f"  -> Autocompleted OldTagNo: {old_tag_no}")
            
            # 5. Save All
            current_field = "Save All Button"
            sleep(1)
            save_btn = wait.until(EC.element_to_be_clickable((By.XPATH, '//button[contains(translate(text(),"ABCDEFGHIJKLMNOPQRSTUVWXYZ","abcdefghijklmnopqrstuvwxyz"), "save all")] | //button[@id="save_link"]')))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", save_btn)
            sleep(0.5)
            driver.execute_script("arguments[0].click();", save_btn)

            # 6. Verify Success
            current_field = "Success Message Banner"
            expected_msg = 'Successfully'
            
            try:
                banner = wait.until(EC.presence_of_element_located((By.XPATH, '//div[contains(@class, "alert-success")] | //div[@class="toast-message"]')))
                actual_msg = banner.text.strip().lower()
                print(f"  -> {actual_msg}")
                driver.execute_script("arguments[0].scrollIntoView(true);", banner)
                self._take_screenshot(f"Success_{row_data.get('TestCaseId')}")
                print(f"  -> Validated Banner: {actual_msg}")

                msg_matches = expected_msg.lower() 
                
                # 7. Verify Status in Report
                current_field = "Verify Order Status Report"
                try:
                    driver.get(BASE_URL + "index.php/admin_ret_reports/order_status/list")
                    sleep(2)
                    
                    branch = str(row_data.get("Branch", "")).strip()
                    if branch:
                        try:
                            self.fc.dropdown_select2(
                                '//select[contains(@id, "branch") or contains(@name, "branch")]/following-sibling::span | //span[contains(@id, "branch")]',
                                branch,
                                '//span[@class="select2-search select2-search--dropdown"]/input'
                            )
                            sleep(1)
                        except:
                            pass # Branch might default correctly, continue
                            
                    self.fc.click('//button[contains(., "Date range picker")]')
                    sleep(1)
                    self.fc.click('//li[contains(text(), "Last 30 Days")]')
                    sleep(1)
                    
                    self.fc.click('//button[@id="search" or text()="Search"]')
                    sleep(2)
                    
                    order_no = str(row_data.get("OrderNo", "")).strip()
                    search_box = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@type='search']")))
                    search_box.clear()
                    search_box.send_keys(order_no)
                    sleep(2)
                    
                    row_xpath = f"//table/tbody/tr[contains(., '{order_no}')]"
                    row_el = wait.until(EC.presence_of_element_located((By.XPATH, row_xpath)))
                    
                    customer_mobile = ""
                    try:
                        headers = driver.find_elements(By.XPATH, "//table/thead/tr/th")
                        for i, th in enumerate(headers, 1):
                            if "MOBILE" in th.text.upper():
                                customer_mobile = row_el.find_element(By.XPATH, f"./td[{i}]").text.strip()
                                print(f"  -> Extracted Customer Mobile: {customer_mobile}")
                                break
                    except Exception as e:
                        print(f"  -> Warning: Could not extract Customer Mobile ({e})")
                        
                    Web_status=  row_el.text.strip().lower()
                    status_ready = "delivery ready"                     
                    if msg_matches in actual_msg:
                        if status_ready in Web_status:
                            status=(f"  -> Order '{order_no}' Orderlink Done and status verified as 'Delivery Ready'.")
                            return ("Pass", status, customer_mobile)
                        else:
                            status=(f"  -> Order '{order_no}' Orderlink Done but status not 'Delivery Ready'.")
                            return ("Fail", status, customer_mobile)
                    else:
                        status=(f"  -> Order '{order_no}' Orderlink failed")
                        return ("Fail", status, customer_mobile)
                except Exception as e:
                    status=(f"  -> Order '{order_no}' Orderlink failed")
                    return ("Fail", status, "")
            except TimeoutException:
                status=(f"  -> Order '{order_no}' Orderlink failed")
                self._take_screenshot(f"NoBanner_{row_data.get('TestCaseId')}")
                return ("Fail", "Success banner did not appear after save.", "")

        except Exception as e:
            print(f"❌ Error at '{current_field}': {e}")
            self._take_screenshot(f"Error_{row_data.get('TestCaseId', 'Unknown')}")
            return ("Fail", f"Error in [{current_field}]: {str(e)}")

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name):
        """Write Pass/Fail + actual status back to Excel."""
        try:
            workbook = load_workbook(FILE_PATH)
            sheet    = workbook[sheet_name]
            color    = "00B050" if test_status == "Pass" else "FF0000"
            
            sheet.cell(row=row_num, column=2, value=test_status).font  = Font(bold=True, color=color)
            sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            
            workbook.save(FILE_PATH)
            workbook.close()
        except Exception as e:
            print(f"⚠️ Excel update error: {e}")

    def _take_screenshot(self, name):
        """Helper to capture full screenshots during critical execution states."""
        try:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"OrderLink_{name}_{timestamp}.png"
            os.makedirs("Screenshots/OrderLink", exist_ok=True)
            path = os.path.join("Screenshots/OrderLink", filename)
            self.driver.save_screenshot(path)
            print(f"📸 Screen saved: {filename}")
        except Exception as e:
            print(f"⚠️ Screenshot Failed: {e}")

    def _resolve_tags(self, sheet_name, valid_rows):
        """Resolves tags from Tag_Detail or Purchase_TagDetail using Product, Design, SubDesign."""
        self.used_tags_coords = getattr(self, 'used_tags_coords', [])
        wb = load_workbook(FILE_PATH)
        ol_sheet = wb[sheet_name]
        
        detail_sheets = {}
        for ds in ["Tag_Detail", "Purchase_TagDetail"]:
            if ds in wb.sheetnames:
                detail_sheets[ds] = wb[ds]

        today_str = datetime.datetime.now().strftime("%Y-%m-%d")

        updated = False
        for r in range(2, valid_rows):
            tag_no = str(ol_sheet.cell(row=r, column=7).value or "").strip()
            
            # If TagNo is empty, try to resolve
            if not tag_no or tag_no == "None":
                prod = str(ol_sheet.cell(row=r, column=9).value or "").strip().lower()
                dsgn = str(ol_sheet.cell(row=r, column=10).value or "").strip().lower()
                subd = str(ol_sheet.cell(row=r, column=11).value or "").strip().lower()
                order_no = str(ol_sheet.cell(row=r, column=6).value or "").strip()
                
                if not prod or not dsgn or prod == "none":
                    continue
                    
                matched = False
                for d_name, d_sheet in detail_sheets.items():
                    headers = {str(cell.value).strip().lower(): i for i, cell in enumerate(d_sheet[1], 1) if cell.value}
                    prod_col = headers.get("product")
                    dsgn_col = headers.get("design")
                    sub_col = headers.get("sub design") or headers.get("subdesign")
                    status_col = headers.get("status")
                    tag_col = headers.get("tag no")
                    
                    if not all([prod_col, dsgn_col, sub_col, status_col]) or not tag_col:
                        continue
                    
                    for dr in range(2, d_sheet.max_row + 1):
                        r_prod = str(d_sheet.cell(row=dr, column=prod_col).value or "").strip().lower()
                        r_dsgn = str(d_sheet.cell(row=dr, column=dsgn_col).value or "").strip().lower()
                        r_sub = str(d_sheet.cell(row=dr, column=sub_col).value or "").strip().lower()
                        r_status = str(d_sheet.cell(row=dr, column=status_col).value or "").strip().lower()
                        r_tag = str(d_sheet.cell(row=dr, column=tag_col).value or "").strip()
                        s_est_info = str(d_sheet.cell(row=dr, column=status_col + 1).value or "")
                        
                        if not r_tag or r_tag == "None":
                            continue
                            
                        if r_status in ["estimated", "billed", "tagreserve"]:
                            print(f"⏭️ Skipping Tag '{r_tag}' from {d_name} (Row {dr}) because status is '{r_status}'.")
                            continue
                        
                        if r_prod == prod and r_dsgn == dsgn and r_sub == subd:
                            if (d_name, dr) not in self.used_tags_coords:
                                self.used_tags_coords.append((d_name, dr))
                                ol_sheet.cell(row=r, column=7).value = r_tag
                                
                                updated = True
                                matched = True
                                print(f"✅ Found matching Tag '{r_tag}' for Row {r} from {d_name} (Row {dr}). Will mark 'Tagreserve' on Pass.")
                                break
                    if matched:
                        break
                
                if not matched:
                    print(f"⚠️ No matching tag found for Row {r} (Product: {prod}, Design: {dsgn})")
        
        if updated:
            wb.save(FILE_PATH)
        wb.close()

    def _update_detail_sheet_tag_status(self, tag_no, order_no):
        """Updates the tag status to 'Tagreserve' only after successful Order Link."""
        wb = load_workbook(FILE_PATH)
        updated = False
        for ds in ["Tag_Detail", "Purchase_TagDetail"]:
            if ds in wb.sheetnames:
                d_sheet = wb[ds]
                headers = {str(cell.value).strip().lower(): i for i, cell in enumerate(d_sheet[1], 1) if cell.value}
                tag_col = headers.get("tag no")
                status_col = headers.get("status")
                if not tag_col or not status_col:
                    continue
                    
                for dr in range(2, d_sheet.max_row + 1):
                    r_tag = str(d_sheet.cell(row=dr, column=tag_col).value or "").strip()
                    if r_tag == tag_no:
                        order_col = status_col + 1
                        if not d_sheet.cell(row=1, column=order_col).value:
                            d_sheet.cell(row=1, column=order_col).value = "Ordernumber"
                        
                        d_sheet.cell(row=dr, column=status_col).value = "Tagreserve"
                        d_sheet.cell(row=dr, column=order_col).value = order_no
                        updated = True
                        print(f"✅ Post-Pass: Updated {ds} for Tag '{tag_no}' -> 'Tagreserve'")
                        break
            if updated:
                break
                
        if updated:
            wb.save(FILE_PATH)
        wb.close()

    def _update_billing_sheet(self, row_data, customer_mobile):
        """Appends a new row to the Billing sheet after successful OrderLink."""
        sheet_name = "Billing"
        try:
            wb = load_workbook(FILE_PATH)
            if sheet_name not in wb.sheetnames:
                print(f"⚠️ Sheet '{sheet_name}' not found in Excel.")
                wb.close()
                return

            sh = wb[sheet_name]
            next_row = sh.max_row + 1
            
            def get_next_tc_id():
                last_id = sh.cell(row=next_row - 1, column=1).value
                if not last_id or not isinstance(last_id, str):
                    return "TC002" if next_row > 2 else "TC001"
                import re
                match = re.search(r"(\d+)", last_id)
                if match:
                    num_str = match.group(0)
                    new_num = int(num_str) + 1
                    return last_id.replace(num_str, str(new_num).zfill(len(num_str)))
                return "TC001"


            new_tc_id = get_next_tc_id()

            mappings = {
                1: new_tc_id,
                4: row_data.get("Branch", ""),
                5: "Customer",
                6: "111-Developer Logimax",
                7: customer_mobile,
                9: "Show Room",
                10: "ORDER DELIVERY",
                11: "No"
            }
            for col, value in mappings.items():
                if value is not None:
                    sh.cell(row=next_row, column=col, value=value).font = Font(bold=True)

            wb.save(FILE_PATH)
            wb.close()
            print(f"✅ Billing sheet appended at row {next_row} with ID {new_tc_id} and Mobile '{customer_mobile}'")
        except Exception as e:
            print(f"❌ Failed to append to Billing sheet: {e}")