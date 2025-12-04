from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import  sleep
import unittest
from Utils.Excel import ExcelUtils
from openpyxl.drawing.image import Image
from Test_vendor.Stone import stone_Price
from openpyxl import load_workbook
from openpyxl.styles import Font
from Utils.Function import Function_Call
import re


FILE_PATH = ExcelUtils.file_path
class Contract_Price(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)
        

    def test_contractprice(self,test_case_id):
        driver = self.driver
        wait = self.wait
        sleep(3)
        Sheet_name = 'Contract_Price'
        test_case_id = test_case_id
        value =ExcelUtils.Test_case_id_count(FILE_PATH, Sheet_name,test_case_id)
        print(value)
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, Sheet_name)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]
        row=1
        print(row)
        count = value
        for row_num in range(2, valid_rows):
            current_id = sheet.cell(row=row_num, column=1).value  # Column 1 = Test Case Id
            if current_id == test_case_id:
                data = {
                    "Test Case Id": 1,
                    "Test Status": 2,
                    "Actual Status": 3,
                    "metal_type": 4,
                    "Category": 5,
                    "Purity": 6,
                    "Product": 7,
                    "Design": 8,
                    "subDesign": 9,
                    "Kar_Calc_Type": 10,
                    "Touch": 11,
                    "Calc_Type": 12,
                    "V.A_Type": 13,
                    "V.A(%)": 14,
                    "V.A_Wgt": 15,
                    "MC_Type": 16,
                    "Mc": 17,
                    "Charges": 18,
                    "Char_Calc Type":19,
                    "Charge_Amt":20,
                    "Stone Calc Type": 21,
                    "UOM": 22,
                    "Quality Code": 23,
                    "From Cent": 24,
                    "To Cent": 25,
                    "Rate": 26,
                    "Image": 27,
                    "Action": 28,
                    "Stone":29,
                    "Remove":30,
                    "Field_validation_status":31
                }
                row_data = {
                    key: sheet.cell(row=row_num, column=col).value
                    for key, col in data.items()
                }
                print(row_data)
                
                if 1<row:
                   before_purity = sheet.cell(row=row_num, column=6).value
                   before_subdesign = sheet.cell(row=row_num, column=9).value
                else:
                    before_purity = None
                    before_subdesign =None
                
                # Call your 'create' method
                Create_data = Contract_Price.create(self,row_data, row_num, Sheet_name, row, count,before_purity,before_subdesign)
                print(Create_data)
                # Remove processed customer from the list
                
                if isinstance(Create_data, int):
                    count=Create_data
                    row=row+1
                    continue
                if not isinstance(Create_data, int):
                    Test_Status,Actual_Status= Create_data
                    row=1
                    self.update_excel_status(row_num, Test_Status, Actual_Status, Sheet_name)
                    
                
    def create(self, row_data, row_num, Sheet_name, row, count,before_purity,before_subdesign):      
        driver, wait = self.driver, self.wait
        Mandatory_field=[]   
        Error_field_val=[]        
        Function_Call.click(self,'//span[@class="add_wastage"]') 
        # category 
        if row_data["Category"] is not None:
            if row >= 2:
                driver.execute_script("window.scrollBy(700, 0);")  # scroll left by 300px
            Function_Call.dropdown_select(self,f'//*[@id="select2-category{row}-container"]', row_data["Category"],"/html/body/span/span/span[1]/input")
        else:
            msg = f"'{None}' → Category field is mandatory ⚠️"
            Mandatory_field.append("Category"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        if row_data["Purity"] is not None:
            sleep(2)
            Function_Call.dropdown_select(self,f'//*[@id="select2-purity{row}-container"]', row_data["Purity"],"/html/body/span/span/span[1]/input")
        else:
            msg = f"'{None}' → Purity field is mandatory ⚠️"
            Mandatory_field.append("Purity"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        if row_data["Product"] is not None:
            Function_Call.dropdown_select(self,f'//*[@id="select2-product{row}-container"]', row_data["Product"],"/html/body/span/span/span[1]/input")
        else:
            msg = f"'{None}' → Product field is mandatory ⚠️"
            Mandatory_field.append("Product"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
                
        # Design 
        if row_data["Design"] is not None:
            Function_Call.dropdown_select(self,f'//*[@id="select2-design{row}-container"]', row_data["Design"],"/html/body/span/span/span[1]/input")
        else:
            msg = f"'{None}' → Design field is mandatory ⚠️"
            Mandatory_field.append("Design"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        # subDesign 
        if row_data["subDesign"] is not None:
            purity=row_data["Purity"]
            error=Function_Call.dropdown_subdesign_val(
                self,f'//*[@id="select2-sub_design{row}-container"]', 
                row_data["subDesign"],
                '//span[@class="select2-search select2-search--dropdown"]/input',
                purity,before_purity,before_subdesign)
            if error:
                msg = (f"⚠️ Found the message:'{error}'")
                Function_Call.Remark(self,row_num, msg, Sheet_name)
                Error_field_val.append("Subdesign")
                print(Error_field_val)
        else:
            msg = f"'{None}' → subDesign field is mandatory ⚠️"
            Mandatory_field.append("subDesign"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
    
        if row_data["metal_type"]=='Gold':
              
            # Kar_Calc_Type 
            if row_data["Kar_Calc_Type"] is not None:
                Function_Call.dropdown_select(self,f"//span[starts-with(@id,'select2-o_item[{row}]') and contains(@id,'[kar_calc_type]')]", row_data["Kar_Calc_Type"],"/html/body/span/span/span[1]/input")
            else:
                msg = f"'{None}' → Kar_Calc_Type field is mandatory ⚠️"
                Mandatory_field.append("Kar_Calc_Type"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
                    
            if row_data["Kar_Calc_Type"]=='Purchase Touch':
                if row_data["Touch"]:
                    errors=Function_Call.fill_input(
                        self,wait,
                        locator=(By.XPATH, f'//input[@name="o_item[{row}][pur_touch]"]'),
                        value=row_data["Touch"],
                        #Allow exactly 10 digits (mobile number)
                        pattern = r"^\d{2}\.\d{2}$",
                        field_name="Touch",
                        screenshot_prefix="Touch",
                        range_check = lambda v: 10 <= float(v) <= 120,
                        row_num=row_num,
                        Sheet_name=Sheet_name
                    )
                    Error_field_val.extend(errors)
                    print(Error_field_val)
                else:
                    msg = f"'{None}' → Touch field is mandatory ⚠️"
                    Mandatory_field.append("Touch"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
               
                
            # Calc_Type 
            if row_data["Calc_Type"] is not None:
                Function_Call.dropdown_select(self,f"//span[starts-with(@id,'select2-o_item[{row}]') and contains(@id,'[calc_type]')]", row_data["Calc_Type"],"/html/body/span/span/span[1]/input")
            else:
                msg = f"'{None}' → Calc_Type field is mandatory ⚠️"
                Mandatory_field.append("Calc_Type"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
                    
            # V.A_Type 
            if row_data["Kar_Calc_Type"]!='Purchase Touch':
                if row_data["V.A_Type"] is not None:
                    
                    Function_Call.dropdown_select(self,f"//span[starts-with(@id,'select2-o_item[{row}]') and contains(@id,'[va_type]')]", row_data["V.A_Type"],"/html/body/span/span/span[1]/input")
                else:
                    Function_Call.click(self,f"//span[starts-with(@id,'select2-o_item[{row}]') and contains(@id,'[va_type]')]/span") 
                    msg = f"'{None}' → V.A_Type field is mandatory ⚠️"
                    Mandatory_field.append("V.A_Type"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
                
                
                if row_data["V.A_Type"]=="Percentage":    
                    if row_data["V.A(%)"]is not None:
                        errors=Function_Call.fill_input(
                            self,wait,
                            locator=(By.XPATH, f'//input[@name="o_item[{row}][wast_percent]"]'),
                            value=row_data["V.A(%)"],
                            #Allow exactly 10 digits (mobile number)
                            pattern=r"^\d+(\.\d{1,2})?$",
                            field_name="V.A(%)",
                            screenshot_prefix="V.A(%)",
                            range_check = lambda v: 0 <= float(v) <= 100,
                            row_num=row_num,
                            Sheet_name=Sheet_name
                        )
                        Error_field_val.extend(errors)
                        print(Error_field_val)
                    else:
                        msg = f"'{None}' → V.A(%) field is mandatory ⚠️"
                        Mandatory_field.append("V.A(%)"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
                        
                if row_data["V.A_Type"]=="Weight":
                    if row_data["V.A_Wgt"]is not None:
                        errors=Function_Call.fill_input(
                            self,wait,
                            locator=(By.XPATH, f'//input[@name="o_item[{row}][wast_wgt]"]'),
                            value=row_data["V.A_Wgt"],
                            #Allow exactly 10 digits (mobile number)
                            pattern=r"^\d+(\.\d{1,2})?$",
                            field_name="V.A_Wgt",
                            screenshot_prefix="V.A_Wgt",
                            range_check = lambda v: 0 <= float(v) <= 100,
                            row_num=row_num,
                            Sheet_name=Sheet_name
                        )
                        Error_field_val.extend(errors)
                        print(Error_field_val)
                    else:
                        msg = f"'{None}' → V.A_Wgt field is mandatory ⚠️"
                        Mandatory_field.append("V.A_Wgt"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
        
                # MC_Type 
                if row_data["MC_Type"] is not None:
                    Function_Call.dropdown_select(self,f"//span[starts-with(@id,'select2-o_item[{row}]') and contains(@id,'[id_mc_type]')]", row_data["MC_Type"],"/html/body/span/span/span[1]/input")
                else:
                    Function_Call.click(self,f"//span[starts-with(@id,'select2-o_item[{row}]') and contains(@id,'[id_mc_type]')]/span")
                    msg = f"'{None}' → MC_Type field is mandatory ⚠️"
                    Mandatory_field.append("MC_Type"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)    
                    
                # Mc
                if row_data["Mc"]is not None:
                    errors=Function_Call.fill_input(
                        self,wait,
                        locator=(By.XPATH, f'//input[@name="o_item[{row}][mc]"]'),
                        value=row_data["Mc"],
                        #Allow exactly 10 digits (mobile number)
                        pattern=r"^\d+(\.\d{1,2})?$",
                        field_name="Mc",
                        screenshot_prefix="Mc",
                        row_num=row_num,
                        Sheet_name=Sheet_name
                    )
                    Error_field_val.extend(errors)
                    print(Error_field_val)
                else:
                    msg = f"'{None}' → Mc field is mandatory ⚠️"
                    Mandatory_field.append("Mc"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
                    
        if row_data["metal_type"]=="Stone" or row_data["metal_type"]=="Diamond": 
        
            if row_data["Stone Calc Type"] is not None:
                Function_Call.dropdown_select(self,f"//span[starts-with(@id,'select2-o_item[{row}]') and contains(@id,'[stn_calc_type]')]", row_data["Stone Calc Type"],'//span[@class="select2-search select2-search--dropdown"]/input')
            else:
                msg = f"'{None}' → Stone Calc Type field is mandatory ⚠️"
                Mandatory_field.append("Stone Calc Type"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name) 
                
            if row_data["UOM"] is not None:
                Function_Call.dropdown_select(self,f"//span[@id='select2-prod_uom_id{row}-container']", row_data["UOM"],'//span[@class="select2-search select2-search--dropdown"]/input')
            else:
                msg = f"'{None}' → UOM field is mandatory ⚠️"
                Mandatory_field.append("UOM"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)  
        
        if row_data["metal_type"]=="Diamond":
                
            if row_data["Quality Code"] is not None:
                Function_Call.dropdown_select(self,f"//span[@id='select2-prod_stn_quality{row}-container']", row_data["Quality Code"],'//span[@class="select2-search select2-search--dropdown"]/input')
            else:
                msg = f"'{None}' → Quality Code field is mandatory ⚠️"
                Mandatory_field.append("Quality Code"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)  
                
            # From Cent	
            if row_data["From Cent"]is not None:
                errors=Function_Call.fill_input(
                    self,wait,
                    locator=(By.XPATH, f'//input[@name="o_item[{row}][from_wt]"]'),
                    value=row_data["From Cent"],
                    #Allow exactly 10 digits (mobile number)
                    pattern=r"^\d+(\.\d{1,2})?$",
                    field_name="From Cent",
                    screenshot_prefix="From Cent",
                    row_num=row_num,
                    Sheet_name=Sheet_name
                )
                Error_field_val.extend(errors)
                print(Error_field_val)
            else:
                msg = f"'{None}' → From Cent field is mandatory ⚠️"
                Mandatory_field.append("From Cent"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
                
            # To  Cent	
            if row_data["To Cent"]is not None:
                errors=Function_Call.fill_input(
                    self,wait,
                    locator=(By.XPATH, f'//input[@name="o_item[{row}][to_wt]"]'),
                    value=row_data["To Cent"],
                    #Allow exactly 10 digits (mobile number)
                    pattern=r"^\d+(\.\d{1,2})?$",
                    field_name="To Cent",
                    screenshot_prefix="To Cent",
                    row_num=row_num,
                    Sheet_name=Sheet_name
                )
                Error_field_val.extend(errors)
                print(Error_field_val)
            else:
                msg = f"'{None}' → To Cent field is mandatory ⚠️"
                Mandatory_field.append("To Cent"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
       
        if row_data["metal_type"]=="Stone" or row_data["metal_type"]=="Diamond":        
            # Rate
            if row_data["Rate"]is not None:
                errors=Function_Call.fill_input(
                    self,wait,
                    locator=(By.XPATH, f'//input[@name="o_item[{row}][stone_rate]"]'),
                    value=row_data["Rate"],
                    #Allow exactly 10 digits (mobile number)
                    pattern=r"^\d+(\.\d{1,2})?$",
                    field_name="Rate",
                    screenshot_prefix="Rate",
                    row_num=row_num,
                    Sheet_name=Sheet_name
                )
                Error_field_val.extend(errors)
                print(Error_field_val)
            else:
                msg = f"'{None}' → Rate field is mandatory ⚠️"
                Mandatory_field.append("Rate"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
        remove_map = {
            "Category": f'//*[@id="select2-category{row}-container"]/span',
            "Purity": f'//*[@id="select2-purity{row}-container"]/span',
            "Product": f'//*[@id="select2-product{row}-container"]/span',
            "Design": f'//*[@id="select2-design{row}-container"]/span',
            "subDesign": f'//*[@id="select2-sub_design{row}-container"]/span'
        }
        
        remove=row_data["Remove"]
        if remove:
            xpath = remove_map[remove]
            if xpath:
                close=wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
                # Scroll element into view (works for left/right as well as up/down)
                driver.execute_script("arguments[0].scrollIntoView({block: 'nearest', inline: 'center'});", close)
                close.click()
                Mandatory_field.append(remove)
                
        print(Mandatory_field)
        print(Error_field_val)
        failed_fields = []
        if Mandatory_field:
            failed_fields.extend(Mandatory_field)
        if Error_field_val:
            failed_fields.extend(Error_field_val)
    
        if failed_fields:           
            error = Function_Call.alert1(self, '//span[@class="add_wastage"]')
            if error:
                Test_Status="WARANING"
                Actual_Status= (f"⚠️ Found the message:'{error}'") # prints: Select Order Branch
                print(Actual_Status)
                Function_Call.click(self,'(//button[@class="btn btn-default btn-cancel"])[2]')
            else:
                Test_Status="Fail"
                Actual_Status= (f"'{failed_fields}'-> This field is None, but the vendor was saved successfully ❌") 
                Function_Call.click(self,'(//button[@class="btn btn-default btn-cancel"])[2]')
            return Test_Status,Actual_Status 
        if count > 1:
            row += 1
            count -= 1
            return count
        else:
            #submit
            print(row_data["Stone"])
            if row_data["Stone"]=='Yes':
                result = Function_Call.alert1(self, '//button[@id="add_contractPriceNext"]')
                if result:
                    Test_Status="Pass"
                    Actual_Status= (f"✅ Found the message:'{result}'")
                    print(Actual_Status)
                else:
                    Test_Status="Fail"
                    Actual_Status= (f"'{failed_fields}'-> This field is None, but the vendor was saved successfully ❌")
                self.update_excel_status(row_num, Test_Status, Actual_Status, Sheet_name)
                test_case_id=row_data['Test Case Id']
                stone_Price.test_stoneprice(self,test_case_id)
                return
            else:
                result = Function_Call.alert1(self, '//button[@id="add_NewcontractPrice"]')
                if result:
                    Test_Status="Pass"
                    Actual_Status= (f"✅ Found the message:'{result}'")
                    print(Actual_Status)
                else:
                    Test_Status="Fail"
                    Actual_Status=('New contract Price page save not successfully')
                self.update_excel_status(row_num, Test_Status, Actual_Status, Sheet_name)
                

    def update_excel_status(self,row_num, Test_Status, Actual_Status, Sheet_name):
        print(Sheet_name)
        # Load the workbook
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]  # or workbook["SheetName"]
        
        if Test_Status== 'Pass':
            # Write Test_Status into column 2
            sheet.cell(row=row_num, column=2, value=Test_Status).font=Font(bold=True, color="00B050")
            
            # Write Actual_Status in col 3 
            sheet.cell(row=row_num, column=3, value=Actual_Status).font = Font(bold=True, color="00B050")
        if Test_Status=='Fail':
            # Write Test_Status into column 2
            sheet.cell(row=row_num, column=2, value=Test_Status).font=Font(bold=True, color="FF0000")
            # Write Actual_Status in col 3 
            sheet.cell(row=row_num, column=3, value=Actual_Status).font = Font(bold=True, color="FF0000")
        # Save workbook
        workbook.save(FILE_PATH)
        # Get status from ExcelUtils
        Status = ExcelUtils.get_Status(FILE_PATH, Sheet_name)
        # Print and return status
        print(Status)
                                 
                
                    
                
            
                        
                        
  