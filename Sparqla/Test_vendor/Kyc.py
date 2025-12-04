from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import  sleep
import unittest
from Utils.Excel import ExcelUtils
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font
from Utils.Function import Function_Call
import re


FILE_PATH = ExcelUtils.file_path
class KYC(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 10)

    def test_Kyc(self,test_case_id):
        driver = self.driver
        wait = self.wait
        sleep(3)
        Sheet_name = 'Kyc'
        test_case_id = test_case_id
        value =ExcelUtils.Test_case_id_count(FILE_PATH, Sheet_name,test_case_id)
        print(value)
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, Sheet_name)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]
        row=1
        count = value
        for row_num in range(2, valid_rows):
            current_id = sheet.cell(row=row_num, column=1).value  # Column 1 = Test Case Id
            if current_id == test_case_id:
                data = {
                    "Test Case Id": 1,
                    "Test Status": 2,
                    "Actual Status": 3,
                    "Bank":4,
                    "Bank Name": 5,
                    "Account Holder Name": 6,
                    "Account Number": 7,
                    "IFSC Code": 8,
                    "KYC": 9,
                    "Proof Name": 10,
                    "ID": 11,
                    "ImagesFront": 12,
                    "ImagesBack": 13,
                    "Document": 14,
                    "Field_validation_status": 15,
                }
                row_data = {
                    key: sheet.cell(row=row_num, column=col).value
                    for key, col in data.items()
                }
                print(row_data)
                Create_data=KYC.create(self, row_data, row_num, Sheet_name, row, count)
                print(Create_data)
                
                if Create_data is int:
                    Create_data = count
                
                if Create_data:
                    Test_Status,Actual_Status= Create_data
                    row=1
                    self.update_excel_status(row_num, Test_Status, Actual_Status, Sheet_name)
                    break
                
    def create(self, row_data, row_num, Sheet_name, row, count):  
        driver, wait = self.driver, self.wait         
        Mandatory_field=[]
        Error_field_val=[]
        if row_data["Bank"] == "Yes":
            Function_Call.click(self,'//span[@class="add_bank"]')
            if row_data["Bank Name"]is not None:
                errors=Function_Call.fill_input(
                    self,wait,
                    locator=(By.XPATH, f'//input[@id="bank_name{row}"]'),
                    value=row_data["Bank Name"],
                    #Allow exactly 10 digits (mobile number)
                    pattern = r"^[A-Za-z\s]{3,50}$",
                    field_name="Bank Name",
                    screenshot_prefix="Bank Name",
                    row_num=row_num,
                    Sheet_name=Sheet_name
                )
                Error_field_val.extend(errors)
                print(Error_field_val)
            else:
                msg = f"'{None}' → Bank Name field is mandatory ⚠️"
                Mandatory_field.append(msg); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)

            if row_data["Account Holder Name"]is not None:
                errors=Function_Call.fill_input(
                    self,wait,
                    locator=(By.XPATH, f'//input[@id="acc_holder{row}"]'),
                    value=row_data["Account Holder Name"],
                    #Allow exactly 10 digits (mobile number)
                    pattern = r"^\d{2}\.\d{2}$",
                    field_name="Account Holder Name",
                    screenshot_prefix="Account Holder Name",
                    row_num=row_num,
                    Sheet_name=Sheet_name
                )
                Error_field_val.extend(errors)
                print(Error_field_val)
            else:
                msg = f"'{None}' → Account Holder Name field is mandatory ⚠️"
                Mandatory_field.append(msg); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        
            if row_data["Account Number"]is not None:
                errors=Function_Call.fill_input(
                    self,wait,
                    locator=(By.XPATH, f'//input[@id="acc_number{row}"]'),
                    value=row_data["Account Number"],
                    #Allow exactly 10 digits (mobile number)
                    pattern = r"^\d{9,18}$",
                    field_name="Account Number",
                    screenshot_prefix="Account Number",
                    row_num=row_num,
                    Sheet_name=Sheet_name
                )
                Error_field_val.extend(errors)
                print(Error_field_val)
            else:
                msg = f"'{None}' → Account Number field is mandatory ⚠️"
                Mandatory_field.append(msg); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
            
            if row_data["IFSC Code"]is not None:
                errors=Function_Call.fill_input(
                    self,wait,
                    locator=(By.XPATH, f'//input[@id="ifsc_code{row}"]'),
                    value=row_data["IFSC Code"],
                    #Allow exactly 10 digits (mobile number)
                    pattern = r"^[A-Z]{4}0[A-Z0-9]{6}$",
                    field_name="IFSC Code",
                    screenshot_prefix="IFSC Code",
                    row_num=row_num,
                    Sheet_name=Sheet_name
                )
                Error_field_val.extend(errors)
                print(Error_field_val)
            else:
                msg = f"'{None}' → IFSC Code field is mandatory ⚠️"
                Mandatory_field.append(msg); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
    
        if row_data["KYC"]=="Yes":
            Function_Call.click(self,'//span[@class="add_kyc"]')
            # Code 
            if row_data["Proof Name"] is not None:
                Function_Call.dropdown_select(self,f'//*[@id="select2-proof_name{row+1}-container"]', row_data["Proof Name"],"/html/body/span/span/span[1]/input")
            else:
                msg = f"'{None}' → Proof Name field is mandatory ⚠️"
                Mandatory_field.append(msg); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
            # Dynamic pattern selection based on Proof Name
            proof_name = row_data["Proof Name"].lower()
            pattern = None

            if proof_name == "pan":
                pattern = r"^[A-Z]{5}[0-9]{4}[A-Z]$"
            elif proof_name == "aadhar":
                pattern = r"^[2-9]{1}[0-9]{11}$"
            elif proof_name == "gst":
                pattern = r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[A-Z0-9]{1}Z[A-Z0-9]{1}$"
            elif proof_name == "company reg cert":
                pattern = r"^[LU]{1}[0-9A-Z]{5}[0-9]{2}[0-9]{4}[A-Z]{3}[0-9]{6}$"
            
            if row_data["ID"]is not None:
                errors=Function_Call.fill_input(
                    self,wait,
                    locator=(By.XPATH, f'//input[@name="kyc_det[{row+1}][id_reg_exp]"]'),
                    value=row_data["ID"],
                    #Allow exactly 10 digits (mobile number)
                    pattern = pattern,
                    field_name="ID",
                    screenshot_prefix="ID",
                    row_num=row_num,
                    Sheet_name=Sheet_name
                )
                Error_field_val.extend(errors)
                print(Error_field_val)
            else:
                msg = f"'{None}' → ID field is mandatory ⚠️"
                Mandatory_field.append(msg); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
                
            ImagesFront = row_data["ImagesFront"].strip()
            Path = r"D:\Retail_Testing\Image_all_Format"
            Image_path = fr"{Path}\{ImagesFront}"
            print (Image_path)
                
            if row_data["ImagesFront"]is not None:
                Image_fr = row_data["ImagesFront"].upper()
                Function_Call.click(self,'//a[@class="btn btn-default btn-sm fr_img"]')
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.ID, "order_images_front"))).send_keys(Image_path)
               
        
                # Function_Call.Image_upload(self,'//input[@id="order_images_front"]', Image_path)
                if "JPG" in Image_fr or "PNG" in Image_fr:
                   Function_Call.click(self,'//button[@id="update_img_front"]')
                else:
                    Function_Call.alert(self)   
                    Function_Call.click(self,'(//button[@id="close_stone_details"])[2]')
            else:
                msg = f"'{None}' → ImagesFront field is mandatory ⚠️"
                Mandatory_field.append(msg); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
            ImagesBack = row_data["ImagesBack"]
            Path = r"D:\Retail_Testing\Image_all_Format"
            Image_path = f"{Path}\{ImagesBack}"
                
            if row_data["ImagesBack"]is not None:
                Image_Ba=row_data["ImagesBack"].upper()
                Function_Call.click(self,'//a[@class="btn btn-default btn-sm bk_img"]')
                Function_Call.Image_upload(self,'//input[@id="order_images_back"]',Image_path)
                if "JPG" in Image_Ba or "PNG" in Image_Ba:
                    Function_Call.click(self,'//button[@id="update_img_back"]')
                else:
                    Function_Call.alert(self)
                    Function_Call.click(self,'(//button[@id="close_stone_details"])[3]')
                    
            else:
                msg = f"'{None}' → ImagesBack field is mandatory ⚠️"
                Mandatory_field.append(msg); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
        
        if Mandatory_field or Error_field_val:           
            error = Function_Call.alert1(self, '//button[@id="add_NewKyc"]')
            if error is not None:
                Test_Status="Pass"
                Actual_Status= (f"⚠️ Found the message:'{error}'") 
                print(error)
                wait.until(EC.element_to_be_clickable((By.XPATH,'(//button[@class="btn btn-default btn-cancel"])[4]'))).click()
                
            else:
                Test_Status="Fail"
                Actual_Status= (f"'{Mandatory_field}' but Kyc save successfully") 
            return Test_Status,Actual_Status 
                
                   
        if count > 1:
            row += 1
            count -= 1
            return count
        else:
            #submit
            Function_Call.click(self,'//button[@id="add_NewKyc"]')
            sleep(3)
            added = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[1]/section/div/div/div/div[2]/div[1]"))).text
            added = added.replace("×", "").replace("\n", "")
            expected_Result = 'Add User!New User added successfully!..'
                       
    def update_excel_status(self,row_num, Test_Status, Actual_Status, Sheet_name):
        print(Sheet_name)
        # Load the workbook
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]  # or workbook["SheetName"]
        
        if Test_Status== 'Pass':
            # Write Test_Status into column 2
            sheet.cell(row=row_num, column=2, value=Test_Status).font=Font(bold=True, color="00B050")
            
            # Write Actual_Status in col 3 
            sheet.cell(row=row_num, column=3, value=Actual_Status).font = Font(bold=True, color="00B050")
        if Test_Status=='Fail':
            # Write Test_Status into column 2
            sheet.cell(row=row_num, column=2, value=Test_Status).font=Font(bold=True, color="FF0000")
            # Write Actual_Status in col 3 
            sheet.cell(row=row_num, column=3, value=Actual_Status).font = Font(bold=True, color="FF0000")
        # Save workbook
        workbook.save(FILE_PATH)
        # Get status from ExcelUtils
        Status = ExcelUtils.get_Status(FILE_PATH, Sheet_name)
        # Print and return status
        print(Status)
                                       
