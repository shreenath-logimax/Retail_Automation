from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import  sleep
import unittest
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from Test_vendor .Contract_price import Contract_Price
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font
import re

FILE_PATH = ExcelUtils.file_path
class VendorRegistration(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)

    def test_vendor_registration(self):
        driver = self.driver
        wait = self.wait
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Toggle navigation"))).click()
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Purchase Module"))).click()
        # Locate the sidebar (left navigation panel)
        # sidebar = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "main-sidebar"))) # adjust if needed
        # # Scroll inside the sidebar
        # driver.execute_script(f"window.scrollBy(0, 400);",sidebar)
        # driver.execute_script("arguments[0].scrollTop = 600", sidebar)
        
        vendor=wait.until(EC.invisibility_of_element_located((By.XPATH,"//span[contains(text(), 'Vendor Registration')]")))
        driver.execute_script("arguments[0].scrollIntoView({block: 'nearest', inline: 'center'});", vendor)
        vendor.click()
        
        Sheet_name = "Vendor"                                        
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, Sheet_name)
        print(f"'{valid_rows}': valid rows")
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]
        for row_num in range(2, valid_rows):  
            data = {
                    "TestCaseId": 1,
                    "TestStatus": 2,
                    "ActualStatus": 3,
                    "userType": 4,
                    "usedFor": 5,
                    "calcType": 6,
                    "TCS": 7,
                    "TCS%": 8,
                    "TDS": 9,
                    "TDS%": 10,
                    "openingBalanceAmount": 11,
                    "financialYear": 12,
                    "firstName": 13,
                    "lastName": 14,
                    "mobileNumber": 15,
                    "email": 16,
                    "phoneNumber": 17,
                    "userName": 18,
                    "password": 19,
                    "address1": 20,
                    "address2": 21,
                    "address3": 22,
                    "country": 23,
                    "state": 24,
                    "city": 25,
                    "pincode": 26,
                    "Remarks":27,
                    "Image":28,
                    "companyName": 29,
                    "gstNumber": 30,
                    "Contract Price": 31,
                }
            row_data = {key: sheet.cell(row=row_num, column=col).value 
                            for key, col in data.items()}
            print(row_data)
            # Call you 'create' method
            Create_data = self.create(row_data, row_num, Sheet_name)
            print(Create_data)
            # Remove processed customer from the list
            if Create_data:
                Test_Status,Actual_Status= Create_data
                self.update_excel_status(row_num, Test_Status, Actual_Status, Sheet_name)
                
            
        
    def create(self, row_data, row_num, Sheet_name):        
        driver, wait = self.driver, self.wait
        Mandatory_field=[]
        Error_field_val = []
        Function_Call.click(self,'//a[@id="karigar_add"]')
        
        #User Type
        user_type_map = {
                "company": '//*[@id="click_label_company"]',
                "individual": '//*[@id="click_label_individual"]'
            }
        Function_Call.click(self,user_type_map[row_data["userType"]])
        
          
        # usedFor 
        if row_data["usedFor"] is not None:
           Function_Call.dropdown_select(self,'//span[@id="select2-karigar_for-container"]', row_data["usedFor"],'/html/body/span[1]/span/span[1]/input')
           
        else:
            Function_Call.click(self,'//span[@id="select2-karigar_for-container"]/span')
            Function_Call.click(self,'//span[@id="select2-karigar_for-container"]')
            msg = f"'{None}' → usedFor field is mandatory ⚠️"
            Mandatory_field.append("usedFor"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            

        # TCS
        if row_data["TCS"] == "Yes":
           Function_Call.click(self,'//*[@id="click_label_tcsyes"]') 
        else :
            Function_Call.click(self,'//*[@id="click_label_tcsno"]') 
        
        print(Sheet_name)
        if row_data["TCS"]=="Yes":
            if row_data["TCS%"] is not None:
                errors=Function_Call.fill_input(self,
                    wait,
                    locator=(By.XPATH, '//*[@id="tcs"]'),
                    value=row_data["TCS%"],
                    pattern=r"^\d+(\.\d{1,2})?$",
                    field_name="TCS%",
                    screenshot_prefix="TCS%",
                    range_check=lambda v: 0 <= float(v) <= 100,
                    row_num=row_num,
                    Sheet_name=Sheet_name)
                Error_field_val.extend(errors)
                print(Error_field_val)
            else:
                msg = f"'{None}' → TCS% field is mandatory ⚠️"
                Mandatory_field.append("TCS%"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        
        
        # Opening balance
        if row_data["openingBalanceAmount"] is not None:
            Function_Call.fill_input2(self,'//*[@id="opening_bal_amt"]', row_data["openingBalanceAmount"])
        else:
            msg = f"'{None}' → openingBalance field is mandatory ⚠️"
            Mandatory_field.append("openingBalance"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
       
        # Financial year
        Select(wait.until(EC.element_to_be_clickable((By.ID, "order_fin_year_select")))
        ).select_by_visible_text(row_data["financialYear"])
        
         # TDS
        Function_Call.click(self,'//*[@id="click_label_tdsyes"]' if row_data["TDS"] == "Yes" else '//*[@id="click_label_tdsno"]')
        if row_data["TDS"] == "Yes":
            if row_data["TDS%"] is not None:
                errors=Function_Call.fill_input(
                self, wait,
                    locator=(By.XPATH, '//*[@id="tds_tax"]'),
                    value=row_data["TDS%"],
                    pattern=r"^\d+(\.\d{1,2})?$",
                    field_name="TDS%",
                    screenshot_prefix="TDS%",
                    range_check=lambda v: 0 <= float(v) <= 100,
                    row_num=row_num,
                    Sheet_name=Sheet_name)
                Error_field_val.extend(errors)
                print(Error_field_val)
            else:
                msg = f"'{None}' → TDS% field is mandatory ⚠️"
                Mandatory_field.append("TDS%"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name) 
        
        
        # calcType
        # sleep(2)
        # Function_Call.dropdown_select(self,'//*[@id="tab_1"]/div/div[3]/div[4]/div/div/span/span[1]/span/span[2]', row_data["calcType"],'/html/body/span[1]/span/span[1]/input')
  
        #firstName
        if row_data["firstName"] is not None:
            errors=Function_Call.fill_input(
                self, wait,
                locator=(By.XPATH, '//*[@id="first_name"]'),
                value=row_data["firstName"],
                #Allow only letters/spaces, minimum 3 characters
                pattern = r"^[A-Za-z\s]{2,}$",
                field_name="firstName",
                screenshot_prefix="firstName",
                row_num=row_num,
                Sheet_name=Sheet_name)
            Error_field_val.extend(errors)
            print(Error_field_val)
        else:
            msg = f"'{None}' → firstName field is mandatory ⚠️"
            Mandatory_field.append("firstName"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
       
        #lastName
        errors=Function_Call.fill_input(
        self, wait,
        locator=(By.XPATH, '//*[@id="last_name_karigar"]'),
        value=row_data["lastName"],
        #Allow only letters/spaces, minimum 3 characters
        pattern = r"^[A-Za-z\s]{3,}$",
        field_name="lastName",
        screenshot_prefix="lastName",
        row_num=row_num,
        Sheet_name=Sheet_name)
        Error_field_val.extend(errors)
        print(Error_field_val)
        
        if row_data["userType"]=="company": 
            #company
            if row_data["companyName"] is not None:
                errors=Function_Call.fill_input(
                    self, wait,
                    locator=(By.XPATH, '//*[@id="company_karigar"]'),
                    value=row_data["companyName"],
                    pattern = r"^[A-Za-z\s]{3,}$",
                    field_name="companyName",
                    screenshot_prefix="companyName",
                    row_num=row_num,
                    Sheet_name=Sheet_name)
                Error_field_val.extend(errors)
                print(Error_field_val)
            else:
                msg = f"'{None}' → companyName field is mandatory ⚠️"
                Mandatory_field.append("companyName"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
    
            #gstNumber
            if row_data["gstNumber"] is not None:    
                errors=Function_Call.fill_input(
                    self, wait,
                    locator=(By.XPATH, '//*[@id="gst_number_karigar"]'),
                    value=row_data["gstNumber"],
                    pattern = r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[A-Z0-9]{1}Z[A-Z0-9]{1}$",
                    field_name="gstNumber",
                    screenshot_prefix="gstNumber",
                    row_num=row_num,
                    Sheet_name=Sheet_name)
                Error_field_val.extend(errors)
                print(Error_field_val)
            else:
                msg = f"'{None}' → gstNumber field is mandatory ⚠️"
                Mandatory_field.append("gstNumber"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
       
        #mobileNumber
        if row_data["mobileNumber"] is not None:
            #fill_input(self,'//*[@id="tds_tax"]', row_data["tds%"])
            errors=Function_Call.fill_input3(
                self, wait,
                locator=(By.XPATH, '//*[@id="karigar_mobile"]'),
                value=row_data["mobileNumber"],
                #Allow exactly 10 digits (mobile number)
                pattern = r"^\d{10}$",  # ✅ now it's a string, not a tuple
                field_name="mobileNumber",
                screenshot_prefix="mobileNumber",
                row_num=row_num,
                extra_keys = (Keys.SHIFT,Keys.TAB),
                Sheet_name=Sheet_name)
            Error_field_val.extend(errors)
            print(Error_field_val)
        else:
            msg = f"'{None}' → mobileNumber field is mandatory ⚠️"
            Mandatory_field.append("mobileNumber"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        #email   
        if row_data["email"] is not None:
            #fill_input(self,'//*[@id="tds_tax"]', row_data["tds%"])
            errors=Function_Call.fill_input3(
               self, wait,
                locator=(By.XPATH, '//*[@id="email_karigar"]'),
                value=row_data["email"],
                #Allow exactly 10 digits (mobile number)
                pattern = r"^[\w\.-]+@[\w\.-]+\.\w+$",  # ✅ now it's a string, not a tuple
                field_name="email",
                screenshot_prefix="email",
                row_num=row_num,
                extra_keys = Keys.TAB,
                Sheet_name=Sheet_name)
            Error_field_val.extend(errors)
            print(Error_field_val)
        else:
            msg = f"'{None}' → email field is mandatory ⚠️"
            Mandatory_field.append("email"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        #phoneNumber    
        
        errors=Function_Call.fill_input(
            self, wait,
            locator=(By.XPATH, '//*[@id="phone_karigar"]'),
            value=row_data["phoneNumber"],
            #Allow exactly 10 digits (mobile number)
            pattern = r"^\d{10}$",  # ✅ now it's a string, not a tuple
            field_name="phoneNumber",
            screenshot_prefix="phoneNumber",
            row_num=row_num,
            Sheet_name=Sheet_name)
            
         # Basic details
        for field, key in {
            '//*[@id="user_name_karigar"]': "userName",
            '//*[@id="password"]': "password",
            }.items():
            Function_Call.fill_input2(self,field, row_data[key])
            
        #address1
        if row_data["address1"] is not None:
            #fill_input(self,'//*[@id="tds_tax"]', row_data["tds%"])
            errors=Function_Call.fill_input(
               self, wait,
                locator=(By.XPATH, '//*[@id="address1"]'),
                value=row_data["address1"],
                #Allow exactly 10 digits (mobile number)
                pattern = r"^[A-Za-z0-9\s,.-]+$",  # ✅ now it's a string, not a tuple
                field_name="address1",
                screenshot_prefix="address1",
                row_num=row_num,
                Sheet_name=Sheet_name)
            Error_field_val.extend(errors)
            print(Error_field_val)
        else:
            msg = f"'{None}' → address1 field is mandatory ⚠️"
            Mandatory_field.append("address1"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        #address2    
        errors=Function_Call.fill_input(
        self, wait,
        locator=(By.XPATH, '//*[@id="address2_karigar"]'),
        value=row_data["address2"],
        #Allow exactly 10 digits (mobile number)
        pattern =r"^[A-Za-z0-9\s,.-]+$",  # ✅ now it's a string, not a tuple
        field_name="address2",
        screenshot_prefix="address2",
        row_num=row_num,
        Sheet_name=Sheet_name)
        
        #address3
        errors=Function_Call.fill_input(
        self, wait,
        locator=(By.XPATH, '//*[@id="address3"]'),
        value=row_data["address3"],
        #Allow exactly 10 digits (mobile number)
        pattern = r"^[A-Za-z0-9\s,.-]+$",  # ✅ now it's a string, not a tuple
        field_name="address3",
        screenshot_prefix="address3",
        row_num=row_num,
        Sheet_name=Sheet_name)
        
        # Country 
        if row_data["country"] is not None:
           country = row_data["country"]
           Function_Call.dropdown_country(self,'//button[@data-id="country"]', row_data["country"],'//button[@data-id="country"]/following-sibling::div[@class="dropdown-menu open"]//input',f"//li[@class='selected']/a/span[text()='{country}']")
        else:
            msg = f"'{None}' → country field is mandatory ⚠️"
            Mandatory_field.append("country"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        # State 
        if row_data["state"] is not None:
           Function_Call.dropdown_select(self,'//button[@data-id="state"]', row_data["state"],'//button[@data-id="state"]/following-sibling::div[@class="dropdown-menu open"]//input')
        else:
            msg = f"'{None}' → state field is mandatory ⚠️"
            Mandatory_field.append("state"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
        
        # city 
        if row_data["city"] is not None:
           Function_Call.dropdown_select(self,'//button[@data-id="city"]', row_data["city"],'//button[@data-id="city"]/following-sibling::div[@class="dropdown-menu open"]//input')
        else:
            msg = f"'{None}' → city field is mandatory ⚠️"
            Mandatory_field.append("city"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        # Pincode 
        if row_data["pincode"] is not None:
           Function_Call.fill_input2(self,'//*[@id="kar_pincode"]', row_data["pincode"],)
        else:
            msg = f"'{None}' → pincode field is mandatory ⚠️"
            Mandatory_field.append("pincode"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        Function_Call.fill_input2(self,'//textarea[@id="remarks"]', row_data["Remarks"])
        if row_data["Image"] is not None:
           Function_Call.fill_input2(self,'//input[@id="user_img"]',row_data["Image"])
        else:
             pass  
         
        print(Mandatory_field)
        print(Error_field_val)
        failed_fields = []
        if Mandatory_field:
            failed_fields.extend(Mandatory_field)
        if Error_field_val:
            failed_fields.extend(Error_field_val)   
               
        #submit
        if row_data["Contract Price"]=='Yes':
            if failed_fields:
                error = Function_Call.alert1(self,'//*[@id="add_karNext"]')
                if error:
                    Test_Status="WARANING"
                    Actual_Status= (f"⚠️ Found the message:'{error}'") # prints: Select Order Branch
                    print(Actual_Status)
                    
                    Function_Call.click(self,'//button[@class="btn btn-default btn-cancel"]')
                else:
                    Test_Status="Fail"
                    Actual_Status= (f"'{failed_fields}'-> This field is None, but the vendor was saved successfully ❌") 
                return Test_Status,Actual_Status 
            else:    
                result=Function_Call.alert1(self,'//*[@id="add_karNext"]')
                if result:
                    Test_Status="Pass"
                    Actual_Status =(f"✅ Found the message:'{result}'") 
                    print(result)
                    self.update_excel_status(row_num, Test_Status, Actual_Status, Sheet_name)
                    # Function_Call.click(self,'//span[@class="add_wastage"]')
                    test_case_id=row_data['TestCaseId']
                    Contract_Price.test_contractprice(self,test_case_id)
                    return 
                else:
                    Test_Status="Fail"
                    Actual_Status= (f"'{result}' Vendor not save successfully ❌") 
                    return Test_Status,Actual_Status                
        else:
            if failed_fields:    
                result=Function_Call.alert1(self,'//*[@id="add_newkar"]')
                if result:
                    Test_Status="WARANING"
                    Actual_Status =(f"⚠️ Found the message:'{result}'") 
                    print(result)
                    Function_Call.click(self,'//button[@class="btn btn-default btn-cancel"]')
                else: 
                    
                    Test_Status="Fail"
                    Actual_Status= (f"'{failed_fields}'-> This field is None, but the vendor was saved successfully ❌") 
                return Test_Status,Actual_Status
            else:
                result=Function_Call.alert1(self,'//*[@id="add_newkar"]')
                if result:
                    Test_Status="Pass"
                    Actual_Status =(f"✅ Found the message:'{result}'") 
                    print(result)
                else:
                    Test_Status="Fail"
                    Actual_Status= ("Vendor save not successfully ❌") 
                return Test_Status,Actual_Status
                
        
      
      
      
        

    def update_excel_status(self,row_num, Test_Status, Actual_Status, Sheet_name):
        print(Sheet_name)
        # Load the workbook
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]  # or workbook["SheetName"]
        
        if Test_Status== 'Pass':
            # Write Test_Status into column 2
            sheet.cell(row=row_num, column=2, value=Test_Status).font= Font(bold=True, color="00B050")
            # Write Actual_Status in col 3 
            sheet.cell(row=row_num, column=3, value=Actual_Status).font = Font(bold=True, color="00B050")
        if Test_Status=='Fail':
            # Write Test_Status into column 2
            sheet.cell(row=row_num, column=2, value=Test_Status).font= Font(bold=True, color="FF0000")
            # Write Actual_Status in col 3 
            sheet.cell(row=row_num, column=3, value=Actual_Status).font = Font(bold=True, color="FF0000")
        if  Test_Status == 'WARANING':
            
            sheet.cell(row=row_num, column=2, value=Test_Status).font= Font(bold=True, color="FF8000")

            sheet.cell(row=row_num, column=3, value=Actual_Status).font = Font(bold=True, color="FF8000")
            
        # Save workbook
        workbook.save(FILE_PATH)
        # Get status from ExcelUtils
        Status = ExcelUtils.get_Status(FILE_PATH, Sheet_name)
        # Print and return status
        print(Status)
        return Status
    

            
           
   
                
     
            
                
            

            
                

            
            

           
           
            
        

            
        
        
#     def is_element_present(self, how, what):
#         try: Function_Call.driver.find_element(by=how, value=what)
#         except NoSuchElementException as e: return False
#         return True
    
#     def is_alert_present(self):
#         try: Function_Call.driver.switch_to_alert()
#         except NoAlertPresentException as e: return False
#         return True
    
#     def close_alert_and_get_its_text(self):
#         try:
#             alert = Function_Call.driver.switch_to_alert()
#             alert_text = alert.text
#             if Function_Call.accept_next_alert:
#                 alert.accept()
#             else:
#                 alert.dismiss()
#             return alert_text
#         finally: Function_Call.accept_next_alert = True
    
#     def tearDown(self):
#         Function_Call.driver.quit()
#         Function_Call.assertEqual([], Function_Call.verificationErrors)

# if __name__ == "__main__":
#     unittest.main()
