from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import  sleep
import unittest
from Utils.Excel import ExcelUtils
from openpyxl.drawing.image import Image
from Test_vendor.Kyc import KYC
from openpyxl import load_workbook
from openpyxl.styles import Font
from Utils.Function import Function_Call
import re


FILE_PATH = ExcelUtils.file_path
class stone_Price(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)

    def test_stoneprice(self,test_case_id):
        driver = self.driver
        wait = self.wait
        sleep(3)
        Sheet_name = 'Stone_price'
        test_case_id = test_case_id
        value =ExcelUtils.Test_case_id_count(FILE_PATH, Sheet_name,test_case_id)
        print(value)
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, Sheet_name)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]
        row=1
        count = value
        for row_num in range(2, valid_rows):
            current_id = sheet.cell(row=row_num, column=1).value  # Column 1 = Test Case Id
            if current_id == test_case_id:
                data = {
                    "Test Case Id": 1,
                    "Test Status": 2,
                    "Actual Status": 3,
                    "Stone Type": 4,
                    "Stone Name": 5,
                    "UOM": 6,
                    "Calc Type": 7,
                    "Code": 8,
                    "From Cent": 9,
                    "To Cent": 10,
                    "Rate": 11,
                    "Kyc":12,
                    "Field_validation_status": 13,
                }

                row_data = {
                    key: sheet.cell(row=row_num, column=col).value
                    for key, col in data.items()
                }
                print(row_data)
                # Call your 'create' method
                Create_data = stone_Price.create(self, row_data, row_num, Sheet_name, row, count)
                print(Create_data)
                # Remove processed customer from the list
                
                if Create_data is int:
                    Create_data = count
                
                if Create_data:
                    Test_Status,Actual_Status= Create_data
                    row=1
                    self.update_excel_status(row_num, Test_Status, Actual_Status, Sheet_name)
                    return
                
    def create(self,row_data, row_num, Sheet_name, row, count):  
        driver, wait = self.driver, self.wait         
        Mandatory_field=[]
        Function_Call.click(self,'//span[@class="add_stone"]')
        # Stone Type 
        if row_data["Stone Type"] is not None:
            Function_Call.dropdown_select(self,f'//*[@id="select2-stn_type{row}-container"]', row_data["Stone Type"],"/html/body/span/span/span[1]/input")
            Function_Call.dropdown_select
        else:
            msg = f"'{None}' → Stone Type field is mandatory ⚠️"
            Mandatory_field.append(msg); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        if row_data["Stone Name"] is not None:
            sleep(2)
            Function_Call.dropdown_select(self,f'//*[@id="select2-stn_name{row}-container"]', row_data["Stone Name"],"/html/body/span/span/span[1]/input")
        else:
            msg = f"'{None}' → Stone Name field is mandatory ⚠️"
            Mandatory_field.append(msg); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        if row_data["UOM"] is not None:
            Function_Call.dropdown_select(self,f'//*[@id="select2-uom_id{row}-container"]', row_data["UOM"],"/html/body/span/span/span[1]/input")
        else:
            msg = f"'{None}' → UOM field is mandatory ⚠️"
            Mandatory_field.append(msg); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
                
        # Design 
        if row_data["Calc Type"] is not None:
            Function_Call.dropdown_select(self,f'//*[@id="select2-stn_calc_type{row}-container"]', row_data["Calc Type"],"/html/body/span/span/span[1]/input")
        else:
            msg = f"'{None}' → Calc Type field is mandatory ⚠️"
            Mandatory_field.append(msg); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        # Code 
        if row_data["Code"] is not None:
            Function_Call.dropdown_select(self,f'//*[@id="select2-id_stn_quality{row}-container"]', row_data["Code"],"/html/body/span/span/span[1]/input")
        else:
            msg = f"'{None}' → Code field is mandatory ⚠️"
            Mandatory_field.append(msg); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)


        Error_field_val=[]    
        if row_data["From Cent"]is not None:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, f'//input[@name="stn[{row}][from_wt]"]'),
                value=row_data["From Cent"],
                pattern = r"^\d{1,3}$",
                field_name="From Cent",
                screenshot_prefix="From Cent",
                range_check = lambda v: 10 <= float(v) <= 120,
                row_num=row_num,
                Sheet_name=Sheet_name
            )
            Error_field_val.extend(errors)
            print(Error_field_val)
        else:
            msg = f"'{None}' → From Cent field is mandatory ⚠️"
            Mandatory_field.append(msg); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        if row_data["To Cent"]is not None:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, f'//input[@name="stn[{row}][to_wt]"]'),
                value=row_data["To Cent"],
                pattern = r"^\d{1,3}$",
                field_name="To Cent",
                screenshot_prefix="To Cent",
                range_check = lambda v: 10 <= float(v) <= 120,
                row_num=row_num,
                Sheet_name=Sheet_name
            )
            Error_field_val.extend(errors)
            print(Error_field_val)
        else:
            msg = f"'{None}' → To Cent field is mandatory ⚠️"
            Mandatory_field.append(msg); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        if row_data["Rate"]is not None:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, f'//input[@name="stn[{row}][stone_rate]"]'),
                value=row_data["Rate"],
                pattern = r"^\d{1,7}(\.\d{1,2})?$",
                field_name="Rate",
                screenshot_prefix="Rate",
                row_num=row_num,
                Sheet_name=Sheet_name
            )
            Error_field_val.extend(errors)
            print(Error_field_val)
        else:
            msg = f"'{None}' → Rate field is mandatory ⚠️"
            Mandatory_field.append(msg); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
        print(Mandatory_field)
        print(Error_field_val)
        if Mandatory_field or Error_field_val:           
            error = Function_Call.alert1(self, '//span[@class="add_stone"]')
            Test_Status="Pass"
            Actual_Status= (f"⚠️ Found the message:'{error}'") # prints: Select Order Branch
            print(error)
            wait.until(EC.element_to_be_clickable((By.XPATH,'(//button[@class="btn btn-default btn-cancel"])[3]'))).click()
            return Test_Status,Actual_Status 
            
        if count > 1:
            row += 1
            count -= 1
            return count
        else:
            #submit
            if row_data["Kyc"]=='Yes':
                Function_Call.click(self,'//button[@id="add_StoneNext"]')
                Test_Status="Pass"
                Actual_Status = "Stone Added Successfully"
                self.update_excel_status(row_num, Test_Status, Actual_Status, Sheet_name)
                test_case_id=row_data['Test Case Id']
                KYC.test_Kyc(self,test_case_id)
                
                
                
    def update_excel_status(self,row_num, Test_Status, Actual_Status, Sheet_name):
        print(Sheet_name)
        # Load the workbook
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]  # or workbook["SheetName"]
        
        if Test_Status== 'Pass':
            # Write Test_Status into column 2
            sheet.cell(row=row_num, column=2, value=Test_Status).font=Font(bold=True, color="00B050")
            
            # Write Actual_Status in col 3 
            sheet.cell(row=row_num, column=3, value=Actual_Status).font = Font(bold=True, color="00B050")
        if Test_Status=='Fail':
            # Write Test_Status into column 2
            sheet.cell(row=row_num, column=2, value=Test_Status).font=Font(bold=True, color="FF0000")
            # Write Actual_Status in col 3 
            sheet.cell(row=row_num, column=3, value=Actual_Status).font = Font(bold=True, color="FF0000")
        # Save workbook
        workbook.save(FILE_PATH)
        # Get status from ExcelUtils
        Status = ExcelUtils.get_Status(FILE_PATH, Sheet_name)
        # Print and return status
        print(Status)
                        