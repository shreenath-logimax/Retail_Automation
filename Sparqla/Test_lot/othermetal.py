from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
import unittest
from Utils.Excel import ExcelUtils
from openpyxl import load_workbook
from time import sleep

FILE_PATH = ExcelUtils.file_path
class Othermetal(unittest.TestCase):  
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 100)
        
    def test_othermetal(self,Sheet_name,test_case_id):
        driver = self.driver
        wait = self.wait
      
        function_name = Sheet_name
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, function_name)
        test_case_id = test_case_id
        value =ExcelUtils.Test_case_id_count(FILE_PATH, function_name,test_case_id)
        print(value)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[function_name]
        row=1
        count = value
        for row_num in range(2, valid_rows):
            current_id = sheet.cell(row=row_num, column=1).value  # Column 1 = Test Case Id
            if current_id == test_case_id:
                data = {
                        "Test Case Id": 1,
                        "Other Metals": 2,
                        "Metal": 3,
                        "Purity" : 4,
                        "Pcs": 5,
                        "Nwt": 6,
                        "V.A(%)": 7,
                        "Mc Type": 8,
                        "Mc": 9,
                        "Rate": 10
                    }
                row_metaldata = {key: sheet.cell(row=row_num, column=col).value 
                            for key, col in data.items()}
                print(row_metaldata)   
                wait.until(EC.element_to_be_clickable((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td/select)".format(row)))).click()
                Select(wait.until(EC.element_to_be_clickable((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td/select)".format(row))))).select_by_visible_text(row_metaldata["Metal"])
            
                wait.until(EC.element_to_be_clickable((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[2]/select)".format(row)))).click()
                Select(wait.until(EC.element_to_be_clickable((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[2]/select)".format(row))))).select_by_visible_text(row_metaldata["Purity"])
                
                # wait.until(EC.visibility_of_element_located((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[3]/select)".format(row)))).clear()
                wait.until(EC.presence_of_element_located((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[3]/input)".format(row)))).send_keys(row_metaldata["Pcs"])
                
                wait.until(EC.visibility_of_element_located((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[4]/input)".format(row)))).clear()
                wait.until(EC.visibility_of_element_located((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[4]/input)".format(row)))).send_keys(row_metaldata["Nwt"])
                
                wait.until(EC.visibility_of_element_located((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[5]/input)".format(row)))).clear()
                wait.until(EC.visibility_of_element_located((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[5]/input)".format(row)))).send_keys(row_metaldata["V.A(%)"])
                print(row_metaldata["Mc Type"])
                wait.until(EC.element_to_be_clickable((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[6]/select)".format(row)))).click()
                Select(wait.until(EC.element_to_be_clickable((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[6]/select)".format(row))))).select_by_visible_text(row_metaldata["Mc Type"])
                    
          
                wait.until(EC.visibility_of_element_located((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[7]/input)".format(row)))).clear()
                wait.until(EC.visibility_of_element_located ((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[7]/input)".format(row)))).send_keys(row_metaldata["Mc"])
                
                wait.until(EC.visibility_of_element_located((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[8]/input)".format(row)))).clear()
                wait.until(EC.visibility_of_element_located ((By.XPATH,"(//table[@id='other_metal_table']/tbody/tr[{0}]/td[8]/input)".format(row)))).send_keys(row_metaldata["Rate"])
                sleep(2)
                if count != 1:
                    wait.until(EC.element_to_be_clickable((By.ID,'create_other_metal_item_details'))).click()
                    row=row+1
                    count=value-1
                    sleep(4)
                    
                else:
                    break      
            else:
                print(f'This row is not {test_case_id}') 
        TotalAmount =wait.until(EC.visibility_of_element_located((By.XPATH,"//table[@id='other_metal_table']/tfoot/tr/td[9]"))).text
        TotalAmount = float(TotalAmount)
        print(TotalAmount)                           
        wait.until(EC.element_to_be_clickable((By.ID,"update_other_metal_details"))).click()
        sleep(4)
        OtherMetal = "Other Metal detail Add successfully"       
        return OtherMetal,TotalAmount
             