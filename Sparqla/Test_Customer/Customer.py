from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import UnexpectedAlertPresentException
from time import sleep
import os
import unittest
import re
import math
import random
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from Test_gettag.getttag import GetTag
from Test_Customer.less import Stone
from openpyxl import load_workbook
from openpyxl.styles import Font
from PIL import ImageGrab

FILE_PATH = ExcelUtils.file_path

class CustomerOrderTR(unittest.TestCase):
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)
        self.Board_Rate = []

    def test_customer_order_t_r(self):
        """Main test entry point for Customer Order automation."""
        sheet_name = "Customer"

        # Read Excel data
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
        customer_list = ExcelUtils.customer_details(FILE_PATH, sheet_name)
        tag_count = ExcelUtils.Tag_reserve(FILE_PATH, sheet_name)

        if tag_count != 0:
            tags = GetTag.test_gettag(self, tag_count)
            ExcelUtils.update_tag_id(FILE_PATH, sheet_name, tags)
        else:
            print("All tags available")

        # Get Gold rate
        self.Board_Rate = []
        Function_Call.click(self, "//span[@class='header_rate']/b[contains(text(),'INR')]")
        rate_text1 = self.wait.until(EC.presence_of_element_located((By.XPATH, "//li[@class='user-body rate_block_body']//tr[th[contains(text(),'Gold 22KT 1gm')]]/td"))).text
        rate_text2 = self.wait.until(EC.presence_of_element_located((By.XPATH, "//li[@class='user-body rate_block_body']//tr[th[contains(text(),'Gold 18KT 1gm')]]/td"))).text
        rate_text3 = self.wait.until(EC.presence_of_element_located((By.XPATH, "//li[@class='user-body rate_block_body']//tr[th[contains(text(),'Silver 1gm')]]/td"))).text
        
        gold_rate22KT = int(float(rate_text1.replace("INR", "").strip()))
        self.Board_Rate.append(gold_rate22KT)
        gold_rate18KT = int(float(rate_text2.replace("INR", "").strip()))
        self.Board_Rate.append(gold_rate18KT)
        silver_rate = int(float(rate_text3.replace("INR", "").strip()))
        self.Board_Rate.append(silver_rate)
        
        print(f"Rates - 22KT: {gold_rate22KT}, 18KT: {gold_rate18KT}, Silver: {silver_rate}")

        # Navigate to Customer Orders page
        self.fc.click("//a[@class='sidebar-toggle']")
        self.fc.click("//span[contains(text(), 'Customer Orders')]")
        self.fc.click("//span[contains(text(), 'Create Order')]")

        # Load Excel sheet
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[sheet_name]
        prev_customer = ''
        row_counter = 1

        for row_num in range(2, valid_rows + 1):
            if not customer_list:
                break
                
            current_cus_no = customer_list[0]
            
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3, "Customer Number": 4,
                "Customer Name": 5, "OrderBranch": 6, "Employee": 7, "BalanceType": 8,
                "OrderType": 9, "RateType": 10, "TagScan": 11, "Product": 12,
                "Design": 13, "SubDesign": 14, "Purity": 15, "GrossWt": 16,
                "LessWt": 17, "Size": 18, "Pcs": 19, "Wast%": 20,
                "Wast_Wgt": 21, "MC_Type": 22, "MC_Value": 23, "OtherCharge": 24,
                "ChargeName": 25, "Rate": 26, "Description": 27, "OrderEdit": 28,
                "UpdateGwt": 29, "Remove": 30, "Field_validation_satus": 31
            }

            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            next_cus_no = sheet.cell(row=row_num + 1, column=4).value
            
            keys_to_check = ["Customer Number", "OrderBranch", "Employee", "Product", "Purity", "GrossWt", "Pcs", "Rate", "Description"]

            result = self.create(row_data, current_cus_no, next_cus_no, prev_customer, row_counter, keys_to_check, row_num, sheet_name)
            
            prev_customer = current_cus_no
            print('prev_customer-------------------',prev_customer)
            customer_list.pop(0)
            print('customer_list-------------------',customer_list)

            if result == current_cus_no:
                row_counter += 1
                print('row_counter-------------------',row_counter)
            elif isinstance(result, tuple):
                test_status, actual_status = result
                row_counter = 1
                self.update_excel_status(row_num, test_status, actual_status, sheet_name)

        workbook.close()

    def create(self, row_data, cus_no, next_no, prev_cus, row_idx, keys_to_check, row_num, sheet_name):
        """Processes a single order record."""
        if prev_cus != cus_no:
            self.fc.click("//a[@id='add_Order']")
            if not self._handle_customer_header(row_data, row_num, sheet_name):
                return "Fail", "Header information missing or invalid."
            self._handle_order_types_selection(row_data)

            if row_idx >= 5:
                self.driver.execute_script(f"window.scrollBy(0, -{row_idx*100});")
        
        if row_data["OrderType"] == "Tag Reserve":
            data=self.handle_tag_reserve(row_data, row_idx, row_num, keys_to_check, next_no, cus_no, sheet_name)
            if data:
                return data
            return "Fail", "Tag Reserve failed."
            
        
        if row_data["OrderType"] == "Customized Order":
            print('row_idx-------------------',row_idx)
            data=self.handle_customized_order(row_data, row_idx, row_num, next_no, cus_no, sheet_name)
            if data:
                return data
            return "Fail", "Customized Order failed."
            

        return "Fail", f"Unknown Order Type: {row_data['OrderType']}"

    def _handle_customer_header(self, row_data, row_num, sheet_name):
        """Fills header information: Customer, Branch, Employee."""
        # Customer
        if row_data.get("Customer Number"):
            self.fc.fill_autocomplete_field("cus_name", str(row_data["Customer Number"]))
        else:
            self.fc.Remark(row_num, "Customer field is mandatory ⚠️", sheet_name)
            Function_Call.click(self,'//button[@class="btn btn-default btn-cancel"]')
            return False

        # Branch
        if row_data.get("OrderBranch"):
            self.fc.dropdown_select("//span[@id='select2-branch_select-container']", 
                                   row_data["OrderBranch"], 
                                   "//span[@class='select2-search select2-search--dropdown']/input")
        else:
            self.fc.Remark(row_num, "Order Branch field is mandatory ⚠️", sheet_name)
            Function_Call.click(self,'//button[@class="btn btn-default btn-cancel"]')
            return False

        # Employee
        if row_data.get("Employee"):
            self.fc.dropdown_select("//span[@id='select2-issue_employee-container']", 
                                   row_data["Employee"], 
                                   "//span[@class='select2-search select2-search--dropdown']/input")
        else:
            self.fc.Remark(row_num, "Employee field is mandatory ⚠️", sheet_name)
            Function_Call.click(self,'//button[@class="btn btn-default btn-cancel"]')
            return False
            
        return True

    def _handle_order_types_selection(self, row_data):
        """Selects Balance, Order, and Rate types."""
        # Balance Type
        self.fc.click("//input[@id='metal_bal_type']" if row_data.get("BalanceType") == "Metal Balance" else "//input[@id='cash_bal_type']")
        # Order Type
        self.fc.click("//input[@id='tag_order']" if row_data.get("OrderType") == "Tag Reserve" else "//input[@id='customer_order']")
        # Rate Type
        self.fc.click("//input[@id='order_rate']" if row_data.get("RateType") == "Order Rate(Fixed)" else "//input[@id='delivery_rate']")

    def handle_customized_order(self, row_data, row_idx, row_num, next_no, cus_no, sheet_name):
        """Processes a Customized Order line item."""
        test_case_id = row_data["TestCaseId"]
        self.fc.click("//button[@id='add_order_item']")

        if row_idx >= 4:
            self.driver.execute_script(f"window.scrollBy(0, {row_idx*100});")
            sleep(1)

        # Dropdowns
        for field, value in [("prod", row_data["Product"]), ("dsgn", row_data["Design"]), ("sub_design", row_data["SubDesign"])]:
            self.fc.dropdown_select(f"//span[@id='select2-{field}_{row_idx}-container']", value, "//span[@class='select2-search select2-search--dropdown']/input")

        if row_data.get("Purity"):
            self.fc.select_visible_text(f"//select[@id='purity{row_idx}']", row_data["Purity"])

        # Inputs
        if row_data.get("GrossWt"):
            self.fc.fill_input(self.wait, (By.ID, f"weight_{row_idx}"), row_data["GrossWt"], "GrossWt", row_num, pattern=r"\d+(\.\d{1,3})?", screenshot_prefix="GrossWt", Sheet_name=sheet_name)
        else:
            self.fc.Remark(row_num, "GrossWt field is mandatory ⚠️", sheet_name)
            Function_Call.click(self,'//button[@class="btn btn-default btn-cancel"]')
            return False
        
        if row_data.get("LessWt") == "Yes":
            self.fc.click(f"//input[@name='o_item[{row_idx}][less_wt]']")
            Stone.test_tagStone(self, "Customer_LWT", test_case_id)

        self.fc.fill_input(self.wait, (By.ID, f"qty_{row_idx}"), row_data["Pcs"], "Pcs", row_num, pattern=r"^\d+$", screenshot_prefix="Pcs", extra_keys=Keys.ENTER, Sheet_name=sheet_name)
        self.fc.fill_input(self.wait, (By.NAME, f"o_item[{row_idx}][wast_percent]"), row_data["Wast%"], "Wast%", row_num, pattern=r"^\d+(\.\d{1,2})?$", screenshot_prefix="Wast%", range_check=lambda v: 0 <= float(v) <= 100, Sheet_name=sheet_name)

        # MC
        self.fc.click(f"(//span[contains(@id, 'id_mc_type')])[{row_idx}]")
        sleep(0.5)
        mc_search = self.wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/span[2]/span/span[1]/input")))
        mc_search.clear()
        mc_search.send_keys(row_data["MC_Type"], Keys.ENTER)
        self.fc.fill_input(self.wait, (By.NAME, f"o_item[{row_idx}][mc]"), row_data["MC_Value"], "MC_Value", row_num, pattern=r"^\d+(\.\d{1,2})?$", screenshot_prefix="MC", Sheet_name=sheet_name)

        if row_data.get("OtherCharge") == "Yes":
            self._handle_other_charges(row_data, row_idx)
            
        gold_rate = self._get_dynamic_gold_rate(row_idx)
        self.fc.fill_input(self.wait, (By.XPATH, f"//*[@id='detail{row_idx}']/td[20]/input"), gold_rate, "Rate", row_num, pattern=r"^\d+(\.\d{1,2})?$", screenshot_prefix="Rate", Sheet_name=sheet_name)

        if row_data.get("Description"):
            self._fill_description(row_data, row_idx)

        if row_data.get("Remove"):
            self._handle_item_removal(row_data["Remove"], row_idx)

        return self.check_calculation_and_submit(row_data, row_idx, next_no, cus_no)

    def handle_tag_reserve(self, row_data, row_idx, row_num, keys_to_check, next_no, cus_no, sheet_name):
        """Handles the Tag Reserve order type."""
        self.fc.fill_input2("//input[@id='est_tag_scan']", row_data["TagScan"])
        self.fc.click("//button[@id='tag_search']")

        if row_data.get("Description"):
            self._fill_description(row_data, row_idx)
        else:
            self.fc.Remark(row_num, "Description field is mandatory ⚠️", sheet_name)

        return self.check_calculation_and_submit(row_data, row_idx, next_no, cus_no)

    def _fill_description(self, row_data, row_idx):
        """Fills the Description Modal."""
        self.fc.click(f"//tr[@id='detail{row_idx}']/td[24]/a")
        self.fc.fill_input2("//textarea[@id='description']", str(row_data["Description"]))
        sleep(1); self.fc.click("//a[contains(text(), 'Add')]")

    def _handle_other_charges(self, row_data, row_idx):
        """Handles adding 'Other Charges'."""
        self.fc.click(f"//*[@id='detail{row_idx}']/td[17]/a")
        charges_raw = row_data.get("ChargeName")
        if not charges_raw: return

        charges_list = [s.strip() for s in str(charges_raw).split(",")]
        for idx, charge in enumerate(charges_list):
            if idx > 0: self.fc.click("//button[@id='add_new_charge']")
            self.fc.select_visible_text(f"(//select[@name='est_stones_item[id_charge][]'])[{idx+1}]", charge)
            val_xpath = f"(//input[@name='est_stones_item[value_charge][]'])[{idx+1}]"
            if self.fc.get_value(val_xpath) in ["0.00", "0", ""]: self.fc.fill_input2(val_xpath, "500")

        self.fc.click("//button[@id='update_charge_details']")

    def _handle_item_removal(self, remove_field, row_idx):
        """Removes a field value as specified in Excel."""
        remove_map = {"Product": f"prod_{row_idx}", "Design": f"dsgn_{row_idx}", "SubDesign": f"sub_design_{row_idx}"}
        xpath = f'//span[@id="select2-{remove_map.get(remove_field)}-container"]/span'
        if xpath: self.fc.click(xpath)

    def check_calculation_and_submit(self, row_data, row_idx, next_no, cus_no):
        """Verifies calculations and submits the order."""
        cal_val_raw = self.fc.get_value(f"//input[@name='o_item[{row_idx}][calculation_based_on]']")
        
        # Get dynamic gold rate based on purity
        gold_rate = self._get_dynamic_gold_rate(row_idx)
        
        calc_result = self.calculation(cal_val_raw, row_idx, gold_rate)
        ceil_val, table_val, cal_type = calc_result
        
        if abs(float(ceil_val or 0) - float(table_val or 0)) > 1:
            self.driver.save_screenshot(os.path.join(ExcelUtils.SCREENSHOT_PATH, f"CalcError_{row_data['TestCaseId']}.png"))
            Function_Call.click(self,'//button[@class="btn btn-default btn-cancel"]')
            print(f"Row {row_idx} - {cal_type} | Expected: {ceil_val}, Found: {table_val}")
            return "Fail", f"❌ {cal_type} mismatch. Expected: {ceil_val}, Found: {table_val}"
        
        print(f"✅ {cal_type} calculation verified.")
        if next_no == cus_no: 
            return cus_no
        
        self.fc.click("//button[@id='create_order']")
        try:
            msg = self.fc.get_text("/html/body/div[1]/div[1]/section[2]/div/div/div/div[2]/div[1]/div/div").replace("×", "").replace("\n", "").strip()
            if "successfully" in msg.lower():
                sleep(2)  # Wait for page stabilization/redirect
                return ("Pass", f"✅ {msg}")
            return ("Fail", f"❌ {msg}")
        except UnexpectedAlertPresentException as e:
            # Capturing Desktop screenshot to include the alert pop-up
            try:
                screenshot_path = os.path.join(ExcelUtils.SCREENSHOT_PATH, "SubmitError.png")
                ImageGrab.grab().save(screenshot_path)
                print(f"✅ Desktop screenshot captured: {screenshot_path}")
            except Exception as se:
                print(f"⚠️ Desktop screenshot failed: {se}")
            
            alert_text = e.alert_text or "Required fields missing"
            print(f"⚠️ Caught Alert: {alert_text}")
            try:
                self.driver.switch_to.alert.accept()
                self.driver.switch_to.alert.accept()
            except:
                pass
            Function_Call.click(self, '//button[@class="btn btn-default btn-cancel"]')
            Function_Call.click(self, '//button[@class="btn btn-default btn-cancel"]')
            return "Fail", f"❌ Alert: {alert_text}"
        except Exception as e:
            print(f"⚠️ Submit Error: {e}")
            self.driver.save_screenshot(os.path.join(ExcelUtils.SCREENSHOT_PATH, "SubmitError.png"))
            Function_Call.click(self, '//button[@class="btn btn-default btn-cancel"]')
            return "Fail", "❌ Order submission failed."

    def _get_dynamic_gold_rate(self, row_idx):
        """Extracts purity from UI and returns the matching board rate."""
        try:
            from selenium.webdriver.support.ui import Select
            purity_dropdown = Select(self.wait.until(EC.presence_of_element_located((By.ID, f"purity{row_idx}"))))
            purity = purity_dropdown.first_selected_option.text.strip()
            print(f"Selected Purity: {purity}")

            if "916" in purity:
                rate = self.Board_Rate[0]
            elif "75" in purity:
                rate = self.Board_Rate[1]
            elif "92.5" in purity:
                rate = self.Board_Rate[2]
            else:
                rate = self.Board_Rate[0]  # Default to 22KT
            
            print(f"Using Dynamic Gold Rate: {rate}")
            return rate
        except Exception as e:
            print(f"⚠️ Error fetching dynamic rate: {e}")
            return self.Board_Rate[0] if self.Board_Rate else 0

    def calculation(self, cal_type_idx, row_idx, gold_rate):
        """Helper to compute expected Taxable Amount with detailed logic."""
        wait = self.wait
        data_map = {
            "0": "Mc & Wast On Gross",
            "1": "Mc & Wast On Net",
            "2": "Mc on Gross, Wast On Net",
            "3": "Fixed Rate",
            "4": "Fixed Rate based on Weight"
        }
        cal_type = data_map.get(str(cal_type_idx), "Unknown")
        print(f"Calculation Type: {cal_type}")

        def get_v(by, locator, cast=float, default=0.0):
            try:
                el = wait.until(EC.presence_of_element_located((by, locator)))
                val = el.get_attribute("value")
                return cast(val) if val else default
            except:
                return default

        # Fetch all required fields
        gwt = get_v(By.NAME, f"o_item[{row_idx}][weight]")
        nwt = get_v(By.NAME, f"o_item[{row_idx}][net_wt]")
        stone_amt = get_v(By.NAME, f"o_item[{row_idx}][stn_amt]")
        wast_per = get_v(By.NAME, f"o_item[{row_idx}][wast_percent]")
        mc_rate = get_v(By.NAME, f"o_item[{row_idx}][mc]")
        other_amt = get_v(By.NAME, f"o_item[{row_idx}][value_charge]")
        table_taxable = get_v(By.NAME, f"o_item[{row_idx}][taxable]")

        # Fetch MC type from select2 container
        try:
            mc_type = wait.until(
                EC.presence_of_element_located((By.XPATH, f'//span[contains(@id,"o_item[{row_idx}][id_mc_type]")]'))
            ).get_attribute("title")
        except:
            mc_type = "Gram"

        print(f"Gwt={gwt}, Nwt={nwt}, Stone={stone_amt}, Wast%={wast_per}, MC={mc_rate}, MC_Type={mc_type}, Other={other_amt}, Table_Taxable={table_taxable}")

        ceil_value = 0.0

        if cal_type == "Mc on Gross, Wast On Net":
            # Making cost on Gross, Wastage on Net
            mc = mc_rate if mc_type == 'Piece' else math.ceil(mc_rate * gwt)
            va = round((wast_per / 100) * nwt, 3)
            total_wt = round(nwt + va, 3)
            calc_val = (total_wt * gold_rate) + stone_amt + mc + other_amt
            ceil_value = math.ceil(calc_val)

        elif cal_type == "Mc & Wast On Net":
            # Making cost & Wastage on Net
            mc = mc_rate if mc_type == 'Piece' else math.ceil(mc_rate * nwt)
            va = round((wast_per / 100) * nwt, 3)
            total_wt = round(nwt + va, 3)
            calc_val = (total_wt * gold_rate) + mc + stone_amt + other_amt
            ceil_value = math.ceil(calc_val)

        elif cal_type == "Mc & Wast On Gross":
            # Making cost & Wastage on Gross
            mc = mc_rate if mc_type == 'Piece' else math.ceil(mc_rate * gwt)
            va = round((wast_per / 100) * gwt, 3)
            total_wt = round(nwt + va, 3)
            calc_val = (total_wt * gold_rate) + mc + stone_amt + other_amt
            ceil_value = math.ceil(calc_val)

        elif cal_type == "Fixed Rate based on Weight":
            # Making cost on weight, Wastage on Gross
            mc = mc_rate if mc_type == 'Piece' else math.ceil(mc_rate * gwt)
            va = round((wast_per / 100) * gwt, 3)
            total_wt = round(nwt + va, 3)
            calc_val = (total_wt * gold_rate) + mc + stone_amt + other_amt
            ceil_value = math.ceil(calc_val)

        elif cal_type == "Fixed Rate":
            ceil_value = table_taxable

        print(f"Calculated Ceil Value: {ceil_value}")
        return ceil_value, table_taxable, cal_type

    def update_excel_status(self, row_num, test_status, actual_status, sheet_name):
        """Writes the test result back to Excel."""
        workbook = load_workbook(FILE_PATH); sheet = workbook[sheet_name]
        color = "00B050" if test_status == "Pass" else "FF0000"
        sheet.cell(row=row_num, column=2, value=test_status).font = Font(bold=True, color=color)
        sheet.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
        workbook.save(FILE_PATH)