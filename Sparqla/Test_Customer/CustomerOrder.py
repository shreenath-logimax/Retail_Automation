from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import UnexpectedAlertPresentException
from time import sleep
import os
import unittest
import math
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from Test_gettag.getttag import GetTag
from Test_Customer.less import Stone
from openpyxl import load_workbook
from openpyxl.styles import Font
from PIL import ImageGrab
import re

FILE_PATH = ExcelUtils.file_path
BASE_URL = ExcelUtils.BASE_URL

class CustomerOrder(unittest.TestCase):
    """
    Restructured Customer Order Module
    Follows Sparqla System Brain standards.
    """
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)
        self.Board_Rate = []

    def test_customer_order(self):
        """Main entry point for Customer Order automation."""
        sheet_name = "CustomerOrder"

        # 1. Get Initial Data & Rates
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            customer_list = ExcelUtils.customer_details(FILE_PATH, sheet_name)
            # tag_count = ExcelUtils.Tag_reserve(FILE_PATH, sheet_name,valid_rows)

            # if tag_count != 0:
            #     tags = GetTag.test_gettag(self, tag_count)
            #     ExcelUtils.update_tag_id(FILE_PATH, sheet_name, tags)
            
            CustomerOrder._resolve_tags(self, sheet_name, valid_rows)
            
            # Fetch Gold Rates
            self.Board_Rate = self._get_header_rates()
            print(f"✅ Rates Loaded - 22KT: {self.Board_Rate[0]}, 18KT: {self.Board_Rate[1]}, Silver: {self.Board_Rate[2]}")
        except Exception as e:
            print(f"❌ Initialization failed: {e}")
            return

        # 2. Navigation
        try:
            self.fc.click("//a[@class='sidebar-toggle']")
            self.fc.click("//span[contains(text(), 'Customer Orders')]")
            self.fc.click("//span[contains(text(), 'Create Order')]")
            sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            self.driver.get(BASE_URL + "index.php/admin_ret_customer_order/customer_order/add")
            sleep(2)
        # 3. Main Loop
        prev_customer = ''
        
        row_counter = 1
        current_order_tags = []
        
        for row_num in range(2, valid_rows):
            # Reload workbook each row as per standard

            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]
            
            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,"OrderFlow":4, "Customer Number": 5,
                "Customer Name": 6, "OrderBranch": 7, "Employee": 8, "BalanceType": 9,
                "OrderType": 10, "RateType": 11, "TagScan": 12, "Product": 13,
                "Design": 14, "SubDesign": 15, "Purity": 16, "GrossWt": 17,
                "LessWt": 18, "Size": 19, "Pcs": 20, "Wast%": 21,
                "Wast_Wgt": 22, "MC_Type": 23, "MC_Value": 24, "OtherCharge": 25,
                "ChargeName": 26, "Rate": 27, "Description": 28, "DueDate":29, "OrderEdit": 30,
                "UpdateGwt": 31, "Remove": 32, "Field_validation_status":33,"Order NO":34
            }
            
            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            next_cus_no = sheet.cell(row=row_num + 1, column=4).value
            workbook.close()

            # if str(row_data["TestStatus"]).strip().lower() != "enable":
            #     continue

            current_cus_no = row_data["Customer Number"]
            
            if row_data.get("OrderType") == "Tag Reserve" and row_data.get("TagScan"):
                current_order_tags.append(str(row_data["TagScan"]).strip())
            
            print(f"\n🧪 Processing Row {row_num} | Case: {row_data['TestCaseId']}")
            
            try:
                result = CustomerOrder._process_row(self,row_data, current_cus_no, next_cus_no, prev_customer, row_counter, row_num, sheet_name)
                
                if result == "Continue":
                    row_counter += 1
                    prev_customer = current_cus_no
                else:
                    # Handle extra return value (captured_id) if present
                    test_status = result[0]
                    actual_status = result[1]
                    captured_id = result[2] if len(result) > 2 else None
                    
                    row_counter = 1
                    prev_customer = '' # Reset after submit
                    CustomerOrder._update_excel_status(self,row_num, test_status, actual_status, sheet_name, captured_id)
                    
                    if test_status == "Pass" and captured_id:
                        CustomerOrder._update_billing_sheet(self, row_data, captured_id, row_num)
                        CustomerOrder._update_purchase_po_sheet(self, row_data, captured_id)
                        CustomerOrder._update_tag_details_with_order(self, current_order_tags, captured_id)
                    current_order_tags = []
            except Exception as e:
                print(f"❌ Row {row_num} failed: {e}")
                CustomerOrder._update_excel_status(self,row_num, "Fail", f"Error: {str(e)}", sheet_name)
                current_order_tags = []

    def _process_row(self, row_data, cus_no, next_no, prev_cus, row_idx, row_num, sheet_name):
        """Logic for a single row processing."""
        current_field = "Init"
        try:
            if prev_cus != cus_no:
                current_field = "Header Click"
                # If we are on a new customer, we might need to click 'Add' or start the form
                # Based on original code:
                self.fc.click("//a[@id='add_Order']")
                # But sometimes it's already open if it's the first one.
                # Let's check or just click add if needed.
                # self.fc.click("//a[@id='add_Order']")
                pass 

            # Header filling (Customer, Branch, Employee)
            if prev_cus != cus_no:
                current_field = "Filling Header"
                if not CustomerOrder._fill_header(self,row_data, row_num, sheet_name):
                    return "Fail", "Header missing or invalid"
                CustomerOrder._select_types(self,row_data)

            # Scroll handling
            if row_idx >= 5:
                self.driver.execute_script(f"window.scrollBy(0, 100);")

            # Item processing
            if row_data["OrderType"] == "Tag Reserve":
                current_field = "Tag Reserve Flow"
                return CustomerOrder._handle_tag_reserve(self,row_data, row_idx, row_num, next_no, cus_no, sheet_name)
            
            elif row_data["OrderType"] == "Customized Order":
                current_field = "Customized Order Flow"
                return CustomerOrder._handle_customized_order(self,row_data, row_idx, row_num, next_no, cus_no, sheet_name)
            
            return "Fail", f"Unknown Order Type: {row_data['OrderType']}"
        except Exception as e:
            return "Fail", f"Exception in {current_field}: {str(e)}"

    def _fill_header(self, row_data, row_num, sheet_name):
        """Standardized header filling."""
        # Customer
        if row_data["Customer Number"]:
            self.fc.fill_autocomplete_field("cus_name", str(row_data["Customer Number"]))
        else:
            self.fc.Remark(row_num, "Customer is mandatory", sheet_name)
            return False

        # Branch
        if row_data["OrderBranch"]:
            self.fc.dropdown_select("//span[@id='select2-branch_select-container']", 
                                   row_data["OrderBranch"], 
                                   "//span[@class='select2-search select2-search--dropdown']/input")
        else:
            self.fc.Remark(row_num, "Branch is mandatory", sheet_name)
            return False

        # Employee
        if row_data["Employee"]:
            self.fc.dropdown_select("//span[@id='select2-issue_employee-container']", 
                                   row_data["Employee"], 
                                   "//span[@class='select2-search select2-search--dropdown']/input")
        else:
            self.fc.Remark(row_num, "Employee is mandatory", sheet_name)
            return False
            
        return True

    def _select_types(self, row_data):
        """Radios for Balance, Order, Rate."""
        # Balance Type
        self.fc.click("//input[@id='metal_bal_type']" if row_data.get("BalanceType") == "Metal Balance" else "//input[@id='cash_bal_type']")
        # Order Type
        self.fc.click("//input[@id='tag_order']" if row_data.get("OrderType") == "Tag Reserve" else "//input[@id='customer_order']")
        # Rate Type
        self.fc.click("//input[@id='order_rate']" if row_data.get("RateType") == "Order Rate(Fixed)" else "//input[@id='delivery_rate']")

    def _handle_customized_order(self, row_data, row_idx, row_num, next_no, cus_no, sheet_name):
        """Individual Customized Order item entry."""
        self.fc.click("//button[@id='add_order_item']")
        sleep(1)

        # Dropdowns
        for field, value in [("prod", row_data["Product"]), ("dsgn", row_data["Design"]), ("sub_design", row_data["SubDesign"])]:
            if value:
                self.fc.dropdown_select(f"//span[@id='select2-{field}_{row_idx}-container']", value, "//span[@class='select2-search select2-search--dropdown']/input")

        if row_data.get("Purity"):
            self.fc.select_visible_text(f"//select[@id='purity{row_idx}']", row_data["Purity"])

        # Inputs
        if row_data.get("GrossWt"):
            self.fc.fill_input(self.wait, (By.ID, f"weight_{row_idx}"), row_data["GrossWt"], "GrossWt", row_num, Sheet_name=sheet_name)
        
        if row_data.get("LessWt") == "Yes":
            self.fc.click(f"//input[@name='o_item[{row_idx}][less_wt]']")
            Stone.test_tagStone(self, "Customer_LWT", row_data["TestCaseId"])

        self.fc.fill_input(self.wait, (By.ID, f"qty_{row_idx}"), row_data["Pcs"], "Pcs", row_num, Sheet_name=sheet_name)
        self.fc.fill_input(self.wait, (By.NAME, f"o_item[{row_idx}][wast_percent]"), row_data["Wast%"], "Wast%", row_num, Sheet_name=sheet_name)

        # MC
        self.fc.click(f"(//span[contains(@id, 'id_mc_type')])[{row_idx}]")
        sleep(0.5)
        mc_search = self.wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/span[2]/span/span[1]/input")))
        mc_search.send_keys(row_data["MC_Type"], Keys.ENTER)
        self.fc.fill_input(self.wait, (By.NAME, f"o_item[{row_idx}][mc]"), row_data["MC_Value"], "MC", row_num, Sheet_name=sheet_name)

        if row_data.get("OtherCharge") == "Yes":
            self._fill_other_charges(row_data, row_idx)
            
        rate = self._get_dynamic_rate(row_idx)
        self.fc.fill_input(self.wait, (By.XPATH, f"//*[@id='detail{row_idx}']/td[20]/input"), rate, "Rate", row_num, Sheet_name=sheet_name)

        if row_data.get("DueDate"):
            self._fill_due_date(row_data, row_idx)

        if row_data.get("Description"):
            self._fill_description(row_data, row_idx)

        return self._verify_and_submit(row_data, row_idx, next_no, cus_no)

    def _handle_tag_reserve(self, row_data, row_idx, row_num, next_no, cus_no, sheet_name):
        """Tag Reserve entry."""
        self.fc.fill_input2("//input[@id='est_tag_scan']", row_data["TagScan"])
        self.fc.click("//button[@id='tag_search']")
        sleep(1)

        if row_data.get("DueDate"):
            self._fill_due_date(row_data, row_idx)

        if row_data.get("Description"):
            self._fill_description(row_data, row_idx)

        return self._verify_and_submit(row_data, row_idx, next_no, cus_no)

    def _fill_due_date(self, row_data, row_idx):
        due_date_val = row_data.get("DueDate")
        if not due_date_val: return
        
        if hasattr(due_date_val, 'strftime'):
            due_date_str = due_date_val.strftime("%d-%m-%Y")
        else:
            due_date_str = str(due_date_val).split(" ")[0].replace("/", "-")
            # Convert YYYY-MM-DD to DD-MM-YYYY if needed
            if "-" in due_date_str and len(due_date_str) == 10 and due_date_str[4] == "-":
                due_date_str = f"{due_date_str[8:10]}-{due_date_str[5:7]}-{due_date_str[0:4]}"

        try:
            self.driver.execute_script(f"document.getElementsByName('o_item[{row_idx}][due_date]')[0].removeAttribute('readonly');")
            due_input = self.wait.until(EC.presence_of_element_located((By.NAME, f"o_item[{row_idx}][due_date]")))
            due_input.clear()
            due_input.send_keys(due_date_str)
            # Trigger change event to ensure any JS listeners fire
            self.driver.execute_script(f"document.getElementsByName('o_item[{row_idx}][due_date]')[0].dispatchEvent(new Event('change'));")
        except Exception as e:
            print(f"⚠️ Failed to set Due Date: {e}")

    def _fill_description(self, row_data, row_idx):
        self.fc.click(f"//tr[@id='detail{row_idx}']/td[24]/a")
        self.fc.fill_input2("//textarea[@id='description']", str(row_data["Description"]))
        sleep(0.5)
        self.fc.click("//a[contains(text(), 'Add')]")

    def _fill_other_charges(self, row_data, row_idx):
        self.fc.click(f"//*[@id='detail{row_idx}']/td[17]/a")
        charges_raw = row_data.get("ChargeName")
        if not charges_raw: return

        charges_list = [s.strip() for s in str(charges_raw).split(",")]
        for idx, charge in enumerate(charges_list):
            if idx > 0: self.fc.click("//button[@id='add_new_charge']")
            self.fc.select_visible_text(f"(//select[@name='est_stones_item[id_charge][]'])[{idx+1}]", charge)
            # Default value if 0
            val_xpath = f"(//input[@name='est_stones_item[value_charge][]'])[{idx+1}]"
            if self.fc.get_value(val_xpath) in ["0.00", "0", ""]: 
                self.fc.fill_input2(val_xpath, "500")

        self.fc.click("//button[@id='update_charge_details']")

    def _verify_and_submit(self, row_data, row_idx, next_no, cus_no):
        """Calculation check and Save."""
        cal_val_raw = self.fc.get_value(f"//input[@name='o_item[{row_idx}][calculation_based_on]']")
        rate = self._get_dynamic_rate(row_idx)
        
        calc_res = self._calculation(cal_val_raw, row_idx, rate)
        ceil_val, table_val, cal_type = calc_res
        
        if abs(float(ceil_val or 0) - float(table_val or 0)) > 1:
            print(f"❌ Calculation mismatch in {cal_type}: Expected {ceil_val}, Table {table_val}")
            # Optional: Screenshot
            return "Fail", f"Calc mismatch: {cal_type}"
            
        try:
            ui_rate = self.fc.get_value(f"//input[@name='o_item[{row_idx}][order_rate]']")
            row_data["Captured_UI_Rate"] = ui_rate if ui_rate else None
        except Exception:
            row_data["Captured_UI_Rate"] = None
        
        # If there are more rows for this customer, don't submit yet
        if next_no == cus_no:
            return "Continue"
        
        # Final Submit
        self.fc.click("//button[@id='create_order']")
        sleep(2)
        
        try:
            msg_xpath = "//div[contains(@class, 'alert-success')] | //div[contains(@id, 'toast-container')]"
            msg = self.fc.get_text(msg_xpath).replace("×", "").strip()
            if "successfully" in msg.lower():
                # Capture Order Number from listing page
                captured_id = self._capture_order_no(row_data)
                return "Pass", msg, captured_id
            return "Fail", msg
        except UnexpectedAlertPresentException as e:
            alert_text = e.alert_text or "Required fields missing"
            self.driver.switch_to.alert.accept()
            return "Fail", f"Alert: {alert_text}"
        except Exception as e:
            return "Fail", f"Submission failed or Capture Error: {str(e)}"

    def _capture_order_no(self, row_data):
        """Filters the list and captures the Order Number."""
        try:
            # 1. Select Branch
            branch = str(row_data.get("OrderBranch", ""))
            if branch:
                self.fc.dropdown_select("//span[@id='select2-branch_select-container']", 
                                       branch, 
                                       "//span[@class='select2-search select2-search--dropdown']/input")
                sleep(1)

            # 2. Select Today Date Range
            self.fc.click("//button[contains(., 'Date range picker')]")
            sleep(1)
            self.fc.click("//li[contains(text(), 'Today')]")
            sleep(1)

            # 3. Click Search
            self.fc.click("//button[@id='search' or text()='Search']")
            sleep(2)

            # 4. Extract Order NO (2nd column in the first row)
            order_no = self.fc.get_text("//table[@id='order_list']//tbody/tr[1]/td[2]")
            print(f"🎯 Captured Order NO: {order_no}")
            return order_no
        except Exception as e:
            print(f"⚠️ Error capturing Order NO: {e}")
            return None

    def _get_dynamic_rate(self, row_idx):
        """Get board rate matching selected purity."""
        try:
            from selenium.webdriver.support.ui import Select
            purity = Select(self.driver.find_element(By.ID, f"purity{row_idx}")).first_selected_option.text.strip()
            if "916" in purity: return self.Board_Rate[0]
            if "75" in purity: return self.Board_Rate[1]
            if "92.5" in purity: return self.Board_Rate[2]
            return self.Board_Rate[0] # Default
        except:
            return self.Board_Rate[0] if self.Board_Rate else 0

    def _get_header_rates(self):
        """Read standard board rates from header."""
        self.fc.click("//span[@class='header_rate']/b[contains(text(),'INR')]")
        sleep(0.5)
        r1 = self.wait.until(EC.presence_of_element_located((By.XPATH, "//li[@class='user-body rate_block_body']//tr[th[contains(text(),'Gold 22KT 1gm')]]/td"))).text
        r2 = self.wait.until(EC.presence_of_element_located((By.XPATH, "//li[@class='user-body rate_block_body']//tr[th[contains(text(),'Gold 18KT 1gm')]]/td"))).text
        r3 = self.wait.until(EC.presence_of_element_located((By.XPATH, "//li[@class='user-body rate_block_body']//tr[th[contains(text(),'Silver 1gm')]]/td"))).text
        
        return [
            int(float(r1.replace("INR", "").strip())),
            int(float(r2.replace("INR", "").strip())),
            int(float(r3.replace("INR", "").strip()))
        ]

    def _calculation(self, cal_type_idx, row_idx, gold_rate):
        """Ported calculation logic from original file."""
        data_map = {"0": "Mc & Wast On Gross", "1": "Mc & Wast On Net", "2": "Mc on Gross, Wast On Net", "3": "Fixed Rate", "4": "Fixed Rate based on Weight"}
        cal_type = data_map.get(str(cal_type_idx), "Unknown")
        
        def gv(name):
            try: return float(self.fc.get_value(f"//input[@name='o_item[{row_idx}][{name}]']"))
            except: return 0.0

        gwt = gv("weight"); nwt = gv("net_wt"); stone = gv("stn_amt")
        wast = gv("wast_percent"); mc_rate = gv("mc"); other = gv("value_charge")
        table_taxable = gv("taxable")
        
        try:
            mc_type = self.driver.find_element(By.XPATH, f'//span[contains(@id,"o_item[{row_idx}][id_mc_type]")]').get_attribute("title")
        except:
            mc_type = "Gram"

        res = 0.0
        if cal_type == "Mc on Gross, Wast On Net":
            mc = mc_rate if mc_type == 'Piece' else math.ceil(mc_rate * gwt)
            va = round((wast / 100) * nwt, 3)
            res = ((nwt + va) * gold_rate) + stone + mc + other
        elif cal_type == "Mc & Wast On Net":
            mc = mc_rate if mc_type == 'Piece' else math.ceil(mc_rate * nwt)
            va = round((wast / 100) * nwt, 3)
            res = ((nwt + va) * gold_rate) + mc + stone + other
        elif cal_type == "Mc & Wast On Gross":
            mc = mc_rate if mc_type == 'Piece' else math.ceil(mc_rate * gwt)
            va = round((wast / 100) * gwt, 3)
            res = ((nwt + va) * gold_rate) + mc + stone + other
        elif cal_type == "Fixed Rate based on Weight":
            mc = mc_rate if mc_type == 'Piece' else math.ceil(mc_rate * gwt)
            va = round((wast / 100) * gwt, 3)
            res = ((nwt + va) * gold_rate) + mc + stone + other
        elif cal_type == "Fixed Rate":
            res = table_taxable

        return math.ceil(res), table_taxable, cal_type

    def _update_excel_status(self, row_num, status, message, sheet_name, captured_id=None):
        wb = load_workbook(FILE_PATH); sh = wb[sheet_name]
        color = "00B050" if status == "Pass" else "FF0000"
        sh.cell(row=row_num, column=2, value=status).font = Font(bold=True, color=color)
        sh.cell(row=row_num, column=3, value=message).font = Font(bold=True, color=color)
        
        # Save captured Order NO to column 32 (AF)
        if captured_id:
            sh.cell(row=row_num, column=34, value=captured_id).font = Font(bold=True)
            
        else:
            print('Order NO not captured')
        wb.save(FILE_PATH); wb.close()

    def _update_purchase_po_sheet(self, row_data, captured_id):
        """Appends the 'PurchasePO' or 'CustomerOrderKarigarAllotment' sheet based on OrderFlow."""
        if row_data.get("OrderType") != "Customized Order":
            return
            
        order_flow = str(row_data.get("OrderFlow", "")).strip()
        if order_flow != "PurchaseOrder":
            try:
                allotment_sheet = "CustomerOrderKarigarAllotment"
                wb = load_workbook(FILE_PATH)
                if allotment_sheet not in wb.sheetnames:
                    print(f"⚠️ Sheet '{allotment_sheet}' not found in Excel.")
                    wb.close()
                    return

                sh = wb[allotment_sheet]
                
                next_row = sh.max_row + 1
                
                def get_next_tc_id():
                    last_id = sh.cell(row=next_row - 1, column=1).value
                    if not last_id or not isinstance(last_id, str):
                        return "TC002" if next_row > 2 else "TC001"
                    
                    match = re.search(r"(\d+)", last_id)
                    if match:
                        num_str = match.group(0)
                        new_num = int(num_str) + 1
                        return last_id.replace(num_str, str(new_num).zfill(len(num_str)))
                    return "TC001"

                new_tc_id = get_next_tc_id()

                mappings = {
                    1: new_tc_id,
                    4: captured_id,
                    7: row_data.get("DueDate"),
                    9: "Assign"
                }

                for col, value in mappings.items():
                    if value is not None:
                        if col == 7 and hasattr(value, 'strftime'):
                            value = value.strftime("%d-%m-%Y")
                        sh.cell(row=next_row, column=col, value=value).font = Font(bold=True)

                wb.save(FILE_PATH)
                wb.close()
                print(f"✅ CustomerOrderKarigarAllotment sheet appended at row {next_row} with ID {new_tc_id} and Order {captured_id}")
            except Exception as e:
                print(f"❌ Failed to append to CustomerOrderKarigarAllotment sheet: {e}")
                
            try:
                orderlink_sheet = "OrderLink"
                wb = load_workbook(FILE_PATH)
                if orderlink_sheet not in wb.sheetnames:
                    print(f"⚠️ Sheet '{orderlink_sheet}' not found in Excel.")
                    wb.close()
                else:
                    sh = wb[orderlink_sheet]
                    next_row = sh.max_row + 1
                    
                    def get_next_tc_id_ol():
                        last_id = sh.cell(row=next_row - 1, column=1).value
                        if not last_id or not isinstance(last_id, str):
                            return "TC002" if next_row > 2 else "TC001"
                        match = re.search(r"(\d+)", last_id)
                        if match:
                            num_str = match.group(0)
                            new_num = int(num_str) + 1
                            return last_id.replace(num_str, str(new_num).zfill(len(num_str)))
                        return "TC001"

                    new_tc_id_ol = get_next_tc_id_ol()

                    mappings_ol = {
                        1: new_tc_id_ol,
                        4: row_data.get("OrderBranch"),
                        5: "FY 26 - 27",
                        6: captured_id,
                        9: row_data.get("Product"),
                        10: row_data.get("Design"),
                        11: row_data.get("SubDesign")
                    }

                    for col, value in mappings_ol.items():
                        if value is not None:
                            sh.cell(row=next_row, column=col, value=value).font = Font(bold=True)

                    wb.save(FILE_PATH)
                    wb.close()
                    print(f"✅ OrderLink sheet appended at row {next_row} with ID {new_tc_id_ol} and Order {captured_id}")
            except Exception as e:
                print(f"❌ Failed to append to OrderLink sheet: {e}")

            return
            
        try:
            purchase_po_sheet = "PurchasePO"
            wb = load_workbook(FILE_PATH)
            if purchase_po_sheet not in wb.sheetnames:
                print(f"⚠️ Sheet '{purchase_po_sheet}' not found in Excel.")
                wb.close()
                return

            sh = wb[purchase_po_sheet]
            
            # Find the next available row (append mode)
            next_row = sh.max_row + 1
            
            # Generate sequential Test Case ID (from TC001, TC002...)
            def get_next_tc_id():
                last_id = sh.cell(row=next_row - 1, column=1).value
                if not last_id or not isinstance(last_id, str):
                    return "TC002" if next_row > 2 else "TC001" # Handle header only case
                
                match = re.search(r"(\d+)", last_id)
                if match:
                    num_str = match.group(0)
                    new_num = int(num_str) + 1
                    return last_id.replace(num_str, str(new_num).zfill(len(num_str)))
                return "TC001"

            new_tc_id = get_next_tc_id()
            
            rate_type_val = "Fixed" if str(row_data.get("RateType")).strip() == "Order Rate(Fixed)" else "Unfixed"

            # Rate logic for col 9
            final_rate = row_data.get("Captured_UI_Rate") if rate_type_val == "Fixed" else None

            # Mapping based on user request
            mappings = {
                1: new_tc_id,                             # Col 1: Test Case ID
                4: row_data.get("OrderType"),             # Col 4: Order For update OrderType
                6: row_data.get("DueDate"),               # Col 6: Due Date
                7: captured_id,                           # Col 7: CustomerOrderNo
                8: rate_type_val,                         # Col 8: Rate Type
                9: final_rate                             # Col 9: Rate
            }

            for col, value in mappings.items():
                if value is not None:
                    # Store dates natively or formatted for clarity
                    if col == 6 and hasattr(value, 'strftime'):
                        value = value.strftime("%d-%m-%Y")
                    sh.cell(row=next_row, column=col, value=value).font = Font(bold=True)

            wb.save(FILE_PATH)
            wb.close()
            print(f"✅ PurchasePO sheet appended at row {next_row} with ID {new_tc_id} and Order {captured_id}")
        except Exception as e:
            print(f"❌ Failed to append to PurchasePO sheet: {e}")

    def _update_billing_sheet(self, row_data, captured_id, row_num):
        """Appends the 'Billing' sheet with sequential TC ID and order details."""
        try:
            billing_sheet = "Billing"
            wb = load_workbook(FILE_PATH)
            if billing_sheet not in wb.sheetnames:
                print(f"⚠️ Sheet '{billing_sheet}' not found in Excel.")
                wb.close()
                return

            sh = wb[billing_sheet]
            
            # Find the next available row (append mode)
            next_row = sh.max_row + 1
            
            # Generate sequential Test Case ID (from TC001, TC002...)
            def get_next_tc_id():
                last_id = sh.cell(row=next_row - 1, column=1).value
                if not last_id or not isinstance(last_id, str):
                    return "TC002" if next_row > 2 else "TC001" # Handle header only case
                
                match = re.search(r"(\d+)", last_id)
                if match:
                    num_str = match.group(0)
                    new_num = int(num_str) + 1
                    return last_id.replace(num_str, str(new_num).zfill(len(num_str)))
                return "TC001"

            new_tc_id = get_next_tc_id()

            # Mapping based on user request and screenshots
            mappings = {
                1: new_tc_id,                             # A: TestCaseId (Incremented)
                4: row_data.get("OrderBranch"),          # D: Cost Centre
                5: "Customer",                            # E: Billing To
                6: "111-Developer Logimax",               # F: Employee
                7: row_data.get("Customer Number"),       # G: Customer Number
                9: "CALICUT",                             # I: Delivery Location
                10: "ORDER ADVANCE",                     # J: Bill Type
                11: "No",                                 # K: direct
                24: 10000,                                # X: Cash
                31: captured_id,                          # AE: OrderNo
                32: 10000                                 # AF: Amount
            }

            for col, value in mappings.items():
                if value is not None:
                    sh.cell(row=next_row, column=col, value=value).font = Font(bold=True)

            wb.save(FILE_PATH)
            wb.close()
            print(f"✅ Billing sheet appended at row {next_row} with ID {new_tc_id}")
        except Exception as e:
            print(f"❌ Failed to append to Billing sheet: {e}")

    def _resolve_tags(self, sheet_name, valid_rows):
        """Resolves tags for 'Tag Reserve' items from detail sheets."""
        self.used_tags_coords = getattr(self, 'used_tags_coords', [])
        wb = load_workbook(FILE_PATH)
        co_sheet = wb[sheet_name]
        
        detail_sheets = {}
        for ds in ["Tag_Detail", "NonTag_Detail"]:
            if ds in wb.sheetnames:
                detail_sheets[ds] = wb[ds]

        updated = False
        for r in range(2, valid_rows):
            order_type = str(co_sheet.cell(row=r, column=9).value or "").strip()
            tag_scan = str(co_sheet.cell(row=r, column=11).value or "").strip()
            
            if order_type == "Tag Reserve" and not tag_scan:
                prod = str(co_sheet.cell(row=r, column=12).value or "").strip().lower()
                dsgn = str(co_sheet.cell(row=r, column=13).value or "").strip().lower()
                subd = str(co_sheet.cell(row=r, column=14).value or "").strip().lower()
                
                matched = False
                for d_name, d_sheet in detail_sheets.items():
                    headers = {str(cell.value).strip().lower(): i for i, cell in enumerate(d_sheet[1], 1) if cell.value}
                    prod_col = headers.get("product")
                    dsgn_col = headers.get("design")
                    sub_col = headers.get("sub design") or headers.get("subdesign")
                    status_col = headers.get("status")
                    tag_col = headers.get("tag no")
                    
                    if not all([prod_col, dsgn_col, sub_col, status_col]) or not tag_col:
                        continue
                    
                    for dr in range(2, d_sheet.max_row + 1):
                        r_prod = str(d_sheet.cell(row=dr, column=prod_col).value or "").strip().lower()
                        r_dsgn = str(d_sheet.cell(row=dr, column=dsgn_col).value or "").strip().lower()
                        r_sub = str(d_sheet.cell(row=dr, column=sub_col).value or "").strip().lower()
                        r_status = str(d_sheet.cell(row=dr, column=status_col).value or "").strip().lower()
                        r_tag = str(d_sheet.cell(row=dr, column=tag_col).value or "").strip()
                        
                        if not r_tag:
                            continue
                            
                        # Condition: no estimate and billed
                        if "estimate" in r_status or "billed" in r_status or "tagreserve" in r_status:
                            continue
                        
                        if r_prod == prod and r_dsgn == dsgn and r_sub == subd:
                            if (d_name, dr) not in self.used_tags_coords:
                                self.used_tags_coords.append((d_name, dr))
                                co_sheet.cell(row=r, column=11).value = r_tag
                                updated = True
                                matched = True
                                print(f"✅ Found matching Tag '{r_tag}' for Row {r} from {d_name} (Row {dr})")
                                break
                    if matched:
                        break
                
                if not matched:
                    print(f"⚠️ No matching tag found for Row {r} (Product: {prod}, Design: {dsgn})")
        
        if updated:
            wb.save(FILE_PATH)
        wb.close()

    def _update_tag_details_with_order(self, tags_list, order_number):
        if not tags_list:
            return
            
        wb = load_workbook(FILE_PATH)
        updated = False
        for ds in ["Tag_Detail", "NonTag_Detail"]:
            if ds in wb.sheetnames:
                sheet = wb[ds]
                headers = {str(cell.value).strip().lower(): i for i, cell in enumerate(sheet[1], 1) if cell.value}
                tag_col = headers.get("tag no")
                status_col = headers.get("status")
                
                if not tag_col or not status_col: continue
                
                for dr in range(2, sheet.max_row + 1):
                    r_tag = str(sheet.cell(row=dr, column=tag_col).value or "").strip()
                    if r_tag in tags_list:
                        order_col = status_col + 1
                        if not sheet.cell(row=1, column=order_col).value:
                            sheet.cell(row=1, column=order_col).value = "Ordernumber"
                        
                        sheet.cell(row=dr, column=status_col).value = "Tagreserve"
                        sheet.cell(row=dr, column=order_col).value = order_number
                        updated = True
                        print(f"✅ Updated {ds} Row {dr} for Tag {r_tag} -> Tagreserve, {order_number}")
        
        if updated:
            wb.save(FILE_PATH)
        wb.close()

if __name__ == "__main__":
    unittest.main()
