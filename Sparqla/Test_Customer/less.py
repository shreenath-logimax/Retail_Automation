from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from time import sleep
import math
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call

FILE_PATH = ExcelUtils.file_path

class Stone:
    """Helper class for handling Less Weight / Stone details."""
    
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 10)
        self.fc = Function_Call(driver)

    def test_tagStone(self, sheet_name, test_case_id):
        """Main method to fill stone details for a specific test case."""
        wait = self.wait
        sleep(2)
        
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            count = ExcelUtils.Test_case_id_count(FILE_PATH, sheet_name, test_case_id)
            workbook = load_workbook(FILE_PATH)
            sheet = workbook[sheet_name]
        except Exception as e:
            print(f"❌ Failed to load stone data for {sheet_name}: {e}")
            return None

        row_idx = 1
        processed_count = 0
        all_stones_data = []

        for row_num in range(2, valid_rows + 1):
            current_id = sheet.cell(row=row_num, column=1).value
            if str(current_id).strip() == str(test_case_id).strip():
                data_map = {
                    "Type": 3, "Name": 4, "Code": 5, "Pcs": 6, 
                    "Wt": 7, "Wt Type": 8, "Cal.Type": 9, "Rate": 10, "Amount": 11
                }
                row_data = {k: sheet.cell(row=row_num, column=v).value for k, v in data_map.items()}
                
                print(f"💎 Adding Stone: {row_data['Name']} (Row {row_idx})")
                
                # Fill Dropdowns
                self._select_dropdown(f"(//select[@name='est_stones_item[stones_type][]'])[{row_idx}]", row_data["Type"])
                self._select_dropdown(f"(//select[@name='est_stones_item[stone_id][]'])[{row_idx}]", row_data["Name"])

                # Fill Inputs
                self._fill_input(f"(//input[@name='est_stones_item[stone_pcs][]'])[{row_idx}]", row_data["Pcs"])
                self._fill_input(f"(//input[@name='est_stones_item[stone_wt][]'])[{row_idx}]",  row_data["Wt"])
                self._select_dropdown(f"(//select[@name='est_stones_item[stone_uom_id][]'])[{row_idx}]", row_data["Wt Type"])

                # Cal.Type Radio
                stone_row_param = row_idx - 1
                cal_type_val = 1 if str(row_data["Cal.Type"]).strip().lower() == "wt" else 2
                cal_xpath = f"(//input[@name='est_stones_item[cal_type][{stone_row_param}]' and @value='{cal_type_val}'])"
                wait.until(EC.element_to_be_clickable((By.XPATH, cal_xpath))).click()

                # Rate
                self._fill_input(f"(//input[@name='est_stones_item[stone_rate][]'])[{row_idx}]", row_data["Rate"])
                
                # Validation
                amt_xpath = f'(//input[@name="est_stones_item[stone_price][]"])[{row_idx}]'
                table_amt = self.fc.get_value(amt_xpath)
                self._validate_calc(row_data, table_amt)

                processed_count += 1
                all_stones_data.append(row_data)

                # Add more rows if needed
                if processed_count < count:
                    wait.until(EC.element_to_be_clickable((By.ID, "create_stone_item_details"))).click()
                    row_idx += 1
                    sleep(1)

        # Save and return
        if processed_count > 0:
            wait.until(EC.element_to_be_clickable((By.ID, "update_stone_details"))).click()
            print(f"✅ Successfully added {processed_count} stone(s)")
            
            # Return totals (Wt, Amt) if needed
            wt_total = self._get_table_total("est_stones_item[stone_wt][]")
            amt_total = self._get_table_total("est_stones_item[stone_price][]")
            
            return True, wt_total, amt_total, all_stones_data
        
        workbook.close()
        return None

    def _select_dropdown(self, locator, value):
        if value is None: return
        try:
            el = self.wait.until(EC.presence_of_element_located((By.XPATH, locator)))
            Select(el).select_by_visible_text(str(value))
        except Exception as e:
            print(f"⚠️ Dropdown failed: {locator} | {e}")

    def _fill_input(self, locator, value):
        if value is None: return
        try:
            el = self.wait.until(EC.visibility_of_element_located((By.XPATH, locator)))
            el.clear()
            el.send_keys(str(value))
        except Exception as e:
            print(f"⚠️ Input failed: {locator} | {e}")

    def _validate_calc(self, data, table_amt):
        try:
            rate = float(data["Rate"] or 0)
            pcs = float(data["Pcs"] or 0)
            wt = float(data["Wt"] or 0)
            cal_type = str(data["Cal.Type"]).strip().lower()
            
            expected = math.ceil(wt * rate) if cal_type == "wt" else math.ceil(pcs * rate)
            found = math.ceil(float(table_amt or 0))
            
            if abs(expected - found) > 1:
                print(f"❌ Stone Calc mismatch: Expected {expected}, Found {found}")
        except:
            pass

    def _get_table_total(self, name):
        """Sum up values from all inputs with given name."""
        sleep(1)
        els = self.driver.find_elements(By.NAME, name)
        total = 0.0
        for el in els:
            try:
                val = el.get_attribute("value")
                if val: total += float(val)
            except: pass
        return round(total, 3)