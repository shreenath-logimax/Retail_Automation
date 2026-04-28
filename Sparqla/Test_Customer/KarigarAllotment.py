from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException
from Utils.Function import Function_Call
from Utils.Excel import ExcelUtils
from openpyxl import load_workbook
from openpyxl.styles import Font
from time import sleep
import unittest

FILE_PATH = ExcelUtils.file_path

class CustomerOrderKarigarAllotment(unittest.TestCase):
    """
    Customer Order Karigar Allotment Module Automation.
    Assigns customer orders to a Karigar or Employee in bulk list page.
    """

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)
        self.fc = Function_Call(driver)

    def test_customer_order_allotment(self):
        driver = self.driver
        wait = self.wait
        sheet_name = "CustomerOrderKarigarAllotment"

        # Read Excel data
        try:
            valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
            print(f"✅ Found {valid_rows - 1} test cases for {sheet_name}")
        except Exception as e:
            print(f"❌ Failed to read Excel: {e}")
            return

        # Navigation
        try:
            if "admin_ret_order/customer_order/neworders" not in driver.current_url:
                wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
                sleep(1)
                self.fc.click("//span[contains(text(), 'Customer Orders')]")
                sleep(1)
                # Looking for the Karigar Allotment or New Order list
                # Based on common patterns:
                self.fc.click("(//span[contains(text(), 'Karigar Allotment')])[1]")
                sleep(2)
        except Exception as e:
            print(f"⚠️ Navigation failed: {e}")
            driver.get(ExcelUtils.BASE_URL + "index.php/admin_ret_order/customer_neworders/list")
            sleep(3)

        for row_num in range(2, valid_rows):
            try:
                workbook = load_workbook(FILE_PATH)
                sheet = workbook[sheet_name]
            except Exception as e:
                print(f"Error loading workbook for row {row_num}: {e}")
                continue

            data_map = {
                "TestCaseId": 1, "TestStatus": 2, "ActualStatus": 3,
                "OrderNo": 4, "AssignTo": 5, "AssignName": 6,
                "SmithDueDate": 7, "Action": 8, "Remark": 9
            }

            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            workbook.close()

            # if str(row_data["TestStatus"]).strip().lower() != "run":
            #     continue

            print(f"🚀 Running TC: {row_data['TestCaseId']} - Order: {row_data['OrderNo']}")

            try:
                result = self.test_allotment_flow(row_data)
                self._update_excel_status(row_num, result[0], result[1], sheet_name)
            except Exception as e:
                self._update_excel_status(row_num, "Fail", str(e), sheet_name)

    def test_allotment_flow(self, row_data):
        driver = self.driver
        wait = self.wait
        
        order_no = str(row_data["OrderNo"]) if row_data.get("OrderNo") else None
        assign_to = str(row_data["AssignTo"]).strip().lower() if row_data.get("AssignTo") else ""
        assign_name = str(row_data["AssignName"]) if row_data.get("AssignName") else None
        smith_due_date = str(row_data.get("SmithDueDate", "")).strip()[:10] if row_data.get("SmithDueDate") else None
        action = str(row_data["Action"]).strip().lower() if row_data.get("Action") else "assign"

        if not order_no:
            return ("Fail", "OrderNo is required")

        # 1. Search Order
        search_box = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@type='search']")))
        search_box.clear()
        search_box.send_keys(order_no)
        sleep(2)

        # 2. Select Row
        table_id = "neworder_list" # Common table ID or customerorder_list
        row_xpath = f"//table[contains(@id, 'order_list') or contains(@id, 'neworder_list')]/tbody/tr[contains(., '{order_no}')]//input[@type='checkbox']"
        try:
            checkbox = wait.until(EC.presence_of_element_located((By.XPATH, row_xpath)))
            if not checkbox.is_selected():
                try:
                    driver.execute_script("arguments[0].click();", checkbox)
                except:
                    label_xpath = f"//table[contains(@id, 'order_list') or contains(@id, 'neworder_list')]/tbody/tr[contains(., '{order_no}')]//label"
                    self.fc.click(label_xpath)
            sleep(1)
        except:
            return ("Fail", f"Order {order_no} not found")

        # 3. Assign Role
        if assign_to == "employee":
            self.fc.click("//input[@name='order[assign_to]' and @value='2']")
            target_select_id = "employee_sel"
        else:
            self.fc.click("//input[@name='order[assign_to]' and @value='1']")
            target_select_id = "karigar_sel"
        sleep(1)

        # 4. Pick Worker
        if assign_name:
            trigger_xpath = f"//select[@id='{target_select_id}']/following-sibling::span"
            search_xpath = "//span[@class='select2-search select2-search--dropdown']/input"
            self.fc.dropdown_select(trigger_xpath, assign_name, search_xpath)
            sleep(1)

        # 5. Set Due Date (JS Bypass)
        if smith_due_date and smith_due_date != "None":
            row_input_xpath = f"//table[contains(@id, 'order_list') or contains(@id, 'neworder_list')]/tbody/tr[contains(., '{order_no}')]//input[@type='text' or contains(@class, 'datepicker')]"
            try:
                date_el = driver.find_element(By.XPATH, row_input_xpath)
                driver.execute_script("arguments[0].value = arguments[1];", date_el, smith_due_date)
                driver.execute_script("arguments[0].dispatchEvent(new Event('change'));", date_el)
            except:
                pass

        # 6. Submit
        if action == "reject":
            self.fc.click("//label[@id='reject'] or //button[@id='reject']")
        else:
            # Try both label (Repair style) and direct button ID
            try:
                self.fc.click("//label[@id='approve']")
            except:
                self.fc.click("//button[@id='approve']")

        # 7. Verification
        try:
            success_xpath = "//div[contains(@class, 'alert-success')]"
            msg = wait.until(EC.presence_of_element_located((By.XPATH, success_xpath))).text.replace("×", "").strip()
            print(f"🔔 Toaster Message: {msg}")
            if "success" in msg.lower():
                try:
                    search_box = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@type='search']")))
                    search_box.clear()
                    search_box.send_keys(order_no)
                    sleep(2)
                    
                    row_xpath = f"//table[contains(@id, 'order_list') or contains(@id, 'neworder_list')]/tbody/tr[contains(., '{order_no}')]"
                    row_el = wait.until(EC.presence_of_element_located((By.XPATH, row_xpath)))
                    
                    if "Work in Progress" in row_el.text:
                        return ("Pass", f"Assigned {order_no} successfully")
                    else:
                        return ("Fail", f"Status is not Work in Progress")
                except Exception as e:
                    return ("Fail", f"Failed to verify Work in Progress status: {e}")
            return ("Fail", f"Alert: {msg}")
        except TimeoutException:
            return ("Fail", "Action submitted but no success message (Toaster timeout)")

    def _update_excel_status(self, row_num, status, message, sheet_name):
        wb = load_workbook(FILE_PATH)
        sheet = wb[sheet_name]
        color = "00B050" if status == "Pass" else "FF0000"
        sheet.cell(row=row_num, column=2, value=status).font = Font(bold=True, color=color)
        sheet.cell(row=row_num, column=3, value=message).font = Font(bold=True, color=color)
        wb.save(FILE_PATH)
        wb.close()
