# -*- coding: utf-8 -*-
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import unittest
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from time import sleep

FILE_PATH = ExcelUtils.file_path
class Login(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver
        self.wait = WebDriverWait(driver, 30)
    def test_login(self):
        #self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        #self.driver.implicitly_wait(30)
        self.verificationErrors = []
        self.accept_next_alert = True
        driver = self.driver
        wait = self.wait
        function_name = "Login"
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, function_name)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[function_name]
        print(valid_rows)
        for row_num in range(2, valid_rows):
                # Define columns and dynamically fetch their values
                data = {
                    "Url":5,
                    "Username": 6,
                    "PassWord":7,
                }
                row_data = {key: sheet.cell(row=row_num, column=col).value 
                            for key, col in data.items()}
                print(row_data)
                # Url = LoginAutomation.url(self) 
                # Call add_Payment_Mode
        driver.get(row_data["Url"])
        wait.until(EC.element_to_be_clickable((By.ID,"username"))).click()
        wait.until(EC.element_to_be_clickable((By.ID,"username"))).clear()
        wait.until(EC.element_to_be_clickable((By.ID,"username"))).send_keys(row_data["Username"])
        wait.until(EC.element_to_be_clickable((By.ID,"password"))).click()
        wait.until(EC.element_to_be_clickable((By.ID,"password"))).clear()
        wait.until(EC.element_to_be_clickable((By.ID,"password"))).send_keys(row_data["PassWord"])
        wait.until(EC.element_to_be_clickable((By.ID,"submit_login"))).click()
        Full_text = wait.until(EC.element_to_be_clickable((By.XPATH, "//h1"))).text.strip()
        dashboard_text = Full_text.split()[0]
        print(dashboard_text)
        if dashboard_text==('Dashboard') :
            message="Login page Open successfully "
            Test_status= "Pass"
            print(message,Test_status)
        else:
            message="Login page Not successfully"
            Test_status= "Fail"
            print(message,Test_status)   
        sheet.cell(row=row_num, column=2).value = Test_status 
        
        # Get existing value from cell
        existing_value = sheet.cell(row=row_num, column=3).value

        if existing_value:  # If something already exists
            sheet.cell(row=row_num, column=3).value = str(existing_value) + ", " + message
        else:  # If empty
            sheet.cell(row=row_num, column=3).value = message

        # Save workbook   
        workbook.save(FILE_PATH)    
        sleep(2)
        Status = ExcelUtils.get_Status(FILE_PATH,function_name)  
        print(Status)
        Update_master = ExcelUtils.update_master_status(FILE_PATH,Status,function_name)
        
    def is_element_present(self, how, what):
        try: self.driver.find_element(by=how, value=what)
        except NoSuchElementException as e: return False
        return True
    
    def is_alert_present(self):
        try: self.driver.switch_to_alert()
        except NoAlertPresentException as e: return False
        return True
    
    def close_alert_and_get_its_text(self):
        try:
            alert = self.driver.switch_to_alert()
            alert_text = alert.text
            if self.accept_next_alert:
                alert.accept()
            else:
                alert.dismiss()
            return alert_text
        finally: self.accept_next_alert = True

if __name__ == "__main__":
    unittest.main()
