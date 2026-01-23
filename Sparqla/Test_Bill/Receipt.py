from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import  sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from Utils.Board_rate import Boardrate
from Test_Bill.Sales import SALES
from Test_Bill.Credit_Card import CreditCard
from Test_Bill.Cheque import Cheque
from Test_Bill.NetBanking import NetBanking
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
import re
import unittest

FILE_PATH = ExcelUtils.file_path 
class Receipt(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)

    def test_Receipt(self):
        driver = self.driver
        wait = self.wait 
        
        Rate=Boardrate.Todayrate(self)
        print(Rate)  
        
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Toggle navigation"))).click()
        Function_Call.click(self,"//span[contains(text(), 'Billing')]")
        Function_Call.click(self,"//span[contains(text(), 'Receipt')]")
        
        Sheet_name = "Receipt"                                        
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, Sheet_name)
        print(f"'{valid_rows}': valid rows")
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]
        for row_num in range(2, valid_rows):  
            data = {
                    "Test Case Id": 1,
                    "TestStatus": 2,
                    "ActualStatus": 3,
                    "Select Branch": 4,
                    "Customer": 5,
                    "Receipt Type": 6,
                    "Against Est": 7,
                    "Esti No": 8,
                    "Amount": 9,
                	"Receipt As": 10,
                    "Store As":11,
                    "Rate Calculation From":12,
                    "Amount":13,
                    "Weight":14,
                    "Date":15,
                    "Employee":16,
                    "NetBanking":17
                    
                }
            row_data = {key: sheet.cell(row=row_num, column=col).value 
                            for key, col in data.items()}
            print(row_data)
            # Call you 'create' method
            Create_data = self.create(row_data, row_num, Sheet_name,Rate)
            print(Create_data)
            
    def create(self,row_data, row_num, Sheet_name,Board_Rate):
        driver = self.driver
        wait = self.wait
        driver.refresh()
        Mandatory_field=[]        
        #Cost Centre
        if row_data['Select Branch'] is not None:
            Function_Call.dropdown_select(self,f"//span[@id='select2-branch_select-container']", row_data["Select Branch"],'//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Select Branch field is mandatory ⚠️"
            Mandatory_field.append("Select Branch"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
        
        # Name
        if row_data["Customer"]:
            Function_Call.fill_autocomplete_field(self,"name", row_data["Customer"])
        else:
            msg = f"'{None}' → Customer field is mandatory ⚠️"
            Mandatory_field.append(msg)
            print(msg)
            Function_Call.Remark(row_num, msg)
            sleep(3)
        
        Receipt_Type = {
            "Credit Collection": '//input[@id="receipt_type1"]',
            "Advance": '//input[@id="receipt_type2"]',
            "Chit Close": '//input[@id="receipt_type6"]',
            "Petty Cash Receipt":'//input[@id="receipt_type8"]'
        } 
        Function_Call.click(self,Receipt_Type[row_data["Receipt Type"]])
        
        if row_data["Against Est"]=="Yes":
            Function_Call.click(self,'//input[@id="is_aganist_est_yes"]')
        else:
            Function_Call.click(self,'//input[@id="is_aganist_est_no"]')
        
        if row_data["Esti No"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, '//input[@name="billing[tot_amt_received]"]'),
                value=row_data["Esti No"],
                pattern = r"^(\d{1,2})?)?$",
                field_name="Received",
                screenshot_prefix="Received",
                row_num=row_num,
                Sheet_name=Sheet_name) 
        else:
            pass
        
        if row_data["Receipt As"]:  
            Receipt_As = {
            "Amount": '//input[@id="receipt_as1"]',
            "Weight": '//input[@id="receipt_as2"]',
        } 
        Function_Call.click(self,Receipt_As[row_data["Receipt As"]])
        
             
        if row_data["Store As"]:  
            Store_As = {
            "Amount": '(//input[@id="store_receipt_as_1"])[1]',
            "Weight": '(//input[@id="store_receipt_as_2"])[1]',
        } 
        Function_Call.click(self,Store_As[row_data["Store As"]])
        
        if row_data["Rate Calculation From"]:  
            Rate_Calculation_From = {
            "Gold": '(//input[@id="store_receipt_as_1"])[2]',
            "Silver": '(//input[@id="store_receipt_as_2"])[2]',
        }
        Function_Call.click(self,Rate_Calculation_From[row_data["Rate Calculation From"]])
        
        if row_data["Esti No"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, '//input[@id="amount"]'),
                value=row_data["Esti No"],
                pattern = r"^(\d{1,2})?)?$",
                field_name="Received",
                screenshot_prefix="Received",
                row_num=row_num,
                Sheet_name=Sheet_name)
         
        
        
        
        
        
        