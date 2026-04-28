from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import  sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from Utils.Board_rate import Boardrate
from Test_Bill.Sales import SALES
from Test_Bill.Credit_Card import CreditCard
from Test_Bill.Cheque import Cheque
from Test_Bill.NetBanking import NetBanking
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
import unittest
import random
import string

FILE_PATH = ExcelUtils.file_path 
class Billing(unittest.TestCase):
    def __init__(self, driver):
        self.driver = driver   
        self.wait = WebDriverWait(driver, 30)

    

    def test_Billing(self):
        """Main test entry point for Billing automation."""
        driver = self.driver
        wait = self.wait 
        
        rate = Boardrate.Todayrate(self)
        print(f"Today's Rate: {rate}")  
        
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Toggle navigation"))).click()
        Function_Call.click(self, "//span[contains(text(), 'Billing')]")
        Function_Call.click(self, "//span[contains(text(), 'New Bill')]")
        
        sheet_name = "Billing"                                        
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
        print(f"Processing {valid_rows - 2} rows from {sheet_name}")

        workbook = load_workbook(FILE_PATH)
        sheet = workbook[sheet_name]

        for row_num in range(2, valid_rows):  
            data_map = {
                "Test Case Id": 1, "TestStatus": 2, "ActualStatus": 3, "Cost Centre": 4,
                "Billing To": 5, "Employee": 6, "Customer Number": 7, "Customer Name": 8,
                "Delivery Location": 9, "Bill Type": 10, "driect": 11, "EstNo": 12,
                "SGST": 13, "CGST": 14, "Total": 15, "Discount": 16,
                "Handling_Charges": 17, "Return_Charges": 18, "Is Credit": 19,
                "Is Tobe": 20, "Received": 21, "Credit Due Date": 22,
                "Gift Voucher": 23, "Cash": 24, "Creditcard": 25, "Cheque": 26,
                "NetBanking": 27, "BillNo": 28, "Keep_it_As": 29, "Store As": 30,
                "OrderNo": 31, "Amount": 32, "IGST":33, "RepairAmount": 34
            }
            row_data = {key: sheet.cell(row=row_num, column=col).value for key, col in data_map.items()}
            
            if row_data.get("BillNo"):
                print(f"Row {row_num}: Bill No {row_data['BillNo']} already generated successfully. skipping... ✅")
                continue
            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['Test Case Id']}")
            print(f"{'='*80}")
            self.repair_total = 0  # Reset for each row
            print(f"Starting test case: {row_data['Test Case Id']}")
            
            self.create(row_data, row_num, sheet_name, rate)

        workbook.close()

    def create(self, row_data, row_num, sheet_name, board_rate):
        """Processes a single billing record."""
        self.driver.refresh()
        
        # 1. Handle Header Information
        if not self._handle_bill_header(row_data, row_num, sheet_name):
            return

        # 2. Handle Bill Type Specifics
        if not self._handle_bill_type(row_data, row_num, sheet_name):
            return

        # 3. Handle Calculations and Summary
        if not self._handle_total_summary(row_data, row_num, sheet_name):
            return

        # 4. Process Payments
        self._handle_payments(row_data, row_num, sheet_name)

    def _handle_bill_header(self, row_data, row_num, sheet_name):
        """Fills Cost Centre, Bill To, Employee, and Customer details."""
        # Cost Centre
        if row_data.get('Cost Centre'):
            Function_Call.select_visible_text(self, '//select[@id="id_branch"]', 
                                          value=row_data['Cost Centre'])
        else:
            msg = "Cost Centre field is mandatory ⚠️"
            Function_Call.Remark(self, row_num, msg, sheet_name)
            return False

        # Billing To
        bill_to_map = {
            "Customer": '//input[@id="billing_for1"]',
            "Company": '//input[@id="billing_for2"]',
            "Supplier": '//input[@id="billing_for3"]'
        }
        target_bill_to = row_data.get("Billing To", "Customer")
        Function_Call.click(self, bill_to_map.get(target_bill_to, bill_to_map["Customer"]))

        # Employee
        if row_data.get("Employee"):
            Function_Call.dropdown_select(self, "//span[@id='select2-emp_select-container']", 
                                         row_data["Employee"], 
                                         '//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = "Employee field is mandatory ⚠️"
            Function_Call.Remark(self, row_num, msg, sheet_name)
            return False

        # Customer
        if row_data.get("Customer Number"):
            Function_Call.fill_autocomplete_field(self, "bill_cus_name", str(row_data["Customer Number"]))
        else:
            msg = "Customer field is mandatory ⚠️"
            Function_Call.Remark(self, row_num, msg, sheet_name)
            return False

        Function_Call.click2(self, '(//button[@class="btn btn-close btn-warning"])[11]')
        return True

    def _handle_bill_type(self, row_data, row_num, sheet_name):
        """Selects and configures the bill type."""
        bill_type = row_data.get("Bill Type")
        
        match bill_type:
            case "SALES":
                Function_Call.click(self, '//input[@id="bill_typesales"]')
            case "SALES & PURCHASE":
                Function_Call.click(self, '//input[@id="bill_type_salesPurch"]')
            case "SALES PURCHASE & RETURN":
                Function_Call.click(self, '//input[@id="bill_type_saleRet"]')
                return self._select_last_filter_bill(row_num, sheet_name)
            case "PURCHASE":
                Function_Call.click(self, '//input[@id="bill_type_purchase"]')
            case "ORDER ADVANCE":
                Function_Call.click(self, '//input[@id="bill_type_order_advance"]')
                Function_Call.fill_input2(self, '//input[@id="filter_order_no"]', row_data.get("OrderNo"))
                Function_Call.click(self, '//button[@id="search_order_no"]')
                Function_Call.click2(self, '(//button[@class="btn btn-close btn-warning"])[11]')
                Function_Call.fill_input2(self, '//input[@name="billing[bill_amount]"]', row_data.get("Amount"))
                Function_Call.click(self, '//li[@id="tab_make_pay"]')
            case "SALES RETURN":
                Function_Call.click(self, '//input[@id="bill_type_sales_return"]')
                return self._select_last_filter_bill(row_num, sheet_name)
            case "CREDIT COLLECTION":
                Function_Call.click(self, '//input[@id="bill_type_credit_bill"]')
            case "ORDER DELIVERY":
                Function_Call.click(self, '//input[@id="bill_type_order_del"]')
            case "Repair Order Delivery":
                Function_Call.click(self, '//input[@id="repair_order_delivery"]')
                Function_Call.fill_input2(self, '//input[@id="filter_order_no"]', row_data.get("OrderNo"))
                Function_Call.click(self, '//button[@id="search_order_no"]')
                Function_Call.click2(self, '(//button[@class="btn btn-close btn-warning"])[11]')
                Function_Call.click(self, '//a[@id="tab_items"]')
                # Validate Sales Items Amount and Repair Item Details Amount
                self._validate_repair_order_amounts(row_data, row_num, sheet_name)
        return True

    def _select_last_filter_bill(self, row_num, sheet_name):
        """Helper to select the last bill in the filter dropdown."""
        try:
            sleep(5)
            element = self.wait.until(EC.element_to_be_clickable((By.ID, "filter_bill_no")))
            dropdown = Select(element)
            options_count = len(dropdown.options)
            
            if options_count > 1:
                dropdown.select_by_index(options_count - 1)
                Function_Call.click(self, '//button[@id="search_bill_no"]')
                Function_Call.click(self, '//input[@class="select_est_details"]')
                Function_Call.click(self, '//a[@id="update_bill_return"]')
                return True
            else:
                msg = "No bills available to return ⚠️"
                print(msg)
                Function_Call.Remark(self, row_num, msg, sheet_name)
                return False
        except Exception as e:
            print(f"Error in bill selection: {e}")
            return False

    def _handle_total_summary(self, row_data, row_num, sheet_name):
        """Handles estimations, calculations, and charges in the summary tab."""
        if row_data.get("driect") == 'No' and row_data.get("Bill Type") != "ORDER ADVANCE":
            if row_data.get("Bill Type") != "Repair Order Delivery":
                Function_Call.fill_input2(self, '//input[@id="filter_est_no"]', row_data["EstNo"])
                Function_Call.click(self, '//button[@id="search_est_no"]')
                sleep(2)
                Function_Call.click(self, '(//button[@class="btn btn-close btn-warning"])[11]')
                sleep(3)
                Function_Call.click(self, "//a[normalize-space()='Total Summary']")
            
            
                # Fetch summary details
                summary_labels = {
                    "Taxable Sale Amount": "//span[@class='summary_lbl summary_sale_amt']",
                    "CGST": '//span[@class="summary_lbl sales_cgst"]',
                    "SGST": '//span[@class="summary_lbl sales_sgst"]',
                    "IGST": '//span[@class="summary_lbl sales_igst"]',
                    "Sale Amount": '//span[@class="summary_lbl sale_amt_with_tax"]',
                    "Purchase Amount": '//span[@class="summary_lbl summary_pur_amt"]'
                }
                for label, xpath in summary_labels.items():
                    val = Function_Call.get_text(self, xpath)
                    print(f"{label}: {val}")

                if Function_Call.get_text(self, summary_labels["Purchase Amount"]) == '0.00':
                    self._fill_summary_charges(row_data, row_num, sheet_name)
            else:
                Function_Call.click(self, "//a[normalize-space()='Total Summary']")
            # Navigate to the next tab (Make Payment)
            try:
                Function_Call.click(self, '(//button[@class="btn btn-warning next-tab"])[2]')
            except:
                Function_Call.click(self, '//li[@id="tab_make_pay"]')
        else:
            print(f"Direct Billing or Order Advance detected: Skipping Estimation search for {row_data.get('Bill Type')}")
        return True

    def _fill_summary_charges(self, row_data, row_num, sheet_name):
        """Fills discount, handling, and return charges."""
        charge_map = {
            "Discount": ('//input[@id="summary_discount_amt"]', r"^(\d{1,7}(\.\d{1,2})?)?$"),
            "Handling_Charges": ('//input[@id="handling_charges"]', r"^(\d{1,3}(\.\d{1,2})?)?$"),
            "Return_Charges": ('//input[@id="return_charges"]', r"^(\d{1,3}(\.\d{1,2})?)?$")
        }
        for field, (xpath, pattern) in charge_map.items():
            if row_data.get(field):
                Function_Call.fill_input(
                    self, self.wait, locator=(By.XPATH, xpath),
                    value=row_data[field], pattern=pattern, field_name=field,
                    screenshot_prefix=field, row_num=row_num, Sheet_name=sheet_name
                )

    def _handle_payments(self, row_data, row_num, sheet_name):
        """Manages received amount, credit settings, and specific payment methods."""
        # Ensure we are on the Make Payment tab
        try:
            Function_Call.click(self, '//li[@id="tab_make_pay"]')
        except:
            pass

        # Check for PAN mandatory threshold (2 Lakh)
        # For Repair Order Case, use calculated total; otherwise use Excel
        if row_data.get("Bill Type") == "Repair Order Delivery" and getattr(self, "repair_total", 0):
            total_val = self.repair_total
            print(f"Using calculated Repair Order Total: {total_val}")
        else:
            total_val = row_data.get("Total")

        if total_val:
            try:
                # Normalize total value
                val = float(str(total_val).replace(',', '').strip() if total_val and str(total_val).lower() != 'none' else 0)
                if val >= 200000:
                    pan = self._generate_pan()
                    print(f"Total Amount {val} is >= 2 Lakh, Entering Random PAN: {pan}")
                    Function_Call.fill_input(self, self.wait, (By.ID, "pan_no"), pan, field_name="PAN NO", screenshot_prefix="PAN", row_num=row_num, Sheet_name=sheet_name)
                else:
                    print("Total Amount value is less than 2 Lakh")
            except Exception as e:
                print(f"Error checking Total for PAN: {e}")  

        received_str = Function_Call.get_value(self, '//input[@name="billing[tot_amt_received]"]')
        received_value = float(received_str if received_str and str(received_str).lower() != 'none' else 0)
        cash_val = float(row_data.get("Cash") if row_data.get("Cash") and str(row_data.get("Cash")).lower() != 'none' else 0)
        final_received = received_value

        if row_data.get('Received'):
            percentage = float(row_data['Received'] if row_data['Received'] and str(row_data['Received']).lower() != 'none' else 0)
            credit_val = (received_value * percentage) / 100
            final_received = credit_val
            
            Function_Call.click(self, '//input[@id="is_credit_yes"]')
            if row_data.get('Is Credit') == 'Yes':
                Function_Call.fill_input(
                    self, self.wait, locator=(By.XPATH, '//input[@name="billing[tot_amt_received]"]'),
                    value=credit_val, pattern=r"^(\d{1,7}(\.\d{1,2})?)?$",
                    field_name="Received", screenshot_prefix="Received",
                    row_num=row_num, Sheet_name=sheet_name
                )
            
            if row_data.get('Is Tobe') == 'Yes':
                Function_Call.click(self, '//input[@id="is_to_be_yes"]')
            
            if row_data.get("Credit Due Date"):
                Function_Call.fill_input(
                    self, self.wait, locator=(By.XPATH, '//input[@id="credit_due_date"]'),
                    value=row_data["Credit Due Date"],
                    pattern=r"^(0[1-9]|[12][0-9]|3[01])-(0[1-9]|1[0-2])-\d{4}$",
                    field_name="Credit_Due_Date", screenshot_prefix="Credit_Due_Date",
                    row_num=row_num, Sheet_name=sheet_name, extra_keys=Keys.TAB, Date_range="Yes"
                )

        # Determine adjustable amount
        adjustable_amount = final_received
        if received_value == 0:
            pay_to_cus = float(Function_Call.get_value(self, '//input[@name="billing[pay_to_cus]"]') or 0)
            adjustable_amount = pay_to_cus

        pending_payment = adjustable_amount - cash_val
        
        # Process Cash
        if cash_val > 0:
            Function_Call.fill_input(
                self, self.wait, locator=(By.XPATH, '//input[@id="make_pay_cash"]'),
                value=int(cash_val), pattern=r"^(?:[0-9]{1,5}|1[0-9]{5})$", field_name="Cash",
                screenshot_prefix="Cash", row_num=row_num, Sheet_name=sheet_name
            )

        # Track amounts for summary
        amounts = {"Cash": cash_val, "Card": 0, "Cheque": 0, "NetBanking": 0, "Adjustable": adjustable_amount}
        
        # Process Other Methods
        test_case_id = row_data.get('Test Case Id')
        
        # Prepare sub-amounts based on sub-sheet logic
        if row_data.get('Creditcard') == 'Yes':
            pct = self._get_sub_label("Credit_Card", test_case_id, 6) # Col 6 is Amount%
            amounts["Card"] = (pending_payment * float(pct or 0)) / 100
            # Pull Type (CC/DC) from Col 3
            amounts["CardName"] = self._get_sub_label("Credit_Card", test_case_id, 3) 

        if row_data.get('Cheque') == 'Yes':
            pct = self._get_sub_label("Cheque", test_case_id, 5)
            amounts["Cheque"] = (pending_payment * float(pct or 0)) / 100

        if row_data.get('NetBanking') == 'Yes':
            pct = self._get_sub_label("NetBanking", test_case_id, 6) # Col 6 is Amount%
            amounts["NetBanking"] = (pending_payment * float(pct or 0)) / 100
            # Pull Type (RTGS/IMPS/etc) from Col 2
            amounts["NBName"] = self._get_sub_label("NetBanking", test_case_id, 2)

        payment_methods = {
            'Creditcard': CreditCard.test_Credit_Card,
            'Cheque': Cheque.test_Cheque,
            'NetBanking': NetBanking.test_NetBanking
        }
        for key, method in payment_methods.items():
            if row_data.get(key) == 'Yes':
                method(self, test_case_id, pending_payment)
        
        balance=Function_Call.get_text(self, '//table[@id="payment_modes"]//tfoot//tr[2]//th[3]')
        print(f"Balance: {balance}")
        
        # Click submit and handle the process
        original_window = self.driver.current_window_handle
        Function_Call.click(self, '//button[@id="pay_submit"]')
        
        captured_id = ""
        try:
            # Wait for new tab and switch
            WebDriverWait(self.driver, 10).until(EC.number_of_windows_to_be(2))
            for window_handle in self.driver.window_handles:
                if window_handle != original_window:
                    self.driver.switch_to.window(window_handle)
                    # Extract ID from URL (e.g. .../billing_invoice/3072)
                    sleep(2) # Give it 0.5s to load URL content
                    captured_id = self.driver.current_url.split('/')[-1]
                    self.driver.close()
                    break
        except Exception as e:
            print(f"Could not extract Bill ID: {e}")
        finally:
            self.driver.switch_to.window(original_window)

        # Verify success message and update status
        try:
            msg_el = self.wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'alert-success')]")))
            if "successfully" in msg_el.text.lower():
                 self._update_excel_status(row_num, "Pass", "Success", sheet_name, captured_id)
                 # Update Bill Details Summary
                 self._update_bill_summary(row_data, captured_id, amounts)
                 # Update tags to 'Billed' if this was from an estimation
                 est_no = row_data.get("EstNo")
                 if est_no:
                     self._update_tag_source_sheets(est_no, captured_id)
            else:
                 self._update_excel_status(row_num, "Fail", f"Message: {msg_el.text}", sheet_name)
        except:
             self._update_excel_status(row_num, "Fail", "Success message not found", sheet_name)

    def _update_tag_source_sheets(self, est_no, bill_no):
        """Finds rows with the given Estimation No in Tag_Detail and Purchase_TagDetail and updates Status to 'Billed'."""
        try:
            wb = load_workbook(FILE_PATH)
            sheets_to_update = ["Tag_Detail", "Purchase_TagDetail"]
            updated = False
            for sheet_name in sheets_to_update:
                if sheet_name in wb.sheetnames:
                    sh = wb[sheet_name]
                    # Loop through rows to find matching EstNo in column 15
                    for r in range(2, sh.max_row + 1):
                        est_info = str(sh.cell(row=r, column=15).value or "")
                        # est_info is like "12 - 2026-04-28 12:58:19"
                        if str(est_no) and est_info.startswith(str(est_no) + " -"):
                            sh.cell(row=r, column=14, value="Billed").font = Font(bold=True, color="0000FF")
                            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            sh.cell(row=r, column=15, value=f"{bill_no} - {timestamp}")
                            updated = True
            if updated:
                wb.save(FILE_PATH)
                print(f"✅ Updated source tags to 'Billed' for EstNo {est_no} with BillNo {bill_no}")
            wb.close()
        except Exception as e:
            print(f"❌ Failed to update tag source sheets for billing: {e}")

  
    def _get_network_response_data(self):
        """Parse Chrome performance log to extract 'respondedata'.
        Supports two modes:
          1. Socket.IO (WebSocket FrameReceived)
          2. Standard HTTP (Network.responseReceived + CDP getResponseBody)
        """
        import json as _json
        try:
            logs = self.driver.get_log('performance')
        except Exception as e:
            print(f"⚠️ Performance log not available: {e}")
            return {}

        # Scan newest frames first
        for entry in reversed(logs):
            try:
                msg = _json.loads(entry['message'])['message']
                method = msg.get('method', '')

                # Mode 1: Standard HTTP POST (e.g. getEstimationDetails)
                if method == 'Network.responseReceived':
                    url = msg['params']['response']['url']
                    if 'getEstimationDetails' in url:
                        req_id = msg['params']['requestId']
                        body_data = self.driver.execute_cdp_cmd('Network.getResponseBody', {'requestId': req_id})
                        body = _json.loads(body_data.get('body', '{}'))
                        resp = body.get('responsedata', {})
                        if resp:
                            print(f"✅ Captured data via HTTP: {url.split('/')[-1]}")
                            return resp
            except Exception:
                continue

        return {}

    def _validate_repair_order_amounts(self, row_data, row_num, sheet_name):
        """After Repair Order Delivery loads, validates:
        1. Sales Items table calculation (Taxable, IGST, Total Amount).
        2. Repair Item Details amount (DOM vs Network vs Excel).
        """
        
        sleep(2)  # Let page render

        CALC_TYPE_MAP = {
            "0": "Mc & Wast On Gross",
            "1": "Mc & Wast On Net",
            "2": "Mc on Gross, Wast On Net",
            "3": "Fixed Rate",
            "4": "Fixed Rate based on Weight",
        }

        # Initialize totals for grand Total calculation
        calc_total_3 = 0
        excel_repair_amt = 0

        # ── Fetch data from network (primary source) ─────────────────────────────
        resp_data = self._get_network_response_data()
        net_items = resp_data.get('item_details', [])
        net_orders = resp_data.get('order_details', [])

        print(f"✅ Network Data: {len(net_items)} sale item(s), {len(net_orders)} repair order(s)")

        # ── 1. Sales Items Table — calculation_based_on formula check ─────────────
        try:
            # Confirm Sales Items table is present (table id confirmed from DevTools)
            self.wait.until(EC.presence_of_element_located(
                (By.XPATH, '//table[@id="billing_sale_details"]')))
            print("✅ Sales Items table (billing_sale_details) is available")

            def _safe_float(xpath, default=0.0):
                """Read value/text from element and parse as float."""
                try:
                    val = Function_Call.get_value(self, xpath)
                    if val and str(val).strip() not in ("", "0"):
                        return round(float(str(val).replace(',', '').strip()), 2)
                except Exception:
                    pass
                try:
                    val = Function_Call.get_text(self, xpath)
                    if val and str(val).strip() not in ("", "0"):
                        return round(float(str(val).replace(',', '').strip()), 2)
                except Exception:
                    pass
                return default

            def _safe_str(xpath, default=""):
                """Read value attribute (works for select and input)."""
                try:
                    val = Function_Call.get_value(self, xpath)
                    if val is not None:
                        return str(val).strip()
                except Exception:
                    pass
                try:
                    val = Function_Call.get_text(self, xpath)
                    if val:
                        return str(val).strip()
                except Exception:
                    pass
                return default

            # ── Read calc_type and mc_type: network first, DOM fallback ───────────
            _ROW = '//table[@id="billing_sale_details"]//tbody//tr[1]//'  # User preferred scope
            
            # Use data from network if available
            net_item = net_items[0] if net_items else {}
            calc_type = str(net_item.get('calculation_based_on', '')).strip() if net_item else ""
            if not calc_type:
                calc_type = _safe_str(f'{_ROW}select[@name="sale[calculation_based_on][]"]')
            
            # MC Type ID mapping (1="Gram", 2="Pcs")
            mc_type_raw = _safe_str(f'{_ROW}select[@name="sale[bill_mctype][]"]')
            if mc_type_raw == '1':
                mc_type = "pcs" 
            else:
                mc_type = "gram"

            print(f"Calc Type used: '{calc_type}' ({CALC_TYPE_MAP.get(calc_type, 'Unknown')}) | MC Type: {mc_type}") 

            # ── Read row fields (Exact Name Attributes) ───────────────────────────
            n_wt     = _safe_float(f'{_ROW}input[@name="sale[net][]"]')
            wast_wt  = _safe_float(f'{_ROW}input[@class="form-control bill_wastage_wt"]')
            mc_val   = _safe_float(f'{_ROW}input[@name="sale[mc][]"]')
            pcs      = _safe_float(f'{_ROW}input[@name="sale[pcs][]"]', default=1.0)
            rate     = _safe_float(f'{_ROW}input[@class="form-control bill_rate_per_grm per_grm_amount"]')
            discount = _safe_float(f'{_ROW}input[@name="sale[discount][]"]')

            # ── Read displayed values ─────────────────────────────────────────────
            taxable_ui = _safe_float(f'{_ROW}input[@class="form-control bill_taxable_amt"]')
            tax_amt    = _safe_float(f'{_ROW}td[@class="tax_amt"]')
            metal_amt  = 0
            ui_amount  = _safe_float(f'{_ROW}input[@name="sale[billamount][]"]')

            g_wt = round(n_wt + wast_wt, 3) 

            # ── Compute MC Amount based on calculation_based_on ───────────────────
            if calc_type in ("0", "2"):
                mc_amt = round(pcs * mc_val, 2) if mc_type == "pcs" else round(g_wt * mc_val, 2)
            elif calc_type == "1":
                mc_amt = round(pcs * mc_val, 2) if mc_type == "pcs" else round(n_wt * mc_val, 2)
            elif calc_type == "3":
                mc_amt = round(mc_val, 2)
            elif calc_type == "4":
                mc_amt = round(n_wt * mc_val, 2)
            else:
                mc_amt = round(pcs * mc_val, 2)

            # ── Expected Formulas ─────────────────────────────────────────────────
            metal_cost      = round(g_wt * rate, 2)
            expected_taxable = round(metal_cost + mc_amt - discount, 2)
            expected_amount  = round(taxable_ui + tax_amt + metal_amt, 2)
            print(expected_amount)

            calc_label = CALC_TYPE_MAP.get(calc_type, f"Unknown({calc_type})")
            print(f"\n Sales Items Calculation Verification:")
            print(f"   [{calc_label}] | N.Wt={n_wt}, Wast_Wt={wast_wt}, G.Wt={g_wt}")
            print(f"   Rate={rate}, Disc={discount}, MC_Val={mc_val} ({mc_type})")
            print(f"   Expected Taxable: ({g_wt} * {rate}) + {mc_amt} - {discount} = {expected_taxable}")
            print(f"   UI Taxable: {taxable_ui}")

            # ── Check 1: Taxable Amt formula ──────────────────────────────────────
            if expected_taxable == taxable_ui:
                print(f" ✅ Taxable Amt PASS: {expected_taxable} == UI")
            else:
                msg = f"Taxable Amt Mismatch: Expected {expected_taxable} != UI {taxable_ui} ⚠️"
                print(f" ⚠️ {msg}")
                Function_Call.Remark(self, row_num, msg, sheet_name)

            # ── Check 2: 3% GST Verification ──────────────────────────────────────
            calc_tax_3 = round(expected_taxable * 0.03, 2)
            calc_total_3 = round(expected_taxable + calc_tax_3, 2)

            if calc_total_3 == ui_amount:
                print(f" ✅ 3% GST PASS: {expected_taxable} + {calc_tax_3} (3%) = {ui_amount}")
            else:
                msg = f"3% GST Mismatch: {expected_taxable} + 3%({calc_tax_3}) = {calc_total_3} != UI={ui_amount} ⚠️"
                print(f" ⚠️ {msg}")
                Function_Call.Remark(self, row_num, msg, sheet_name)
            
                
                

        except Exception as e:
            msg = f"Sales Items table not found or calculation unreadable ⚠️: {e}"
            print(f"❌ {msg}")
            Function_Call.Remark(self, row_num, msg, sheet_name)

        # ── 2. Repair Item Details Table ───────────────────────────────────────────
        try:
            # Confirm Repair Item Details section is present
            self.wait.until(EC.presence_of_element_located(
                (By.XPATH, '//table[@id="billing_repair_order_details"]')))
            print("✅ Repair Item Details section is available")

            # ── Read Amount from Repair Items table ───────────────────────────────
            repair_amt_ui_str = ""
            try:
                repair_amt_ui_str = Function_Call.get_value(
                    self, '//table[@id="billing_repair_order_details"]//tbody//tr//input[@name="order[amount][]"]')
            except Exception:
                pass

            repair_amt_ui = round(float(str(repair_amt_ui_str).replace(',', '').strip() or 0), 2)

            excel_repair_amount = row_data.get("RepairAmount")
            excel_repair_amt = round(float(str(excel_repair_amount).replace(',', '').strip() or 0), 2)

            # ── Validation 1: UI Amount vs Excel ──────────────────────────────────
            if repair_amt_ui == excel_repair_amt:
                print(f"✅ Repair Item UI Amount PASS: {repair_amt_ui} == Excel")
            else:
                msg = f"Repair Amount Mismatch: UI={repair_amt_ui} != Excel={excel_repair_amt} ⚠️"
                print(f"⚠️ {msg}")
                Function_Call.Remark(self, row_num, msg, sheet_name)

        except Exception as e:
            msg = f"Repair Item Details section error ⚠️: {e}"
            print(f"❌ {msg}")
            Function_Call.Remark(self, row_num, msg, sheet_name)

        # ── Update Grand Total in Excel Billing Sheet (Col 15: Total) ────────────
        Total = round(calc_total_3 + excel_repair_amt, 2)
        self.repair_total = Total  # Store for payment handling
        print(f"✅ Grand Total: {calc_total_3} (Sales) + {excel_repair_amt} (Repair) = {Total}")
        Function_Call.update_excel_data(self, row_num, 15, Total, sheet_name)

    def _generate_pan(self):
        """Generates a random PAN number in standard format [A-Z]{5}[0-9]{4}[A-Z]{1}"""
        return ''.join(random.choices(string.ascii_uppercase, k=5)) + \
               ''.join(random.choices(string.digits, k=4)) + \
               ''.join(random.choices(string.ascii_uppercase, k=1))

    def _update_excel_status(self, row_num, test_status, actual_status, sheet_name, bill_no=None):
        """Updates test results and bill number in the Excel sheet."""
        try:
            wb = load_workbook(FILE_PATH)
            sh = wb[sheet_name]
            color = "00B050" if test_status == "Pass" else "FF0000"
            sh.cell(row=row_num, column=2, value=test_status).font = Font(bold=True, color=color)
            sh.cell(row=row_num, column=3, value=actual_status).font = Font(bold=True, color=color)
            if bill_no:
                # Based on data_map, BillNo is column 28
                sh.cell(row=row_num, column=28, value=bill_no)
            wb.save(FILE_PATH)
            wb.close()
            print(f"✅ Excel Updated: Row {row_num}, Status={test_status}, BillNo={bill_no}")
        except Exception as e:
            print(f"❌ Failed to update Excel: {e}")

    def _update_bill_summary(self, row_data, invoice_no, amounts):
        """Populates the 'Bill details' sheet with transaction summary."""
        try:
            wb = load_workbook(FILE_PATH)
            sn = "Bill details"
            if sn not in wb.sheetnames:
                sh = wb.create_sheet(sn)
                # Updated 13-column header structure
                headers = ["Bill Type", "InvoiceNo", "Cash", "Carddetails", "Amount", "Cheque", "Amount", "Net Banking", "Amount", "Bill Amount", "Creidt Amount", "Credit Due Date", "Credit Balance Amt"]
                for i, h in enumerate(headers, 1):
                    sh.cell(row=1, column=i, value=h).font = Font(bold=True)
            else:
                sh = wb[sn]

            # Find next empty row
            row = 2
            while sh.cell(row=row, column=2).value is not None:
                row += 1

            sh.cell(row=row, column=1, value=row_data.get("Bill Type"))
            sh.cell(row=row, column=2, value=invoice_no)
            sh.cell(row=row, column=3, value=amounts.get("Cash", 0)) # Cash (Col 3)

            # Card (Col 4, 5)
            if amounts.get("Card", 0) > 0:
                sh.cell(row=row, column=4, value=amounts.get("CardName", "CC"))
                sh.cell(row=row, column=5, value=amounts["Card"])
            
            # Cheque (Col 6, 7)
            if amounts.get("Cheque", 0) > 0:
                sh.cell(row=row, column=6, value="Cheque")
                sh.cell(row=row, column=7, value=amounts["Cheque"])

            # Net Banking (Col 8, 9)
            if amounts.get("NetBanking", 0) > 0:
                sh.cell(row=row, column=8, value=amounts.get("NBName", "NetBanking"))
                sh.cell(row=row, column=9, value=amounts["NetBanking"])

            collected = amounts.get("Card", 0) + amounts.get("Cheque", 0) + amounts.get("NetBanking", 0) + amounts.get("Cash", 0)
            total_val = row_data.get("Total", 0)
            total = float(str(total_val).replace(',', '').strip() if total_val and str(total_val).lower() != 'none' else 0)
            
            # Calculations
            sh.cell(row=row, column=10, value=collected) # Bill Amount
            sh.cell(row=row, column=11, value=max(0, total - collected)) # Creidt Amount
            sh.cell(row=row, column=12, value=row_data.get("Credit Due Date"))
            
            # Credit Balance Amt (Col 13)
            if row_data.get("Is Credit") == "Yes":
                # Use the actual adjustable_amount (received_val) passed from _handle_payments
                actual_received = amounts.get("Adjustable", 0)
                sh.cell(row=row, column=13, value=total - actual_received)

            wb.save(FILE_PATH)
            wb.close()
            print(f"✅ Bill Details Updated: Invoice {invoice_no}")
        except Exception as e:
            print(f"❌ Failed Bill Details update: {e}")

    def _get_sub_label(self, sheet_name, test_case_id, col):
        """Helper to read detail label from sub-sheets."""
        try:
            wb = load_workbook(FILE_PATH); sh = wb[sheet_name]
            for r in range(2, sh.max_row + 1):
                if str(sh.cell(row=r, column=1).value) == str(test_case_id):
                    return sh.cell(row=r, column=col).value
        except: return None
        return None