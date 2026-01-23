from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import unittest
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from time import sleep

FILE_PATH = ExcelUtils.file_path
class Cheque(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)

    def test_Cheque(self,test_case_id,Received):
        driver = self.driver
        wait = self.wait  
        Sheet_name = "Cheque"         
        count = ExcelUtils.Test_case_id_count(FILE_PATH, Sheet_name, test_case_id)                               
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, Sheet_name)
        print(f"'{valid_rows}': valid rows")
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]
        row=1
        for row_num in range(2, valid_rows):  
            current_id = sheet.cell(row=row_num, column=1).value  # Column 1 = Test Case Id
            if current_id == test_case_id:
                data = {
                        "Test Case Id": 1,
                        "ChequeDate": 2,			
                        "Bank": 3,
                        "ChequeNo": 4,
                        "Amount": 5,
                        "Action": 6,
                        "field_validation_status":7
                    }
                row_data = {key: sheet.cell(row=row_num, column=col).value 
                                for key, col in data.items()}
                print(row_data)
                Function_Call.click(self,'//a[@id="cheque_modal"]')
                if row_data["ChequeDate"]:
                    errors=Function_Call.fill_input(
                        self,wait,
                        locator=(By.XPATH, '//input[@name="cheque_details[cheque_date][]"]'),
                        value=row_data["ChequeDate"],
                        pattern=r"^(0[1-9]|[12][0-9]|3[01])-(0[1-9]|1[0-2])-\d{4}$",
                        field_name="ChequeDate",
                        screenshot_prefix="ChequeDate",
                        row_num=row_num,
                        Sheet_name=Sheet_name,
                        extra_keys = Keys.TAB,
                        Date_range="Yes"
                        )   
                else:
                    pass
                
                Function_Call.select_visible_text(self,f"(//select[@name='cheque_details[id_bank][]'])[{row}]", row_data["Bank"])
                
                if row_data["ChequeNo"]:
                    errors=Function_Call.fill_input(
                        self,wait,
                        locator=(By.XPATH, '//input[@name="cheque_details[cheque_no][]"]'),
                        value=row_data["ChequeDate"],
                        pattern = r"\d{1,11}?$",
                        field_name="ChequeNo",
                        screenshot_prefix="ChequeNo",
                        row_num=row_num,
                        Sheet_name=Sheet_name,
                        extra_keys = Keys.TAB,
                        Date_range="Yes"
                        )   
                else:
                    pass
                
                if row_data["Amount"]:
                    Cheque_percent = row_data["Amount"]
                    Cheque_amount = float((Received * Cheque_percent) / 100)
                    errors=Function_Call.fill_input(
                        self,wait,
                        locator=(By.XPATH, '//input[@name="cheque_details[payment_amount][]"]'),
                        value=Cheque_amount,
                        pattern = r"\d{1,11}?$",
                        field_name="Amount",
                        screenshot_prefix="Amount",
                        row_num=row_num,
                        Sheet_name=Sheet_name)   
                else:
                    pass
                
                Function_Call.click(self,'//a[@id="add_newchq"]')