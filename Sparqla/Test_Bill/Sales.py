from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import  sleep
import unittest
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from Utils.Board_rate import Boardrate
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font
import re

FILE_PATH = ExcelUtils.file_path
class SALES(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)

    def test_Sales(self,test_case_id):
        driver = self.driver
        wait = self.wait 
        Rate=Boardrate.Todayrate(self)
        print(Rate)        
        Sheet_name = "SALES"                                        
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, Sheet_name)
        print(f"'{valid_rows}': valid rows")
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]
        for row_num in range(2, valid_rows):  
            current_id = sheet.cell(row=row_num, column=1).value  # Column 1 = Test Case Id
            if current_id == test_case_id:
                data = {
                            "Test Case Id": 1,
                            "Test Status": 2,
                            "Actual Status": 3,
                            "EstNo": 4,
                            "SGST": 5,
                            "CGST": 6,
                            "Total": 7,
                            "TagNo": 8,
                            "Old TagNo": 9,
                            "Home Bill": 10,
                            "Non_tagged": 11,
                            "Employee": 12,
                            "Is Partly": 13,
                            "Section": 14,
                            "Product": 15,
                            "Design": 16,
                            "Sub Design": 17,
                            "Pcs": 18,
                            "Purity": 19,
                            "Size": 20,
                            "G.Wt": 21,
                            "Wast(%)": 22,
                            "Wast Wt(g)": 23,
                            "MC Type": 24,
                            "MC": 25,
                            "Rate": 26,
                            "Discount": 27,
                            "Taxable Amt": 28,
                            "Charges": 29
                        }

                row_data = {key: sheet.cell(row=row_num, column=col).value 
                                for key, col in data.items()}
                print(row_data)
                # Call you 'create' method
                Create_data = SALES.create(self,row_data, row_num, Sheet_name,Rate)
                print(Create_data)
             
    def create(self,row_data, row_num, Sheet_name,Rate):
        driver = self.driver
        wait = self.wait
        Mandatory_field=[]        
        Function_Call.fill_input2(self,'//input[@id="filter_est_no"]',row_data["EstNo"])
        Function_Call.click(self,'//button[@id="search_est_no"]')
        Function_Call.click(self,'(//button[@class="btn btn-close btn-warning"])[11]')
        sleep(3)
        Function_Call.click(self,"//a[normalize-space()='Total Summary']")
        Taxable_Sale_Amount = Function_Call.get_text(self, "//span[@class='summary_lbl summary_sale_amt']")
        print(Taxable_Sale_Amount)
        CGST=Function_Call.get_text(self,'//span[@class="summary_lbl sales_cgst"]')
        print(CGST)
        SGST=Function_Call.get_text(self,'//span[@class="summary_lbl sales_sgst"]') 
        print(SGST)
        IGST=Function_Call.get_text(self,'//span[@class="summary_lbl sales_igst"]')
        print(IGST)
        Sale_Amount=Function_Call.get_text(self,'//span[@class="summary_lbl sale_amt_with_tax"]')
        print(Sale_Amount)
        Purchase_Amount =Function_Call.get_text(self,'//span[@class="summary_lbl summary_pur_amt"]')
        print(Purchase_Amount)
        Discount =Function_Call.fill_input(self,'//div[@id="sale_discount"]')
        print(Discount)
        if row_data["Discount"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, '//div[@id="sale_discount"]'),
                value=row_data["Discount"],
                pattern = r"\d{1,2}?$",
                field_name="Discount",
                screenshot_prefix="Discount",
                row_num=row_num,
                Sheet_name=Sheet_name)   
        else:
            pass
       
        if row_data["Handling_Charges"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, '//input[@id="handling_charges"]'),
                value=row_data["Handling_Charges"],
                pattern = r"\d{1,2}?$",
                field_name="Handling_Charges",
                screenshot_prefix="Handling_Charges",
                row_num=row_num,
                Sheet_name=Sheet_name)   
        else:
            pass
        if row_data["Return_Charges"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, '//input[@id="return_charges"]'),
                value=row_data["Return_Charges"],
                pattern = r"\d{1,2}?$",
                field_name="Return_Charges",
                screenshot_prefix="Return_Charges",
                row_num=row_num,
                Sheet_name=Sheet_name)   
        else:
            pass
        if row_data["Return_Charges"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, '//input[@id="return_charges"]'),
                value=row_data["Return_Charges"],
                pattern = r"\d{1,2}?$",
                field_name="Return_Charges",
                screenshot_prefix="Return_Charges",
                row_num=row_num,
                Sheet_name=Sheet_name)   
        else:
            pass
        Function_Call.click(self,'//li[@id="tab_make_pay"]')
        if row_data['Is_Credit'] == 'Yes':
           Function_Call.click(self,'//input[@id="is_credit_yes"]')
        if row_data['Is_Tobe'] == 'Yes':
           Function_Call.click(self,'//input[@id="is_to_be_yes"]')
           
        if row_data["Credit_Due_Date"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, '//input[@id="credit_due_date"]'),
                value=row_data["Credit_Due_Date"],
                pattern = r"\d{1,2}?$",
                field_name="Credit_Due_Date",
                screenshot_prefix="Credit_Due_Date",
                row_num=row_num,
                Sheet_name=Sheet_name)   
        else:
            pass   
           
        if row_data["Cash"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, '//input[@id="make_pay_cash"]'),
                value=row_data["Cash"],
                pattern = r"\d{1,2}?$",
                field_name="Cash",
                screenshot_prefix="Cash",
                row_num=row_num,
                Sheet_name=Sheet_name)   
        else:
            pass
        
        Function_Call.click(self,'//input[@id="card_detail_modal"]')
        Function_Call.select(self, '//select[@name="card_details[card_name][]"]',row_data[''])
        Function_Call.select(self, '//select[@name="card_details[card_type][]"]',row_data[''])
        
        
      
        