from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import  sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from Utils.Board_rate import Boardrate
from Test_Bill.Sales import SALES
from Test_Bill.Credit_Card import CreditCard
from Test_Bill.Cheque import Cheque
from Test_Bill.NetBanking import NetBanking
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
import re
import unittest

FILE_PATH = ExcelUtils.file_path 
class Issue(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)

    def test_Issue(self):
        driver = self.driver
        wait = self.wait 
        
        Rate=Boardrate.Todayrate(self)
        print(Rate)  
        
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Toggle navigation"))).click()
        Function_Call.click(self,"//span[contains(text(), 'Billing')]")
        Function_Call.click(self,"//span[contains(text(), 'Issue')]")
        
        Sheet_name = "Issuse"                                        
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, Sheet_name)
        print(f"'{valid_rows}': valid rows")
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]
        for row_num in range(2, valid_rows):  
            data = {
                    "Test Case Id": 1,
                    "TestStatus": 2,
                    "ActualStatus": 3,
                    "Select Branch": 4,
                    "Issue To": 5,
                    "Issue Type": 6,
                    "Name": 7,
                    "Mobile": 8,
                    "Amount": 9,
                	"Reference No": 10,
                    "Employee":11,
                    "PAN No":12,
                    "Adhaar No":13,
                    "Cash":14,
                    "Creditcard":15,
                    "Cheque":16,
                    "NetBanking":17
                    
                }
            row_data = {key: sheet.cell(row=row_num, column=col).value 
                            for key, col in data.items()}
            print(row_data)
            # Call you 'create' method
            Create_data = self.create(row_data, row_num, Sheet_name,Rate)
            print(Create_data)
            
    def create(self,row_data, row_num, Sheet_name,Board_Rate):
        driver = self.driver
        wait = self.wait
        driver.refresh()
        Mandatory_field=[]        
        #Cost Centre
        if row_data['Select Branch'] is not None:
            Function_Call.dropdown_select(self,f"//span[@id='select2-branch_select-container']", row_data["Select Branch"],'//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Select Branch field is mandatory ⚠️"
            Mandatory_field.append("Select Branch"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        Issue_To = {
            "Employee": '//input[@id="issue_to1"]',
            "Customer": '//input[@id="issue_to2"]',
            "Karigar": '//input[@id="issue_to4"]'
        }
        Function_Call.click(self,Issue_To[row_data["Issue To"]])
        
        Issue_Type = {
            "Employee": '//input[@id="issue_type1"]',
            "Customer": '//input[@id="issue_type2"]',
            "Karigar": '//input[@id="issue_type3"]'
        }
        Function_Call.click(self,Issue_Type[row_data["Issue Type"]])
        
        # Name
        if row_data["Name"]:
            Function_Call.fill_autocomplete_field(self,"name", row_data["Name"])
        else:
            msg = f"'{None}' → Name field is mandatory ⚠️"
            Mandatory_field.append(msg)
            print(msg)
            Function_Call.Remark(row_num, msg)
            sleep(3)
        
        if row_data["Amount"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, '//input[@id="issue_amount"]'),
                value=row_data["Amount"],
                pattern = r"\d{1,3}?$",
                field_name="Amount",
                screenshot_prefix="Amount",
                row_num=row_num,
                Sheet_name=Sheet_name)
        else:
            pass
        
        if row_data["Reference No"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, '//input[@id="refno"]'),
                value=row_data["Reference No"],
                pattern = r"\d{1,13}?$",
                field_name="Reference No",
                screenshot_prefix="Reference No",
                row_num=row_num,
                Sheet_name=Sheet_name)
        
        if row_data["Employee"] is not None:
            Function_Call.dropdown_select(self,f"//span[@id='select2-emp_select-container']", row_data["Employee"],'//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Employee field is mandatory ⚠️"
            Mandatory_field.append("Employee"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
        
        if row_data["Cash"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, '//input[@id="cash_pay"]'),
                value=row_data["Cash"],
                pattern = r"\d{1,3}?$",
                field_name="Cash",
                screenshot_prefix="Cash",
                row_num=row_num,
                Sheet_name=Sheet_name)
        else:
            pass
        received = Function_Call.get_value(self, '//input[@id="issue_amount"]')
        received_value = float(received)
        if row_data['Cheque']=='Yes':
            print(received_value)
            test_case_id=row_data['Test Case Id']
            pay = Cheque.test_Cheque(self,test_case_id,received_value)
        
        if row_data['NetBanking']=='Yes':
            print(received_value)
            test_case_id=row_data['Test Case Id']
            pay = NetBanking.test_NetBanking(self,test_case_id,received_value)
            
        Balance = Function_Call.get_value(self, '//table[@id="payment_modes"]/tfoot/tr[2]/th[3]')
        
        