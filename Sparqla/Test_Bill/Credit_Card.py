from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import unittest
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook
from time import sleep

FILE_PATH = ExcelUtils.file_path
class CreditCard(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)

    def test_Credit_Card(self,test_case_id,Received):
        driver = self.driver
        wait = self.wait  
        Sheet_name = "Credit_Card"         
        count = ExcelUtils.Test_case_id_count(FILE_PATH, Sheet_name, test_case_id)                               
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, Sheet_name)
        print(f"'{valid_rows}': valid rows")
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]
        row=1
        for row_num in range(2, valid_rows):  
            current_id = sheet.cell(row=row_num, column=1).value  # Column 1 = Test Case Id
            if current_id == test_case_id:
                data = {
                        "Test Case Id": 1,
                        "CardName": 2,						
                        "Type": 3,
                        "Device": 4,
                        "CardNo": 5,
                        "Amount": 6,
                        "ApprovalNo": 7,
                        "Action": 8
                    }
                row_data = {key: sheet.cell(row=row_num, column=col).value 
                                for key, col in data.items()}
                print(row_data)
                # Call you 'create' method
                # Create_data = self.create(row_data, row_num, Sheet_name)
                # print(Create_data)id="card_detail_modal"
                Function_Call.click(self,'//a[@id="card_detail_modal"]')
                
                Function_Call.select_visible_text(self,f"(//select[@name='card_details[card_name][]'])[{row}]", row_data["CardName"])
                
                Function_Call.select_visible_text(self,f"(//select[@name='card_details[card_type][]'])[{row}]", row_data["Type"])
                
                Function_Call.select_visible_text(self,f"(//select[@name='card_details[id_device][]'])[{row}]", row_data["Device"])
                
                if row_data["CardNo"]:
                    errors=Function_Call.fill_input(
                        self,wait,
                        locator=(By.XPATH, '//input[@name="card_details[card_no][]"]'),
                        value=row_data["CardNo"],
                        pattern = r"\d{1,11}?$",
                        field_name="CardNo",
                        screenshot_prefix="CardNo",
                        row_num=row_num,
                        Sheet_name=Sheet_name)   
                else:
                    pass
                
                if row_data["Amount"]:
                    credit_card_percent = row_data["Amount"]
                    credit_card_amount = float((Received * credit_card_percent) / 100)
                    errors=Function_Call.fill_input(
                        self,wait,
                        locator=(By.XPATH, '//input[@name="card_details[card_amt][]"]'),
                        value=credit_card_amount,
                        pattern = r"\d{1,11}?$",
                        field_name="Amount",
                        screenshot_prefix="Amount",
                        row_num=row_num,
                        Sheet_name=Sheet_name)   
                else:
                    pass
                
                if row_data["ApprovalNo"]:
                    errors=Function_Call.fill_input(
                        self,wait,
                        locator=(By.XPATH, '//input[@name="card_details[ref_no][]"]'),
                        value=row_data["ApprovalNo"],
                        pattern = r"\d{1,11}?$",
                        field_name="ApprovalNo",
                        screenshot_prefix="ApprovalNo", 
                        row_num=row_num,
                        Sheet_name=Sheet_name)   
                else:
                    pass
                Function_Call.click(self,'//a[@id="add_newcc"]')