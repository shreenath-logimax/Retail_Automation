from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import  sleep
import unittest
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from Test_EST.Homebill_Lwt import stone
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font
import re
import random
import math

FILE_PATH = ExcelUtils.file_path
class ESTIMATION_Homebill(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)

    def test_estimation_Homebill(self,test_case_id,Board_Rate):
        driver = self.driver
        wait = self.wait
        sleep(3)
        Sheet_name = 'HomeBill_Est'
        test_case_id = test_case_id
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, Sheet_name)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]
        row=1
        print(row)
        salevalue=0
        for row_num in range(2, valid_rows):
            current_id = sheet.cell(row=row_num, column=1).value  # Column 1 = Test Case Id
            if current_id == test_case_id:
                data = {
                    "Test Case Id": 1,
                    "Test Status": 2,
                    "Actual Status": 3,
                    "Tag": 4,
                    "Old Tag": 5,
                    "Section": 6,
                    "Product": 7,
                    "Design": 8,
                    "Sub Design": 9,
                    "Employee": 10,
                    "Purity": 11,
                    "Size": 12,
                    "Pcs": 13,
                    "G.Wt": 14,
                    "L.Wt": 15,
                    "MC Type": 16,
                    "MC Value": 17,
                    "Wastage%": 18,
                    "Other Charge": 19,
                    "Field_validation_status": 20
                }
                row_data = {
                    key: sheet.cell(row=row_num, column=col).value
                    for key, col in data.items()
                }
                print(row_data)
                # Call your 'create' method
                Create_data = ESTIMATION_Homebill.create(self,row_data, row_num, Sheet_name, row, Board_Rate)
                print(Create_data)
                row=row+1
                if Create_data:
                    ceil_value,Test_Status,Actual_Status= Create_data
                    ESTIMATION_Homebill.update_excel_status(self,row_num, Test_Status, Actual_Status, Sheet_name)                    
                    salevalue = salevalue + float(ceil_value)                   
        return salevalue 
                
    def create(self,row_data, row_num, Sheet_name, row, Board_Rate):
        wait = self.wait
        Mandatory_field=[]
        Error_field_val=[]
        sleep(3)
        #Tag Check box selected
        if row > 1:
            sleep(4)
            Function_Call.click(self,'//button[@id="create_custom_details"]')
        else:
            sleep(4)
            Function_Call.click2(self,'//input[@id="select_custom_details"]')
        # Tag 
        if row_data["Tag"] is not None:
            Function_Call.fill_input2(self, f'(//input[@name="est_custom[tag_name][]"])[{row}]'),row_data["Tag"] 
        # Old Tag 
        if row_data["Old Tag"] is not None:
            Function_Call.fill_input2(self, f'(//input[@name="est_custom[old_tag_code][]"])[{row}]'),row_data["Old Tag"] 
        # Section
        sleep(3)
        if row_data["Section"]:
            Function_Call.dropdown_select(
                self,f"(//span[starts-with(@id,'select2-est_custom') and contains(@id,'[id_section]')])[{row}]", 
                row_data["Section"],
                '//span[@class="select2-search select2-search--dropdown"]/input')
              
        if row_data["Product"] is not None:
            Function_Call.dropdown_select(
                self,f"(//span[starts-with(@id,'select2-est_custom') and contains(@id,'[product]')])[{row}]", 
                row_data["Product"],
                '//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Product field is mandatory ⚠️"
            Mandatory_field.append("Product"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
                
        # Design 
        if row_data["Design"] is not None:
            Function_Call.dropdown_select(
                self,f"(//span[starts-with(@id,'select2-est_custom') and contains(@id,'[design]')])[{row}]", 
                row_data["Design"],
                '//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Design field is mandatory ⚠️"
            Mandatory_field.append("Design"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
        
        #Sub Design 
        if row_data["Sub Design"] is not None:
            Function_Call.dropdown_select(
                self,f"(//span[starts-with(@id,'select2-est_custom') and contains(@id,'[sub_design]')])[{row}]", 
                row_data["Sub Design"],
                '//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Sub Design field is mandatory ⚠️"
            Mandatory_field.append("Sub Design"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
        
        # Employee 
        if row_data["Employee"] is not None:
            Function_Call.dropdown_select(
                self,f"(//span[starts-with(@id,'select2-est_custom') and contains(@id,'[item_emp_id]')])[{row}]", 
                row_data["Employee"],
                '//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Employee field is mandatory ⚠️"
            Mandatory_field.append("Employee"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)    
        
        # Purity
        if row_data["Purity"] is not None:
            sleep(2)
            Function_Call.dropdown_select(
                self,f"(//span[starts-with(@id,'select2-est_custom') and contains(@id,'[purity]')]/span)[{row}]", 
                row_data["Purity"],
                '//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Purity field is mandatory ⚠️"
            Mandatory_field.append("Purity"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        if row_data["Size"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, f'(//input[@name="est_custom[size][]"])[{row}]'),
                value=row_data["Size"],
                pattern = r"\d{1,2}(\.\d{1,2})?$",
                field_name="Size",
                screenshot_prefix="Size",
                row_num=row_num,
                Sheet_name=Sheet_name)   
        else:
            pass
        
        if row_data["Pcs"]:
            errors=Function_Call.fill_input(
            self,wait,
            locator=(By.XPATH, f'(//input[@name="est_custom[pcs][]"])[{row}]'),
            value=row_data["Pcs"],
            pattern = r"^\d{1,3}$",
            field_name="Pcs",
            screenshot_prefix="Pcs",
            row_num=row_num,
            Sheet_name=Sheet_name
            )
            Error_field_val.extend(errors)
            print(Error_field_val)
        else:
            msg = f"'{None}' → Pcs field is mandatory ⚠️"
            Mandatory_field.append("Pcs"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
                
        if row_data["G.Wt"]:
            errors=Function_Call.fill_input(
            self,wait,
            locator=(By.XPATH, f'(//input[@name="est_custom[gwt][]"])[{row}]'),
            value=row_data["G.Wt"], 
            pattern = r"^\d{1,3}(\.\d{1,3})?$",
            field_name="G.Wt",
            screenshot_prefix="G.Wt",
            row_num=row_num,
            Sheet_name=Sheet_name
            )
            Error_field_val.extend(errors)
            print(Error_field_val)
        else:
            msg = f"'{None}' → G.Wt field is mandatory ⚠️"
            Mandatory_field.append("G.Wt"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        if row_data["L.Wt"]=='Yes':
            Function_Call.click(self,f'(//input[@name="est_custom[lwt][]"])[{row}]')
            test_case_id=row_data["Test Case Id"]
            sheet_name ='HomeBill_Lwt'
            Lwt =stone.test_tagStone(self, sheet_name, test_case_id)
        else:
            Lwt=0            
            Function_Call.select_visible_text(self, '//select[@class="form-control cus_mc_type"]',value='Gram')
            
        if row_data['MC Type']:
            Function_Call.select_visible_text(self, '//select[@class="form-control cus_mc_type"]',value=row_data['MC Type'])
                    
        if row_data["MC Value"]:
            errors=Function_Call.fill_input(
            self,wait,
            locator=(By.XPATH, f'(//input[@name="est_custom[mc][]"])[{row}]'),
            value=row_data["MC Value"],
            pattern = r"^\d{1,5}$",
            field_name="MC Value",
            screenshot_prefix="MC Value",
            row_num=row_num,
            Sheet_name=Sheet_name)
            Error_field_val.extend(errors)
            print(Error_field_val)
        else:
            msg = f"'{None}' → MC Value field is mandatory ⚠️"
            Mandatory_field.append("MC Value"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
                
        if row_data["Wastage%"]:
            errors=Function_Call.fill_input(
            self,wait,
            locator=(By.XPATH, f'(//input[@name="est_custom[mc][]"])[{row}]'),
            value=row_data["Wastage%"],
            pattern = r"^\d{1,2}(\.\d{2})?$",
            field_name="Wastage%",
            screenshot_prefix="Wastage%",
            range_check = lambda v: 0 <= float(v) <= 99,
            row_num=row_num,
            Sheet_name=Sheet_name)
            Error_field_val.extend(errors)
            print(Error_field_val)
        else:
            msg = f"'{None}' → Wastage% Vlue field is mandatory ⚠️"
            Mandatory_field.append("Wastage%"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
                
        # Open Other Charge section
        Function_Call.click(self, f"(//table[@id='estimation_custom_details']//td[19]/a)[{row}]")

        charges_raw = row_data["Other Charge"]
        if not charges_raw:
            print("⚠️ OtherCharge flag is Yes but no ChargeName provided")
            return
        charges_list = [s.strip() for s in charges_raw.split(",")]
        for idx, charge in enumerate(charges_list):
            # For the 2nd, 3rd, ... charges → click +Add
            if idx > 0:
                Function_Call.click(self, '//button[@id="add_new_charge"]')
        
            # Select charge type
            sleep(3)
            Function_Call.select(self,f'(//select[@name="est_stones_item[id_charge][]"])[{idx+1}]',charge)
         
            # Locate corresponding value field (same row as idx+1)
            value_input = wait.until(EC.presence_of_element_located(
                (By.XPATH, f"(//input[@name='est_stones_item[value_charge][]'])[{idx+1}]")
            ))
            current_value = value_input.get_attribute("value").strip()

            # If empty or "0.00" → auto-fill random multiple of 100
            if current_value == "0.00":
                random_value = random.randint(1, 10) * 100
                sleep(1)
                value_input.clear()
                value_input.send_keys(str(random_value))
                print(f"⚡ Added random value {random_value} for {charge}")
            else:
                print(f"✅ Auto value {current_value} kept for {charge}")

        # Save button
        wait.until(EC.element_to_be_clickable((By.ID, "update_charge_details"))).click()
        print("Field✅ OtherCharges added:", charges_list)
        
        # Fetch values with one-liners
        Gwt   = ESTIMATION_Homebill.get_val(self, f'(//input[@name="est_custom[gwt][]"])[{row}]')
        # Lwt   = ESTIMATION_Homebill.get_val(self, f'(//input[@name="est_custom[lwt][]"])[{row}]')
        Nwt   = ESTIMATION_Homebill.get_val(self, f'(//input[@name="est_custom[nwt][]"])[{row}]')
        PCS   = ESTIMATION_Homebill.get_val(self, f'(//input[@name="est_custom[pcs][]"])[{row}]', cast=float)
        Wast_per = ESTIMATION_Homebill.get_val(self, f'(//input[@name="est_custom[wastage][]"])[{row}]')
        Wast  = ESTIMATION_Homebill.get_val(self, f'(//input[@class="form-control cus_wastage_wt"])[{row}]')
        Mc    = ESTIMATION_Homebill.get_val(self, f'(//input[@name="est_custom[mc][]"])[{row}]')
        other_Amt=ESTIMATION_Homebill.get_val(self,f'(//input[@name="est_custom[value_charge][]"])[{row}]')
        
        Stone=Lwt
    
        # Taxable amount kept as string (not converted to float)
        Taxable_Amt = Function_Call.get_value(self,f'(//input[@name="est_custom[amount][]"])[{row}]')
        print(Taxable_Amt)
        
        # MC type  
        mc_type_dropdown = wait.until(EC.presence_of_element_located((By.XPATH, '//select[@class="form-control cus_mc_type"]')))
        select = Select(mc_type_dropdown)
        selected_text = select.first_selected_option.text.strip()
        print("Selected MC Type:", selected_text)
        Mc_type =selected_text
        
        Cal_current_value=Function_Call.get_value(self, f'(//input[@class="cus_calculation_based_on"])[{row}]')
        tax_percent_el = wait.until(EC.presence_of_element_located((By.XPATH, "//h3[text()='Home Bill Details']/following::input[contains(@class,'tax_percentage')][1]")))
        tax_percentage = tax_percent_el.get_attribute("value")
        Taxvalue=float(tax_percentage)
        print(Taxvalue)
        
        value = wait.until(EC.presence_of_element_located((By.XPATH, f"(//span[starts-with(@id,'select2-est_custom') and contains(@id,'[purity]')])[{row}]")))
        purity = value.get_attribute("title")
        print(purity) 
        if purity =='916.0000':
           gold_rate=Board_Rate[0]
        if purity == '75.0000':
           gold_rate=Board_Rate[1]
        if purity == '92.5000':
            gold_rate=Board_Rate[2]
                
        # Debug print all values
        print(f"Gwt={Gwt}, Lwt={Lwt}, Nwt={Nwt}, PCS={PCS}, Stone={Stone},"
            f"Wast_per={Wast_per}, Wast={Wast}, Mc={Mc}, Mc_type={Mc_type},"
            f"Taxable={Taxable_Amt}")
       
        Result=ESTIMATION_Homebill.calculation(self,Cal_current_value,gold_rate,Gwt,Nwt,Wast_per,Mc,Stone,other_Amt,Mc_type,Taxvalue)
        ceil_value,Cal_type,IGst,SGst =Result
        
        if ceil_value==Taxable_Amt:
            Test_Status= "Pass"
            Actual_Status =(f"✅ Calculation Value is correct {ceil_value}")
        else:
            Test_Status= "Fail"
            Actual_Status =(f"❌ Calculation Error in {ceil_value} | Web Value={Taxable_Amt}")
        if Mandatory_field or Error_field_val:
            Test_Status= "Fail"
            if Mandatory_field:
               Actual_Status=f"{Mandatory_field} These field is mandatory"
            if Error_field_val:
               Actual_Status=f"{Error_field_val} These field is wrong data to save."
        return ceil_value,Test_Status,Actual_Status
            
    def calculation(self,Cal_current_value,gold_rate,Gwt,Nwt,Wast_per,Mc,Stone,other_Amt,Mc_type,Taxvalue):
        wait = self.wait 
        data = {
            "0": "Mc & Wast On Gross",
            "1": "Mc & Wast On Net",
            "2": "Mc on Gross, Wast On Net",
            "3": "Fixed Rate",
            "4": "Fixed Rate based on Weight"
        }
        Cal_type = data[str(Cal_current_value)]   # convert int → str because keys are strings
        print(Cal_type)
        gross_weight=Gwt
        net_weight = Nwt  
        wastage_percentage = Wast_per 
        Making_cost_pergram = Mc 
        diamond_cost =Stone
        Charge_Amt = other_Amt
        Tax = float(Taxvalue)
        # initialize
        ceil_value = None
        if Cal_type=="Mc on Gross, Wast On Net":
        # calculation making cost on gross Wastage% on Net  
            if Mc_type == 'Piece':
                mc =Making_cost_pergram
            else:    
                Mc=Making_cost_pergram*gross_weight
                mc=float('{:.2f}'.format(math.ceil(Mc)))
            Va = (wastage_percentage/100)*net_weight
            Va = round(Va, 3)
            total = net_weight+Va
            total = round(total, 3)
            Cal = (total*gold_rate)+diamond_cost+mc+Charge_Amt
            ceil_value=("{:.2f}".format(math.ceil(Cal)))
            print(ceil_value)
            
        if  Cal_type=="Mc & Wast On Net":
            # calculation making cost & Wastage% on Net  
            if Mc_type == 'Piece':
                mc =Making_cost_pergram
            else:    
                Mc=Making_cost_pergram*net_weight
                mc=float("{:.2f}".format(math.ceil(Mc)))
            Va = (wastage_percentage/100)*net_weight
            Va= round(Va, 3)
            total = net_weight+Va
            total = round(total, 3)
            Cal2 = total*gold_rate+mc+diamond_cost+Charge_Amt
            ceil_value=("{:.2f}".format(math.ceil(Cal2)))
                        
        if  Cal_type == "Mc & Wast On Gross":
            # calculation making cost & Wastage% on Gross 81148.00
            if Mc_type == 'Piece':
                mc =Making_cost_pergram
            else:    
                Mc=Making_cost_pergram*gross_weight
                Mc= gross_weight*Making_cost_pergram
                mc=float("{:.2f}".format(math.ceil(Mc)))
        
            Va = (wastage_percentage/100)*gross_weight
            Va = round(Va, 3)
            total= net_weight+Va
            total = round(total, 3)
            cal3 = total*gold_rate+mc+diamond_cost+Charge_Amt
            ceil_value=("{:.2f}".format(math.ceil(cal3)))
            
        if Cal_type== "Fixed Rate based on Weight":
            if Mc_type=='Piece':
                mc = Making_cost_pergram
            else:
                Mc=Making_cost_pergram*gross_weight
                mc=float('{:.2f}'.format(math.ceil(Mc)))
            Va = (wastage_percentage/100)*gross_weight
            Va = round(Va, 3)
            total= net_weight+Va
            total = round(total, 3)
            cal3 = total*gold_rate+mc+diamond_cost+Charge_Amt
            ceil_value=("{:.2f}".format(math.ceil(cal3)))
        
        if Cal_type == "Fixed Rate":
            ceil_value        
        if Tax:  
           salevalue=float(ceil_value)
           Find_Tax=(salevalue*Tax)/100
           Tol_Amt = salevalue+Find_Tax
           Gst=Find_Tax/2
           IGst=("{:.2f}".format(Gst))
           SGst=("{:.2f}".format(Gst))
           ceil_value=("{:.2f}".format(Tol_Amt))        
        print(ceil_value)
        print(type(ceil_value))
        return ceil_value,Cal_type,IGst,SGst
                                                         
      # --- Helper to fetch field values safely ---
    def get_val(self,locator, cast=float, default=0):
        wait = self.wait
        el = wait.until(EC.presence_of_element_located((By.XPATH, locator)))
        val = el.get_attribute("value")
        if not val:  
            return default
        return cast(val)
    def update_excel_status(self,row_num, Test_Status, Actual_Status, function_name):
        print(function_name)
        # Load the workbook
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[function_name]  # or workbook["SheetName"]        
        if Test_Status== 'Pass':
            # Write Test_Status into column 2
            sheet.cell(row=row_num, column=2, value=Test_Status).font=Font(bold=True, color="00B050")            
            # Write Actual_Status in col 3 
            sheet.cell(row=row_num, column=3, value=Actual_Status).font = Font(bold=True, color="00B050")
        if Test_Status=='Fail':
            # Write Test_Status into column 2
            sheet.cell(row=row_num, column=2, value=Test_Status).font=Font(bold=True, color="FF0000")
            # Write Actual_Status in col 3 
            sheet.cell(row=row_num, column=3, value=Actual_Status).font = Font(bold=True, color="FF0000")
        # Save workbook
        workbook.save(FILE_PATH)
        # Get status from ExcelUtils
        Status = ExcelUtils.get_Status(FILE_PATH, function_name)
        # Print and return status
        print(Status)
        return Status








