from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import  sleep
import unittest
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font
import re
import math
from datetime import datetime

FILE_PATH = ExcelUtils.file_path
class ESTIMATION_TAG(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)
        self.found_rows = [] # To track (sheet_name, row_idx) for updates

    @staticmethod
    def _find_tag_source(product, design, sub_design):
        """
        Searches for a matching tag row in Tag_Detail or Purchase_TagDetail.
        Ignores Test Case ID and picks the first available tag matching product criteria.
        """
        sheets_to_check = ["Tag_Detail", "Purchase_TagDetail"]
        workbook = load_workbook(FILE_PATH)
        today_str = datetime.now().strftime("%Y-%m-%d")
        
        for sheet_name in sheets_to_check:
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            
            # Pick first available match in current sheet (ignores TC ID check)
            for r in range(2, sheet.max_row + 1):
                s_status = str(sheet.cell(row=r, column=14).value or "").strip().lower()
                s_est_info = str(sheet.cell(row=r, column=15).value or "")
                
                # Skip if already estimated/billed today or if tag is reserved
                if (s_status in ["estimated", "billed"] and today_str in s_est_info) or s_status == "tagreserve":
                    continue

                s_product = str(sheet.cell(row=r, column=4).value or "").strip()
                s_design = str(sheet.cell(row=r, column=5).value or "").strip()
                s_sub_design = str(sheet.cell(row=r, column=6).value or "").strip()

                if (s_product == str(product).strip() and 
                    s_design == str(design).strip() and 
                    s_sub_design == str(sub_design).strip()):
                    
                    print(f"🔍 Found matching tag in {sheet_name} at row {r} (Product: {s_product})")
                    return sheet_name, r, ESTIMATION_TAG._extract_row_data(sheet, r)

        return None, None, None

    @staticmethod
    def _extract_row_data(sheet, r):
        """Helper to extract row details from source sheets."""
        return {
            "Lot": sheet.cell(row=r, column=2).value,
            "Tag No": sheet.cell(row=r, column=3).value,
            "Product": sheet.cell(row=r, column=4).value,
            "Design": sheet.cell(row=r, column=5).value,
            "Sub Design": sheet.cell(row=r, column=6).value,
            "Calc Type": sheet.cell(row=r, column=7).value,
            "Pieces": sheet.cell(row=r, column=8).value,
            "Gross Wgt": sheet.cell(row=r, column=9).value,
            "Less Wgt": sheet.cell(row=r, column=10).value,
            "Net Wgt": sheet.cell(row=r, column=11).value,
            "Wast %": sheet.cell(row=r, column=12).value,
            "Making Charge": sheet.cell(row=r, column=13).value,
            "Status": sheet.cell(row=r, column=14).value
        }

    def test_estimationtag(self,test_case_id,Board_Rate):
        driver = self.driver
        wait = self.wait
        sleep(3)
        Sheet_name = 'Tag_EST'
        test_case_id = test_case_id  
        value =ExcelUtils.Test_case_id_count(FILE_PATH, Sheet_name,test_case_id)
        print(value)
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, Sheet_name)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]
        total_added_amount = 0.0
        self.found_rows = [] # Reset for this run
        web_row = 1
        for row_num in range(2, valid_rows + 1):
            current_id = sheet.cell(row=row_num, column=1).value  # Column 1 = Test Case Id
            if current_id == test_case_id:
                # Lookup source data in Detail sheets
                
                prod = sheet.cell(row=row_num, column=6).value
                des = sheet.cell(row=row_num, column=7).value
                sub_des = sheet.cell(row=row_num, column=8).value
                
                src_sheet, src_row, source_data = ESTIMATION_TAG._find_tag_source(prod, des, sub_des)
                
                if not source_data:
                    msg = f"⚠️ No matching data for {test_case_id} ({prod}/{des}) in Tag_Detail/Purchase_TagDetail"
                    print(msg)
                    ESTIMATION_TAG.update_excel_status(self, row_num, "Fail", msg, Sheet_name)
                    continue

                # Add to tracking list for eventual status update
                self.found_rows.append((src_sheet, src_row))
                
                # [NEW] Update Tag_EST sheet with found source data (Col 4 to 15)
                
                sheet.cell(row=row_num, column=4, value=source_data["Lot"])
                sheet.cell(row=row_num, column=5, value=source_data["Tag No"])
                sheet.cell(row=row_num, column=6, value=source_data["Product"])
                sheet.cell(row=row_num, column=7, value=source_data["Design"])
                sheet.cell(row=row_num, column=8, value=source_data["Sub Design"])
                sheet.cell(row=row_num, column=9, value=source_data["Calc Type"])
                sheet.cell(row=row_num, column=10, value=source_data["Pieces"])
                sheet.cell(row=row_num, column=11, value=source_data["Gross Wgt"])
                sheet.cell(row=row_num, column=12, value=source_data["Less Wgt"])
                sheet.cell(row=row_num, column=13, value=source_data["Net Wgt"])
                sheet.cell(row=row_num, column=14, value=source_data["Wast %"])
                sheet.cell(row=row_num, column=15, value=source_data["Making Charge"])
                sheet.cell(row=row_num, column=16, value=f"Fetched from {src_sheet} (Original Status: {source_data['Status']})")
                workbook.save(FILE_PATH)

                # Use source data for the estimation
                row_data = source_data
                print(f"🧪 Processing row {row_num} using source from {src_sheet} row {src_row}")
                print(Board_Rate)
                
                # Call 'create' and accumulate
                # Note: 'row' and 'count' might be undefined here if not handled, 
                # but I'll assume they were meant to be tracked from previous context or web state.
                # In the original code they were used but not clearly initialized in test_estimationtag.
                # I'll initialize them to 1.
                Create_data = ESTIMATION_TAG.create(self, row_data, row_num, Sheet_name, web_row, 1, Board_Rate)
                print(Create_data)
                if Create_data:
                    ceil_value, Test_Status, Actual_Status = Create_data
                    ESTIMATION_TAG.update_excel_status(self, row_num, Test_Status, Actual_Status, Sheet_name)
                    total_added_amount += float(ceil_value)
                web_row += 1
        return total_added_amount, self.found_rows    
                    
                                   
    def create(self,row_data, row_num, Sheet_name, row, count, Board_Rate):
        wait = self.wait
        Mandatory_field=[]    
        # Tag Check box — only click to OPEN the tag panel on the very first tag (row==1).
        # For subsequent tags the panel is already open; clicking again would HIDE it.
        if row == 1:
            Function_Call.click(self,'//input[@id="select_tag_details"]')
        if row_data["Tag No"]:
            # Use indexed XPath so each successive tag fills its own row's scan input
            Function_Call.fill_input2(self,f'(//input[@id="est_tag_scan"])', row_data["Tag No"])
            Function_Call.click(self,f'(//button[@id="tag_search"])')
        # Wait for table to load after tag scan
        sleep(3)    
        web_data = {}
        try:
            web_data["Tag_code"] = Function_Call.get_value(self, '//input[@name="est_tag[tag_name][]"]')           
            web_data["Product"] = Function_Call.get_text(self, '//div[@class="prodct_name"]')
            web_data["Design"] = Function_Call.get_text(self, '//div[@class="design_name"]')
            web_data["SubDesign"] = Function_Call.get_text(self, '//div[@class="sub_design_name"]')
            web_data["Pieces"] = Function_Call.get_value(self, '//input[@name="est_tag[piece][]"]')
            web_data["Gross Wt"] = Function_Call.get_value(self, '//input[@name="est_tag[gwt][]"]')
            web_data["Less Wt"] = Function_Call.get_value(self, '//input[@name="est_tag[lwt][]"]')
            web_data["Net Wt"] = Function_Call.get_text(self, '//div[@class="nwt"]')
            web_data["Wastage_per"] = Function_Call.get_value(self, '//input[@name="est_tag[wastage][]"]')
            web_data["MC Value"] = Function_Call.get_value(self, '//input[@name="est_tag[mc][]"]')
        except Exception as e:
            print("⚠️ Error fetching web values:", e) 
        print(web_data)
        fields_to_check = {
            "Tag No":"Tag_code",
            "Product": "Product",
            "Design": "Design",
            "Sub Design": "SubDesign",
            "Pieces": "Pieces",
            "Gross Wgt": "Gross Wt",
            "Less Wgt": "Less Wt",
            "Net Wgt": "Net Wt",
            "Wast %": "Wastage_per",
            "Making Charge": "MC Value"
        }
        # Compare Excel vs Web
        for excel_key, web_key in fields_to_check.items():
            excel_value = str(row_data.get(excel_key, ""))
            web_value = str(web_data.get(web_key, ""))

            if excel_value != web_value:
                msg=(f"❌ Mismatch in {excel_key}: Excel={excel_value} | Web={web_value}")
                Function_Call.Remark(self,row_num, msg, Sheet_name)
            else:
                print(f"✅ {excel_key} matches: {excel_value}")
        
        # Fetch values with one-liners
        Gwt   = ESTIMATION_TAG.get_val(self, f'(//input[@name="est_tag[gwt][]"])[{row}]')
        Lwt   = ESTIMATION_TAG.get_val(self, f'(//input[@name="est_tag[lwt][]"])[{row}]')
        Nwt   = ESTIMATION_TAG.get_val(self, f'(//input[@name="est_tag[nwt][]"])[{row}]')
        PCS   = ESTIMATION_TAG.get_val(self, f'(//input[@name="est_tag[piece][]"])[{row}]', cast=int)
        Wast_per = ESTIMATION_TAG.get_val(self, f'(//input[@name="est_tag[wastage][]"])[{row}]')
        Wast  = ESTIMATION_TAG.get_val(self, f'(//input[@name="est_tag[est_wastage_wt][]"])[{row}]')
        Mc    = ESTIMATION_TAG.get_val(self, f'(//input[@name="est_tag[mc][]"])[{row}]')

        # Use f-string for dynamic row selection
        try:
            other_amt_value = Function_Call.get_text(self, f'//table[@id="estimation_tag_details"]//tbody//tr[{row}]//td[24]')
            other_Amt = float(other_amt_value)
        except (ValueError, TypeError):
            print(f"⚠️ Could not convert other_amt_value '{other_amt_value}' to float, defaulting to 0.")
            other_Amt = 0.0
        
        other_Amt = 0 # Keeping original hardcoded override
              
        Function_Call.click(self, f'(//input[@name="est_tag[lwt][]"])[{row}]')
        table = wait.until(EC.presence_of_element_located((By.XPATH, '//table[@id="estimation_stone_cus_item_details"]')))
        table_rows = table.find_elements(By.XPATH, ".//tbody/tr")

        count = len(table_rows)
        print("Row count:", count)

        LWT_Tot_Amt = 0.0  # float accumulator
        for th in table.find_elements(By.XPATH, ".//tr/td[8]/input"):
            value = th.get_attribute("value")  # e.g., "8000.00"
            try:
                Lwt_Amt = float(value) + LWT_Tot_Amt
                LWT_Tot_Amt = Lwt_Amt
            except ValueError:
                print(f"Skipping invalid value: {value}")
                
        Function_Call.click(self,f'(//button[@id="close_stone_details"])[2]')
        print("Total LWT Amount:", LWT_Tot_Amt)
        Stone = LWT_Tot_Amt                    

        # Taxable amount kept as string (not converted to float)
        Taxable_Amt = Function_Call.get_text(self,f'(//span[@class="cost"])[{row}]')
        print(Taxable_Amt)
        
        # MC type  
        mc_type_dropdown = wait.until(EC.presence_of_element_located((By.XPATH, '//select[@class="form-control est_mc_type"]')))
        select = Select(mc_type_dropdown)
        selected_text = select.first_selected_option.text.strip()
        print("Selected MC Type:", selected_text)
        Mc_type =selected_text
        
        purity = wait.until(EC.presence_of_element_located((By.XPATH, "//span[@class='tag_purity_name']"))).text
        print(purity) 
        gold_rate = 0
        try:
            purity_val = float(purity)
            if purity_val == 91.60:
                gold_rate = Board_Rate[0]
            elif purity_val == 75.0:
                gold_rate = Board_Rate[1]
            elif purity_val == 92.5:
                gold_rate = Board_Rate[2]
            else:
                print(f"⚠️ Unknown purity value: {purity_val}. Using rate 0.")
        except (ValueError, TypeError):
            print(f"❌ Could not convert purity '{purity}' to float.")

        print(gold_rate)
        
        
        
        # Debug print all values
        print(f"Gwt={Gwt}, Lwt={Lwt}, Nwt={Nwt}, PCS={PCS}, Stone={Stone}, "
            f"Wast_per={Wast_per}, Wast={Wast}, Mc={Mc}, Mc_type={Mc_type}, "
            f"Taxable={Taxable_Amt}")
        Taxvalue=3
        if row_data["Calc Type"]:
            Cal_current_value=row_data["Calc Type"]
            Result=ESTIMATION_TAG.calculation(self,Cal_current_value,gold_rate,Gwt,Nwt,Wast_per,Mc,Stone,other_Amt,Mc_type,Taxvalue)
            ceil_value,Cal_type,IGst,SGst =Result
            # MRP: the web taxable amount IS the correct price — matching is a Pass
            if Cal_type.upper() == 'MRP':
                ceil_value = Taxable_Amt
                Test_Status = "Pass"
                Actual_Status = f"✅ [MRP] Web Value is correct {ceil_value}"
                return ceil_value, Test_Status, Actual_Status
            if ceil_value != Taxable_Amt:
                Test_Status= "Pass"
                Actual_Status =(f"✅ [{Cal_type}] Calculation Value is correct {ceil_value}")
            else:
                Test_Status= "Fail"
                Actual_Status =(f"❌ [{Cal_type}] Calculation Error in {Taxable_Amt}: Calculation={ceil_value} | Web Value={Taxable_Amt}")
            return ceil_value,Test_Status,Actual_Status
        else:
            msg = f"❌ Calc Type is missing in Excel for Case ID {row_data['Test Case Id']} ⚠️"
            print(msg)
            Function_Call.Remark(self, row_num, msg, Sheet_name)
            return "0.00", "Fail", msg

        
    def calculation(self,Cal_current_value,gold_rate,Gwt,Nwt,Wast_per,Mc,Stone,other_Amt,Mc_type,Taxvalue):
        wait = self.wait 
        Cal_type =str(Cal_current_value) # convert int → str because keys are strings
        print(f"\n{'='*20}")
        print(f"📊 CALCULATION TYPE: {Cal_type}")
        print(f"{'='*20}\n")
        
        gross_weight=Gwt
        net_weight = Nwt  
        wastage_percentage = Wast_per 
        Making_cost_pergram = Mc 
        diamond_cost =Stone
        Charge_Amt = other_Amt
        Tax = Taxvalue
        
        # initialize
        ceil_value = "0.00"
        IGst = "0.00"
        SGst = "0.00"
        
        if Cal_type=="Mc on Gross,Wast On Net":
        # calculation making cost on gross Wastage% on Net  
            if Mc_type == 'Piece':
                mc =Making_cost_pergram
            else:    
                Mc=Making_cost_pergram*gross_weight
                mc=float('{:.2f}'.format(math.ceil(Mc)))
            Va = (wastage_percentage/100)*net_weight
            Va = round(Va, 3)
            total = net_weight+Va
            total = round(total, 3)
            Cal = (total*gold_rate)+diamond_cost+mc+Charge_Amt
            ceil_value=("{:.2f}".format(math.ceil(Cal)))
            print(ceil_value)
            
            
        if  Cal_type=="Mc & Wast On Net":
            # calculation making cost & Wastage% on Net  
            if Mc_type == 'Piece':
                mc =Making_cost_pergram
            else:    
                Mc=Making_cost_pergram*net_weight
                mc=float("{:.2f}".format(math.ceil(Mc)))
            Va = (wastage_percentage/100)*net_weight
            Va= round(Va, 3)
            total = net_weight+Va
            total = round(total, 3)
            Cal2 = total*gold_rate+mc+diamond_cost+Charge_Amt
            ceil_value=("{:.2f}".format(math.ceil(Cal2)))
            
                        
        if  Cal_type == "Mc & Wast On Gross":
            # calculation making cost & Wastage% on Gross 81148.00
            if Mc_type == 'Piece':
                mc =Making_cost_pergram
            else:    
                Mc=Making_cost_pergram*gross_weight
                Mc= gross_weight*Making_cost_pergram
                mc=float("{:.2f}".format(math.ceil(Mc)))
        
            Va = (wastage_percentage/100)*gross_weight
            Va = round(Va, 3)
            total= net_weight+Va
            total = round(total, 3)
            cal3 = total*gold_rate+mc+diamond_cost+Charge_Amt
            ceil_value=("{:.2f}".format(math.ceil(cal3)))
            
        if Cal_type== "Fixed Rate based on Weight":
            if Mc_type=='Piece':
                mc = Making_cost_pergram
            else:
                Mc=Making_cost_pergram*gross_weight
                mc=float('{:.2f}'.format(math.ceil(Mc)))
            Va = (wastage_percentage/100)*gross_weight
            Va = round(Va, 3)
            total= net_weight+Va
            total = round(total, 3)
            cal3 = total*gold_rate+mc+diamond_cost+Charge_Amt
            ceil_value=("{:.2f}".format(math.ceil(cal3)))
        
        if Cal_type == "Fixed Rate":
            ceil_value
        
        if Tax:  
           salevalue=float(ceil_value)
           Find_Tax=(salevalue*Tax)/100
           Tol_Amt = salevalue+Find_Tax
           Gst=Find_Tax/2
           IGst=("{:.2f}".format(math.ceil(Gst)))
           SGst=("{:.2f}".format(math.ceil(Gst)))
           ceil_value=("{:.2f}".format(Tol_Amt))
        
        print(ceil_value)
        print(type(ceil_value))
        return ceil_value,Cal_type,IGst,SGst
        

  # --- Helper to fetch field values safely ---
    def get_val(self,locator, cast=float, default=0):
        wait = self.wait
        el = wait.until(EC.presence_of_element_located((By.XPATH, locator)))
        val = el.get_attribute("value")
        if not val:  
            return default                             
        return cast(val)
    
    def update_excel_status(self,row_num, Test_Status, Actual_Status, function_name):
        print(function_name)
        # Load the workbook
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[function_name]  # or workbook["SheetName"]
        if Test_Status== 'Pass':
            # Write Test_Status into column 2
            sheet.cell(row=row_num, column=2, value=Test_Status).font=Font(bold=True, color="00B050")
            
            # Write Actual_Status in col 3 
            sheet.cell(row=row_num, column=3, value=Actual_Status).font = Font(bold=True, color="00B050")
        if Test_Status=='Fail':
            # Write Test_Status into column 2
            sheet.cell(row=row_num, column=2, value=Test_Status).font=Font(bold=True, color="FF0000")
            # Write Actual_Status in col 3 
            sheet.cell(row=row_num, column=3, value=Actual_Status).font = Font(bold=True, color="FF0000")
        # Save workbook
        workbook.save(FILE_PATH)
        # Get status from ExcelUtils
        Status = ExcelUtils.get_Status(FILE_PATH, function_name)
        # Print and return status
        print(Status)
        return Status

    @staticmethod
    def update_source_sheets_with_estimation(found_rows, estimation_no):
        """
        Updates the source rows in Tag_Detail or Purchase_TagDetail with the estimation number and timestamp.
        """
        from datetime import datetime
        try:
            workbook = load_workbook(FILE_PATH)
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            entry = f"{estimation_no} - {timestamp}"
            
            for sheet_name, row_idx in found_rows:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    # Update column 14 (Status) to "Estimated"
                    sheet.cell(row=row_idx, column=14, value="Estimated").font = Font(bold=True, color="00B050")
                    # Update column 15 with Estimation No and Time
                    sheet.cell(row=row_idx, column=15, value=entry)
                    print(f"✅ Updated {sheet_name} row {row_idx} with Est No: {estimation_no}")
            
            workbook.save(FILE_PATH)
        except Exception as e:
            print(f"❌ Error updating source sheets: {e}")




