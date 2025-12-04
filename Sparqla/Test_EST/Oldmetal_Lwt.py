from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from time import sleep
import unittest, math
from Utils.Function import Function_Call
from Utils.Excel import ExcelUtils

FILE_PATH = ExcelUtils.file_path


class stone(unittest.TestCase):
    def __init__(self, driver):
        super().__init__()  # ensure unittest init is called
        self.driver = driver
        self.wait = WebDriverWait(driver, 10)

    # ---------- Main Test ----------
    def test_tagStone(self, sheet_name, test_case_id):
        wait = self.wait
        sleep(2)
        # Excel setup
        count = ExcelUtils.Test_case_id_count(FILE_PATH, sheet_name, test_case_id)
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, sheet_name)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[sheet_name]
        row = 1
        for row_num in range(2, valid_rows + 1):  # include last row
            current_id = sheet.cell(row=row_num, column=1).value  # Column 1 = Test Case Id
            if current_id == test_case_id:
                # continue
                # Map Excel row to dict
                data_map = {
                    "Test Case Id": 1, "Less Weight": 2, "Type": 3, "Name": 4,
                    "Pcs": 5, "Wt": 6, "Wt Type": 7,
                    "Cal.Type": 8, "Rate": 9, "Amount": 10
                }
                row_Stonedata = {
                    k: sheet.cell(row=row_num, column=v).value for k, v in data_map.items()
                }
                print("Row Data:", row_Stonedata)
                
                Mandatory_field=[]
                Error_field_val=[]
                
                # --- Fill dropdowns ---
                Function_Call.select_visible_text(self,f"(//select[@name='est_stones_item[stones_type][]'])[{row}]", row_Stonedata["Type"])
                Function_Call.select_visible_text(self,f"(//select[@name='est_stones_item[stone_id][]'])[{row}]", row_Stonedata["Name"])

                # --- Fill inputs ---
                if row_Stonedata["Pcs"]:
                    errors=Function_Call.fill_input(
                    self,wait,
                    locator=(By.XPATH, (f"(//input[@name='est_stones_item[old_stone_pcs][]'])[{row}]")),
                    value=row_Stonedata["Pcs"],
                    pattern = r"^\d{1,3}$",
                    field_name="Pcs",
                    screenshot_prefix="Pcs",
                    row_num=row_num,
                    Sheet_name=sheet_name
                    )
                    Error_field_val.extend(errors)
                    print(Error_field_val)
                else:
                    msg = f"'{None}' → Pcs field is mandatory ⚠️"
                    Mandatory_field.append("Pcs"); print(msg); Function_Call.Remark(self,row_num, msg,sheet_name)
                    
                
                if row_Stonedata["Wt"]:
                    errors=Function_Call.fill_input(
                    self,wait,
                    locator=(By.XPATH, f'(//input[@name="est_stones_item[old_stone_wt][]"])[{row}]'),
                    value=row_Stonedata["Wt"], 
                    pattern = r"^\d{1,3}(\.\d{1,3})?$",
                    field_name="Wt",
                    screenshot_prefix="Wt",
                    row_num=row_num,
                    Sheet_name=sheet_name
                    )
                    Error_field_val.extend(errors)
                    print(Error_field_val)
                else:
                    msg = f"'{None}' → Wt field is mandatory ⚠️"
                    Mandatory_field.append("Wt"); print(msg); Function_Call.Remark(self,row_num, msg,sheet_name)

                
                Function_Call.select_visible_text(self,f"(//select[@name='est_stones_item[uom_id][]'])[{row}]", row_Stonedata["Wt Type"])

                # --- Cal.Type radio button ---
                stonerow = row - 1
                cal_type = 1 if row_Stonedata["Cal.Type"] == "Wt" else 2
                Function_Call.click( self, f"(//input[@name='est_stones_item[cal_type][{stonerow}]' and @value='{cal_type}'])")

                # --- Rate & amount validation ---
                if row_Stonedata["Rate"]:
                    errors=Function_Call.fill_input(
                    self,wait,
                    locator=(By.XPATH, f'(//input[@name="est_stones_item[old_stone_rate][]"])[{row}]'),
                    value=row_Stonedata["Rate"], 
                    pattern = r"^\d{1,7}(\.\d{1,2})?$",
                    field_name="Rate",
                    screenshot_prefix="Rate",
                    row_num=row_num,
                    Sheet_name=sheet_name
                    )
                    Error_field_val.extend(errors)
                    print(Error_field_val)
                else:
                    msg = f"'{None}' → Rate field is mandatory ⚠️"
                    Mandatory_field.append("Rate"); print(msg); Function_Call.Remark(self,row_num, msg,sheet_name)

                
                
                
                amt_element = wait.until(EC.visibility_of_element_located(
                    (By.XPATH, f'(//input[@name="est_stones_item[stone_price][]"])[{row}]'))
                )
                table_amt = f"{float(amt_element.get_attribute('value')):.2f}"
                stone.calc_and_validate(self,row_Stonedata, table_amt)

                # --- Add multiple rows if needed ---
                if count > 1:
                    Function_Call.click(self, '(//button[@id="create_stone_old"])')
                    row += 1
                    count -= 1
    
                else:
                    wt_total = stone.test_tablewtvalue(self,'//input[@name="est_stones_item[old_stone_wt][]"]')
                    print("Total Wt:", wt_total)
                    amt_total = stone.test_tablevalue(self,'//input[@name="est_stones_item[stone_price][]"]')
                    print("Total Amount:", amt_total)

                    # Save stone details
                    Function_Call.click(self, '(//button[@id="update_stone_details"])[3]')
                    
                    
                    print("✅ Less weight detail added successfully")
                    return amt_total,wt_total
            


    def calc_and_validate(self, row_Stonedata, table_amt):
        """Validate stone amount calculation"""
        try:
            if row_Stonedata["Cal.Type"] == "Wt":
                total = float(row_Stonedata["Wt"]) * float(row_Stonedata["Rate"])
            else:
                total = float(row_Stonedata["Pcs"]) * float(row_Stonedata["Rate"])
            expected_amt = f"{math.ceil(total):.2f}"
            if expected_amt == table_amt:
                print(f"✅ Stone Rate Calculation correct, Got {table_amt}")
            else:
                print(f"❌ Stone Rate Calculation incorrect → Expected {expected_amt}, Got {table_amt}")
        except Exception as e:
            print(f"⚠️ Calculation skipped due to missing/invalid data: {e}")

    def test_tablewtvalue(self, xpath):
        """Get total from input fields"""
        wait = self.wait
        driver = self.driver

        rows = wait.until(EC.presence_of_all_elements_located((By.XPATH, xpath)))
        types = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//select[@name="est_stones_item[uom_id][]"]')))

        values = []

        for row, type_el in zip(rows, types):
            val = row.get_attribute("value")
            select = Select(type_el)
            unit = select.first_selected_option.text.strip().lower()

            num = float(val)
            if unit == "carat":
                num = round(num / 5, 3) # convert carat and round to 3 decimals
            values.append(num)

        if values :
            total_Wtvalue = round(sum(values), 3)
        else:
            total_Wtvalue=0.0
 
        print("Total Value:", total_Wtvalue)
        return total_Wtvalue
    
    def test_tablevalue(self, xpath):
        """Get total from input fields"""
        wait = self.wait
        driver = self.driver
        rows = wait.until(EC.presence_of_all_elements_located((By.XPATH, xpath)))
        value=[]
        sleep(4)
        for row in rows:
            val = row.get_attribute("value")  
            if val and val.strip():  # Ensure it's not empty
                value.append(float(val.strip()))
            else:
                print("No value found in input.")  
        print("Collected Values:", value)
        if value:
            total_value = round(sum(value), 3)
        else:
            total_value = 0.0
        return total_value