from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import  sleep
import unittest
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from Test_EST.Oldmetal_Lwt import stone
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font
import re
import random
import math

FILE_PATH = ExcelUtils.file_path
class ESTIMATION_Oldmetal(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)

    def test_estimation_Oldmetal(self,test_case_id):
        driver = self.driver
        wait = self.wait
        sleep(3)
        Sheet_name = 'OldMetal_Est'
        test_case_id = test_case_id
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, Sheet_name)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]
        row=1
        print(row)
        for row_num in range(2, valid_rows):
            current_id = sheet.cell(row=row_num, column=1).value  # Column 1 = Test Case Id
            if current_id == test_case_id:
                data = {
                    "Test Case Id": 1,
                    "Test Status": 2,
                    "Actual Status": 3,
                    "Metals": 4,
                    "Metal Type": 5,
                    "Category": 6,
                    "Product": 7,
                    "Purity": 8,
                    "Pcs": 9,
                    "G_Wt": 10,
                    "Dust_Wt": 11,
                    "Wastage%": 12,
                    "Rate": 13,
                    "Exchange Value%": 14,
                    "Stone": 15,
                    "Purpose": 16,
                    "Remarks": 17,
                    "Amount": 18,
                    "Field_validation_status": 19
                }
                row_data = {
                    key: sheet.cell(row=row_num, column=col).value
                    for key, col in data.items()
                }
                print(row_data)
                # Call your 'create' method
                Create_data = ESTIMATION_Oldmetal.create(self,row_data, row_num, Sheet_name, row)
                print(Create_data)
                row = row+1
                if Create_data:
                    Cal_Amt,Test_Status,Actual_Status= Create_data
                    ESTIMATION_Oldmetal.update_excel_status(self,row_num, Test_Status, Actual_Status, Sheet_name)                
        return Cal_Amt
                
    def create(self,row_data, row_num, Sheet_name, row):
        wait = self.wait
        Mandatory_field=[]
        Error_field_val=[]
        sleep(3)
        #Tag Check box selected
        if row > 1:
            sleep(4)
            Function_Call.click(self,'//button[@id="create_old_matel_details"]')        
        else:
            sleep(4)
            Function_Call.click2(self,'//input[@id="select_oldmatel_details"]')
           
        # Section
        sleep(3)
        if row_data["Metals"]:
            Function_Call.dropdown_select(
                self,f"(//span[starts-with(@id,'select2-est_oldmatel') and contains(@id,'[id_category]')])[{row}]", 
                row_data["Metals"],
                '//span[@class="select2-search select2-search--dropdown"]/input')
            
              
        if row_data["Metal Type"] is not None:
            Function_Call.dropdown_select(
                self,f"(//span[starts-with(@id,'select2-est_oldmatel') and contains(@id,'[id_old_metal_type]')])[{row}]", 
                row_data["Metal Type"],
                '//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Metal Type field is mandatory ⚠️"
            Mandatory_field.append("Metal Type"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
                
        # Design 
        if row_data["Category"] is not None:
            Function_Call.dropdown_select(
                self,f"(//span[starts-with(@id,'select2-est_oldmatel') and contains(@id,'[id_old_metal_category]')])[{row}]", 
                row_data["Category"],
                '//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Category field is mandatory ⚠️"
            Mandatory_field.append("Category"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
        
        #Product 
        if row_data["Product"] is not None:
            Function_Call.dropdown_select(
                self,f"(//span[starts-with(@id,'select2-est_oldmatel') and contains(@id,'[old_metal_product]')])[{row}]", 
                row_data["Product"],
                '//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Product field is mandatory ⚠️"
            Mandatory_field.append("Product"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
        
         
        # Purity
        if row_data["Purity"] is not None:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, f'(//input[@name="est_oldmatel[purity][]"])[{row}]'),
                value=row_data["Purity"],
                pattern = r"\d{1,2}(\.\d{1,4})?$",
                field_name="Purity",
                screenshot_prefix="Purity",
                row_num=row_num,
                Sheet_name=Sheet_name) 
        else:
            msg = f"'{None}' → Purity field is mandatory ⚠️"
            Mandatory_field.append("Purity"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
                    
        if row_data["Pcs"]:
            errors=Function_Call.fill_input(
            self,wait,
            locator=(By.XPATH, f'(//input[@name="est_oldmatel[pcs][]"])[{row}]'),
            value=row_data["Pcs"],
            pattern = r"^\d{1,3}$",
            field_name="Pcs",
            screenshot_prefix="Pcs",
            row_num=row_num,
            Sheet_name=Sheet_name
            )
            Error_field_val.extend(errors)
            print(Error_field_val)
        else:
            msg = f"'{None}' → Pcs field is mandatory ⚠️"
            Mandatory_field.append("Pcs"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
        
        
        if row_data["G_Wt"]:
            errors=Function_Call.fill_input(
            self,wait,
            locator=(By.XPATH, f'(//input[@name="est_oldmatel[gwt][]"])[{row}]'),
            value=row_data["G_Wt"], 
            pattern = r"^\d{1,3}(\.\d{1,3})?$",
            field_name="G.Wt",
            screenshot_prefix="G_Wt",
            row_num=row_num,
            Sheet_name=Sheet_name
            )
            Error_field_val.extend(errors)
            print(Error_field_val)
        else:
            msg = f"'{None}' → G_Wt field is mandatory ⚠️"
            Mandatory_field.append("G_Wt"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
       
        if row_data["Dust_Wt"]:
            errors=Function_Call.fill_input(
            self,wait,
            locator=(By.XPATH, f'(//input[@name="est_oldmatel[dwt][]"])[{row}]'),
            value=row_data["Dust_Wt"],
            pattern = r"^\d{1,5}$",
            field_name="Dust_Wt",
            screenshot_prefix="Dust_Wt",
            row_num=row_num,
            Sheet_name=Sheet_name,
            extra_keys=Keys.TAB
            )
            Error_field_val.extend(errors)
            print(Error_field_val)
        else:
            msg = f"'{None}' → Dust_Wt field is mandatory ⚠️"
            Mandatory_field.append("Dust_Wt"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
        
        
        if row_data["Wastage%"]:
            errors=Function_Call.fill_input(
            self,wait,
            locator=(By.XPATH, f'(//input[@name="est_oldmatel[wastage][]"])[{row}]'),
            value=row_data["Wastage%"],
            pattern = r"^\d{1,2}(\.\d{1,2})?$",
            field_name="Wastage%",
            screenshot_prefix="Wastage%",
            range_check = lambda v: 0 <= float(v) <= 99,
            row_num=row_num,
            Sheet_name=Sheet_name)
            Error_field_val.extend(errors)
            print(Error_field_val)
        else:
            msg = f"'{None}' → Wastage% Vlue field is mandatory ⚠️"
            Mandatory_field.append("Wastage%"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
        
               
        # --- Rate & amount validation ---
        if row_data["Rate"]:
            errors=Function_Call.fill_input(
            self,wait,
            locator=(By.XPATH, f'(//input[@name="est_oldmatel[rate][]"])[{row}]'),
            value=row_data["Rate"], 
            pattern = r"^\d{1,7}(\.\d{1,2})?$",
            field_name="Rate",
            screenshot_prefix="Rate",
            row_num=row_num,
            Sheet_name=Sheet_name
            )
            Error_field_val.extend(errors)
            print(Error_field_val)
        else:
            msg = f"'{None}' → Rate field is mandatory ⚠️"
            Mandatory_field.append("Rate"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
        
        if row_data["Exchange Value%"]:
            errors=Function_Call.fill_input(
            self,wait,
            locator=(By.XPATH, f'(//input[@name="est_oldmatel[touch][]"])[{row}]'),
            value=row_data["Exchange Value%"], 
            pattern = r"^\d{1,2}(\.\d{1,2})?$",
            field_name="Exchange Value%",
            screenshot_prefix="Exchange Value%",
            range_check = lambda v: 0 <= float(v) <= 99,
            row_num=row_num,
            Sheet_name=Sheet_name
            )
            Error_field_val.extend(errors)
            print(Error_field_val)
            
        if row_data["Stone"] =="Yes":
            Function_Call.click(self,f'(//table[@id="estimation_old_matel_details"]//td[15]/a)[{row}]')
            test_case_id=row_data["Test Case Id"]
            sheet_name ='Oldmetal_Lwt'
            data =stone.test_tagStone(self, sheet_name, test_case_id)
            Lwt,Stone_wt=data
        else:
            Lwt=0
            Stone_wt =0
            
            
      
        if row_data["Purpose"] is not None:
            Function_Call.dropdown_select(
                self,f"(//span[starts-with(@id,'select2-est_oldmatel') and contains(@id,'[id_old_metal_type]')])[{row}]", 
                row_data["Purpose"],
                '//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Purpose field is mandatory ⚠️"
            Mandatory_field.append("Purpose"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
        
        if row_data["Remarks"]:
            errors=Function_Call.fill_input2(
                self, f'(//input[@name="est_oldmatel[touch][]"])[{row}]',
                row_data["Remarks"])
        else:
            msg = f"'{None}' → Remarks field is mandatory ⚠️"
            Mandatory_field.append("Remarks"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)

        # Fetch values with one-liners
        Gwt   = ESTIMATION_Oldmetal.get_val(self, f'(//input[@name="est_oldmatel[gwt][]"])[{row}]')
        dwt  = ESTIMATION_Oldmetal.get_val(self, f'(//input[@name="est_oldmatel[dwt][]"])[{row}]')
        swt  = ESTIMATION_Oldmetal.get_val(self, f'(//input[@name="est_oldmatel[swt][]"])[{row}]')
        Nwt   = ESTIMATION_Oldmetal.get_val(self, f'(//input[@name="est_oldmatel[nwt][]"])[{row}]')
        Wast_per = ESTIMATION_Oldmetal.get_val(self, f'(//input[@name="est_oldmatel[wastage][]"])[{row}]')
        Rate    = ESTIMATION_Oldmetal.get_val(self, f'(//input[@name="est_oldmatel[rate][]"])[{row}]')
        Exchange_value =ESTIMATION_Oldmetal.get_val(self,f'(//input[@name="est_oldmatel[touch][]"])[{row}]')
        Taxable_Amt =ESTIMATION_Oldmetal.get_val(self,f'(//input[@name="est_oldmatel[amount][]"])[{row}]')
        print(Taxable_Amt)
        if swt == Stone_wt:
            pass
        else:
            msg = f"'{Stone_wt}' → Actual total Stone weight | Web page showing{swt}⚠️"
            Mandatory_field.append("Tol_StoneWt"); print(msg); Function_Call.Remark(self,row_num, msg,sheet_name)

 
        Cal_Amt=ESTIMATION_Oldmetal.calculation(self,Gwt,dwt,swt,Nwt,Wast_per,Rate,Exchange_value,Lwt)
        

        if Cal_Amt==Taxable_Amt:
            Test_Status = 'Pass'
            Actual_status=(f'Calculation Amount is correct {Cal_Amt}')
        else:    
            Test_Status = 'Fail'
            Actual_status = f"'{Cal_Amt}' → Actual total Calculation Amount | Web page showing Amount{Taxable_Amt}⚠️"
        return Cal_Amt,Test_Status,Actual_status
        
       
    def calculation(self,Gwt,dwt,swt,Nwt,Wast_per,Rate,Exchange_value,Lwt):
        wait = self.wait 
        gross_weight=Gwt
        dust_weight =dwt
        Stone_weight = swt
        net_weight = Nwt  
        wastage_percentage = Wast_per 
        Gold_Rate = Rate 
        Exchangevalue =Exchange_value
        Stone_rate =Lwt
        
        Find_Wt= gross_weight-dust_weight-Stone_weight
        Find_Wt=round(Find_Wt, 3)
        Va = (wastage_percentage/100)*Find_Wt
        Va = round(Va, 3)
        Find_NWt = Find_Wt - Va
        Find_NWt=round(Find_NWt, 3)

        if Find_NWt == net_weight:
            val=(net_weight*Gold_Rate)/100*Exchangevalue
            Amt = val+Stone_rate
            value=round(Amt, 2)
            print(f"cal amount {value}")
        else:
            print('Net Value Not match')
        return value
    
 
    # --- Helper to fetch field values safely ---
    def get_val(self,locator, cast=float, default=0):
        wait = self.wait
        el = wait.until(EC.presence_of_element_located((By.XPATH, locator)))
        val = el.get_attribute("value")
        if not val:  
            return default
        return cast(val)
    
    def update_excel_status(self,row_num, Test_Status, Actual_Status, function_name):
        print(function_name)
        # Load the workbook
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[function_name]  # or workbook["SheetName"]
        
        if Test_Status== 'Pass':
            # Write Test_Status into column 2
            sheet.cell(row=row_num, column=2, value=Test_Status).font=Font(bold=True, color="00B050")
            
            # Write Actual_Status in col 3 
            sheet.cell(row=row_num, column=3, value=Actual_Status).font = Font(bold=True, color="00B050")
        if Test_Status=='Fail':
            # Write Test_Status into column 2
            sheet.cell(row=row_num, column=2, value=Test_Status).font=Font(bold=True, color="FF0000")
            # Write Actual_Status in col 3 
            sheet.cell(row=row_num, column=3, value=Actual_Status).font = Font(bold=True, color="FF0000")
        # Save workbook
        workbook.save(FILE_PATH)
        # Get status from ExcelUtils
        Status = ExcelUtils.get_Status(FILE_PATH, function_name)
        # Print and return status
        print(Status)
        return Status




