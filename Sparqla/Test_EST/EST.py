from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import  sleep
import unittest
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from Test_EST.EST_Tag import ESTIMATION_TAG
from Test_EST.EST_Nontag import ESTIMATION_NonTag
from Test_EST.EST_Homebill import ESTIMATION_Homebill
from Test_EST.EST_oldmetal import ESTIMATION_Oldmetal
from Test_EST.EST_No import EstimationExtractor
from openpyxl.drawing.image import Image
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from openpyxl.styles import Font
from pathlib import Path
import base64, re, time

FILE_PATH = ExcelUtils.file_path
BASE_URL=ExcelUtils.BASE_URL
class ESTIMATION(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)


    def test_estimation(self):
        driver = self.driver
        wait = self.wait
        Board_Rate=[]
        Function_Call.click(self,"//span[@class='header_rate']/b[contains(text(),'INR')]")
        rate_text1 = wait.until(EC.presence_of_element_located((By.XPATH, "//li[@class='user-body rate_block_body']//tr[th[contains(text(),'Gold 22KT 1gm')]]/td"))).text
        rate_text2 = wait.until(EC.presence_of_element_located((By.XPATH, "//li[@class='user-body rate_block_body']//tr[th[contains(text(),'Gold 18KT 1gm')]]/td"))).text
        rate_text3 = wait.until(EC.presence_of_element_located((By.XPATH, "//li[@class='user-body rate_block_body']//tr[th[contains(text(),'Silver 1gm')]]/td"))).text
        # Example: "INR 9500"
        gold_rate22KT = int(float(rate_text1.replace("INR", "").strip()))
        Board_Rate.append(gold_rate22KT)
        print(gold_rate22KT)  
        gold_rate18KT = int(float(rate_text2.replace("INR", "").strip()))
        Board_Rate.append(gold_rate18KT)
        print(gold_rate18KT)  
        Silver_rate = int(float(rate_text3.replace("INR", "").strip()))
        Board_Rate.append(Silver_rate)
        print(Silver_rate)  
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Toggle navigation"))).click()
        Function_Call.click(self,"//span[contains(text(), 'Estimation')]")
        Function_Call.click(self,"//span[contains(text(), 'Add Estimation')]")
        # module=wait.until(EC.invisibility_of_element_located((By.XPATH,"//span[contains(text(), 'Estimation')]")))
        # driver.execute_script("arguments[0].scrollIntoView({block: 'nearest', inline: 'center'});", module)
        # module.click()
        # Estimation=wait.until(EC.invisibility_of_element_located((By.XPATH,"//span[contains(text(), 'Add Estimation')]")))
        # driver.execute_script("arguments[0].scrollIntoView({block: 'nearest', inline: 'center'});", Estimation)
        # Estimation.click()
        
        Sheet_name = "EST"                                        
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, Sheet_name)
        print(f"'{valid_rows}': valid rows")
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]
        for row_num in range(2, valid_rows):  
            data = {
                    "Test Case Id": 1,
                    "TestStatus": 2,
                    "ActualStatus": 3,
                    "Branch": 4, # Sheet has 'Branch\xa0' (non-breaking space)
                    "Sales Employee": 5,
                    "Esti For": 6,
                    "Customer": 7,
                    "Estimation TAG": 8,
                    "Estimation Non-Tag": 9, # Sheet has 'Estimation Non-Tag\n'
                	"Estimation Home Bill": 10, # Sheet has 'Estimation Home Bill\n'
                	"Estimation Old Metal": 11, # Sheet has 'Estimation Old Metal\n'
                }
            row_data = {key: sheet.cell(row=row_num, column=col).value 
                            for key, col in data.items()}
            print(row_data)
            if str(row_data.get("TestStatus", "")).strip().lower() != "run":
                print(f"⏭️ Skipping row {row_num} (Status={row_data.get('TestStatus')})")
                continue


            print(f"\n{'='*80}")
            print(f"🧪 Running Test Case: {row_data['Test Case Id']}")
            print(f"{'='*80}")
            # Call you 'create' method
            Create_data = self.create(row_data, row_num, Sheet_name,Board_Rate)
            print(Create_data)
            
            if Create_data:
                Test_Status,Actual_Status= Create_data
                ESTIMATION.update_excel_status(self,row_num, Test_Status, Actual_Status,Sheet_name)
                
    def create(self,row_data, row_num, Sheet_name,Board_Rate):
        driver = self.driver
        wait = self.wait
        driver.refresh()
        Mandatory_field=[]
        
        if row_num != 2:
            sleep(2)
            driver.get(BASE_URL + "index.php/admin_ret_estimation/estimation/add")
            
        #Branch
        if row_data["Branch"] is not None:
           Function_Call.dropdown_select(self,'//span[@id="select2-branch_select-container"]',row_data['Branch'],'//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Branch field is mandatory ⚠️"
            Mandatory_field.append("Branch"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
            
        #Sales Employes
        if row_data["Sales Employee"] is not None:
            Function_Call.dropdown_select(self,f"//span[@id='select2-emp_select-container']", row_data["Sales Employee"],'//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Sales Employee field is mandatory ⚠️"
            Mandatory_field.append("Sales Employee"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
        
        #Esti For
        Esti_For = {
                "Customer": '//input[@id="type1"]',
                "Branch Transfer": '//input[@id="type2"]',
                "Company":'//input[@id="type3"]'
            }
        print(Esti_For[row_data["Esti For"]])
        Function_Call.click(self,Esti_For[row_data["Esti For"]])
        
        # Customer
        if row_data["Customer"]:
            Function_Call.fill_autocomplete_field(self,"est_cus_name", row_data["Customer"])
        else:
            msg = f"'{None}' → Customer field is mandatory ⚠️"
            Mandatory_field.append(msg)
            print(msg)
            Function_Call.Remark(row_num, msg)
        sleep(3)
        Function_Call.click(self,'(//button[@class="btn btn-close btn-warning "])[6]')
        Error=[]
        Actual_Status=[]
        Total_amount=[]
        old_amount=[]
        bill_type=[]
        Row_No=1
        if row_data["Estimation TAG"]=="Yes":
            test_case_id=row_data["Test Case Id"]
            # Capture both the total amount and the list of source rows found
            Call_Tag, tag_found_rows = ESTIMATION_TAG.test_estimationtag(self,test_case_id,Board_Rate)     
            print(Call_Tag)
            Total_amount.append(Call_Tag)
            bill_type.append("SALES")
            No=1
            Row_No=Row_No+1
        if row_data["Estimation Non-Tag"]=="Yes":
            test_case_id=row_data["Test Case Id"]
            Call_Non_Tag, nontag_found_rows = ESTIMATION_NonTag.test_estimation_Nontag(self,test_case_id,Board_Rate)
            print(Call_Non_Tag)
            Total_amount.append(Call_Non_Tag)
            if bill_type:
                bill_type.pop()
            bill_type.append("SALES")
            
        if row_data["Estimation Home Bill"]=="Yes":
            test_case_id=row_data["Test Case Id"]
            Call_HomeBill=ESTIMATION_Homebill.test_estimation_Homebill(self,test_case_id,Board_Rate)
            print(Call_HomeBill)
            Total_amount.append(Call_HomeBill)
            if bill_type:
                bill_type.pop()
            bill_type.append("SALES")
            
        if row_data["Estimation Old Metal"]=="Yes":
            test_case_id=row_data["Test Case Id"]
            Call_OldMetal=ESTIMATION_Oldmetal.test_estimation_Oldmetal(self,test_case_id)
            print(Call_OldMetal)
            old_amount.append(Call_OldMetal)
            if bill_type:
                bill_type.pop()
                bill_type.append("SALES & PURCHASE")
            else:
                bill_type.append("PURCHASE")
                
        if Total_amount:           
            print('Done')
            print("Total_amount list:", Total_amount)
            total = sum(float(v) for v in Total_amount if v is not None)
            value = 0.0   # Initialize to avoid NameError in fail path
            print("Total =", total)
            if total!=0:
                elements = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//span[@class='summary_lbl summary_pur_amt']")))
                value = elements[0].text.strip()
                value = float(value) if value else 0.0
                print("Purchase Amount:", value)
            else:
                Test_Status='Fail'   
                Actual_Status= f'❌ Calculation result: {total:.2f} | Inventory Failure or No Data Added'
                Function_Call.click(self, '//button[@class="btn btn-default btn-cancel"]') 
                return Test_Status, Actual_Status
        else:
            total=0
            value=0
        if old_amount:
            print('Done')
            print("Total_amount list:", old_amount)
            Old_Amt = sum(float(v) for v in old_amount if v is not None)
            print("Old Purchase Total =", Old_Amt)
            elements = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//span[@class='summary_lbl summary_sale_amt']")))
            old_value = elements[0].text.strip()
            old_value = float(old_value) if old_value else 0.0
            print("Purchase Amount:", old_value)
        else:
            Old_Amt=0
            old_value=0
        
        
           
        total_str = "{:.2f}".format(total)
        value_str = "{:.2f}".format(value)
        Old_Amt_str = "{:.2f}".format(Old_Amt)
        old_value_str = "{:.2f}".format(old_value)
           
        #if total_str == value_str and Old_Amt_str == old_value_str:
        if total_str:
            Test_Status='Pass'
            Actual_Status='✅Calculation is correct'
            old_tabs = driver.window_handles
            Function_Call.click(self,'//button[@id="est_print"]')
            # wait for new tab
            for _ in range(20):
                new_tabs = driver.window_handles
                if len(new_tabs) > len(old_tabs):
                    new_tab = [t for t in new_tabs if t not in old_tabs][0]
                    driver.switch_to.window(new_tab)
                    break
                time.sleep(0.3)

            # get the PDF URL
            url = driver.current_url
            print("PDF URL:", url)
            viewer_url = driver.current_url  # after print button opens PDF tab
            extractor = EstimationExtractor(driver)
            EST_details = extractor.save_and_extract(out_pdf="est_3.pdf", viewer_url=viewer_url)
            print(EST_details)
            ESTIMATION.update_EST_Details(self,EST_details,row_data,bill_type,row_num)
            
            # [NEW] Update source sheets (Tag_Detail / Purchase_TagDetail) with Estimation No
            if row_data["Estimation TAG"] == "Yes" and 'tag_found_rows' in locals() and tag_found_rows:
                ESTIMATION_TAG.update_source_sheets_with_estimation(tag_found_rows, EST_details["Estimate"])
            
            # [NEW] Update source sheets (NonTag_Detail / Purchase_NonTagDetail) with Inventory tracking
            if row_data["Estimation Non-Tag"] == "Yes" and 'nontag_found_rows' in locals() and nontag_found_rows:
                ESTIMATION_NonTag.update_source_inventory(nontag_found_rows, EST_details["Estimate"])

            # driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
            windows = driver.window_handles
            sleep(3)
            driver.switch_to.window(old_tabs[0])
           
            # pass URL directly to your class
            # Estimation_number(self.driver).url(url)
        else:
            Test_Status='Fail'   
            Actual_Status= f'❌ Calculation Error in {total_str} | Web Value={value_str}'
            Function_Call.click(self, '//button[@class="btn btn-default btn-cancel"]') 
        return Test_Status, Actual_Status
  
        
    def update_excel_status(self,row_num, Test_Status, Actual_Status, function_name):
        print(function_name)
        # Load the workbook
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[function_name]  # or workbook["SheetName"]
        row=sheet.max_row+1
        if Test_Status== 'Pass':
            # Write Test_Status into column 2
            sheet.cell(row=row_num, column=2, value=Test_Status).font=Font(bold=True, color="00B050")
            
            # Write Actual_Status in col 3 
            sheet.cell(row=row_num, column=3, value=Actual_Status).font = Font(bold=True, color="00B050")
        if Test_Status=='Fail':
            # Write Test_Status into column 2
            sheet.cell(row=row_num, column=2, value=Test_Status).font=Font(bold=True, color="FF0000")
            # Write Actual_Status in col 3 
            sheet.cell(row=row_num, column=3, value=Actual_Status).font = Font(bold=True, color="FF0000")
        # Save workbook
        workbook.save(FILE_PATH)
        # Get status from ExcelUtils
        Status = ExcelUtils.get_Status(FILE_PATH, function_name)
        # Print and return status
        print(Status)
        return Status
    
    def update_EST_Details(self,EST_details,row_data,bill_type,row_num):
        function_name='Billing'
        type=str(bill_type[0])
        # 
        # Load the workbook
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[function_name]  # or workbook["SheetName"]
        # Find the first available row by checking where Column 1 (Test Case Id) is empty
        row = 2
        while sheet.cell(row=row, column=1).value is not None:
            row += 1
        
        # Update details in the Billing sheet
        # Column mapping aligned with Bill.py data_map
        # Auto-generate Test Case Id based on previous row (Format: TC001)
        prev_id_full = sheet.cell(row=row - 1, column=1).value
        try:
            if prev_id_full and str(prev_id_full).startswith("TC"):
                # Extract digits from the TC format (e.g., 'TC005' -> 5)
                prev_num = int(re.search(r'\d+', str(prev_id_full)).group())
                new_id_num = prev_num + 1
            else:
                new_id_num = 1
        except Exception:
            new_id_num = 1
            
        new_id = f"TC{new_id_num:03d}"
        
        sheet.cell(row=row, column=1, value=new_id)
        sheet.cell(row=row, column=4, value=row_data["Branch"])
        sheet.cell(row=row, column=5, value=row_data["Esti For"])
        sheet.cell(row=row, column=6, value=row_data["Sales Employee"])
        sheet.cell(row=row, column=7, value=row_data["Customer"]) # Customer Number
        sheet.cell(row=row, column=9, value=row_data["Branch"])   # Delivery Location (using Branch)
        sheet.cell(row=row, column=10, value=type)               # Bill Type
        sheet.cell(row=row, column=11, value="No")               # driect
        
        # Estimation Details (from PDF extraction: {Estimate, cgst, sgst, igst, total})
        sheet.cell(row=row, column=12, value=EST_details["Estimate"]) # EstNo
        sheet.cell(row=row, column=13, value=EST_details["sgst"])     # SGST
        sheet.cell(row=row, column=14, value=EST_details["cgst"])     # CGST
        sheet.cell(row=row, column=15, value=EST_details["total"])    # Total
        sheet.cell(row=row, column=33, value=EST_details["igst"])     # IGST
            
        workbook.save(FILE_PATH)
        workbook.close()
        