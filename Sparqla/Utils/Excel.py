import pandas as pd
import re
import win32com.client
import os
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from pathlib import Path     # ← ADD THIS
import shutil 


class ExcelUtils:
    # Base folder = Sparqla (since this file is Sparqla/Utils/Excel.py)
    _base_dir = Path(__file__).resolve().parents[1]

    # Original template in repo
    template_path = _base_dir / "C:\Users\Dell\Desktop\sqrqlas\Sqarqla_Retail_data2.xlsx"

    # Output folder & writable copy (for Jenkins run)
    output_dir = _base_dir / "output"
    output_dir.mkdir(exist_ok=True)

    # This is the file we will read & write in tests
    file_path = output_dir / "C:\Users\Dell\Desktop\sqrqlas\Sqarqla_Retail_data2.xlsx"

    @staticmethod
    def ensure_copy_exists():
        """
        Make sure we have a writable copy of the template for this run.
        """
        if not ExcelUtils.file_path.exists():
            shutil.copy2(ExcelUtils.template_path, ExcelUtils.file_path)
            print(f"Created Excel copy at: {ExcelUtils.file_path}")
        else:
            print(f"Using existing Excel copy: {ExcelUtils.file_path}")
    #file_path =  "C:\ProgramData\Jenkins\.jenkins\workspace\Retail_Testing\Sparqla\Sqarqla_Retail_data2.xlsx"
    
    # Start Excel
    def ExcelClose(file_path):
        excel_app = win32com.client.Dispatch("Excel.Application")
        excel_app.Visible = False

        # Initialize workbook variable
        workbook = None

        # Check if workbook is already open
        for wb in excel_app.Workbooks:
            if os.path.abspath(wb.FullName) == os.path.abspath(file_path):
                workbook = wb
                break

        if workbook:
            workbook.Save()   # Save changes
            workbook.Close()  # Close workbook
            print("Workbook saved and closed.")
        else:
            print("Workbook is not currently open.")

        excel_app.Quit()

    # open Excel file
    def read_excel(file_path):
        df = pd.read_excel(file_path)
        return df.fillna("")
    
    #To get sheet names in a list from workbook
    def get_sheet_names(file_path):
        excel_file = pd.ExcelFile(file_path)
        File = excel_file.sheet_names
        noOfSheets = len(excel_file.sheet_names)
        sheet_names = list(File)
        #print(noOfSheets)
        return sheet_names

    #Read & Get value from master sheet data
    def get_master_sheet_data(file_path):
        """
        Reads the 'Master' sheet and returns the functions marked for execution.
        """
        df = pd.read_excel(file_path, sheet_name="Master")
        print(df)
        functions_to_execute = []
        for index, row in df.iterrows():
            if re.match(r"yes", str(row["Execution"]), re.IGNORECASE):
                functions_to_execute.append(row["Function"])
                print(functions_to_execute)
        return functions_to_execute
    
    #To Fetch Valid Rows
    def get_valid_rows (file_path,function_name):
        workbook = load_workbook(file_path)
            # workbook = load_workbook(file_path)
        sheet = workbook[function_name]
        i =2
        count = 0
        while (i<100):
            cellvalue = sheet.cell(row =i, column=1).value
            if cellvalue is None:
                break
            i=i+1
            count = count+1
        return(count+2)
    
    #Fetch No Of Pass & Fail
    def get_Status(file_path,function_name):
        workbook = load_workbook(file_path)
        sheet = workbook[function_name]
        count = (ExcelUtils.get_valid_rows(file_path,function_name))+1
        Pass = 0
        Fail = 0
        i = 2
        while(i<=count):
            cellvalue =sheet.cell(row =i, column=2).value
            if cellvalue and cellvalue.strip().lower() == "pass":  # Case-insensitive comparison
               Pass += 1
            elif cellvalue and cellvalue.strip().lower() == "fail":  # Case-insensitive comparison
               Fail += 1
            # if cellvalue == "Pass":
            #     Pass = Pass+1
            # elif cellvalue == "Fail":
            #     Fail = Fail+1   
            i=i+1    
        status = (f"Pass {Pass}, Fail {Fail}")
        workbook.close()
        print(status)
        return(status)

    #Update Master Status
    def update_master_status(file_path,Status,function_name):
        
        workbook = load_workbook(file_path)
            # workbook = load_workbook(file_path)
        sheet = workbook["Master"]
        print(Status)
        i =2
        while(i<=100):
            cellvalue =sheet.cell(row =i, column=1).value
            if cellvalue == function_name:
                sheet.cell(row=i,column=3).value = Status 
                workbook.save(file_path)
                break
            i=i+1 
                 
    def Test_case_id_count(file_path,function_name,test_case_id):
        # Load the workbook and sheet
        workbook = load_workbook(file_path)
        sheet = workbook[function_name]
        # Initialize count
        count = 0
        # Loop through rows starting from row 2
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == test_case_id:  # Assuming Test Case ID is in column A (index 0)
                count += 1
        # Store count in a variable
        tc_count = count
        print(f"Test Case ID {test_case_id} count: {tc_count}")
        return tc_count
    
    def Smith_count(file_path,function_name,test_case_id):
        # Load the workbook and sheet
        workbook = load_workbook(file_path)
        sheet = workbook[function_name]
        # Initialize count
        count = 0
        # Loop through rows starting from row 2
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == test_case_id:  # Assuming Test Case ID is in column A (index 0)
                count += 1
        # Store count in a variable
        tc_count = count

        print(f"Test Case ID {test_case_id} count: {tc_count}")
        return tc_count

    def Lot_details(file_path, function_name):
        # Read the Excel file with pandas, specifying the sheet
        df = pd.read_excel(file_path, sheet_name=function_name)

        # Extract the "Lot" column into a list
        lot_list = df["Lot"].dropna().tolist()
        lot_counts = df["Lot"].value_counts().tolist()

        print(lot_list)
        return lot_list
                
                
    def Tag_reserve(file_path, function_name):
         
        workbook = load_workbook(file_path)
        sheet = workbook[function_name]
        count = 0

        # Iterate through rows starting from row 2 (skip header)
        for row in sheet.iter_rows(min_row=2, values_only=False):  
            order_type = row[8].value       # Column J (OrderType)
            tag_scan = row[10].value        # Column K (TagScan)

            if order_type and "Tag Reserve" in str(order_type):  # Check OrderType
                if not tag_scan:  # Empty TagScan
                    count += 1
                else:
                    print ("Tag is avabile") # Reset count if TagScan has value           
            else:
                pass # Not a Tag Reserve row
        print(count)   
        return count


    def update_tag_id(file_path, function_name,TAG):
        workbook = load_workbook(file_path)
        sheet = workbook[function_name]
        available_tags=TAG
        print(available_tags)
        tag_index = 0  
        for row in range(2, sheet.max_row + 1):
            if "Tag Reser" in str(sheet[f"I{row}"].value):   # check OrderType
                if not sheet[f"K{row}"].value:              # if TagScan is empty
                    if tag_index < len(available_tags):
                        sheet[f"K{row}"] = available_tags[tag_index]
                        print(f"Row {row} → {available_tags[tag_index]}")
                        tag_index += 1

        workbook.save(file_path)
        return("✅ Done! Saved as Tag Id updated.xlsx")
    
    
    # def customer_details(file_path, function_name):
    #     # Read the Excel file with pandas, specifying the sheet
    #     df = pd.read_excel(file_path, sheet_name=function_name)

    #     # Extract the "Lot" column into a list
    #     Cus_list = df["Customer Number"].dropna().tolist()
    #     # Cus_Counts = df["Customer Number"].value_counts().tolist()

    #     print(Cus_list)
    #     return Cus_list
    

    def customer_details(file_path, function_name):
        # Read column as string (avoids .0 issue)
        df = pd.read_excel(file_path, sheet_name=function_name)

        # Replace NaN with 0, then cast to int (but safe for big numbers)
        Cus_list = df["Customer Number"].fillna(0).astype("int64").tolist()
    
        print(Cus_list)
        return Cus_list

    def get_column_number(file_path, sheet_name):
        # Load workbook and sheet
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Find the column number of "Status"
        status_col_num = None
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value and cell_value.strip().lower() == "field_validation_status":
                status_col_num = col
                print(f"Status column number is: {status_col_num}")
                print(type(status_col_num))  # <class 'int'>
                return status_col_num
        
        
    def update_Lot_id(file_path, Lot_id, row_count, Pcs_count,workbook):
        sheet = workbook["Tag"]
        Pcs_count = int(Pcs_count)  # ensure it's an integer
               # fallback if something wrong
        for i in range(Pcs_count):     
            row_Num=row_count+i
            before = sheet.cell(row=row_Num, column=5).value 
            print(f'{before} Before Lot')
            sheet.cell(row=row_Num, column=5).value=Lot_id
            After= sheet.cell(row=row_Num, column=5).value
            print(f'{After} After Lot')
        workbook.save(file_path)
        workbook.close()
        return Pcs_count,"Lot ID Added in Tag sheet successfully"    


       
        
           
    
                
        

    








