
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.actions.wheel_input import ScrollOrigin
from time import  sleep
import unittest
from openpyxl.drawing.image import Image
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
from openpyxl.styles import Font
from Utils.Excel import ExcelUtils
from datetime import datetime
import re
import os

FILE_PATH=ExcelUtils.file_path
class Function_Call(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 20)
        
    def click2(self,xpath):
        wait = self.wait
        driver=self.driver
        clicked= wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
        sleep(2)
        ActionChains(driver).move_to_element(clicked).perform()
        sleep(2)
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", clicked)
        sleep(4)
        try:
           clicked.click()
        except:
            
            driver.execute_script("window.scrollBy(0, -700);") 
            sleep(2)
            clicked.click()
            
    def click(self, xpath):
        wait = self.wait
        driver = self.driver 
        try:
            # 1. Wait for element to be present
            element = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
            
            # 2. Try standard scroll and click
            try:
                # Scroll into view (center)
                driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'nearest'});", element)
                sleep(0.5)
                # Traditional click
                wait.until(EC.element_to_be_clickable((By.XPATH, xpath))).click()
                print("✅ Clicked (Standard):", xpath)
                return
            except Exception as e:
                print(f"⚠️ Standard click failed for {xpath}, attempting robust fallback... Error: {str(e)}")
            
            # 3. Robust Fallback: ActionChains Move + JS Click
            sleep(1)
            # Re-locate to avoid stale element
            clicked = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
            
            # Move mouse to element (triggers hover and ensures visibility for some frameworks)
            try:
                ActionChains(driver).move_to_element(clicked).perform()
            except:
                pass
            
            # Use JS to trigger the click with event bubbling
            js_script = """
            var el = arguments[0];
            var ev = document.createEvent('MouseEvents');
            ev.initMouseEvent('mousedown', true, true, window, 0, 0, 0, 0, 0, false, false, false, false, 0, null);
            el.dispatchEvent(ev);
            ev.initMouseEvent('mouseup', true, true, window, 0, 0, 0, 0, 0, false, false, false, false, 0, null);
            el.dispatchEvent(ev);
            el.click();
            """
            driver.execute_script(js_script, clicked)
            print("✅ Clicked (Robust JS):", xpath)
            
        except Exception as e2:
            print(f"❌ All click attempts failed for {xpath}. Error: {str(e2)}")
            # Take screenshot on total failure
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            driver.save_screenshot(os.path.join(ExcelUtils.SCREENSHOT_PATH, f"ClickError_{timestamp}.png"))
            raise e2
        
            
    
    def get_text(self, xpath):
        wait = self.wait
        driver = self.driver
        element = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
        driver.execute_script("arguments[0].scrollIntoView({block: 'nearest', inline: 'center'});", element)
        return element.text.strip()  # ✅ Return the actual text
        
    def get_value(self, xpath):
        wait = self.wait
        driver = self.driver
        element = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
        driver.execute_script("arguments[0].scrollIntoView({block: 'nearest', inline: 'center'});", element)
        value = element.get_attribute("value")
        return value.strip() if value else ""
        # ✅ Return the actual value
        
        
    def dropdown_select(self,xpath, value,text_xpath):
            wait = self.wait
            sleep(2)
            Function_Call.click(self,xpath)
            sleep(1)
            el = wait.until(EC.element_to_be_clickable((By.XPATH, text_xpath)))
            el.clear()
            el.send_keys(value, Keys.ENTER)
            
    def dropdown_select2(self,xpath, value,text_xpath):
            wait = self.wait
            sleep(2)
            Function_Call.click(self,xpath)
            el = wait.until(EC.element_to_be_clickable((By.XPATH, text_xpath)))
            el.clear()
            el.send_keys(value, Keys.ENTER)
            
    def dropdown_country(self,xpath, value,text_xpath,clickable):
        wait = self.wait
        sleep(3)
        Function_Call.click(self,xpath)
        sleep(1)
        el=wait.until(EC.element_to_be_clickable((By.XPATH, text_xpath)))
        el.clear()
        el.send_keys(value)
        wait.until(EC.element_to_be_clickable((By.XPATH,clickable))).click()
            
    def fill_autocomplete_field(self, field_id, value):
        driver, wait = self.driver, self.wait
        field = wait.until(EC.element_to_be_clickable((By.ID, field_id)))
        field.click()
        field.clear()
        field.send_keys(value)
        sleep(2)
        field.send_keys(Keys.BACKSPACE)

        wait.until(EC.presence_of_element_located((By.XPATH, f"//li[contains(text(),'{value}')]"))).click()   
            
    def select_visible_text(self, locator,  value):
        wait = self.wait
        element = self.wait.until(EC.element_to_be_clickable((By.XPATH, locator)))
        element.click()
        Select(element).select_by_visible_text(str(value))
        print(f"✅ Dropdown filled with {value}")
        
    def Image_upload(self,xpath, value):
        wait = self.wait
        wait.until(EC.presence_of_element_located((By.XPATH, xpath))).send_keys(value)
        
        
    def fill_input(self, wait,locator, value, field_name, row_num, pattern=None, screenshot_prefix="", extra_keys=None, range_check=None, Sheet_name="",Date_range=None):
        """Generic handler for text/numeric fields with validation and optional range check."""
        errors = []
        driver, wait = self.driver, self.wait
        test_case_id = row_num
        field = wait.until(EC.element_to_be_clickable(locator))
        # driver.execute_script("arguments[0].scrollIntoView({block: 'nearest', inline: 'center'});", field)
        field.click()
        field.clear()
        
        if value is not None:
            # Handle datetime objects
            if isinstance(value, datetime):
                value = value.strftime("%d-%m-%Y")
            else:
                value = str(value)
                
            field.send_keys(value)
            if extra_keys:
                field.send_keys(extra_keys)
        entered_value = field.get_attribute("value")
        if entered_value == "" or entered_value=='0':
            driver.save_screenshot(os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{screenshot_prefix}_{test_case_id}.png"))
            msg = f"{value} → Not allowed in {field_name} ⚠️"
            Function_Call.Remark(self,row_num, msg, Sheet_name)
            errors.append(field_name)
            return "Fail",msg

        # Regex / Range check
        valid = True
        if pattern:
            valid = re.fullmatch(pattern, entered_value) is not None
        if valid and range_check:
            try:
                valid = range_check(float(entered_value))
            except:
                valid = False
        
        if Date_range:
            try:
                # Normalize separator to '-' for strptime compatibility
                normalized_date = entered_value.replace("/", "-")
                entered_date = datetime.strptime(normalized_date, "%d-%m-%Y").date()
                today = datetime.today().date()
                
                error_msg = None
                
                # Validation based on specific modes
                if Date_range == "future":
                    if entered_date < today:
                        pass
                    else:
                        error_msg = f"{field_name} must be a FUTURE date -> {entered_value}"
                
                if Date_range == "past":
                    if entered_date > today:
                        pass
                    else:
                        error_msg = f"{field_name} must be a PAST date -> {entered_value}"
                
                if Date_range in ["current", "today"]:
                    if entered_date != today:
                        pass
                    else:
                        error_msg = f"{field_name} must be TODAY'S date -> {entered_value}"
                        
                if Date_range == "future_or_current":
                    if entered_date <= today:
                        pass
                    else:
                        error_msg = f"{field_name} must be TODAY or FUTURE date -> {entered_value}"
                        
                if Date_range == "past_or_current":
                    if entered_date >= today:
                        pass
                    else:
                        error_msg = f"{field_name} must be TODAY or PAST date -> {entered_value}"                
                if error_msg:
                    raise ValueError(error_msg)
            except Exception as e:
                driver.save_screenshot(os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{screenshot_prefix}_{test_case_id}.png"))
                msg = str(e) if "must be" in str(e) else f"{field_name} invalid date format or criteria -> {entered_value}"
                Function_Call.Remark(self, row_num, msg, Sheet_name)
                errors.append(field_name)
                return "Fail",msg

        if not valid:
            driver.save_screenshot(os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{screenshot_prefix}_{test_case_id}.png"))
            msg = f"'{entered_value}' -> Invalid data allowed in {field_name} [FAIL]"
            Function_Call.Remark(self,row_num, msg, Sheet_name)
            errors.append(field_name)
            return "Fail",msg
        else:
            print(f"'{entered_value}' → Accepted {field_name} ✅")
        return "Pass"
    
    def fill_input2(self,xpath, value, clear=True):
        wait = self.wait
        el = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
        el.click()
        if clear:
            el.clear()
        el.send_keys(value)
    
    def fill_input3(self, wait,locator, value, field_name, row_num, pattern=None, screenshot_prefix="", extra_keys=None, range_check=None, Sheet_name=""):
        """Generic handler for text/numeric fields with validation and optional range check."""
        errors = []
        driver, wait = self.driver, self.wait
        test_case_id = row_num
        field = wait.until(EC.element_to_be_clickable(locator))
        # driver.execute_script("arguments[0].scrollIntoView({block: 'nearest', inline: 'center'});", field)
        field.click()
        field.clear()
        if value is not None:
            # Handle datetime objects
            if isinstance(value, datetime):
                value = value.strftime("%d-%m-%Y")
            else:
                value = str(value)
                
            field.send_keys(value)
            if extra_keys:
                field.send_keys(extra_keys)
                error=Function_Call.alert2(self,screenshot_prefix,test_case_id)
                if error:
                    msg = (f"⚠️ Found the message:'{error}'")
                    print(msg)
                    Function_Call.Remark(self,row_num, msg, Sheet_name)
                    errors.append(field_name)
                    return  "Fail", msg
        entered_value = field.get_attribute("value")
        if entered_value == "":
            driver.save_screenshot(os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{screenshot_prefix}_{test_case_id}.png"))
            msg = f"{value} → Not allowed in {field_name} ⚠️"
            Function_Call.Remark(self,row_num, msg, Sheet_name)
            errors.append(field_name)
            return "Fail",msg

        # Regex / Range check
        valid = True
        if pattern:
            valid = re.fullmatch(pattern, entered_value) is not None
        if valid and range_check:
            try:
                valid = range_check(float(entered_value))
            except:
                valid = False

        if not valid:
            driver.save_screenshot(os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{screenshot_prefix}_{test_case_id}.png"))
            msg = f"'{entered_value}' -> Not allowed in {field_name} [FAIL]"
            Function_Call.Remark(self,row_num, msg, Sheet_name)
            errors.append(field_name)
            return "Fail",msg
        else:
            print(f"'{entered_value}' -> Accepted {field_name} [OK]")
        return "Pass"
    
    
    def Remark(self,row_num,Field_validation_satus,Sheet_name): 
        print(Sheet_name)
        Col =ExcelUtils.get_column_number(FILE_PATH, Sheet_name,)
        print(type(Col))
        # Load the workbook
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]  # or workbook["SheetName"]
        if Field_validation_satus:
            sheet.cell(row=row_num, column=Col, value=Field_validation_satus).font = Font(bold=True, color="FF8000")
        # Save workbook
        workbook.save(FILE_PATH)
    
    def alert6(self, Xpath):
        old_html = self.driver.page_source
        Function_Call.click(self, Xpath)

        WebDriverWait(self.driver, 5).until(
            lambda d: d.page_source != old_html
        )

        msg = self.driver.execute_script("""
            let el = document.querySelector('#toaster .message');
            return el ? el.innerText.trim() : null;
        """)
        print("TOAST:", msg)
        return msg
       
        
    def alert1(self,Xpath):
        wait = self.wait 
        # Wait for toaster message to appear
        Function_Call.click(self,Xpath)
        try:
            alert_msg = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "(//div[@id='toaster']//span[@class='message'])[1]")))

            # alert_msg = WebDriverWait(self.driver, 5).until(
            #     EC.visibility_of_element_located((By.CSS_SELECTOR, "#toaster .alert"))
            # ).text
            print(alert_msg)
            alert_text = re.sub(r"[x\s]*Close", "", alert_msg).replace("\n", "").strip()
            # alert_text = alert_msg.replace("xClose", "").replace("\n", "").strip()
            Actual_Status= (f"[WARN] Found the message:'{alert_text}'") # prints: Select Order Branch
            print(Actual_Status)
        except:
            alert_text =None
            print(alert_text)
        return alert_text
    
    def alert2(self,screenshot_prefix,test_case_id):
        wait = self.wait 
        driver = self.driver
        try:
            alert_txt = WebDriverWait(self.driver, 2).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#toaster .alert"))
            ).text
            print(alert_txt)
            driver.save_screenshot(os.path.join(ExcelUtils.SCREENSHOT_PATH, f"{screenshot_prefix}_{test_case_id}.png"))
            print(alert_txt)
            alert_text = re.sub(r"[x\s]*Close", "", alert_txt).replace("\n", "").strip()
        except:
            alert_text =None
        return alert_text
    
       
    def alert(self):
        try:
            driver = self.driver
            # Wait until alert is present
            alert = WebDriverWait(driver, 10).until(lambda d: d.switch_to.alert)
            # Get the text from the alert
            alert_text = alert.text
            # Accept the alert (click OK)
            alert.accept()
            return alert_text
        except:
            return None
    
    def dropdown_subdesign_val(self,xpath, value,text_xpath,purity,before_purity,before_subdesign):
            wait = self.wait
            sleep(2)
            Function_Call.click(self,xpath)
            el = wait.until(EC.element_to_be_clickable((By.XPATH, text_xpath)))
            el.clear()
            if before_purity == purity and value == before_subdesign:
              el.send_keys(value, Keys.ENTER)
              error=Function_Call.alert3(self)
              return error 
            else:
              el.send_keys(value, Keys.ENTER)

    
    def alert3(self):
        wait = self.wait 
        driver = self.driver
        try:
            alert_txt = WebDriverWait(self.driver, 2).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#toaster .alert"))
            ).text
            print(alert_txt)
            driver.save_screenshot(os.path.join(ExcelUtils.SCREENSHOT_PATH, "Subdesign.png"))
            print(alert_txt)
            alert_text = re.sub(r"[x\s]*Close", "", alert_txt).replace("\n", "").strip()
        except:
            alert_text =None
        return alert_text
    
    def select(self, xpath,value):
        wait = self.wait
        dropdown = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
        Select(dropdown).select_by_visible_text(value)