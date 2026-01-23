from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import  sleep
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from Utils.Board_rate import Boardrate
from Test_Bill.Sales import SALES
from Test_Bill.Credit_Card import CreditCard
from Test_Bill.Cheque import Cheque
from Test_Bill.NetBanking import NetBanking
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
import re
import unittest

FILE_PATH = ExcelUtils.file_path 
class Grn_Entry(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)

    def test_Receipt(self):
        driver = self.driver
        wait = self.wait 
        
        Rate=Boardrate.Todayrate(self)
        print(Rate)  
        
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Toggle navigation"))).click()
        Function_Call.click(self,"//span[contains(text(), 'Billing')]")
        Function_Call.click(self,"//span[contains(text(), 'GRN Entry')]")
        
        Sheet_name = "Receipt"                                        
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, Sheet_name)
        print(f"'{valid_rows}': valid rows")
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[Sheet_name]
        for row_num in range(2, valid_rows):  
            data = {
                    "Test Case Id": 1,
                    "TestStatus": 2,
                    "ActualStatus": 3,
                    "Type": 4,
                    "Select Karigar": 5,
                    "Ref No": 6,
                    "Ref Date": 7,
                    "E-Way Bill No": 8,
                    "IRN No": 9,
                	"Dispatch Through": 10,
                    "Image":11,
                    "Rate Calculation From":12,
                    "Amount":13,
                    "Weight":14,
                    "Date":15,
                    "Employee":16,
                    "NetBanking":17                    
                }
            row_data = {key: sheet.cell(row=row_num, column=col).value 
                            for key, col in data.items()}
            print(row_data)
            # Call you 'create' method
            Create_data = self.create(row_data, row_num, Sheet_name,Rate)
            print(Create_data)
    def create(self,row_data, row_num, Sheet_name,Rate):
        driver = self.driver
        wait = self.wait
        driver.refresh()
        Mandatory_field=[] 
        Function_Call.click(self,'//a[@id="add_Order"]')
        Type = {
        "Bill": '//input[@id="oranment_type"]',
        "Receipt": '//input[@id="mt_type"]',
        "Charges": '//input[@id="st_type"]',
        } 
        Function_Call.click(self,Type[row_data["Receipt Type"]])
        #Employes
        if row_data["Select Karigar"] is not None:
            Function_Call.dropdown_select(self,f'//span[@id="select2-select_karigar-container"]', row_data["Select Karigar"],'//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Select Karigar field is mandatory ⚠️"
            Mandatory_field.append("Select Karigar"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
    
        if row_data["Ref No"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, '//input[@name="order[po_supplier_ref_no]"]'),
                value=row_data["Ref No"],
                pattern = r"^(\d{1,7})?$",
                field_name="Ref No",
                screenshot_prefix="Ref No",
                row_num=row_num,
                Sheet_name=Sheet_name)  
        
        if row_data["Ref Date"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, '//input[@name="order[po_ref_date]"]'),
                value=row_data["Ref Date"],
                pattern = r"^(\d{1,7})?$",
                field_name="Ref Date",
                screenshot_prefix="Ref Date",
                row_num=row_num,
                Sheet_name=Sheet_name) 
    
        if row_data["E-Way Bill No"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, '//input[@id="ewaybillno"]'),
                value=row_data["E-Way Bill No"],
                pattern = r"^(\d{1,7})?$",
                field_name="Ref Date",
                screenshot_prefix="Ref Date",
                row_num=row_num,
                Sheet_name=Sheet_name)     
            
        if row_data["IRN No"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, '//input[@id="invoice_ref_no"]'),
                value=row_data["E-Way Bill No"],
                pattern = r"^(\d{1,7})?$",
                field_name="Ref Date",
                screenshot_prefix="Ref Date",
                row_num=row_num,
                Sheet_name=Sheet_name) 
        Function_Call.select_visible_text(self,f"//select[@id='despatch_through']", row_data["Dispatch Through"])
        Function_Call.click(self,'//button[@id="add_item_details"]')
        
        if row_data["Select Karigar"] is not None:
            Function_Call.dropdown_select(self,f'//span[@id="select2-select_karigar-container"]', row_data["Select Karigar"],'//span[@class="select2-search select2-search--dropdown"]/input')
        else:
            msg = f"'{None}' → Select Karigar field is mandatory ⚠️"
            Mandatory_field.append("Select Karigar"); print(msg); Function_Call.Remark(self,row_num, msg,Sheet_name)
    
        if row_data["Pcs"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, '//input[@name="item[pcs][]"]'),
                value=row_data["Pcs"],
                pattern = r"^(\d{1,7})?$",
                field_name="Pcs",
                screenshot_prefix="Pcs",
                row_num=row_num,
                Sheet_name=Sheet_name)  
        
        if row_data["G.Wt"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, '//input[@name="item[gross_wt][]"]'),
                value=row_data["G.Wt"],
                pattern = r"^(\d{1,7})?$",
                field_name="G.Wt",
                screenshot_prefix="G.Wt",
                row_num=row_num,
                Sheet_name=Sheet_name)  
            
        
        if row_data["Rate"]:
            errors=Function_Call.fill_input(
                self,wait,
                locator=(By.XPATH, '//input[@name="item[rate_per_gram][]"]'),
                value=row_data["Rate"],
                pattern = r"^(\d{1,7})?$",
                field_name="Rate",
                screenshot_prefix="Rate",
                row_num=row_num,
                Sheet_name=Sheet_name)  
        
        
        
        Function_Call.click(self,'//button[@id="submit_grn_entry"]')
        
        
        
        
        
        
        
        
        
        
        
        