from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
import unittest
from Utils.Excel import ExcelUtils
from openpyxl.drawing.image import Image
from openpyxl import load_workbook

FILE_PATH = ExcelUtils.file_path
class Product(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)
    
    def test_product(self):
        driver = self.driver
        wait = self.wait
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Toggle navigation"))).click()
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Retail Catalog"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH,"(.//*[normalize-space(text()) and normalize-space(.)='Category'])[1]/following::span[1]"))).click()
        function_name = "Product"
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, function_name)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[function_name]
        for row_num in range(2, valid_rows):  
            data = {
                    "TestCaseId": 1,
                    "TestStatus": 2,
                    "ActualStatus": 3,
                    "metal": 4,
                    "category": 5,
                    "taxtype": 6,
                    "Sales Tax": 7,
                    "purchasetax": 8,
                    "section": 9,
                    "productname": 10,
                    "hsn": 11,
                    "shortCode": 12,
                    "noofpce": 13,
                    "minwastage": 14,
                    "maxwastage": 15,
                    "stocktype": 16,
                    "producttype": 17,
                    "uom": 18,
                    "minWastage": 19,
                    "maxWastage": 20,
                    "salesbased": 21,
                    "reorder": 22,
                    "purchasebased": 23,
                    "charges": 24,
                    "displaypurity": 25,
                    "lessstone": 26,
                    "size": 27,
                    "calculation": 28,
                    "chargeneed": 29,
                    "editProduct": 30,
                    "Edit": 31,
                    "Delete": 32
                }

            row_data = {key: sheet.cell(row=row_num, column=col).value 
                            for key, col in data.items()}
            print(row_data)
            wait.until(EC.element_to_be_clickable((By.ID,"add_product"))).click()
            status_list = []
            Actual_list = []
            Create_data=self.create(row_data,row_num)
            Test_Status,Actual_Status = Create_data
            status_list.append(Test_Status)
            Actual_list.append(Actual_Status)
            Edit_data=self.Edit(row_data,row_num)
            Test_Status,Actual_Status = Edit_data
            status_list.append(Test_Status)
            Actual_list.append(Actual_Status)
            Delete_data=self.Delete(row_data,row_num)
            Test_Status,Actual_Status = Delete_data
            status_list.append(Test_Status)
            Actual_list.append(Actual_Status)
            all=True
            for s in status_list:
                    # If any step is not "Pass", mark overall as failed and stop checking
                if s!="Pass":
                    all=False
                    break
            if all:
                Test_Status = "Pass"
            else:
                Test_Status = "Fail"  
                # --- Final Pass/Fail ---
            print(f"Final Status:", Test_Status)
            sheet.cell(row=row_num, column=2).value = Test_Status
            sheet.cell(row=row_num, column=3).value = ", ".join(Actual_list)
            workbook.save(FILE_PATH)
            Status = ExcelUtils.get_Status(FILE_PATH,function_name)  
            print(Status)
            Update_master = ExcelUtils.update_master_status(FILE_PATH,Status,function_name)   
    def create(self,row_data,row_num):
        driver = self.driver
        wait = self.wait
        wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"select2-metal_sel-container\"]"))).click()
        Metal=wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input")))
        Metal.click()
        Metal.send_keys(row_data["metal"],Keys.ENTER)
        sleep(3)
        wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"select2-category_sel-container\"]"))).click()
        category=wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input")))
        category.clear()
        category.send_keys(row_data["category"],Keys.ENTER)
        if row_data["taxtype"] == "inclusive" :
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"tax_type1\"]"))).click()
        elif row_data["taxtype"] == "exclusive" :
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"tax_type2\"]"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"select2-tgrp_sel-container\"]/span"))).click()
        Sale_tax=wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input")))
        Sale_tax.clear()
        Sale_tax.send_keys(row_data['Sales Tax'],(Keys.ENTER))
        wait.until(EC.element_to_be_clickable((By.ID,"select2-pur_tgrp_sel-container"))).click()
        Purchase_tax=wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input")))
        Purchase_tax.click()                                                   
        Purchase_tax.clear()
        Purchase_tax.send_keys(row_data["purchasetax"],Keys.ENTER)
        # Example: Excel "section" column contains values like: "GOLD CHAIN,GOLD PENDANT,GOLD BANGLES"
        sections = row_data["section"]
        if sections:  
            # Split into list if comma separated
            section_list = [s.strip() for s in sections.split(",")]

            dropdown = wait.until(EC.element_to_be_clickable((By.XPATH, '//input[@class="select2-search__field"]')))
            dropdown.click()

            for section in section_list:
                # Locate input box inside dropdown
                input_box = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, '//input[@class="select2-search__field"]')
                ))
                input_box.clear()
                input_box.send_keys(section)
                input_box.send_keys(Keys.ENTER)
            print("✅ Sections added:", section_list)
        product_name=wait.until(EC.element_to_be_clickable((By.ID,"product_name")))
        product_name.click()
        product_name.send_keys(row_data["productname"])
        hsn_Code=wait.until(EC.element_to_be_clickable((By.ID,"hsn_code")))
        hsn_Code.click()
        hsn_Code.send_keys(row_data["hsn"])
        product_short_code=wait.until(EC.element_to_be_clickable((By.ID,"product_short_code")))
        product_short_code.click()
        product_short_code.send_keys(row_data["shortCode"])
        nop = driver.execute_script("return(Math.floor(Math.random()*10))")
        No_of_pcs=wait.until(EC.element_to_be_clickable((By.ID,"no_of_pieces")))
        No_of_pcs.click()
        No_of_pcs.clear()
        No_of_pcs.send_keys(nop)
        
        if row_data["stocktype"]=="tagged" : 
            wait.until(EC.element_to_be_clickable((By.ID,"stock_type1"))).click()
        elif row_data["stocktype"]=="nontagged" :
            wait.until(EC.element_to_be_clickable((By.ID,"stock_type1"))).click()
        Producttype = row_data["producttype"]
        if Producttype=="Oranments" :
            wait.until(EC.element_to_be_clickable((By.ID,"stone_type0"))).click()
        elif Producttype in ["Stone", "Diamond"]:
            if Producttype == "Stone":
                stone_id = "stone_type1" 
            else:
                stone_id = "stone_type2" 
            wait.until(EC.element_to_be_clickable((By.ID, stone_id))).click()
            wait.until(EC.element_to_be_clickable((By.ID, "select2-size_sel-container"))).click()
            uom = wait.until(EC.element_to_be_clickable((By.XPATH,"//span[@class='select2-search select2-search--dropdown']/input")))
            uom.clear()
            uom.send_keys(row_data["uom"], Keys.ENTER)
        min = driver.execute_script("return(Math.floor(Math.random()*2))")
        minimum=wait.until(EC.element_to_be_clickable((By.NAME,"product[min_wastage]")))
        minimum.click()
        minimum.send_keys(min)    
        max = driver.execute_script("return(Math.floor(Math.random()*5))")
        maximum=wait.until(EC.element_to_be_clickable((By.NAME,"product[max_wastage]")))
        maximum.click()
        maximum.send_keys(max)        
        if row_data["salesbased"]=="flexible rate" :
            wait.until(EC.element_to_be_clickable((By.ID,"sales_mode1"))).click()
        else:
            wait.until(EC.element_to_be_clickable((By.ID,"sales_mode2"))).click()
        if row_data["reorder"] ==  "Weight Range":
            wait.until(EC.element_to_be_clickable((By.ID,"reorder_based_on1"))).click()
        elif row_data["reorder"] == "size" :
            wait.until(EC.element_to_be_clickable((By.ID,"reorder_based_on2"))).click()
        if row_data["purchasebased"] == "Flexible Rate" :
            wait.until(EC.element_to_be_clickable((By.ID,"purchase_mode1"))).click()
        elif row_data["purchasebased"]=="fixed rate" :
            wait.until(EC.element_to_be_clickable((By.ID,"purchase_mode2"))).click()
        if row_data["charges"]=="yes" :
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"tab_1\"]/div[6]/legend/span/i"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"tab_1\"]/div[6]/div/div/div[1]/div/span/span[1]/span"))).click()
            charges=wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input")))
            charges.click()
            charges.send_keys(row_data["Select Charge"],Keys.ENTER)
        else: 
            print("\"charge not given\"")
        if row_data["displaypurity"]=="Yes" :
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"display_purity_yes\"]"))).click()
        elif row_data["displaypurity"]=="No" :
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"display_purity_no\"]"))).click()
        if row_data["lessstone"]=="Yes" :
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"less_stone_wt1\"]"))).click()
        elif row_data["lessstone"]=="No" :
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"less_stone_wt2\"]"))).click()
        if row_data["size"]=="Yes" :
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"has_size1\"]"))).click()
        elif row_data["size"]=="No" :
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"has_size2\"]"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"tab_1\"]/div[11]/div[3]/div/div/span/span[1]/span/span[2]"))).click()
        calculation=wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input")))
        calculation.click()
        calculation.send_keys(row_data["calculation"],Keys.ENTER)
        wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/div/div[1]/section[2]/form/div/div[4]/div/div/div/button[1]"))).click()
        driver.save_screenshot('added.png.png')
        # screenshot_path = "D:/Retail_Testing/added.png.png"
        # driver.save_screenshot(screenshot_path)
        # function_name = "Product"
        # workbook = load_workbook(FILE_PATH)
        # sheet = workbook[function_name]
        # img = Image(screenshot_path)
        # cell_address = f"AG{row_num}"   # column 33 = "AG"
        # sheet.add_image(img, cell_address)
        # workbook.save(FILE_PATH)
        try:
            message = wait.until(EC.element_to_be_clickable((By.XPATH,'//div[@class="alert alert-success alert-dismissable"]'))).text
            message = message.replace("×", "").replace("\n", " ").strip()
            if message =="Add Product! New Product added successfully":
                Test_Status="Pass"
                Actual_Status= message
            else:
                Test_Status="Fail"
                Actual_Status= message
        except:
            Test_Status="Fail"
            Actual_Status="Product Not Add Successfully"
        return Test_Status,Actual_Status

    def Edit(self,row_data,row_num):
        driver = self.driver
        wait = self.wait
        if row_data['Edit']=='Yes':
            Sarch_Product=wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"product_list_filter\"]/label/input")))
            Sarch_Product.click()
            Sarch_Product.clear()
            Sarch_Product.send_keys(row_data["productname"])
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"edit\"]"))).click()
            Edit_Product=wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"product_name\"]")))
            Edit_Product.click()
            Edit_Product.clear()
            Edit_Product.send_keys(row_data['editProduct'])
            wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/div/div[1]/section[2]/form/div/div[4]/div/div/div/button[1]"))).click()
            driver.save_screenshot('edit.png.png')
            try:
                editsucess = wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[1]/div[1]/section[2]/div/div/div/div[2]/div[2]/div/div"))).text
                editsucess = editsucess.replace("×", "").replace("\n", " ").strip()
                print(editsucess)
                if editsucess == editsucess:
                    Test_Status="Pass"
                    Actual_Status= editsucess 
                else:
                    Test_Status="Fail"
                    Actual_Status= editsucess
            except:
                Test_Status="Fail"
                Actual_Status= "Edit not updated successfully"                   
        else:
            Test_Status="Pass"
            Actual_Status= "editNotNeeded"
        return Test_Status,Actual_Status
        
    def Delete(self,row_data,row_num):
        driver = self.driver
        wait = self.wait
        if row_data['Delete']=='Yes':
            if row_data['editProduct']=='Yes':
                delete=row_data['editProduct']
            else:
                 delete=row_data['productname']   
            sleep(3)     
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"product_list_filter\"]/label/input"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"product_list_filter\"]/label/input"))).clear()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"product_list_filter\"]/label/input"))).send_keys(delete)
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"product_list\"]/tbody/tr/td[12]/a[2]"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"confirm-delete\"]/div/div/div[3]/a"))).click()
            driver.save_screenshot('delete.png.png')
            try:
                deleted = wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[1]/div[1]/section[2]/div/div/div/div[2]/div[2]/div/div"))).text
                deleted = deleted.replace("×", "").replace("\n", " ").strip()
                print(deleted)
                if deleted=="Delete product! product deleted successfully":
                    Test_Status="Pass"
                    Actual_Status= deleted
                else:
                    Test_Status="Fail"
                    Actual_Status= deleted
            except:
                Test_Status="Pass"
                Actual_Status= "Product Not deleted Successfully"               
        else: 
            Test_Status="Pass"
            Actual_Status= "deleteNotNeeded"
        return Test_Status,Actual_Status

    def is_element_present(self, how, what):
        try: self.driver.find_element(by=how, value=what)
        except NoSuchElementException as e: return False
        return True
    
    def is_alert_present(self):
        try: self.driver.switch_to_alert()
        except NoAlertPresentException as e: return False
        return True
    
    def close_alert_and_get_its_text(self):
        try:
            alert = self.driver.switch_to_alert()
            alert_text = alert.text
            if self.accept_next_alert:
                alert.accept()
            else:
                alert.dismiss()
            return alert_text
        finally: self.accept_next_alert = True
    
    def tearDown(self):
        self.driver.quit()
        self.assertEqual([], self.verificationErrors)

if __name__ == "__main__":
    unittest.main()
