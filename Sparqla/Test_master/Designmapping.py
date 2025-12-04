from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
import unittest
from Utils.Excel import ExcelUtils
from openpyxl import load_workbook

FILE_PATH = ExcelUtils.file_path
class Designmapping(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)
    
    def test_designmapping(self):
        driver = self.driver
        wait = self.wait
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Toggle navigation"))).click()
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Retail Catalog"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH,"(.//*[normalize-space(text()) and normalize-space(.)='Sub Design'])[1]/following::span[1]"))).click()

        function_name = "Designmapping"
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, function_name)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[function_name]
        Window=1
        for row_num in range(2, valid_rows):
            data = {
                "TestCaseId": 1,
                "TestStatus": 2,
                "ActualStatus": 3,
                "selectProduct":4,	
                "selectDesign":5,	
                "deleteDesign":6,	
                
            }
            row_data = {key: sheet.cell(row=row_num, column=col).value 
                            for key, col in data.items()}
            print(row_data)
            status_list = []
            Actual_list = []
            Create_data=self.create(row_data)
            Test_Status,Actual_Status = Create_data
            status_list.append(Test_Status)
            Actual_list.append(Actual_Status)
            Delete_data=self.Delete(row_data)
            Test_Status,Actual_Status = Delete_data
            status_list.append(Test_Status)
            Actual_list.append(Actual_Status)
            all=True
            for s in status_list:
                    # If any step is not "Pass", mark overall as failed and stop checking
                if s!="Pass":
                    all=False
                    break
            if all:
                Test_Status = "Pass"
            else:
                Test_Status = "Fail"  
                # --- Final Pass/Fail ---
            print(f"Final Status:", Test_Status)
            sheet.cell(row=row_num, column=2).value = Test_Status
            sheet.cell(row=row_num, column=3).value = ", ".join(Actual_list)
            workbook.save(FILE_PATH)
            Status = ExcelUtils.get_Status(FILE_PATH,function_name)  
            print(Status)
            Update_master = ExcelUtils.update_master_status(FILE_PATH,Status,function_name)  
            
    def create(self,row_data):   
        driver = self.driver
        wait = self.wait    
        driver.refresh()  
        wait.until(EC.element_to_be_clickable((By.ID,"select2-select_product-container"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH,'(//input[@class="select2-search__field"])[2]'))).clear()
        wait.until(EC.element_to_be_clickable((By.XPATH,'(//input[@class="select2-search__field"])[2]'))).send_keys(row_data["selectProduct"])
        wait.until(EC.element_to_be_clickable((By.XPATH,'(//input[@class="select2-search__field"])[2]'))).send_keys(Keys.ENTER)
        # Example: Excel "selectDesign" column contains values like: "GOLD CHAIN,GOLD PENDANT,GOLD BANGLES"
        Design = row_data["selectDesign"]
        if Design:  
            # Split into list if comma separated
            Design_list = [s.strip() for s in Design.split(",")]

            dropdown = wait.until(EC.element_to_be_clickable((By.XPATH, '(//input[@class="select2-search__field"])[1]')))
            dropdown.click()

            for Design in Design_list:
                # Locate input box inside dropdown
                input_box = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, '(//input[@class="select2-search__field"])[1]')
                ))
                input_box.clear()
                input_box.send_keys(Design)
                input_box.send_keys(Keys.ENTER)
            print("✅ Design added:", Design_list)

        wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@type='search']"))).click()
        wait.until(EC.element_to_be_clickable((By.ID,"update_product_design_mapping"))).click() 
        self.driver.refresh()  
        wait.until(EC.element_to_be_clickable((By.ID,"select2-prod_filter-container"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input"))).clear()
        wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input"))).send_keys(row_data["selectProduct"])
        wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input"))).send_keys(Keys.ENTER)
        sleep(2)
        # Example: Excel "selectDesign" column contains values like: "GOLD CHAIN,GOLD PENDANT,GOLD BANGLES"
        SelectDesign = row_data["selectDesign"]
        if SelectDesign:  
            # Split into list if comma separated
            SelectDesign_list = [s.strip() for s in SelectDesign.split(",")]
            print(SelectDesign_list)
            DisplayDesign=[]
            for SelectDesign in SelectDesign_list:
                # Locate input box inside dropdown
                dropdown = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@id='select2-select_design_fitler-container']/span")))
                dropdown.click()
                input_box = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "/html/body/span/span/span[1]/input")
                ))
                input_box.clear()
                input_box.send_keys(SelectDesign)
                input_box.send_keys(Keys.ENTER)
                sleep(2)
                wait.until(EC.element_to_be_clickable((By.ID,"search_design_maping"))).click()
                sleep(1)
                product = wait.until(EC.element_to_be_clickable((By.XPATH,f"//*[@id=\"subdesign_list\"]/tbody/tr[1]/td[2]"))).text
                print(product)
                design = wait.until(EC.element_to_be_clickable((By.XPATH,f"//*[@id=\"subdesign_list\"]/tbody/tr[1]/td[3]"))).text
                print(design)
                DisplayDesign.append(design)
        if row_data["selectProduct"] ==product and SelectDesign_list==DisplayDesign :
            
            Test_Status="Pass"
            Actual_Status= "design mapping success"
        else:
            Test_Status="Fail"
            Actual_Status= "design mapping not success"
        return Test_Status,Actual_Status 
       

    def Delete(self,row_data):   
        driver = self.driver
        wait = self.wait    
        if row_data['deleteDesign']=='Yes':
            Id=wait.until(EC.presence_of_element_located((By.XPATH,"(//table[@id='subdesign_list']//td)[1]"))).text
            print(Id)
            sleep(2)
            wait.until(EC.element_to_be_clickable((By.XPATH,'//a[@class="btn btn-danger btn-del"]'))).click()
            wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,'Delete'))).click()
            driver.refresh()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@id='subdesign_list_filter']/label/input"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@id='subdesign_list_filter']/label/input"))).clear()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@id='subdesign_list_filter']/label/input"))).send_keys(Id)
            
            message=wait.until(EC.presence_of_element_located((By.XPATH,"(//table[@id='subdesign_list']//td)[1]"))).text
            print(message)
            if message=="No matching records found" :
                Test_Status="Pass"
                Actual_Status= "mapped design deleted successfully"
            else:
                Test_Status="Fail"
                Actual_Status= "mapped deleted not successfully"
            return Test_Status,Actual_Status 
        else:
            Test_Status="Pass"
            Actual_Status="Delete DesignMapping No Needed"
            return Test_Status,Actual_Status         
  
    def is_element_present(self, how, what):
        try: self.driver.find_element(by=how, value=what)
        except NoSuchElementException as e: return False
        return True
    
    def is_alert_present(self):
        try: self.driver.switch_to_alert()
        except NoAlertPresentException as e: return False
        return True
    
    def close_alert_and_get_its_text(self):
        try:
            alert = self.driver.switch_to_alert()
            alert_text = alert.text
            if self.accept_next_alert:
                alert.accept()
            else:
                alert.dismiss()
            return alert_text
        finally: self.accept_next_alert = True
    
    def tearDown(self):
        self.driver.quit()
        self.assertEqual([], self.verificationErrors)

if __name__ == "__main__":
    unittest.main()
