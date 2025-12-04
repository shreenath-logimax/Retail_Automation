from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
import unittest
from Utils.Excel import ExcelUtils
from openpyxl import load_workbook

FILE_PATH = ExcelUtils.file_path
class Metal(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)
    def test_metal(self):
        driver = self.driver
        wait = self.wait
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Toggle navigation"))).click()
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Retail Catalog"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH,"//span[normalize-space(text())='Metal']"))).click()
        function_name = "Metal"
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, function_name)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[function_name]
        for row_num in range(2, valid_rows):
            data={
                "Metal"	:4,
                "shortCode"	:5,
                "selectTaxGroup":6,	
                "editMetal"	:7,
                "updateShortCode":8,	
                "metalDelete":9
            }
            row_data = {key: sheet.cell(row=row_num, column=col).value 
                            for key, col in data.items()}
            print(row_data)
            status_list = []
            Actual_list = []
            Create_data=self.create(row_data)
            Test_Status,Actual_Status = Create_data
            status_list.append(Test_Status)
            Actual_list.append(Actual_Status)
            Edit_data=self.Edit(row_data)
            Test_Status,Actual_Status = Edit_data
            status_list.append(Test_Status)
            Actual_list.append(Actual_Status)
            Delete_data=self.Delete(row_data)
            Test_Status,Actual_Status = Delete_data
            status_list.append(Test_Status)
            Actual_list.append(Actual_Status)
            all=True
            for s in status_list:
                    # If any step is not "Pass", mark overall as failed and stop checking
                if s!="Pass":
                    all=False
                    break
            if all:
                Test_Status = "Pass"
            else:
                Test_Status = "Fail"  
                # --- Final Pass/Fail ---
            print(f"Final Status:", Test_Status)
            sheet.cell(row=row_num, column=2).value = Test_Status
            sheet.cell(row=row_num, column=3).value = ", ".join(Actual_list)
            workbook.save(FILE_PATH)
            Status = ExcelUtils.get_Status(FILE_PATH,function_name)  
            print(Status)
            Update_master = ExcelUtils.update_master_status(FILE_PATH,Status,function_name)   
    
    def create(self,row_data):
        wait = self.wait
        wait.until(EC.element_to_be_clickable((By.ID,"add_metal"))).click()
        Metal=wait.until(EC.element_to_be_clickable((By.ID,"metal_name")))
        Metal.click()
        Metal.send_keys(row_data["Metal"])
        shortcode=wait.until(EC.element_to_be_clickable((By.ID,"metal_code")))
        shortcode.click()
        shortcode.send_keys(row_data["shortCode"])
        wait.until(EC.element_to_be_clickable((By.ID,"select2-tgrp_sel-container"))).click()
        # Wait until the tax option appears and click GST%
        tax = row_data["selectTaxGroup"]
        # Tax=(f"//li[normalize-space()='{tax}']")
        tax_option = wait.until(EC.element_to_be_clickable((By.XPATH, f"//li[normalize-space()='{tax}']")))
        tax_option.click()
        wait.until(EC.element_to_be_clickable((By.ID,'add_newmetal'))).click()
        sleep(2)
        metal =  wait.until(EC.presence_of_element_located((By.XPATH,"//*[@id=\"chit_alert\"]"))).text
        metal = metal.replace("×", "").replace("\n", " ").strip()
        print(metal)
        if metal == "Add Metal : New Metal added successfully!.." :
            Test_Status="Pass"
            Actual_Status="metal added successfully"
            print(Test_Status,Actual_Status)
        else:
            Test_Status="Fail"
            Actual_Status="metal added not successfully" 
            ##status_list.append(Test_Status)
            # #Actual_list.append(Actual_Status)
            print(Test_Status,Actual_Status)
        return Test_Status,Actual_Status
            
    def Edit(self,row_data):  
        wait = self.wait
        if row_data['editMetal'] == 'Yes':
            self.driver.refresh()
            sleep(4)
            wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@type='search']"))).send_keys(row_data["Metal"])
            wait.until(EC.element_to_be_clickable((By.XPATH,"//a[@id='edit']/i"))).click()
            val_name=wait.until(EC.visibility_of_element_located((By.ID,"ed_metal_name")))
            Metal_name = val_name.get_attribute("value")
            # print(Metal_name)
            if Metal_name=="":
                wait.until(EC.visibility_of_element_located((By.ID,"ed_metal_name"))).send_keys(row_data["Metal"])
            else:
                pass    
          
            updateShortCode=wait.until(EC.visibility_of_element_located((By.ID,"ed_metal_code")))
            updateShortCode.clear()
            updateShortCode.send_keys(row_data["updateShortCode"])
            wait.until(EC.element_to_be_clickable((By.ID,"update_metal"))).click()
            try:
                updated = wait.until(EC.presence_of_element_located((By.XPATH,'//div[@class="alert alert-success alert-dismissable"]'))).text
                updated = updated.replace("×", "").replace("\n", " ").strip()
                print(updated)
                if updated in  "Edit Metal! Metal record modified successfully" :
                    Test_Status="Pass"
                    Actual_Status="metal updated success"
                   #status_list.append(Test_Status)
                    #Actual_list.append(Actual_Status)
                    print(Test_Status,Actual_Status)   
                else: 
                    Test_Status="Fail"
                    Actual_Status="metal updated not success" 
                   #status_list.append(Test_Status)
                    #Actual_list.append(Actual_Status)
                    print(Test_Status,Actual_Status)     
            except:
                wait.until(EC.element_to_be_clickable((By.XPATH,'(//button[@type="button"])[8]'))).click()
                Test_Status="Fail"
                Actual_Status="metal updated not success"
               #status_list.append(Test_Status)
                #Actual_list.append(Actual_Status)
                print(Test_Status,Actual_Status)
            return Test_Status,Actual_Status    
        else:
            Test_Status="Pass"
            Actual_Status="Edit Metal No Needed"
            return Test_Status,Actual_Status  
    def Delete(self,row_data):
        driver = self.driver
        wait = self.wait
        if row_data["metalDelete"]=="Yes" :
            sleep(3)
            search=driver.find_element(By.XPATH,"//input[@type='search']")
            search.click()
            search.send_keys(row_data["Metal"])
            wait.until(EC.element_to_be_clickable((By.XPATH,"//table[@id='metal_list']/tbody/tr/td[6]/a[2]/i"))).click()
            wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Delete"))).click()
            metaldeleted = wait.until(EC.presence_of_element_located((By.XPATH,"/html/body/div[1]/div[1]/section[2]/div/div/div/div[2]/div[1]"))).text
            metaldeleted=metaldeleted.replace("×", "").replace("\n", " ").strip()
            print(metaldeleted)
            if metaldeleted == "Delete Metal! Metal deleted successfully" :
                Test_Status="Pass"
                Actual_Status="Metal deleted successfully"
               #status_list.append(Test_Status)
                #Actual_list.append(Actual_Status)
                print(Test_Status,Actual_Status)
            elif metaldeleted == "Delete Metal! Metal Exists in Stock"  :
                Test_Status="Fail"
                Actual_Status="Metal exists in stock, cannot delete"
               #status_list.append(Test_Status)
                #Actual_list.append(Actual_Status)
                print(Test_Status,Actual_Status)
            return Test_Status,Actual_Status  
        else:
            Test_Status="Pass"
            Actual_Status="Delete Metal No Needed"
            return Test_Status,Actual_Status 
    def is_element_present(self, how, what):
        try: self.driver.find_element(by=how, value=what)
        except NoSuchElementException as e: return False
        return True
    
    def is_alert_present(self):
        try: self.driver.switch_to_alert()
        except NoAlertPresentException as e: return False
        return True
    
    def close_alert_and_get_its_text(self):
        try:
            alert = self.driver.switch_to_alert()
            alert_text = alert.text
            if self.accept_next_alert:
                alert.accept()
            else:
                alert.dismiss()
            return alert_text
        finally: self.accept_next_alert = True
    
    def tearDown(self):
        self.driver.quit()
        self.assertEqual([], self.verificationErrors)

if __name__ == "__main__":
    unittest.main()
