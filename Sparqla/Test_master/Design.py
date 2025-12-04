from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
import unittest
from Utils.Excel import ExcelUtils
from openpyxl import load_workbook

FILE_PATH = ExcelUtils.file_path
class Design(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)
    
    def test_design(self):
        driver = self.driver
        wait = self.wait
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Toggle navigation"))).click()
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Retail Catalog"))).click()
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Design"))).click()
        function_name = "Design"
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, function_name)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[function_name]
        Window=1
        for row_num in range(2, valid_rows):
            data = {
                "TestCaseId": 1,
                "TestStatus": 2,
                "ActualStatus": 3,
                "designName":4,	
                "editDesign":5,	
                "designEdit":6,	
                "designDelete":7
            }
            row_data = {key: sheet.cell(row=row_num, column=col).value 
                            for key, col in data.items()}
            print(row_data)
            
            status_list = []
            Actual_list = []
            Create_data=self.create(row_data)
            Test_Status,Actual_Status = Create_data
            status_list.append(Test_Status)
            Actual_list.append(Actual_Status)
            Edit_data=self.Edit(row_data)
            Test_Status,Actual_Status = Edit_data
            status_list.append(Test_Status)
            Actual_list.append(Actual_Status)
            Delete_data=self.Delete(row_data)
            Test_Status,Actual_Status = Delete_data
            status_list.append(Test_Status)
            Actual_list.append(Actual_Status)
            all=True
            for s in status_list:
                    # If any step is not "Pass", mark overall as failed and stop checking
                if s!="Pass":
                    all=False
                    break
            if all:
                Test_Status = "Pass"
            else:
                Test_Status = "Fail"  
                # --- Final Pass/Fail ---
            print(f"Final Status:", Test_Status)
            sheet.cell(row=row_num, column=2).value = Test_Status
            sheet.cell(row=row_num, column=3).value = ", ".join(Actual_list)
            workbook.save(FILE_PATH)
            Status = ExcelUtils.get_Status(FILE_PATH,function_name)  
            print(Status)
            Update_master = ExcelUtils.update_master_status(FILE_PATH,Status,function_name)  
            
    def create(self,row_data):   
        driver = self.driver
        wait = self.wait    
        driver.refresh()
        wait.until(EC.element_to_be_clickable((By.ID,"add_product"))).click()
        wait.until(EC.element_to_be_clickable((By.ID,"design_name"))).clear()
        wait.until(EC.element_to_be_clickable((By.ID,"design_name"))).send_keys(row_data["designName"])
        wait.until(EC.element_to_be_clickable((By.XPATH,"//button[@type='submit']"))).click()
        try:
            design = wait.until(EC.element_to_be_clickable((By.XPATH,"(.//*[normalize-space(text()) and normalize-space(.)='Add'])[1]/following::div[4]"))).text
            design = design.replace("×", "").replace("\n", " ").strip()
            print(design)
            expected_message= "Add Design! New Design added successfully"
            if design == expected_message:
                Test_Status="Pass"
                Actual_Status= design
            else:
                Test_Status="Fail"
                Actual_Status= design
        except:
            Test_Status="Fail"
            Actual_Status="Design Not Add Successfully"
        print(Actual_Status)
        return Test_Status,Actual_Status

    def Edit(self,row_data):   
        driver = self.driver
        wait = self.wait    
        self.driver.refresh() 
        if row_data["designEdit"]== "Yes" :
            wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@type='search']"))).clear()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@type='search']"))).send_keys(row_data["designName"])
            wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@type='search']"))).click()
            wait.until(EC.element_to_be_clickable((By.ID,"edit"))).click()
            wait.until(EC.element_to_be_clickable((By.ID,"design_name"))).click()
            wait.until(EC.element_to_be_clickable((By.ID,"design_name"))).clear()
            wait.until(EC.element_to_be_clickable((By.ID,"design_name"))).send_keys(row_data["editDesign"])
            wait.until(EC.element_to_be_clickable((By.XPATH,"//button[@type='submit']"))).click()
            try:
                modified = wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/div/div[1]/section[2]/div/div/div/div[2]/div[1]/div/div"))).text
                modified = modified.replace("×", "").replace("\n", " ").strip()
                print(modified)
                expected_message ="Edit Design! Design record modified successfully" 
                if modified == expected_message:
                    Test_Status="Pass"
                    Actual_Status= modified 
                else:
                    Test_Status="Fail"
                    Actual_Status= modified
            except:
                Test_Status="Fail"
                Actual_Status= "Design Edit not updated successfully"        
        else:
            Test_Status="Pass"
            Actual_Status= "editNotNeeded"
        return Test_Status,Actual_Status 
           
                
    def Delete(self,row_data):   
        driver = self.driver
        wait = self.wait   
        sleep(2) 
        self.driver.refresh()            
        if row_data["designDelete"]=="Yes" :
            if row_data['editDesign']=="Yes":
                Delete_Design=row_data['editDesign']
            else:
                Delete_Design=row_data["designName"]
            print(Delete_Design)
            wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@type='search']"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@type='search']"))).clear()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@type='search']"))).send_keys(Delete_Design)
            wait.until(EC.element_to_be_clickable((By.XPATH,"//table[@id='design_list']/tbody/tr/td[5]/a[2]/i"))).click()
            wait.until(EC.element_to_be_clickable((By.ID,"remove_design"))).click()
            driver.save_screenshot('deleted.png.png')
            try:
                deleteDesign = wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/div/div[1]/section[2]/div/div/div/div[2]/div[1]/div/div"))).text
                print(deleteDesign)
                expected_message = "delete design record successfully" 
                if deleteDesign == expected_message:
                    Test_Status="Pass"
                    Actual_Status= deleteDesign
                if deleteDesign == deleteDesign:
                    Test_Status="Pass"
                    Actual_Status= deleteDesign   
                else:
                    Test_Status="Fail"
                    Actual_Status= deleteDesign
            except:
                driver.get("https://retail.logimaxindia.com/qa/admin/index.php/admin_ret_catalog/ret_design/list")
                driver.refresh()
                Test_Status="Fail"
                Actual_Status= "Design Not deleted Successfully"               
        else: 
            Test_Status="Pass"
            Actual_Status= "deleteNotNeeded"
        return Test_Status,Actual_Status
    
    def is_element_present(self, how, what):
        try: self.driver.find_element(by=how, value=what)
        except NoSuchElementException as e: return False
        return True
    
    def is_alert_present(self):
        try: self.driver.switch_to_alert()
        except NoAlertPresentException as e: return False
        return True
    
    def close_alert_and_get_its_text(self):
        try:
            alert = self.driver.switch_to_alert()
            alert_text = alert.text
            if self.accept_next_alert:
                alert.accept()
            else:
                alert.dismiss()
            return alert_text
        finally: self.accept_next_alert = True
    
    def tearDown(self):
        self.driver.quit()
        self.assertEqual([], self.verificationErrors)

if __name__ == "__main__":
    unittest.main()
