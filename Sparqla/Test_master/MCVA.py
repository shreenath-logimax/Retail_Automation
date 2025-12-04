from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from time import sleep
import unittest
import re
from Utils.Excel import ExcelUtils
from Utils.Function import Function_Call
from openpyxl import load_workbook

FILE_PATH = ExcelUtils.file_path
class McVa(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)
    def test_mc_va(self):
        driver = self.driver
        wait = self.wait
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Toggle navigation"))).click()
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Retail Catalog"))).click()
        Function_Call.click2(self, "//span[normalize-space(text())='MC & VA Settings']")
        # wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"MC & VA Settings"))).click()
        
        function_name = "MC&VA"
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, function_name)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[function_name]
        Window=1
        for row_num in range(2, valid_rows):
            data = {
                    "TestCaseId": 1,
                    "TestStatus": 2,
                    "ActualStatus": 3,
                    "category": 4,
                    "productname": 5,
                    "design": 6,
                    "subdesign": 7,
                    "type": 8,
                    "wastage type": 9,
                    "From Weight": 10,
                    "To Weight": 11,
                    "Method": 12,
                    "Min_VA%": 13,
                    "Max_VA%": 14,
                    "Min_VA(wt)": 15,
                    "Max_VA(wt)": 16,
                    "UpdateMCType": 17,
                    "Min_MC": 18,
                    "Max_MC": 19,
                    "Margin_MRP%": 20,
                    "catagoryEdit": 21,
                    "productEdit": 22,
                    "designEdit": 23,
                    "subdesignEdit": 24,
                    "EditmcVa": 25,
                    "DeletemcVa": 26
                }

            row_data = {key: sheet.cell(row=row_num, column=col).value 
                            for key, col in data.items()}
            print(row_data)
            status_list = []
            Actual_list = []
            Create_data=self.create(row_data)
            Test_Status,Actual_Status,id = Create_data
            status_list.append(Test_Status)
            Actual_list.append(Actual_Status)
            Edit_data=self.Edit(row_data,id)
            Test_Status,Actual_Status = Edit_data
            status_list.append(Test_Status)
            Actual_list.append(Actual_Status)
            Delete_data=self.Delete(row_data)
            Test_Status,Actual_Status = Delete_data
            status_list.append(Test_Status)
            Actual_list.append(Actual_Status)
            all=True
            for s in status_list:
                    # If any step is not "Pass", mark overall as failed and stop checking
                if s!="Pass":
                    all=False
                    break
            if all:
                Test_Status = "Pass"
            else:
                Test_Status = "Fail"  
                # --- Final Pass/Fail ---
            print(f"Final Status:", Test_Status)
            sheet.cell(row=row_num, column=2).value = Test_Status
            sheet.cell(row=row_num, column=3).value = ", ".join(Actual_list)
            workbook.save(FILE_PATH)
            Status = ExcelUtils.get_Status(FILE_PATH,function_name)  
            print(Status)
            Update_master = ExcelUtils.update_master_status(FILE_PATH,Status,function_name) 
    def create(self,row_data):   
        driver = self.driver
        wait = self.wait    
        driver.refresh()    
        wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id='add_product']"))).click()     
        wait.until(EC.element_to_be_clickable((By.ID,"select2-des_cat_name-container"))).click()
        category=wait.until(EC.visibility_of_element_located((By.XPATH,"/html/body/span/span/span[1]/input")))
        category.click()
        # print(row_data['category'])
        category.send_keys(row_data['category'],Keys.ENTER)
        sleep(2)
        driver.find_element(By.ID,'select2-des_prod_name-container').click()
        productname=wait.until(EC.visibility_of_element_located((By.XPATH,"/html/body/span/span/span[1]/input")))
        productname.click()
        productname.send_keys(row_data['productname'],Keys.ENTER)
        sleep(2)
        driver.find_element(By.ID,"select2-des_des_name-container").click()
        design=wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input")))
        design.click()
        design.send_keys(row_data['design'],Keys.ENTER)
        sleep(2)
        driver.find_element(By.XPATH,'//span[@id="select2-select_sub_design-container"]').click()
        subdesign=wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input")))
        subdesign.click()
        # print(row_data['subdesign'])
        subdesign.send_keys(row_data['subdesign'],Keys.ENTER)
        if row_data['type']=="fixed" :
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"wastage_type_fixed\"]"))).click()
            if row_data['wastage type']=="percentage":
                Percentage=wait.until(EC.element_to_be_clickable((By.ID,"wastag_method")))
                Percentage.click()
                Select(Percentage).select_by_visible_text("Percentage")
                Min_va=wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"min_wastag_valuess\"]")))
                Min_va.click()
                Min_va.clear()
                Min_va.send_keys(row_data['Min_VA%'])
                Max_va=wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"wastag_value\"]")))
                Max_va.click()
                Max_va.clear()
                Max_va.send_keys(row_data['Max_VA%'])
            elif row_data["wastage type"] =="weight" :
                weight=wait.until(EC.element_to_be_clickable((By.ID,"wastag_method")))
                weight.click()
                Select(weight).select_by_visible_text("Weight")
                Min_Va_wt=wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"wastag_min_wt\"]")))
                Min_Va_wt.click()
                Min_Va_wt.clear()
                Min_Va_wt.send_keys(row_data['Min_VA(wt)'])
                Max_va_wt=wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"wastag_wt\"]")))
                Max_va_wt.click()
                Max_va_wt.clear()
                Max_va_wt.send_keys(row_data['Max_VA(wt)'])

            if row_data["UpdateMCType"]=="Per Pcs":
                Per_Pcs=wait.until(EC.element_to_be_clickable((By.ID,"update_mc_type")))
                Per_Pcs.click()
                Select(Per_Pcs).select_by_visible_text("Per Pcs")
            elif row_data["UpdateMCType"]=="Per Grm":
                Per_Gram=wait.until(EC.element_to_be_clickable((By.ID,"update_mc_type")))
                Per_Gram.click()
                Select(Per_Gram).select_by_visible_text("Per Grm")

            Min_Mc=wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"min_mc_value\"]")))
            Min_Mc.click()
            Min_Mc.clear()
            Min_Mc.send_keys(row_data['Min_MC'])
            Max_Mc=wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"mc_value\"]")))
            Max_Mc.click()
            Max_Mc.clear()
            Max_Mc.send_keys(row_data['Max_MC'])
            margin=wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"margin_mrp\"]")))
            margin.click()
            margin.clear()
            margin.send_keys(row_data['Margin_MRP%'])

        if row_data["type"]=="weightRange" :
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"margin_mrp\"]"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"margin_mrp\"]"))).clear()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"margin_mrp\"]"))).send_keys(row_data['Margin_MRP%'])
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"wastage_type_flexi\"]"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"from_weight1\"]"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"from_weight1\"]"))).clear()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"from_weight1\"]"))).send_keys(row_data['From Weight'])
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"to_weight1\"]"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"to_weight1\"]"))).clear()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"to_weight1\"]"))).send_keys(row_data['To Weight'])
            if row_data["Method"]=="Pc" :
                wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"wc_method_pc1\"]"))).click()
                wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"wc_min1\"]"))).click()
                wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"wc_min1\"]"))).clear()
                wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"wc_min1\"]"))).send_keys(row_data['Min_VA%'])
                wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"wastage1\"]"))).click()
                wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"wastage1\"]"))).clear()
                wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"wastage1\"]"))).send_keys(row_data['Max_VA%'])
            elif row_data["Method"] =="Wt" :
                wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"wc_method_wt1\"]"))).click()
                wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"wc_min_wt1\"]"))).click()
                wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"wc_min_wt1\"]"))).clear()
                wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"wc_min_wt1\"]"))).send_keys(row_data['Min_VA(wt)'])
                wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"wc_wt1\"]"))).click()
                wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"wc_wt1\"]"))).clear()
                wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"wc_wt1\"]"))).send_keys(row_data['Max_VA(wt)'])
            if row_data["UpdateMCType"] =="Per Pcs":
                MC_Type=wait.until(EC.element_to_be_clickable((By.ID,"mc_type1")))
                MC_Type.click()
                Select(MC_Type).select_by_visible_text("Per Pcs")
            elif row_data["UpdateMCType"]=="Per Grm":
                MC_Type=wait.until(EC.element_to_be_clickable((By.ID,"mc_type1")))
                MC_Type.click()
                Select(MC_Type).select_by_visible_text("Per Gram")
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"mcrg_min1\"]"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"mcrg_min1\"]"))).clear()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"mcrg_min1\"]"))).send_keys(row_data['Min_MC'])
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"making_charge1\"]"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"making_charge1\"]"))).clear()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"making_charge1\"]"))).send_keys(row_data['Max_MC'])
        Message=Function_Call.alert6(self,"//*[@id='update_weight_range_settings']")  
        print(Message)  
            # wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id='update_weight_range_settings']"))).click()
            # sleep(3)
            
            
            # wait.until(EC.element_to_be_clickable((By.ID,"select2-des_prod_name-container"))).click()
            # product = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@class="select2-search__field"]')))
            # product.click()
            # product.send_keys(row_data['productname'],Keys.ENTER)
            # sleep(2)
            # wait.until(EC.element_to_be_clickable((By.ID,"select2-des_des_name-container"))).click()
            # Design = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@class="select2-search__field"]')))
            # Design.click()
            # Design.send_keys(row_data['design'],Keys.ENTER)
            # sleep(2)
            # wait.until(EC.element_to_be_clickable((By.ID,"select2-select_sub_design-container"))).click()
            # SubDesign = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@class="select2-search__field"]')))
            # SubDesign.click()
            # SubDesign.send_keys(row_data['subdesign'],Keys.ENTER)
            # wait.until(EC.element_to_be_clickable((By.ID,"search_weight_range"))).click()
            # sleep(4)
            # product = wait.until(EC.element_to_be_clickable((By.XPATH,'//table[@id="item_list"]//td[2]'))).text
            # print(product)
            # design = wait.until(EC.element_to_be_clickable((By.XPATH,'//table[@id="item_list"]//td[3]'))).text
            # print(design)
            # sleep(2)
            # subdesign = wait.until(EC.element_to_be_clickable((By.XPATH,'//table[@id="item_list"]//td[4]'))).text
            # print(subdesign)
            # id = wait.until(EC.element_to_be_clickable((By.XPATH,'//table[@id="item_list"]//td[1]'))).text
            # print(id)
            # print(row_data["productname"])
            # print(row_data["design"])
            # print(row_data["subdesign"])
            # conditions = [
            #     re.match(row_data["productname"], product, re.IGNORECASE),
            #     re.match(row_data["design"], design, re.IGNORECASE),
            #     re.match(row_data["subdesign"], subdesign, re.IGNORECASE)
            # ]
        if Message==None:
            Test_Status = "Pass"
            Actual_Status = "V.A and MC Settings Added successfully"
        else:
            Test_Status="Fail"
            Actual_Status= "V.A and MC Settings Not Added successfully"
        return Test_Status,Actual_Status,id
            
    def Edit(self,row_data,id):   
        driver = self.driver
        wait = self.wait   
        driver.refresh()         
        if row_data["EditmcVa"]=="Yes" :
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"item_list_filter\"]/label/input"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"item_list_filter\"]/label/input"))).clear()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"item_list_filter\"]/label/input"))).send_keys(id)
            if row_data["type"]=="weightRange":
                wait.until(EC.element_to_be_clickable((By.XPATH,'(//a[@class="btn btn-primary btn-edit"])[2]'))).click()
            else:  
                wait.until(EC.element_to_be_clickable((By.XPATH,'(//a[@class="btn btn-primary btn-edit"])'))).click()
            sleep(1)
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"select2-des_cat_name-container\"]"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input"))).clear()
            wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input"))).send_keys(row_data['catagoryEdit'])
            wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input"))).send_keys(Keys.ENTER)
            sleep(3)
            wait.until(EC.element_to_be_clickable((By.ID,"select2-des_prod_name-container"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input"))).clear()
            wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input"))).send_keys(row_data['productEdit'])
            wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input"))).send_keys(Keys.ENTER)
            sleep(3)
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"select2-des_des_name-container\"]"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input"))).clear()
            wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input"))).send_keys(row_data['designEdit'])
            wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input"))).send_keys(Keys.ENTER)
            sleep(3)
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"select2-select_sub_design-container\"]"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input"))).clear()
            wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input"))).send_keys(row_data['subdesignEdit'])
            wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/span/span/span[1]/input"))).send_keys(Keys.ENTER)
            try:
                wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id='update_weight_range_settings']"))).click()
                sleep(2)
                # value = int(id)+1
                # print(value)
                # wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"item_list_filter\"]/label/input"))).click()
                # wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"item_list_filter\"]/label/input"))).clear()
                # wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"item_list_filter\"]/label/input"))).send_keys(value)
                sleep(3)
                wait.until(EC.element_to_be_clickable((By.ID,"select2-des_prod_name-container"))).click()
                product = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@class="select2-search__field"]')))
                product.click()
                product.send_keys(row_data['productEdit'],Keys.ENTER)
                sleep(2)
                wait.until(EC.element_to_be_clickable((By.ID,"select2-des_des_name-container"))).click()
                Design = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@class="select2-search__field"]')))
                Design.click()
                Design.send_keys(row_data['designEdit'],Keys.ENTER)
                sleep(2)
                wait.until(EC.element_to_be_clickable((By.ID,"select2-select_sub_design-container"))).click()
                SubDesign = wait.until(EC.element_to_be_clickable((By.XPATH,'//input[@class="select2-search__field"]')))
                SubDesign.click()
                SubDesign.send_keys(row_data['subdesignEdit'],Keys.ENTER)
                wait.until(EC.element_to_be_clickable((By.ID,"search_weight_range"))).click()
                product = wait.until(EC.element_to_be_clickable((By.XPATH,'//table[@id="item_list"]//td[2]'))).text
                print(product)
                design = wait.until(EC.element_to_be_clickable((By.XPATH,'//table[@id="item_list"]//td[3]'))).text
                print(design)
                subdesign = wait.until(EC.element_to_be_clickable((By.XPATH,'//table[@id="item_list"]//td[4]'))).text
                print(subdesign)
                if (row_data["productEdit"].lower() == product.lower() and
                    row_data["designEdit"].lower() == design.lower() and
                    row_data["subdesignEdit"].lower() == subdesign.lower()):
                    Test_Status = "Pass"
                    Actual_Status= "V.A and MC Settings updated successfully"
                else:
                    Test_Status="Fail"
                    Actual_Status= "V.A and MC Settings Not updated successfully"
            except:
                Test_Status="Fail"
                Actual_Status= "V.A and MC Settings Not updated successfully"
        else:
            Test_Status="Pass"
            Actual_Status= "Edit V.A and MC Settings No Needed"              
        return Test_Status,Actual_Status
    
    def Delete(self,row_data):   
        driver = self.driver
        wait = self.wait        
        if row_data["EditmcVa"]=="Yes":
            Delete= row_data['productEdit']
        if row_data["EditmcVa"]=="No" :
            Delete= row_data['productname'] 
        if row_data['DeletemcVa']=="Yes" :
            sleep(3)
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"item_list_filter\"]/label/input"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"item_list_filter\"]/label/input"))).clear()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"item_list_filter\"]/label/input"))).send_keys(Delete)
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"item_list\"]/tbody/tr[1]/td[15]/a[2]/i"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"confirm-delete\"]/div/div/div[3]/a"))).click()
            # wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"item_list\"]/tbody/tr[1]/td[15]/a[2]/i"))).click()
            # wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id=\"confirm-delete\"]/div/div/div[3]/a"))).click()
            try:
                Message=wait.until(EC.element_to_be_clickable((By.XPATH,'//div[@class="alert alert-success alert-dismissable"]'))).text
                Message = Message.replace("×", "").replace("\n", " ").strip()
                print(Message)
                expected_message ="Delete WA Settings! Successfully Delete WA Settings" 
                driver.save_screenshot('delete.png.png')
                if Message == expected_message:
                        Test_Status="Pass"
                        Actual_Status= Message
                else:
                    Test_Status="Fail"
                    Actual_Status= Message
            except:
                driver.save_screenshot('delete.png.png')
                Test_Status="Fail"
                Actual_Status="Sub Design Not Add Successfully"
        else: 
            Test_Status="Pass"
            Actual_Status= "DeleteNotNeeded"
        return Test_Status,Actual_Status 


        #ERROR: Caught exception [ERROR: Unsupported command [endLoadVars |  | ]]
    
    def is_element_present(self, how, what):
        try: self.driver.find_element(by=how, value=what)
        except NoSuchElementException as e: return False
        return True
    
    def is_alert_present(self):
        try: self.driver.switch_to_alert()
        except NoAlertPresentException as e: return False
        return True
    
    def close_alert_and_get_its_text(self):
        try:
            alert = self.driver.switch_to_alert()
            alert_text = alert.text
            if self.accept_next_alert:
                alert.accept()
            else:
                alert.dismiss()
            return alert_text
        finally: self.accept_next_alert = True
    
    def tearDown(self):
        self.driver.quit()
        self.assertEqual([], self.verificationErrors)

if __name__ == "__main__":
    unittest.main()
