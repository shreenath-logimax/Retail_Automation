from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
import unittest
from Utils.Excel import ExcelUtils
from openpyxl import load_workbook
import sys

# Enable UTF-8 safe console output
sys.stdout.reconfigure(encoding='utf-8')

FILE_PATH = ExcelUtils.file_path
class Category(unittest.TestCase):
    def __init__(self,driver):
        self.driver =driver   
        self.wait = WebDriverWait(driver, 30)
    def test_category(self):
        driver = self.driver
        wait = self.wait
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Toggle navigation"))).click()
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Retail Catalog"))).click()
        wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Category"))).click()   
        function_name = "CategoryName"
        valid_rows = ExcelUtils.get_valid_rows(FILE_PATH, function_name)
        workbook = load_workbook(FILE_PATH)
        sheet = workbook[function_name]
        Window=1
        for row_num in range(2, valid_rows):
            data = {
                "TestCaseId": 1,
                "TestStatus": 2,
                "ActualStatus": 3,
                "CategoryName": 4,
                "HSNCode": 5,
                "ShortCode": 6,
                "CategoryType": 7,
                "MultiMetal": 8,
                "Metal": 9,
                "TaxGrp": 10,
                "Purity": 11,
                "CatDesc": 12,
                "CategoryEdit": 13,
                "CategoryEditName": 14,
                "CategoryDelete": 15
            }
            row_data = {key: sheet.cell(row=row_num, column=col).value 
                            for key, col in data.items()}
            print(row_data)
            
            status_list = []
            Actual_list = []
            Create_data=self.create(row_data)
            Test_Status,Actual_Status = Create_data
            status_list.append(Test_Status)
            Actual_list.append(Actual_Status)
            Edit_data=self.Edit(row_data)
            Test_Status,Actual_Status = Edit_data
            status_list.append(Test_Status)
            Actual_list.append(Actual_Status)
            Delete_data=self.Delete(row_data)
            Test_Status,Actual_Status = Delete_data
            status_list.append(Test_Status)
            Actual_list.append(Actual_Status)
            all=True
            for s in status_list:
                    # If any step is not "Pass", mark overall as failed and stop checking
                if s!="Pass":
                    all=False
                    break
            if all:
                Test_Status = "Pass"
            else:
                Test_Status = "Fail"  
                # --- Final Pass/Fail ---
            print(f"Final Status:", Test_Status)
            sheet.cell(row=row_num, column=2).value = Test_Status
            sheet.cell(row=row_num, column=3).value = ", ".join(Actual_list)
            workbook.save(FILE_PATH)
            Status = ExcelUtils.get_Status(FILE_PATH,function_name)  
            print(Status)
            Update_master = ExcelUtils.update_master_status(FILE_PATH,Status,function_name)  
    def create(self,row_data):
        driver = self.driver
        wait = self.wait
        self.driver.refresh()
        wait.until(EC.element_to_be_clickable((By.ID,"add_category"))).click()
        wait.until(EC.element_to_be_clickable((By.ID,"category_name"))).click()
        wait.until(EC.element_to_be_clickable((By.ID,"category_name"))).clear()
        wait.until(EC.element_to_be_clickable((By.ID,"category_name"))).send_keys(row_data["CategoryName"])
        wait.until(EC.element_to_be_clickable((By.ID,"hsn_code"))).click()
        wait.until(EC.element_to_be_clickable((By.ID,"hsn_code"))).clear()
        wait.until(EC.element_to_be_clickable((By.ID,"hsn_code"))).send_keys(row_data["HSNCode"])
        wait.until(EC.element_to_be_clickable((By.ID,"cat_code"))).click()
        wait.until(EC.element_to_be_clickable((By.ID,"cat_code"))).clear()
        wait.until(EC.element_to_be_clickable((By.ID,"cat_code"))).send_keys(row_data["ShortCode"])
        if row_data["CategoryType"] =="Ornament" :
            wait.until(EC.element_to_be_clickable((By.ID,"ornament"))).click()
        elif row_data["CategoryType"] =="Bullion" :
            wait.until(EC.element_to_be_clickable((By.ID,"bullion"))).click()
        elif row_data["CategoryType"] =="Stone" :
            wait.until(EC.element_to_be_clickable((By.ID,"stone"))).click()
        elif row_data["CategoryType"] =="Alloy" :
            wait.until(EC.element_to_be_clickable((By.ID,"alloy"))).click()

        print("CategoryType is ${CategoryType")
        if row_data["MultiMetal"]=="Yes" :
            wait.until(EC.element_to_be_clickable((By.ID,"multimetal_yes"))).click()
        else:       
            wait.until(EC.element_to_be_clickable((By.ID,"multimetal_no"))).click()
        sleep(2)    
        wait.until(EC.element_to_be_clickable((By.XPATH,"//span[@id='select2-metal_category-container']/span"))).click()
        # Wait until the tax option appears and click Metal
        metal_category = row_data["Metal"]
        category_Xpath=(f"//li[normalize-space()='{metal_category}']")
        category_option = wait.until(EC.element_to_be_clickable((By.XPATH,category_Xpath)))
        category_option.click()
        sleep(10)
        wait.until(EC.element_to_be_clickable((By.ID,"select2-tgrp_sel-container"))).click()
        # Wait until the tax option appears and click TaxGrp
        TaxGrp = row_data["TaxGrp"]
        TaxGrp_Xpath=(f"//li[normalize-space()='{TaxGrp}']")
        TaxGrp_Xpath_option = wait.until(EC.element_to_be_clickable((By.XPATH,TaxGrp_Xpath)))
        TaxGrp_Xpath_option.click()
        
        # Example: Excel "Purity" column contains values like: "92.0000,99.0000,94.0000"
        Puritys = row_data["Purity"]

        if Puritys:  
            # Split into list if comma separated
            Purity_list = [s.strip() for s in Puritys.split(",")]

            dropdown = wait.until(EC.element_to_be_clickable((By.XPATH, '(//input[@class="select2-search__field"])[1]')))
            dropdown.click()

            for Purity in Purity_list:
                # Locate input box inside dropdown
                input_box = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, '(//input[@class="select2-search__field"])[1]')
                ))
                input_box.clear()
                input_box.send_keys(Purity)
                input_box.send_keys(Keys.ENTER)
            
            print("✅ Sections added:", Purity_list)
        if row_data["CatDesc"]:
            wait.until(EC.element_to_be_clickable((By.ID,"category_desc"))).click()
            wait.until(EC.element_to_be_clickable((By.ID,"category_desc"))).clear()
            wait.until(EC.element_to_be_clickable((By.ID,"category_desc"))).send_keys(row_data["CatDesc"])
            sleep(10)
        wait.until(EC.element_to_be_clickable((By.ID,"add_newcategory"))).click()
        driver.save_screenshot("error_add_category.png")
        try:
            message=wait.until(EC.presence_of_element_located((By.XPATH,"//*[@id=\"chit_alert\"]/div"))).text
            message = message.replace("×", "").replace("\n", " ").strip()
            print(message)
            expected_message = "Add Category : New Category added successfully!.."
            if message == expected_message:
                Test_Status="Pass"
                Actual_Status= message
            else:
                Test_Status="Fail"
                Actual_Status= message
        except:
            Test_Status="Fail"
            Actual_Status="Category Not Add Successfully"
        return Test_Status,Actual_Status
        
        
    def Edit(self,row_data):
        driver = self.driver
        wait = self.wait   
        if row_data["CategoryEdit"]=="Yes" :
            wait.until(EC.element_to_be_clickable((By.XPATH,"(//input[@type='search'])[1]"))).clear()
            wait.until(EC.element_to_be_clickable((By.XPATH,"(//input[@type='search'])[1]"))).send_keys(row_data["CategoryName"])
            wait.until(EC.element_to_be_clickable((By.XPATH,"//a[@id='edit']/i"))).click()
            wait.until(EC.element_to_be_clickable((By.ID,"ed_category_name"))).clear()
            wait.until(EC.element_to_be_clickable((By.ID,"ed_category_name"))).send_keys(row_data["CategoryEditName"])
            wait.until(EC.element_to_be_clickable((By.ID,"ed_category_name"))).click()
            wait.until(EC.element_to_be_clickable((By.ID,"update_category"))).click()
            driver.save_screenshot("error_add_category.png")
            try:
                sleep(3)
                CategoryEdit = wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[1]/div[1]/section[2]/div/div/div/div[2]/div[1]"))).text
                CategoryEdit = CategoryEdit.replace("×", "").replace("\n", " ").strip()    
                print(CategoryEdit)
                expected_message = "Edit Category! Category record modified successfully"
                if CategoryEdit == expected_message:
                    Test_Status="Pass"
                    Actual_Status= CategoryEdit 
                else:
                    Test_Status="Fail"
                    Actual_Status= CategoryEdit
            except:
                Test_Status="Fail"
                Actual_Status= "Category Edit not updated successfully"                   
        else:
            Test_Status="Pass"
            Actual_Status= "editNotNeeded"
        return Test_Status,Actual_Status
    
    def Delete(self,row_data):
        driver = self.driver
        wait = self.wait          
        if row_data["CategoryDelete"]=="Yes":
            if row_data["CategoryEdit"]=="Yes":
                category =row_data["CategoryEditName"]
            else:
                category =row_data["CategoryName"]   
            sleep(3)      
            wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@type='search']"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@type='search']"))).clear()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@type='search']"))).send_keys(category)
            wait.until(EC.element_to_be_clickable((By.XPATH,"//a[@class='btn btn-danger btn-del']/i"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH,"//a[@class='btn btn-danger btn-confirm']"))).click()
            driver.save_screenshot("error_add_category.png")
            try:
                DeleteSuccessMsg = wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[1]/div[1]/section[2]/div/div/div/div[2]/div[1]"))).text                    
                DeleteSuccessMsg = DeleteSuccessMsg.replace("×", "").replace("\n", " ").strip()
                print(DeleteSuccessMsg)
                expected_message = "Delete category! category deleted successfully"
                if DeleteSuccessMsg == expected_message:
                    Test_Status="Pass"
                    Actual_Status= DeleteSuccessMsg
                else:
                    Test_Status="Fail"
                    Actual_Status= DeleteSuccessMsg
            except:
                Test_Status="Fail"
                Actual_Status= "Product Not deleted Successfully"               
        else: 
            Test_Status="Pass"
            Actual_Status= "deleteNotNeeded"
        return Test_Status,Actual_Status
   
    def is_element_present(self, how, what):
        try: self.driver.find_element(by=how, value=what)
        except NoSuchElementException as e: return False
        return True
    
    def is_alert_present(self):
        try: self.driver.switch_to_alert()
        except NoAlertPresentException as e: return False
        return True
    
    def close_alert_and_get_its_text(self):
        try:
            alert = self.driver.switch_to_alert()
            alert_text = alert.text
            if self.accept_next_alert:
                alert.accept()
            else:
                alert.dismiss()
            return alert_text
        finally: self.accept_next_alert = True
    
    def tearDown(self):
        self.driver.quit()
        self.assertEqual([], self.verificationErrors)

if __name__ == "__main__":
    unittest.main()


